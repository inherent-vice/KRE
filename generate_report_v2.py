import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from statistics import median

# === 전월세 데이터 로드 ===
rent_files = {
    "2026-02": r"C:\Users\leeho22\.claude\projects\C--Devs-KRE\e8639403-5a32-4c2d-9ae2-201d6ba40a44\tool-results\mcp-real-estate-get_apartment_rent-1771478364043.txt",
    "2026-01": r"C:\Users\leeho22\.claude\projects\C--Devs-KRE\e8639403-5a32-4c2d-9ae2-201d6ba40a44\tool-results\mcp-real-estate-get_apartment_rent-1771478366634.txt",
    "2025-12": r"C:\Users\leeho22\.claude\projects\C--Devs-KRE\e8639403-5a32-4c2d-9ae2-201d6ba40a44\tool-results\mcp-real-estate-get_apartment_rent-1771478367444.txt",
}

rent_data = {}
for month, fpath in rent_files.items():
    with open(fpath, "r", encoding="utf-8") as f:
        data = json.load(f)
    rent_data[month] = data
    print(f"{month}: total_count={data['total_count']}, loaded_items={len(data['items'])}, summary={data['summary']}")

# === 매매 데이터 (이전 리포트에서 가져옴) ===
# 매매 중위가 (이전 분석 결과)
trade_medians = {
    "2025-12": 170000,
    "2026-01": 169000,
    "2026-02": 162000,
}

# === 스타일 ===
header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
rent_header_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
sub_header_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
data_font = Font(name="맑은 고딕", size=10)
section_font = Font(name="맑은 고딕", bold=True, size=11, color="C0504D")
title_font = Font(name="맑은 고딕", bold=True, size=14, color="2F5496")
body_font = Font(name="맑은 고딕", size=10)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
center_align = Alignment(horizontal="center", vertical="center")
num_fmt_price = '#,##0'

def style_header_row(ws, row, max_col, fill=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = fill or header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_data_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    cell.font = data_font
    cell.alignment = center_align
    cell.border = thin_border
    return cell

# === 엑셀 로드 ===
wb = load_workbook("C:/Devs/KRE/송파구_잠실_아파트_실거래가_분석.xlsx")

# ===== Sheet: 전월세 거래내역 =====
ws_rent = wb.create_sheet("전월세 거래내역", 1)
headers_rent = ["거래월", "아파트명", "동", "면적(m²)", "층", "보증금(만원)", "보증금(억원)", "월세(만원)", "계약유형", "거래일", "건축년도", "연식"]
for c, h in enumerate(headers_rent, 1):
    ws_rent.cell(row=1, column=c, value=h)
style_header_row(ws_rent, 1, len(headers_rent), rent_header_fill)

row = 2
all_rent_items = []
for month in sorted(rent_data.keys()):
    items = rent_data[month]["items"]
    all_rent_items.extend([(month, t) for t in items])
    for t in sorted(items, key=lambda x: x.get("trade_date", "")):
        deposit_eok = t["deposit_10k"] / 10000
        age = 2026 - t["build_year"]
        ws_rent.cell(row=row, column=1, value=month)
        ws_rent.cell(row=row, column=2, value=t["unit_name"])
        ws_rent.cell(row=row, column=3, value=t["dong"])
        ws_rent.cell(row=row, column=4, value=round(t["area_sqm"], 2))
        ws_rent.cell(row=row, column=5, value=t["floor"])
        c6 = ws_rent.cell(row=row, column=6, value=t["deposit_10k"])
        c6.number_format = num_fmt_price
        ws_rent.cell(row=row, column=7, value=round(deposit_eok, 2))
        c8 = ws_rent.cell(row=row, column=8, value=t["monthly_rent_10k"])
        c8.number_format = num_fmt_price
        ws_rent.cell(row=row, column=9, value=t.get("contract_type", ""))
        ws_rent.cell(row=row, column=10, value=t.get("trade_date", ""))
        ws_rent.cell(row=row, column=11, value=t["build_year"])
        ws_rent.cell(row=row, column=12, value=age)

        for col in range(1, len(headers_rent) + 1):
            style_data_cell(ws_rent, row, col)
        row += 1

for col in range(1, len(headers_rent) + 1):
    ws_rent.column_dimensions[get_column_letter(col)].width = max(12, len(str(headers_rent[col-1])) * 2)
ws_rent.column_dimensions["B"].width = 28
ws_rent.auto_filter.ref = f"A1:{get_column_letter(len(headers_rent))}{row-1}"

# ===== Sheet: 전월세 월별 통계 =====
ws_rent_stat = wb.create_sheet("전월세 월별통계", 2)
headers_rs = ["구분", "2025-12", "2026-01", "2026-02", "추이"]
for c, h in enumerate(headers_rs, 1):
    ws_rent_stat.cell(row=1, column=c, value=h)
style_header_row(ws_rent_stat, 1, len(headers_rs), rent_header_fill)

rent_stats = {}
for month in sorted(rent_data.keys()):
    items = rent_data[month]["items"]
    summary = rent_data[month]["summary"]

    # 전세 (월세 0)
    jeonse = [t for t in items if t["monthly_rent_10k"] == 0]
    jeonse_deposits = [t["deposit_10k"] for t in jeonse]
    jeonse_84 = [t for t in jeonse if 80 <= t["area_sqm"] <= 90]
    jeonse_84_deposits = [t["deposit_10k"] for t in jeonse_84]

    # 월세
    wolse = [t for t in items if t["monthly_rent_10k"] > 0]
    wolse_rents = [t["monthly_rent_10k"] for t in wolse]
    wolse_deposits = [t["deposit_10k"] for t in wolse]

    trade_med = trade_medians.get(month, 0)
    jeonse_med = median(jeonse_deposits) if jeonse_deposits else 0
    jeonse_ratio = round(jeonse_med / trade_med * 100, 1) if trade_med and jeonse_med else 0
    jeonse_84_med = median(jeonse_84_deposits) if jeonse_84_deposits else 0
    jeonse_84_ratio = round(jeonse_84_med / trade_med * 100, 1) if trade_med and jeonse_84_med else 0

    rent_stats[month] = {
        "total": len(items),
        "jeonse_count": len(jeonse),
        "wolse_count": len(wolse),
        "jeonse_median": jeonse_med,
        "jeonse_min": min(jeonse_deposits) if jeonse_deposits else 0,
        "jeonse_max": max(jeonse_deposits) if jeonse_deposits else 0,
        "jeonse_84_median": jeonse_84_med,
        "jeonse_ratio": jeonse_ratio,
        "jeonse_84_ratio": jeonse_84_ratio,
        "wolse_deposit_median": median(wolse_deposits) if wolse_deposits else 0,
        "wolse_rent_median": median(wolse_rents) if wolse_rents else 0,
        "wolse_rent_avg": round(sum(wolse_rents) / len(wolse_rents)) if wolse_rents else 0,
    }

stat_rows = [
    ("총 전월세 건수", "total"),
    ("전세 건수", "jeonse_count"),
    ("월세 건수", "wolse_count"),
    ("", None),
    ("--- 전세 ---", None),
    ("전세 중위보증금(만원)", "jeonse_median"),
    ("전세 최저보증금(만원)", "jeonse_min"),
    ("전세 최고보증금(만원)", "jeonse_max"),
    ("84m² 전세 중위(만원)", "jeonse_84_median"),
    ("전세가율(%)", "jeonse_ratio"),
    ("84m² 전세가율(%)", "jeonse_84_ratio"),
    ("", None),
    ("--- 월세 ---", None),
    ("월세 중위보증금(만원)", "wolse_deposit_median"),
    ("월세 중위월세(만원)", "wolse_rent_median"),
    ("월세 평균월세(만원)", "wolse_rent_avg"),
]

months_sorted = ["2025-12", "2026-01", "2026-02"]
for i, (label, key) in enumerate(stat_rows, 2):
    ws_rent_stat.cell(row=i, column=1, value=label)
    if key:
        for j, m in enumerate(months_sorted, 2):
            val = rent_stats.get(m, {}).get(key, 0)
            c = ws_rent_stat.cell(row=i, column=j, value=val)
            if "만원" in label:
                c.number_format = num_fmt_price
            elif "%" in label:
                c.number_format = '0.0'
        # 추이
        v_dec = rent_stats.get("2025-12", {}).get(key, 0)
        v_feb = rent_stats.get("2026-02", {}).get(key, 0)
        if v_dec and v_dec != 0 and isinstance(v_dec, (int, float)) and isinstance(v_feb, (int, float)):
            change = round((v_feb - v_dec) / v_dec * 100, 1)
            ws_rent_stat.cell(row=i, column=5, value=f"{'+' if change > 0 else ''}{change}%")
    for col in range(1, 6):
        style_data_cell(ws_rent_stat, i, col)

for col in range(1, 6):
    ws_rent_stat.column_dimensions[get_column_letter(col)].width = 22

# ===== Sheet: 전월세 동별 분석 =====
ws_rent_dong = wb.create_sheet("전월세 동별분석", 3)
headers_rd = ["동", "전세건수", "전세중위(만원)", "전세중위(억원)", "월세건수", "월세중위보증금(만원)", "월세중위월세(만원)", "84m²전세중위(만원)"]
for c, h in enumerate(headers_rd, 1):
    ws_rent_dong.cell(row=1, column=c, value=h)
style_header_row(ws_rent_dong, 1, len(headers_rd), rent_header_fill)

dong_rent = {}
for month, data in rent_data.items():
    for t in data["items"]:
        dong = t["dong"]
        if dong not in dong_rent:
            dong_rent[dong] = []
        dong_rent[dong].append(t)

row = 2
for dong in sorted(dong_rent.keys()):
    items = dong_rent[dong]
    jeonse = [t for t in items if t["monthly_rent_10k"] == 0]
    wolse = [t for t in items if t["monthly_rent_10k"] > 0]
    jeonse_deps = [t["deposit_10k"] for t in jeonse]
    jeonse_84 = [t["deposit_10k"] for t in jeonse if 80 <= t["area_sqm"] <= 90]
    wolse_deps = [t["deposit_10k"] for t in wolse]
    wolse_rents = [t["monthly_rent_10k"] for t in wolse]

    ws_rent_dong.cell(row=row, column=1, value=dong)
    ws_rent_dong.cell(row=row, column=2, value=len(jeonse))
    jmed = median(jeonse_deps) if jeonse_deps else 0
    c3 = ws_rent_dong.cell(row=row, column=3, value=jmed); c3.number_format = num_fmt_price
    ws_rent_dong.cell(row=row, column=4, value=round(jmed / 10000, 2) if jmed else 0)
    ws_rent_dong.cell(row=row, column=5, value=len(wolse))
    c6 = ws_rent_dong.cell(row=row, column=6, value=median(wolse_deps) if wolse_deps else 0); c6.number_format = num_fmt_price
    c7 = ws_rent_dong.cell(row=row, column=7, value=median(wolse_rents) if wolse_rents else 0); c7.number_format = num_fmt_price
    c8 = ws_rent_dong.cell(row=row, column=8, value=median(jeonse_84) if jeonse_84 else "N/A")
    if jeonse_84:
        c8.number_format = num_fmt_price

    for col in range(1, len(headers_rd) + 1):
        style_data_cell(ws_rent_dong, row, col)
    row += 1

for col in range(1, len(headers_rd) + 1):
    ws_rent_dong.column_dimensions[get_column_letter(col)].width = 18

# ===== Sheet: 전세가율 분석 =====
ws_gap = wb.create_sheet("전세가율 분석", 4)
headers_gap = ["단지명", "동", "대표면적(m²)", "매매중위(만원)", "전세중위(만원)", "전세가율(%)", "갭(만원)", "갭(억원)", "건수(매매)", "건수(전세)", "리스크"]
for c, h in enumerate(headers_gap, 1):
    ws_gap.cell(row=1, column=c, value=h)
style_header_row(ws_gap, 1, len(headers_gap), PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"))

# 매매 데이터 로드 (기존 엑셀에서)
ws_trade = wb["전체 거래내역"]
trade_by_apt = {}
for r in range(2, ws_trade.max_row + 1):
    apt = ws_trade.cell(row=r, column=2).value
    dong = ws_trade.cell(row=r, column=3).value
    area = ws_trade.cell(row=r, column=4).value
    price = ws_trade.cell(row=r, column=6).value
    if apt and price:
        if apt not in trade_by_apt:
            trade_by_apt[apt] = {"dong": dong, "prices": [], "areas": []}
        trade_by_apt[apt]["prices"].append(price)
        trade_by_apt[apt]["areas"].append(area if area else 0)

# 전세 데이터
rent_by_apt = {}
for month, data in rent_data.items():
    for t in data["items"]:
        if t["monthly_rent_10k"] == 0:  # 전세만
            name = t["unit_name"]
            if name not in rent_by_apt:
                rent_by_apt[name] = {"dong": t["dong"], "deposits": [], "areas": []}
            rent_by_apt[name]["deposits"].append(t["deposit_10k"])
            rent_by_apt[name]["areas"].append(t["area_sqm"])

# 매매+전세 모두 있는 단지
common_apts = set(trade_by_apt.keys()) & set(rent_by_apt.keys())
row = 2
gap_results = []
for apt in sorted(common_apts):
    t_data = trade_by_apt[apt]
    r_data = rent_by_apt[apt]
    trade_med = median(t_data["prices"])
    rent_med = median(r_data["deposits"])
    if trade_med > 0:
        ratio = round(rent_med / trade_med * 100, 1)
        gap = trade_med - rent_med
        gap_results.append((apt, t_data["dong"], median(t_data["areas"]), trade_med, rent_med, ratio, gap, len(t_data["prices"]), len(r_data["deposits"])))

# 전세가율 높은 순 정렬 (위험도 높은 순)
gap_results.sort(key=lambda x: -x[5])

high_risk_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
mid_risk_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
low_risk_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")

for apt, dong, area, trade_med, rent_med, ratio, gap, t_cnt, r_cnt in gap_results:
    risk = "높음" if ratio >= 70 else ("보통" if ratio >= 50 else "낮음")
    risk_fill = high_risk_fill if ratio >= 70 else (mid_risk_fill if ratio >= 50 else low_risk_fill)

    ws_gap.cell(row=row, column=1, value=apt)
    ws_gap.cell(row=row, column=2, value=dong)
    ws_gap.cell(row=row, column=3, value=round(area, 1))
    c4 = ws_gap.cell(row=row, column=4, value=trade_med); c4.number_format = num_fmt_price
    c5 = ws_gap.cell(row=row, column=5, value=rent_med); c5.number_format = num_fmt_price
    ws_gap.cell(row=row, column=6, value=ratio)
    c7 = ws_gap.cell(row=row, column=7, value=gap); c7.number_format = num_fmt_price
    ws_gap.cell(row=row, column=8, value=round(gap / 10000, 2))
    ws_gap.cell(row=row, column=9, value=t_cnt)
    ws_gap.cell(row=row, column=10, value=r_cnt)
    c11 = ws_gap.cell(row=row, column=11, value=risk)
    c11.fill = risk_fill

    for col in range(1, len(headers_gap) + 1):
        style_data_cell(ws_gap, row, col)
    ws_gap.cell(row=row, column=11).fill = risk_fill
    row += 1

ws_gap.column_dimensions["A"].width = 30
for col in range(2, len(headers_gap) + 1):
    ws_gap.column_dimensions[get_column_letter(col)].width = 16
ws_gap.auto_filter.ref = f"A1:{get_column_letter(len(headers_gap))}{row-1}"

# ===== 분석 요약 시트 업데이트 =====
ws5 = wb["분석 요약"]

# 기존 내용 뒤에 전월세 분석 추가
last_row = ws5.max_row + 2

lines = [
    ("", ""),
    ("========================================", ""),
    ("전월세 분석 결과 (데이터 신청 완료 후 업데이트)", "section"),
    ("========================================", ""),
    ("", ""),
    ("6. 전월세 시장 개요", "section"),
]

for m in months_sorted:
    s = rent_stats[m]
    lines.append((f"  - {m}: 전세 {s['jeonse_count']}건, 월세 {s['wolse_count']}건 (총 {s['total']}건)", ""))

lines += [("", ""), ("7. 전세 시세 현황", "section")]
for m in months_sorted:
    s = rent_stats[m]
    lines.append((f"  - {m} 전세 중위보증금: {s['jeonse_median']:,.0f}만원 ({s['jeonse_median']/10000:.1f}억원)", ""))

lines += [("", ""), ("8. 전세가율 분석 (갭투자 리스크)", "section")]
for m in months_sorted:
    s = rent_stats[m]
    lines.append((f"  - {m} 전세가율: {s['jeonse_ratio']}% (84m² 기준: {s['jeonse_84_ratio']}%)", ""))

high_risk = [r for r in gap_results if r[5] >= 70]
if high_risk:
    lines.append(("", ""))
    lines.append(("  [전세가율 70% 이상 고위험 단지]", "section"))
    for apt, dong, area, trade_med, rent_med, ratio, gap, t_cnt, r_cnt in high_risk[:10]:
        lines.append((f"    - {apt}({dong}): 전세가율 {ratio}%, 갭 {gap/10000:.1f}억원", ""))

lines += [
    ("", ""),
    ("9. 월세 시세 현황", "section"),
]
for m in months_sorted:
    s = rent_stats[m]
    lines.append((f"  - {m} 월세 중위보증금: {s['wolse_deposit_median']:,.0f}만원, 중위월세: {s['wolse_rent_median']:,.0f}만원", ""))

lines += [
    ("", ""),
    ("10. 투자 시사점", "section"),
    ("  - 전세가율 70% 이상 단지는 역전세 리스크 주의", ""),
    ("  - 전세가율 50% 이하 단지는 갭투자 여력 존재하나 자기자본 부담 큼", ""),
    ("  - 월세 전환 시 수익률은 보증금 대비 월세 비율로 판단", ""),
]

for text, style in lines:
    ws5.merge_cells(f"A{last_row}:F{last_row}")
    c = ws5.cell(row=last_row, column=1, value=text)
    if style == "section":
        c.font = section_font
    else:
        c.font = body_font
    last_row += 1

# === 저장 ===
output_path = "C:/Devs/KRE/송파구_잠실_아파트_매매전월세_심층분석.xlsx"
wb.save(output_path)
print(f"\n엑셀 파일 업데이트 완료: {output_path}")
print(f"총 시트: {wb.sheetnames}")
