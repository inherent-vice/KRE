"""잠실 재건축 투자 다차원 심층분석 엑셀 생성기"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, RadarChart, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from copy import copy

# ── 스타일 정의 ──
DARK_BLUE = "1B2A4A"
GOLD = "C9A84C"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
MID_GRAY = "D9D9D9"
RED = "E74C3C"
GREEN = "27AE60"
ORANGE = "F39C12"
BLUE = "3498DB"

title_font = Font(name="맑은 고딕", size=16, bold=True, color=WHITE)
header_font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
sub_header_font = Font(name="맑은 고딕", size=10, bold=True, color=DARK_BLUE)
data_font = Font(name="맑은 고딕", size=10)
small_font = Font(name="맑은 고딕", size=9, color="666666")
score_font = Font(name="맑은 고딕", size=12, bold=True)
pct_fmt = '0.0%'
num_fmt = '#,##0'
money_fmt = '#,##0'

title_fill = PatternFill("solid", fgColor=DARK_BLUE)
header_fill = PatternFill("solid", fgColor="2C3E6B")
sub_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
gold_fill = PatternFill("solid", fgColor=GOLD)
green_fill = PatternFill("solid", fgColor="D5F5E3")
red_fill = PatternFill("solid", fgColor="FADBD8")
orange_fill = PatternFill("solid", fgColor="FDEBD0")
white_fill = PatternFill("solid", fgColor=WHITE)

thin_border = Border(
    left=Side(style='thin', color=MID_GRAY),
    right=Side(style='thin', color=MID_GRAY),
    top=Side(style='thin', color=MID_GRAY),
    bottom=Side(style='thin', color=MID_GRAY)
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')


def style_range(ws, row, col_start, col_end, font=None, fill=None, alignment=None, border=None, number_format=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border
        if number_format: cell.number_format = number_format


def write_title(ws, row, title, col_end=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 36
    return row + 1


def write_headers(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[row].height = 28
    return row + 1


def write_data_row(ws, row, data, col_start=1, font=None, fill=None, formats=None):
    for i, val in enumerate(data):
        cell = ws.cell(row=row, column=col_start + i, value=val)
        cell.font = font or data_font
        cell.fill = fill or white_fill
        cell.alignment = center_align
        cell.border = thin_border
        if formats and i < len(formats) and formats[i]:
            cell.number_format = formats[i]
    return row + 1


def score_fill(score):
    if score >= 8: return green_fill
    elif score >= 6: return PatternFill("solid", fgColor="EAFAF1")
    elif score >= 4: return orange_fill
    else: return red_fill


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ========== 단지 데이터 ==========
COMPLEXES = ["잠실주공5단지", "장미1·2차", "장미3차", "우성1·2·3차", "우성4차"]

BASIC_INFO = {
    "잠실주공5단지": {
        "준공년도": 1978, "연식": 48, "세대수": 3930, "동수": 30,
        "최고층": 15, "면적(전용m²)": "76~84", "주소": "잠실동 40",
        "재건축후_세대수": 6491, "재건축후_최고층": 70,
        "시공사": "삼성물산+GS건설+현대산업개발", "브랜드": "래미안+자이+아이파크",
        "용적률_기존": 169, "용적률_계획": 390,
    },
    "장미1·2차": {
        "준공년도": 1979, "연식": 47, "세대수": 2262, "동수": 26,
        "최고층": 15, "면적(전용m²)": "71~82", "주소": "잠실동 16-5",
        "재건축후_세대수": 5165, "재건축후_최고층": 50,
        "시공사": "미선정", "브랜드": "미정",
        "용적률_기존": 179, "용적률_계획": 360,
    },
    "장미3차": {
        "준공년도": 1983, "연식": 43, "세대수": 900, "동수": 10,
        "최고층": 15, "면적(전용m²)": "69~84", "주소": "잠실동 221",
        "재건축후_세대수": 5165, "재건축후_최고층": 50,
        "시공사": "미선정 (장미통합)", "브랜드": "미정",
        "용적률_기존": 185, "용적률_계획": 360,
    },
    "우성1·2·3차": {
        "준공년도": 1981, "연식": 45, "세대수": 1842, "동수": 20,
        "최고층": 15, "면적(전용m²)": "59~121", "주소": "잠실동 22",
        "재건축후_세대수": 2680, "재건축후_최고층": 49,
        "시공사": "미선정 (시공사 총회 예정)", "브랜드": "미정",
        "용적률_기존": 174, "용적률_계획": 350,
    },
    "우성4차": {
        "준공년도": 1982, "연식": 44, "세대수": 555, "동수": 7,
        "최고층": 15, "면적(전용m²)": "59~84", "주소": "잠실동 19",
        "재건축후_세대수": 825, "재건축후_최고층": 32,
        "시공사": "DL이앤씨", "브랜드": "아크로",
        "용적률_기존": 168, "용적률_계획": 299,
    }
}

# ── 재건축 진행 ──
PROGRESS = {
    "잠실주공5단지": {"단계": "사업시행계획인가 신청", "진행률": 65, "착수일": "2019.10", "경과기간": "6년4개월",
                "안전진단": "완료(D등급)", "정비구역지정": "완료", "조합설립인가": "완료(2020.12)",
                "사업시행인가": "신청완료(2025)", "관리처분인가": "미착수(2026말 예상)",
                "착공예정": "2027년", "준공예정": "2030~2031년", "이주예정": "2026말~2027초"},
    "장미1·2차": {"단계": "사업시행인가 진행중", "진행률": 45, "착수일": "2020.06", "경과기간": "5년8개월",
              "안전진단": "완료(D등급)", "정비구역지정": "완료", "조합설립인가": "완료",
              "사업시행인가": "진행중", "관리처분인가": "미착수",
              "착공예정": "2028~2029년", "준공예정": "2031~2032년", "이주예정": "2028년"},
    "장미3차": {"단계": "장미통합 사업시행인가 진행중", "진행률": 40, "착수일": "2020.06", "경과기간": "5년8개월",
             "안전진단": "완료", "정비구역지정": "완료", "조합설립인가": "완료",
             "사업시행인가": "진행중(통합)", "관리처분인가": "미착수",
             "착공예정": "2028~2029년", "준공예정": "2031~2032년", "이주예정": "2028년"},
    "우성1·2·3차": {"단계": "사업시행인가 직전", "진행률": 50, "착수일": "2022.03", "경과기간": "3년11개월",
                "안전진단": "완료(D등급)", "정비구역지정": "완료", "조합설립인가": "완료",
                "사업시행인가": "직전(시공사 선정후)", "관리처분인가": "미착수",
                "착공예정": "2028년", "준공예정": "2031~2032년", "이주예정": "2027~2028년"},
    "우성4차": {"단계": "관리처분인가 준비중", "진행률": 70, "착수일": "2023.11", "경과기간": "2년3개월",
             "안전진단": "완료", "정비구역지정": "완료", "조합설립인가": "완료",
             "사업시행인가": "완료", "관리처분인가": "준비중(2026상반기)",
             "착공예정": "2026.09", "준공예정": "2029년", "이주예정": "2026하반기"},
}

# ── 시세 데이터 (전용 84m² 기준, 억원) ──
PRICE_DATA = {
    "잠실주공5단지": {"현재매매가": 45.0, "전세가": 12.5, "전세가율": 27.8,
                 "1년전매매가": 38.0, "3년전매매가": 25.0, "5년전매매가": 20.0,
                 "신고가": 45.5, "최저가_1년": 36.0, "호가": 47.0,
                 "재건축프리미엄": 15.0, "인근신축대비할인율": -50},
    "장미1·2차": {"현재매매가": 26.5, "전세가": 9.0, "전세가율": 34.0,
              "1년전매매가": 22.0, "3년전매매가": 16.0, "5년전매매가": 13.0,
              "신고가": 26.5, "최저가_1년": 20.0, "호가": 28.0,
              "재건축프리미엄": 6.5, "인근신축대비할인율": -13},
    "장미3차": {"현재매매가": 25.0, "전세가": 8.5, "전세가율": 34.0,
             "1년전매매가": 21.0, "3년전매매가": 15.5, "5년전매매가": 12.0,
             "신고가": 25.0, "최저가_1년": 19.0, "호가": 26.0,
             "재건축프리미엄": 5.0, "인근신축대비할인율": -17},
    "우성1·2·3차": {"현재매매가": 24.0, "전세가": 8.0, "전세가율": 33.3,
                "1년전매매가": 20.0, "3년전매매가": 14.5, "5년전매매가": 11.0,
                "신고가": 25.0, "최저가_1년": 18.0, "호가": 26.0,
                "재건축프리미엄": 4.0, "인근신축대비할인율": -20},
    "우성4차": {"현재매매가": 22.0, "전세가": 7.5, "전세가율": 34.1,
             "1년전매매가": 18.5, "3년전매매가": 13.0, "5년전매매가": 10.0,
             "신고가": 22.5, "최저가_1년": 16.5, "호가": 24.0,
             "재건축프리미엄": 2.0, "인근신축대비할인율": -27},
}

# ── 다차원 팩터 점수 (1-10) ──
FACTOR_SCORES = {
    "잠실주공5단지": {
        "입지 프리미엄": 10, "역세권 접근성": 10, "한강 조망": 9,
        "학군(초중고)": 8, "생활 인프라": 9, "대형 개발호재": 10,
        "재건축 진행도": 7, "시공사 신뢰도": 10, "브랜드 가치": 10,
        "단지 규모": 10, "가격 매력도": 3, "전세가율(갭투자)": 2,
        "기대 수익률": 3, "유동성": 4, "리스크 대비 안정성": 7,
        "커뮤니티/조합 안정성": 8, "정책 수혜 가능성": 8,
        "미래 랜드마크 잠재력": 10, "임대수익 잠재력": 3,
        "총 보유비용 효율": 3,
    },
    "장미1·2차": {
        "입지 프리미엄": 9, "역세권 접근성": 8, "한강 조망": 10,
        "학군(초중고)": 8, "생활 인프라": 8, "대형 개발호재": 8,
        "재건축 진행도": 5, "시공사 신뢰도": 5, "브랜드 가치": 5,
        "단지 규모": 8, "가격 매력도": 6, "전세가율(갭투자)": 5,
        "기대 수익률": 7, "유동성": 6, "리스크 대비 안정성": 5,
        "커뮤니티/조합 안정성": 6, "정책 수혜 가능성": 7,
        "미래 랜드마크 잠재력": 8, "임대수익 잠재력": 5,
        "총 보유비용 효율": 6,
    },
    "장미3차": {
        "입지 프리미엄": 8, "역세권 접근성": 7, "한강 조망": 9,
        "학군(초중고)": 8, "생활 인프라": 8, "대형 개발호재": 8,
        "재건축 진행도": 4, "시공사 신뢰도": 5, "브랜드 가치": 5,
        "단지 규모": 7, "가격 매력도": 7, "전세가율(갭투자)": 5,
        "기대 수익률": 7, "유동성": 5, "리스크 대비 안정성": 5,
        "커뮤니티/조합 안정성": 5, "정책 수혜 가능성": 7,
        "미래 랜드마크 잠재력": 7, "임대수익 잠재력": 5,
        "총 보유비용 효율": 6,
    },
    "우성1·2·3차": {
        "입지 프리미엄": 8, "역세권 접근성": 7, "한강 조망": 6,
        "학군(초중고)": 8, "생활 인프라": 8, "대형 개발호재": 7,
        "재건축 진행도": 5, "시공사 신뢰도": 5, "브랜드 가치": 5,
        "단지 규모": 7, "가격 매력도": 7, "전세가율(갭투자)": 5,
        "기대 수익률": 6, "유동성": 6, "리스크 대비 안정성": 5,
        "커뮤니티/조합 안정성": 6, "정책 수혜 가능성": 7,
        "미래 랜드마크 잠재력": 6, "임대수익 잠재력": 5,
        "총 보유비용 효율": 7,
    },
    "우성4차": {
        "입지 프리미엄": 7, "역세권 접근성": 7, "한강 조망": 5,
        "학군(초중고)": 8, "생활 인프라": 8, "대형 개발호재": 7,
        "재건축 진행도": 8, "시공사 신뢰도": 7, "브랜드 가치": 7,
        "단지 규모": 4, "가격 매력도": 8, "전세가율(갭투자)": 6,
        "기대 수익률": 8, "유동성": 7, "리스크 대비 안정성": 7,
        "커뮤니티/조합 안정성": 7, "정책 수혜 가능성": 7,
        "미래 랜드마크 잠재력": 5, "임대수익 잠재력": 6,
        "총 보유비용 효율": 8,
    },
}

FACTOR_WEIGHTS = {
    "입지 프리미엄": 12, "역세권 접근성": 8, "한강 조망": 6,
    "학군(초중고)": 5, "생활 인프라": 5, "대형 개발호재": 8,
    "재건축 진행도": 10, "시공사 신뢰도": 7, "브랜드 가치": 5,
    "단지 규모": 4, "가격 매력도": 8, "전세가율(갭투자)": 4,
    "기대 수익률": 8, "유동성": 3, "리스크 대비 안정성": 7,
    "커뮤니티/조합 안정성": 4, "정책 수혜 가능성": 4,
    "미래 랜드마크 잠재력": 5, "임대수익 잠재력": 2,
    "총 보유비용 효율": 3,
}

# ── 수익성 시뮬레이션 (억원 단위) ──
ROI_DATA = {
    "잠실주공5단지": {
        "매입가": 45.0, "취득세율": 3.0, "취득세": 1.35, "중개수수료": 0.225,
        "추정분담금_하": 2.0, "추정분담금_상": 5.0, "추정분담금_중": 3.5,
        "이주기간(년)": 4.5, "이주비전세": 10.0, "전세기회비용_연": 0.35,
        "보유세_연": 0.32, "대출이자_연(20억기준)": 0.80,
        "신축취득세율": 2.96, "예상신축가_하": 48.0, "예상신축가_중": 52.0, "예상신축가_상": 58.0,
        "총투자비_중": 53.5, "준공예정": 2031,
    },
    "장미1·2차": {
        "매입가": 26.5, "취득세율": 3.0, "취득세": 0.795, "중개수수료": 0.133,
        "추정분담금_하": 1.5, "추정분담금_상": 4.0, "추정분담금_중": 2.5,
        "이주기간(년)": 4.0, "이주비전세": 9.0, "전세기회비용_연": 0.315,
        "보유세_연": 0.18, "대출이자_연(20억기준)": 0.80,
        "신축취득세율": 2.96, "예상신축가_하": 38.0, "예상신축가_중": 42.0, "예상신축가_상": 48.0,
        "총투자비_중": 32.0, "준공예정": 2032,
    },
    "장미3차": {
        "매입가": 25.0, "취득세율": 3.0, "취득세": 0.75, "중개수수료": 0.125,
        "추정분담금_하": 1.5, "추정분담금_상": 4.0, "추정분담금_중": 2.5,
        "이주기간(년)": 4.0, "이주비전세": 8.5, "전세기회비용_연": 0.298,
        "보유세_연": 0.16, "대출이자_연(20억기준)": 0.80,
        "신축취득세율": 2.96, "예상신축가_하": 36.0, "예상신축가_중": 40.0, "예상신축가_상": 46.0,
        "총투자비_중": 30.5, "준공예정": 2032,
    },
    "우성1·2·3차": {
        "매입가": 24.0, "취득세율": 3.0, "취득세": 0.72, "중개수수료": 0.12,
        "추정분담금_하": 1.5, "추정분담금_상": 4.0, "추정분담금_중": 2.5,
        "이주기간(년)": 4.0, "이주비전세": 8.0, "전세기회비용_연": 0.28,
        "보유세_연": 0.15, "대출이자_연(20억기준)": 0.80,
        "신축취득세율": 2.96, "예상신축가_하": 35.0, "예상신축가_중": 40.0, "예상신축가_상": 45.0,
        "총투자비_중": 30.0, "준공예정": 2032,
    },
    "우성4차": {
        "매입가": 22.0, "취득세율": 3.0, "취득세": 0.66, "중개수수료": 0.11,
        "추정분담금_하": 1.0, "추정분담금_상": 3.0, "추정분담금_중": 2.0,
        "이주기간(년)": 3.0, "이주비전세": 7.5, "전세기회비용_연": 0.263,
        "보유세_연": 0.13, "대출이자_연(20억기준)": 0.80,
        "신축취득세율": 2.96, "예상신축가_하": 30.0, "예상신축가_중": 35.0, "예상신축가_상": 40.0,
        "총투자비_중": 27.0, "준공예정": 2029,
    },
}

# ── 리스크 매트릭스 ──
RISKS = [
    ("추가공사비 발생", {
        "잠실주공5단지": (8, 7), "장미1·2차": (7, 6), "장미3차": (7, 6),
        "우성1·2·3차": (7, 6), "우성4차": (6, 5)}),
    ("사업 지연 (인허가)", {
        "잠실주공5단지": (6, 4), "장미1·2차": (8, 7), "장미3차": (8, 7),
        "우성1·2·3차": (7, 6), "우성4차": (5, 3)}),
    ("조합 내부 갈등", {
        "잠실주공5단지": (7, 4), "장미1·2차": (6, 5), "장미3차": (6, 6),
        "우성1·2·3차": (6, 5), "우성4차": (4, 3)}),
    ("부동산 시장 하락", {
        "잠실주공5단지": (9, 5), "장미1·2차": (7, 5), "장미3차": (7, 5),
        "우성1·2·3차": (7, 5), "우성4차": (6, 5)}),
    ("금리 상승", {
        "잠실주공5단지": (8, 4), "장미1·2차": (6, 4), "장미3차": (6, 4),
        "우성1·2·3차": (6, 4), "우성4차": (5, 4)}),
    ("정책/규제 강화", {
        "잠실주공5단지": (8, 6), "장미1·2차": (7, 6), "장미3차": (7, 6),
        "우성1·2·3차": (7, 6), "우성4차": (6, 5)}),
    ("재초환 부담금", {
        "잠실주공5단지": (9, 8), "장미1·2차": (7, 6), "장미3차": (6, 6),
        "우성1·2·3차": (6, 5), "우성4차": (5, 4)}),
    ("토지거래허가 연장", {
        "잠실주공5단지": (7, 8), "장미1·2차": (5, 6), "장미3차": (5, 6),
        "우성1·2·3차": (5, 6), "우성4차": (4, 5)}),
    ("분담금 급등", {
        "잠실주공5단지": (8, 5), "장미1·2차": (7, 6), "장미3차": (7, 6),
        "우성1·2·3차": (7, 6), "우성4차": (5, 4)}),
    ("시공사 리스크", {
        "잠실주공5단지": (6, 2), "장미1·2차": (7, 7), "장미3차": (7, 7),
        "우성1·2·3차": (7, 7), "우성4차": (5, 3)}),
    ("양도소득세 중과", {
        "잠실주공5단지": (8, 5), "장미1·2차": (6, 5), "장미3차": (6, 5),
        "우성1·2·3차": (6, 5), "우성4차": (5, 5)}),
    ("보유세 부담 증가", {
        "잠실주공5단지": (9, 7), "장미1·2차": (6, 5), "장미3차": (6, 5),
        "우성1·2·3차": (6, 5), "우성4차": (5, 4)}),
]

# ── 벤치마크 (과거 재건축 사례) ──
BENCHMARKS = [
    {"사업명": "둔촌주공→올림픽파크포레온", "위치": "강동구 둔촌동",
     "기존세대": 5930, "신축세대": 12032, "소요기간(년)": 16,
     "분양가(84m²)": 13.0, "현시세(84m²)": 24.0, "시세차익": 11.0,
     "조합원수익률": "200~300%", "추가공사비이슈": "약 1조원"},
    {"사업명": "반포주공1단지→래미안원베일리", "위치": "서초구 반포동",
     "기존세대": 2444, "신축세대": 2990, "소요기간(년)": 15,
     "분양가(84m²)": 20.0, "현시세(84m²)": 42.0, "시세차익": 22.0,
     "조합원수익률": "150~250%", "추가공사비이슈": "일부"},
    {"사업명": "개포주공1단지→디에이치포레센트", "위치": "강남구 개포동",
     "기존세대": 1980, "신축세대": 6702, "소요기간(년)": 14,
     "분양가(84m²)": 15.0, "현시세(84m²)": 38.0, "시세차익": 23.0,
     "조합원수익률": "200~300%", "추가공사비이슈": "해결"},
    {"사업명": "개포주공4단지→개포래미안포레스트", "위치": "강남구 개포동",
     "기존세대": 1100, "신축세대": 3375, "소요기간(년)": 12,
     "분양가(84m²)": 14.0, "현시세(84m²)": 35.0, "시세차익": 21.0,
     "조합원수익률": "180~280%", "추가공사비이슈": "없음"},
    {"사업명": "잠실주공1~4단지→잠실르엘", "위치": "송파구 잠실동",
     "기존세대": 5678, "신축세대": 6300, "소요기간(년)": 13,
     "분양가(84m²)": 13.5, "현시세(84m²)": 30.0, "시세차익": 16.5,
     "조합원수익률": "150~250%", "추가공사비이슈": "일부"},
]

# ── 정책/규제 ──
POLICIES = [
    ("토지거래허가제", "잠실동 전체 지정 (2026.06.22까지)", "매수시 실거주 의무, 전세 놓기 불가",
     "매우 높음", "잠실주공5단지 직접 영향"),
    ("재건축초과이익환수제", "2025년 부활, 2026년 본격 적용", "초과이익 최대 50% 환수",
     "매우 높음", "주공5단지 예상 부담금 수억원"),
    ("분양가상한제", "공공택지 및 과열지역 적용", "일반분양 수익 제한→분담금 영향",
     "높음", "조합 수익구조에 직접 영향"),
    ("안전진단 강화", "2024년 구조안전성 비중 50%→30%", "사업 착수 용이해짐",
     "긍정적", "장미/우성 단지에 유리"),
    ("1기 신도시 특별법", "정비사업 촉진 간접효과", "수도권 공급 확대→잠실 희소성 유지",
     "중립", "공급 분산 효과"),
    ("종부세 공정시장가액비율", "2026년 100% 적용", "보유세 대폭 증가",
     "높음", "고가 주택 보유비용 급증"),
    ("다주택자 양도세 중과", "조정대상지역 +20~30%p", "단기 매도시 세금 부담 극대화",
     "높음", "최소 2년 보유/2년 거주 필요"),
    ("DSR 규제", "총부채원리금상환비율 40%", "대출 한도 제한",
     "중간", "고가 주택 대출 어려움"),
    ("GTX-A 개통", "2025년 말 개통 예정", "수서역 접근성 향상→잠실 간접 수혜",
     "긍정적", "교통 프리미엄 강화"),
    ("MICE 복합단지", "잠실종합운동장 일대 개발", "마이스 산업+상업시설",
     "긍정적", "잠실 일대 미래가치 상승"),
]

# ── 인프라/개발호재 ──
INFRA = {
    "잠실주공5단지": {
        "잠실역(2·8호선)": "도보5분", "종합운동장역": "도보10분",
        "GTX-A(수서)": "환승15분", "위례신사선(예정)": "도보10분",
        "올림픽공원": "도보15분", "석촌호수": "도보10분",
        "롯데월드타워": "도보5분", "MICE단지(예정)": "도보10분",
        "GBC(예정)": "대중교통20분", "탄천": "도보15분",
        "잠실초": "도보5분", "잠신중": "도보7분", "잠실고": "도보10분",
        "대형마트": "도보5분", "대형병원(서울아산)": "대중교통15분",
    },
    "장미1·2차": {
        "잠실나루역(2호선)": "도보7분", "잠실역": "도보15분",
        "GTX-A(수서)": "환승20분", "위례신사선(예정)": "도보15분",
        "한강공원": "도보5분", "석촌호수": "도보15분",
        "롯데월드타워": "도보10분", "MICE단지(예정)": "도보15분",
        "GBC(예정)": "대중교통25분", "탄천": "도보10분",
        "신천초": "도보5분", "잠실중": "도보10분", "잠실고": "도보12분",
        "대형마트": "도보10분", "대형병원(서울아산)": "대중교통20분",
    },
    "우성4차": {
        "잠실역(2·8호선)": "도보10분", "종합운동장역": "도보12분",
        "GTX-A(수서)": "환승18분", "위례신사선(예정)": "도보12분",
        "올림픽공원": "도보20분", "석촌호수": "도보8분",
        "롯데월드타워": "도보8분", "MICE단지(예정)": "도보12분",
        "GBC(예정)": "대중교통22분", "탄천": "도보12분",
        "잠전초": "도보5분", "잠실중": "도보8분", "잠실고": "도보10분",
        "대형마트": "도보7분", "대형병원(서울아산)": "대중교통18분",
    },
}

# ========== 시트 생성 함수 ==========

def create_sheet_scorecard(wb):
    """Sheet 1: 종합 투자 스코어카드"""
    ws = wb.active
    ws.title = "종합 스코어카드"
    set_col_widths(ws, [22, 14, 14, 14, 14, 14, 10, 30])

    r = write_title(ws, 1, "잠실 재건축 다차원 투자 스코어카드", 8)
    r += 1  # blank row

    # 팩터별 점수표
    factors = list(FACTOR_WEIGHTS.keys())
    headers = ["평가 팩터", "가중치(%)"] + COMPLEXES + ["비고"]
    r = write_headers(ws, r, headers)

    for f in factors:
        w = FACTOR_WEIGHTS[f]
        row_data = [f, w]
        for cpx in COMPLEXES:
            row_data.append(FACTOR_SCORES[cpx][f])
        row_data.append("")
        fill_row = sub_fill if factors.index(f) % 2 == 0 else white_fill
        r = write_data_row(ws, r, row_data, fill=fill_row)
        # 점수별 조건부 색상
        for ci, cpx in enumerate(COMPLEXES):
            cell = ws.cell(row=r-1, column=3+ci)
            cell.fill = score_fill(FACTOR_SCORES[cpx][f])
            cell.font = Font(name="맑은 고딕", size=10, bold=True)

    # 가중평균 점수 행
    r += 1
    ws.cell(row=r, column=1, value="가중평균 종합점수").font = Font(name="맑은 고딕", size=12, bold=True)
    ws.cell(row=r, column=1).fill = gold_fill
    ws.cell(row=r, column=2, value="100%").font = score_font
    ws.cell(row=r, column=2).fill = gold_fill
    ws.cell(row=r, column=2).alignment = center_align

    weighted_scores = {}
    for cpx in COMPLEXES:
        total_w = sum(FACTOR_WEIGHTS.values())
        ws_score = sum(FACTOR_SCORES[cpx][f] * FACTOR_WEIGHTS[f] for f in factors) / total_w
        weighted_scores[cpx] = ws_score

    ranked = sorted(weighted_scores.items(), key=lambda x: x[1], reverse=True)
    for ci, cpx in enumerate(COMPLEXES):
        cell = ws.cell(row=r, column=3+ci, value=round(weighted_scores[cpx], 2))
        cell.font = Font(name="맑은 고딕", size=14, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = center_align
        cell.number_format = '0.00'
    ws.cell(row=r, column=8).fill = gold_fill
    ws.row_dimensions[r].height = 32

    # 순위
    r += 1
    ws.cell(row=r, column=1, value="투자 추천 순위").font = Font(name="맑은 고딕", size=11, bold=True)
    ws.cell(row=r, column=1).fill = gold_fill
    ws.cell(row=r, column=2).fill = gold_fill
    for ci, cpx in enumerate(COMPLEXES):
        rank = [i+1 for i, (c, _) in enumerate(ranked) if c == cpx][0]
        medal = {1: "1위", 2: "2위", 3: "3위", 4: "4위", 5: "5위"}
        cell = ws.cell(row=r, column=3+ci, value=medal[rank])
        cell.font = Font(name="맑은 고딕", size=12, bold=True)
        cell.alignment = center_align
        if rank == 1: cell.fill = PatternFill("solid", fgColor="FFD700")
        elif rank == 2: cell.fill = PatternFill("solid", fgColor="C0C0C0")
        elif rank == 3: cell.fill = PatternFill("solid", fgColor="CD7F32")
        else: cell.fill = sub_fill
    ws.cell(row=r, column=8).fill = gold_fill

    # 카테고리별 소계
    r += 2
    categories = {
        "입지/환경": ["입지 프리미엄", "역세권 접근성", "한강 조망", "학군(초중고)", "생활 인프라"],
        "재건축 사업성": ["재건축 진행도", "시공사 신뢰도", "브랜드 가치", "단지 규모", "대형 개발호재"],
        "투자 수익성": ["가격 매력도", "전세가율(갭투자)", "기대 수익률", "유동성", "임대수익 잠재력"],
        "안정성/리스크": ["리스크 대비 안정성", "커뮤니티/조합 안정성", "정책 수혜 가능성", "총 보유비용 효율"],
        "미래 가치": ["미래 랜드마크 잠재력"],
    }
    headers2 = ["카테고리", "가중치합"] + COMPLEXES + ["설명"]
    r = write_headers(ws, r, headers2)
    for cat, cat_factors in categories.items():
        cat_w = sum(FACTOR_WEIGHTS[f] for f in cat_factors)
        row_data = [cat, cat_w]
        for cpx in COMPLEXES:
            cat_score = sum(FACTOR_SCORES[cpx][f] * FACTOR_WEIGHTS[f] for f in cat_factors) / cat_w
            row_data.append(round(cat_score, 1))
        row_data.append(", ".join(cat_factors))
        fill_row = sub_fill if list(categories.keys()).index(cat) % 2 == 0 else white_fill
        r = write_data_row(ws, r, row_data, fill=fill_row)

    # Radar Chart
    r += 2
    chart_start = r
    cat_names = list(categories.keys())
    ws.cell(row=r, column=1, value="카테고리").font = sub_header_font
    for ci, cpx in enumerate(COMPLEXES):
        ws.cell(row=r, column=2+ci, value=cpx).font = sub_header_font
    r += 1
    for cat in cat_names:
        cat_factors = categories[cat]
        cat_w = sum(FACTOR_WEIGHTS[f] for f in cat_factors)
        ws.cell(row=r, column=1, value=cat)
        for ci, cpx in enumerate(COMPLEXES):
            cat_score = sum(FACTOR_SCORES[cpx][f] * FACTOR_WEIGHTS[f] for f in cat_factors) / cat_w
            ws.cell(row=r, column=2+ci, value=round(cat_score, 1))
        r += 1

    radar = RadarChart()
    radar.type = "filled"
    radar.title = "카테고리별 투자 역량 비교"
    radar.style = 26
    radar.width = 20
    radar.height = 14
    labels = Reference(ws, min_col=1, min_row=chart_start+1, max_row=chart_start+len(cat_names))
    for ci in range(len(COMPLEXES)):
        values = Reference(ws, min_col=2+ci, min_row=chart_start, max_row=chart_start+len(cat_names))
        radar.add_data(values, titles_from_data=True)
    radar.set_categories(labels)
    ws.add_chart(radar, f"A{r+1}")

    return ws


def create_sheet_basic_info(wb):
    """Sheet 2: 단지별 기본정보"""
    ws = wb.create_sheet("단지별 기본정보")
    set_col_widths(ws, [20, 14, 14, 14, 14, 14])
    r = write_title(ws, 1, "잠실 재건축 단지 기본정보 비교", 6)
    r += 1

    fields = [
        ("준공년도", lambda d: d["준공년도"]),
        ("연식 (년)", lambda d: d["연식"]),
        ("현재 세대수", lambda d: d["세대수"]),
        ("동 수", lambda d: d["동수"]),
        ("최고층", lambda d: d["최고층"]),
        ("전용면적(m²)", lambda d: d["면적(전용m²)"]),
        ("주소", lambda d: d["주소"]),
        ("기존 용적률(%)", lambda d: d["용적률_기존"]),
        ("─ 재건축 후 ─", lambda d: "──────"),
        ("재건축 세대수", lambda d: d["재건축후_세대수"]),
        ("재건축 최고층", lambda d: d["재건축후_최고층"]),
        ("계획 용적률(%)", lambda d: d["용적률_계획"]),
        ("용적률 증가(%p)", lambda d: d["용적률_계획"] - d["용적률_기존"]),
        ("세대수 증가율", lambda d: f"{(d['재건축후_세대수']/d['세대수']-1)*100:.0f}%"),
        ("시공사", lambda d: d["시공사"]),
        ("브랜드", lambda d: d["브랜드"]),
    ]

    headers = ["항목"] + COMPLEXES
    r = write_headers(ws, r, headers)

    for fname, fn in fields:
        row_data = [fname]
        for cpx in COMPLEXES:
            row_data.append(fn(BASIC_INFO[cpx]))
        fill_row = sub_fill if fields.index((fname, fn)) % 2 == 0 else white_fill
        if fname.startswith("─"):
            fill_row = gold_fill
        r = write_data_row(ws, r, row_data, fill=fill_row)

    return ws


def create_sheet_progress(wb):
    """Sheet 3: 재건축 진행현황"""
    ws = wb.create_sheet("재건축 진행현황")
    set_col_widths(ws, [22, 14, 14, 14, 14, 14])
    r = write_title(ws, 1, "재건축 단계별 진행현황 비교", 6)
    r += 1

    prog_fields = [
        "단계", "진행률", "착수일", "경과기간",
        "안전진단", "정비구역지정", "조합설립인가",
        "사업시행인가", "관리처분인가",
        "착공예정", "준공예정", "이주예정"
    ]
    headers = ["항목"] + COMPLEXES
    r = write_headers(ws, r, headers)

    for field in prog_fields:
        row_data = [field]
        for cpx in COMPLEXES:
            val = PROGRESS[cpx][field]
            if field == "진행률":
                val = f"{val}%"
            row_data.append(val)
        fill_row = sub_fill if prog_fields.index(field) % 2 == 0 else white_fill
        r = write_data_row(ws, r, row_data, fill=fill_row)
        # 진행률 색상
        if field == "진행률":
            for ci, cpx in enumerate(COMPLEXES):
                pv = PROGRESS[cpx]["진행률"]
                cell = ws.cell(row=r-1, column=2+ci)
                if pv >= 70: cell.fill = green_fill
                elif pv >= 50: cell.fill = PatternFill("solid", fgColor="EAFAF1")
                elif pv >= 40: cell.fill = orange_fill
                else: cell.fill = red_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True)

    # 진행률 차트
    r += 2
    chart_r = r
    ws.cell(row=r, column=1, value="단지명").font = sub_header_font
    ws.cell(row=r, column=2, value="진행률(%)").font = sub_header_font
    r += 1
    for cpx in COMPLEXES:
        ws.cell(row=r, column=1, value=cpx)
        ws.cell(row=r, column=2, value=PROGRESS[cpx]["진행률"])
        r += 1

    chart = BarChart()
    chart.type = "col"
    chart.title = "단지별 재건축 진행률"
    chart.y_axis.title = "진행률(%)"
    chart.style = 10
    chart.width = 18
    chart.height = 12
    cats = Reference(ws, min_col=1, min_row=chart_r+1, max_row=chart_r+len(COMPLEXES))
    vals = Reference(ws, min_col=2, min_row=chart_r, max_row=chart_r+len(COMPLEXES))
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, f"D{chart_r}")

    # 재건축 평균 소요기간 참고 테이블
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1, value="[참고] 서울 재건축 평균 소요기간").font = Font(name="맑은 고딕", size=11, bold=True)
    ws.cell(row=r, column=1).fill = sub_fill
    r += 1
    avg_stages = [
        ("구역지정 → 사업시행인가", "2.8년"),
        ("사업시행인가 → 관리처분인가", "2.3년"),
        ("관리처분인가 → 착공", "1.9년"),
        ("착공 → 준공", "3.6년"),
        ("전체 (구역지정~준공)", "평균 10.6년"),
        ("갈등/분쟁 시", "15년 이상"),
    ]
    for stage, dur in avg_stages:
        ws.cell(row=r, column=1, value=stage).font = data_font
        ws.cell(row=r, column=2, value=dur).font = data_font
        ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2).border = thin_border
        r += 1

    return ws


def create_sheet_price(wb):
    """Sheet 4: 시세 및 가격분석"""
    ws = wb.create_sheet("시세 및 가격분석")
    set_col_widths(ws, [24, 14, 14, 14, 14, 14])
    r = write_title(ws, 1, "시세 및 가격 다차원 분석 (전용 84m² 기준, 억원)", 6)
    r += 1

    price_fields = [
        ("현재 매매가", "현재매매가", None),
        ("현재 전세가", "전세가", None),
        ("전세가율(%)", "전세가율", None),
        ("현재 호가", "호가", None),
        ("신고가", "신고가", None),
        ("최저가(1년내)", "최저가_1년", None),
        ("1년전 매매가", "1년전매매가", None),
        ("3년전 매매가", "3년전매매가", None),
        ("5년전 매매가", "5년전매매가", None),
        ("─ 분석 지표 ─", None, None),
        ("1년 상승률(%)", None, lambda d: round((d["현재매매가"]/d["1년전매매가"]-1)*100, 1)),
        ("3년 상승률(%)", None, lambda d: round((d["현재매매가"]/d["3년전매매가"]-1)*100, 1)),
        ("5년 상승률(%)", None, lambda d: round((d["현재매매가"]/d["5년전매매가"]-1)*100, 1)),
        ("연평균 상승률(5Y)(%)", None, lambda d: round(((d["현재매매가"]/d["5년전매매가"])**(1/5)-1)*100, 1)),
        ("재건축 프리미엄(억)", "재건축프리미엄", None),
        ("인근 신축 대비(%)", "인근신축대비할인율", None),
        ("갭(매매-전세)(억)", None, lambda d: round(d["현재매매가"]-d["전세가"], 1)),
        ("갭투자 필요자금(억)", None, lambda d: round(d["현재매매가"]-d["전세가"], 1)),
        ("호가-실거래 괴리(%)", None, lambda d: round((d["호가"]/d["현재매매가"]-1)*100, 1)),
    ]

    headers = ["항목"] + COMPLEXES
    r = write_headers(ws, r, headers)

    for fname, key, fn in price_fields:
        row_data = [fname]
        for cpx in COMPLEXES:
            if fname.startswith("─"):
                row_data.append("──────")
            elif fn:
                row_data.append(fn(PRICE_DATA[cpx]))
            else:
                row_data.append(PRICE_DATA[cpx][key])
        fill_row = sub_fill if price_fields.index((fname, key, fn)) % 2 == 0 else white_fill
        if fname.startswith("─"):
            fill_row = gold_fill
        r = write_data_row(ws, r, row_data, fill=fill_row)

    # 가격 추이 차트
    r += 2
    chart_r = r
    years = ["5년전", "3년전", "1년전", "현재"]
    ws.cell(row=r, column=1, value="시점").font = sub_header_font
    for ci, cpx in enumerate(COMPLEXES):
        ws.cell(row=r, column=2+ci, value=cpx).font = sub_header_font
    r += 1
    for yi, yr in enumerate(years):
        ws.cell(row=r, column=1, value=yr)
        for ci, cpx in enumerate(COMPLEXES):
            d = PRICE_DATA[cpx]
            val = {"5년전": d["5년전매매가"], "3년전": d["3년전매매가"],
                   "1년전": d["1년전매매가"], "현재": d["현재매매가"]}[yr]
            ws.cell(row=r, column=2+ci, value=val)
        r += 1

    line = LineChart()
    line.title = "단지별 매매가 추이 (전용 84m², 억원)"
    line.style = 10
    line.width = 20
    line.height = 12
    line.y_axis.title = "매매가(억원)"
    cats = Reference(ws, min_col=1, min_row=chart_r+1, max_row=chart_r+len(years))
    for ci in range(len(COMPLEXES)):
        vals = Reference(ws, min_col=2+ci, min_row=chart_r, max_row=chart_r+len(years))
        line.add_data(vals, titles_from_data=True)
    line.set_categories(cats)
    ws.add_chart(line, f"A{r+1}")

    return ws


def create_sheet_roi(wb):
    """Sheet 5: 수익성 시뮬레이션"""
    ws = wb.create_sheet("수익성 시뮬레이션")
    set_col_widths(ws, [26, 14, 14, 14, 14, 14])
    r = write_title(ws, 1, "단지별 투자 수익성 시뮬레이션 (전용 84m², 억원)", 6)
    r += 1

    roi_fields = [
        ("── 매입비용 ──", None),
        ("매입가", lambda d: d["매입가"]),
        ("취득세 ({:.1f}%)".format(3.0), lambda d: d["취득세"]),
        ("중개수수료 (0.5%)", lambda d: d["중개수수료"]),
        ("매입 총비용", lambda d: round(d["매입가"]+d["취득세"]+d["중개수수료"], 2)),
        ("── 보유비용 ──", None),
        ("추정 분담금 (하한)", lambda d: d["추정분담금_하"]),
        ("추정 분담금 (중간)", lambda d: d["추정분담금_중"]),
        ("추정 분담금 (상한)", lambda d: d["추정분담금_상"]),
        ("이주기간 (년)", lambda d: d["이주기간(년)"]),
        ("이주 전세비용(억)", lambda d: d["이주비전세"]),
        ("전세 기회비용/년(억)", lambda d: d["전세기회비용_연"]),
        ("보유세/년(억)", lambda d: d["보유세_연"]),
        ("대출이자/년(20억기준)", lambda d: d["대출이자_연(20억기준)"]),
        ("보유기간 총비용(억)", lambda d: round(
            d["추정분담금_중"] +
            d["전세기회비용_연"]*d["이주기간(년)"] +
            d["보유세_연"]*d["이주기간(년)"] +
            d["대출이자_연(20억기준)"]*d["이주기간(년)"], 1)),
        ("── 예상 수익 ──", None),
        ("신축 취득세(2.96%)(억)", lambda d: round(d["예상신축가_중"]*d["신축취득세율"]/100, 2)),
        ("예상 신축가 (비관)", lambda d: d["예상신축가_하"]),
        ("예상 신축가 (중립)", lambda d: d["예상신축가_중"]),
        ("예상 신축가 (낙관)", lambda d: d["예상신축가_상"]),
        ("준공 예정연도", lambda d: d["준공예정"]),
        ("── 수익률 계산 ──", None),
        ("총투자비(중립)", lambda d: round(
            d["매입가"]+d["취득세"]+d["중개수수료"]+d["추정분담금_중"]+
            (d["전세기회비용_연"]+d["보유세_연"]+d["대출이자_연(20억기준)"])*d["이주기간(년)"]+
            d["예상신축가_중"]*d["신축취득세율"]/100, 1)),
        ("예상차익(비관)(억)", lambda d: round(
            d["예상신축가_하"] - (d["매입가"]+d["취득세"]+d["중개수수료"]+d["추정분담금_중"]+
            (d["전세기회비용_연"]+d["보유세_연"]+d["대출이자_연(20억기준)"])*d["이주기간(년)"]+
            d["예상신축가_하"]*d["신축취득세율"]/100), 1)),
        ("예상차익(중립)(억)", lambda d: round(
            d["예상신축가_중"] - (d["매입가"]+d["취득세"]+d["중개수수료"]+d["추정분담금_중"]+
            (d["전세기회비용_연"]+d["보유세_연"]+d["대출이자_연(20억기준)"])*d["이주기간(년)"]+
            d["예상신축가_중"]*d["신축취득세율"]/100), 1)),
        ("예상차익(낙관)(억)", lambda d: round(
            d["예상신축가_상"] - (d["매입가"]+d["취득세"]+d["중개수수료"]+d["추정분담금_중"]+
            (d["전세기회비용_연"]+d["보유세_연"]+d["대출이자_연(20억기준)"])*d["이주기간(년)"]+
            d["예상신축가_상"]*d["신축취득세율"]/100), 1)),
        ("ROI(중립)(%)", lambda d: round(
            (d["예상신축가_중"] - (d["매입가"]+d["취득세"]+d["중개수수료"]+d["추정분담금_중"]+
            (d["전세기회비용_연"]+d["보유세_연"]+d["대출이자_연(20억기준)"])*d["이주기간(년)"]+
            d["예상신축가_중"]*d["신축취득세율"]/100)) /
            (d["매입가"]+d["취득세"]+d["중개수수료"]) * 100, 1)),
        ("연환산 ROI(%)", lambda d: round(
            ((d["예상신축가_중"] / (d["매입가"]+d["취득세"]+d["중개수수료"]+d["추정분담금_중"]+
            (d["전세기회비용_연"]+d["보유세_연"]+d["대출이자_연(20억기준)"])*d["이주기간(년)"]+
            d["예상신축가_중"]*d["신축취득세율"]/100))**(1/max(d["이주기간(년)"]+1,1)) - 1) * 100, 1)),
    ]

    headers = ["항목"] + COMPLEXES
    r = write_headers(ws, r, headers)

    for fname, fn in roi_fields:
        row_data = [fname]
        for cpx in COMPLEXES:
            if fn is None:
                row_data.append("──────")
            else:
                row_data.append(fn(ROI_DATA[cpx]))
        fill_row = sub_fill if roi_fields.index((fname, fn)) % 2 == 0 else white_fill
        if fname.startswith("──"):
            fill_row = gold_fill
        r = write_data_row(ws, r, row_data, fill=fill_row)

    # ROI 비교 차트
    r += 2
    chart_r = r
    ws.cell(row=r, column=1, value="단지명").font = sub_header_font
    ws.cell(row=r, column=2, value="비관ROI").font = sub_header_font
    ws.cell(row=r, column=3, value="중립ROI").font = sub_header_font
    ws.cell(row=r, column=4, value="낙관ROI").font = sub_header_font
    r += 1
    for cpx in COMPLEXES:
        d = ROI_DATA[cpx]
        total_cost_fn = lambda d, nv: d["매입가"]+d["취득세"]+d["중개수수료"]+d["추정분담금_중"]+(d["전세기회비용_연"]+d["보유세_연"]+d["대출이자_연(20억기준)"])*d["이주기간(년)"]+nv*d["신축취득세율"]/100
        base = d["매입가"]+d["취득세"]+d["중개수수료"]
        for scenario, nv_key in [("비관", "예상신축가_하"), ("중립", "예상신축가_중"), ("낙관", "예상신축가_상")]:
            nv = d[nv_key]
            tc = total_cost_fn(d, nv)
            roi = round((nv - tc) / base * 100, 1)
            col = {"비관": 2, "중립": 3, "낙관": 4}[scenario]
            ws.cell(row=r, column=col, value=roi)
        ws.cell(row=r, column=1, value=cpx)
        r += 1

    chart = BarChart()
    chart.type = "col"
    chart.title = "시나리오별 ROI 비교 (%)"
    chart.y_axis.title = "ROI(%)"
    chart.style = 10
    chart.width = 20
    chart.height = 12
    cats = Reference(ws, min_col=1, min_row=chart_r+1, max_row=chart_r+len(COMPLEXES))
    for ci, label in enumerate(["비관ROI", "중립ROI", "낙관ROI"]):
        vals = Reference(ws, min_col=2+ci, min_row=chart_r, max_row=chart_r+len(COMPLEXES))
        chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{r+1}")

    return ws


def create_sheet_cost(wb):
    """Sheet 6: 비용 구조 분석"""
    ws = wb.create_sheet("비용구조 분석")
    set_col_widths(ws, [28, 16, 16, 16])
    r = write_title(ws, 1, "재건축 투자 비용 구조 상세 분석", 4)
    r += 1

    # 취득세 테이블
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1, value="[1] 취득세 세율 구조").font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
    r += 1
    tax_headers = ["구분", "세율", "조건", "비고"]
    r = write_headers(ws, r, tax_headers)
    tax_data = [
        ("1세대 1주택 (6억이하)", "1.0%", "비규제지역", "지방교육세 별도"),
        ("1세대 1주택 (6~9억)", "1.0~3.0%", "비규제지역", "구간별 차등"),
        ("1세대 1주택 (9억초과)", "3.0%", "비규제지역", ""),
        ("2주택자 (조정대상지역)", "8.0%", "규제지역 중과", "지방교육세+농특세 별도"),
        ("3주택 이상", "12.0%", "규제지역 중과", ""),
        ("멸실 후 토지 매수", "4.6%", "토지 취득세율", "주택 아닌 토지로 분류"),
        ("조합원 원시취득 (85m²이하)", "2.8%", "신축 취득시", "+지방교육세 0.16%=2.96%"),
        ("조합원 원시취득 (85m²초과)", "2.8%", "신축 취득시", "+농특세 0.2%=3.16%"),
    ]
    for td in tax_data:
        r = write_data_row(ws, r, td, fill=sub_fill if tax_data.index(td) % 2 == 0 else white_fill)

    # 중개수수료
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1, value="[2] 중개수수료 요율").font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
    r += 1
    fee_headers = ["거래금액", "상한 요율", "예시 (40억 기준)", "실제 협의선"]
    r = write_headers(ws, r, fee_headers)
    fee_data = [
        ("5억 미만", "0.6%", "-", "-"),
        ("5~9억", "0.5%", "-", "-"),
        ("9~12억", "0.5%", "-", "-"),
        ("12~15억", "0.6%", "-", "-"),
        ("15억 이상", "0.7%", "최대 2,800만원", "0.4~0.5% 협의"),
    ]
    for fd in fee_data:
        r = write_data_row(ws, r, fd, fill=sub_fill if fee_data.index(fd) % 2 == 0 else white_fill)

    # 보유세
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1, value="[3] 보유세 (1세대1주택, 공시가 30억 가정)").font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
    r += 1
    tax2_headers = ["항목", "금액(만원/년)", "비고", "변동요인"]
    r = write_headers(ws, r, tax2_headers)
    tax2_data = [
        ("재산세", "800~1,200", "공시가 기준", "공시가 변동"),
        ("종합부동산세", "1,500~2,500", "1세대1주택 공제 12억", "공정시장가액비율"),
        ("합계", "2,300~3,700", "", "2026년 대폭 증가 예상"),
        ("다주택시 종부세", "3,000~6,000+", "다주택 중과세율", "세 부담 상한 폐지"),
    ]
    for td in tax2_data:
        r = write_data_row(ws, r, td, fill=sub_fill if tax2_data.index(td) % 2 == 0 else white_fill)

    # 대출금리
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1, value="[4] 대출 금리 현황 (2026.02)").font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
    r += 1
    rate_headers = ["구분", "금리", "비고", "월이자(20억기준)"]
    r = write_headers(ws, r, rate_headers)
    rate_data = [
        ("한국은행 기준금리", "2.50%", "2026.01 인하", "-"),
        ("시중은행 변동금리", "3.8~4.3%", "주담대 기준", "약 633~717만"),
        ("시중은행 고정금리", "4.0~4.5%", "주담대 기준", "약 667~750만"),
        ("보금자리론", "3.9~4.2%", "일반형", "약 650~700만"),
        ("보금자리론(우대)", "2.9~3.2%", "우대금리 적용", "약 483~533만"),
    ]
    for rd in rate_data:
        r = write_data_row(ws, r, rd, fill=sub_fill if rate_data.index(rd) % 2 == 0 else white_fill)

    return ws


def create_sheet_risk(wb):
    """Sheet 7: 리스크 매트릭스"""
    ws = wb.create_sheet("리스크 매트릭스")
    set_col_widths(ws, [22] + [8, 8]*5 + [12])
    r = write_title(ws, 1, "단지별 리스크 매트릭스 (영향도 x 발생확률)", 12)
    r += 1

    # 헤더
    headers = ["리스크 요인"]
    for cpx in COMPLEXES:
        headers.extend([f"{cpx[:4]}영향", f"{cpx[:4]}확률"])
    headers.append("리스크등급")
    r = write_headers(ws, r, headers)

    for risk_name, risk_data in RISKS:
        row = [risk_name]
        max_score = 0
        for cpx in COMPLEXES:
            impact, prob = risk_data[cpx]
            row.extend([impact, prob])
            max_score = max(max_score, impact * prob)
        # 등급
        if max_score >= 48: grade = "매우높음"
        elif max_score >= 30: grade = "높음"
        elif max_score >= 16: grade = "중간"
        else: grade = "낮음"
        row.append(grade)
        r = write_data_row(ws, r, row)

        # 등급 색상
        cell = ws.cell(row=r-1, column=len(headers))
        if grade == "매우높음": cell.fill = red_fill; cell.font = Font(name="맑은 고딕", size=10, bold=True, color=RED)
        elif grade == "높음": cell.fill = orange_fill; cell.font = Font(name="맑은 고딕", size=10, bold=True, color=ORANGE)
        elif grade == "중간": cell.fill = PatternFill("solid", fgColor="FEF9E7"); cell.font = Font(name="맑은 고딕", size=10, bold=True)
        else: cell.fill = green_fill

    # 단지별 종합 리스크 점수
    r += 1
    ws.cell(row=r, column=1, value="종합 리스크 점수").font = Font(name="맑은 고딕", size=11, bold=True)
    ws.cell(row=r, column=1).fill = gold_fill
    for ci, cpx in enumerate(COMPLEXES):
        total_risk = sum(RISKS[ri][1][cpx][0] * RISKS[ri][1][cpx][1] for ri in range(len(RISKS)))
        avg_risk = round(total_risk / len(RISKS), 1)
        col = 2 + ci*2
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)
        cell = ws.cell(row=r, column=col, value=avg_risk)
        cell.font = Font(name="맑은 고딕", size=12, bold=True)
        cell.alignment = center_align
        if avg_risk >= 40: cell.fill = red_fill
        elif avg_risk >= 30: cell.fill = orange_fill
        else: cell.fill = green_fill

    return ws


def create_sheet_benchmark(wb):
    """Sheet 8: 벤치마크 비교"""
    ws = wb.create_sheet("벤치마크(과거사례)")
    set_col_widths(ws, [30, 14, 12, 12, 12, 14, 14, 14, 16, 16])
    r = write_title(ws, 1, "서울 주요 재건축 완료 사례 벤치마크", 10)
    r += 1

    headers = ["사업명", "위치", "기존세대", "신축세대", "소요기간(년)",
               "분양가(84m²억)", "현시세(84m²억)", "시세차익(억)",
               "조합원수익률", "추가공사비이슈"]
    r = write_headers(ws, r, headers)

    for bm in BENCHMARKS:
        row = [bm["사업명"], bm["위치"], bm["기존세대"], bm["신축세대"],
               bm["소요기간(년)"], bm["분양가(84m²)"], bm["현시세(84m²)"],
               bm["시세차익"], bm["조합원수익률"], bm["추가공사비이슈"]]
        r = write_data_row(ws, r, row, fill=sub_fill if BENCHMARKS.index(bm) % 2 == 0 else white_fill)

    # 시사점 섹션
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.cell(row=r, column=1, value="[벤치마크 시사점]").font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
    r += 1
    insights = [
        "1. 서울 주요 재건축 완료 사례 평균 소요기간: 약 14년 (조합설립~입주)",
        "2. 분양가 대비 현시세 평균 프리미엄: 약 100~200% 상승",
        "3. 추가공사비 이슈는 대규모 단지(5천세대+)에서 빈번 → 잠실주공5단지 주의",
        "4. 조합원 초기 매입가가 낮을수록 수익률 극대화 → 우성4차 유리",
        "5. 한강변/강남권 입지가 완공 후 시세 프리미엄 극대화에 핵심",
        "6. 시공사 브랜드(삼성·현대·GS)가 완공 후 시세에 5~10% 프리미엄 추가",
    ]
    for ins in insights:
        ws.cell(row=r, column=1, value=ins).font = data_font
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        r += 1

    return ws


def create_sheet_scenario(wb):
    """Sheet 9: 시나리오 분석"""
    ws = wb.create_sheet("시나리오 분석")
    set_col_widths(ws, [24, 14, 14, 14, 14, 14])
    r = write_title(ws, 1, "3개 시나리오별 투자 수익 시뮬레이션", 6)
    r += 1

    scenarios = {
        "낙관 (Bull)": {
            "전제": "금리인하, 규제완화, 부동산 상승기",
            "시세상승률": "+20~30%", "분담금": "하한", "이주기간": "계획대로",
            "잠실주공5단지": {"신축가": 58, "분담금": 2.0, "이주기간": 4.0},
            "장미1·2차": {"신축가": 48, "분담금": 1.5, "이주기간": 3.5},
            "장미3차": {"신축가": 46, "분담금": 1.5, "이주기간": 3.5},
            "우성1·2·3차": {"신축가": 45, "분담금": 1.5, "이주기간": 3.5},
            "우성4차": {"신축가": 40, "분담금": 1.0, "이주기간": 2.5},
        },
        "중립 (Base)": {
            "전제": "현 추세 유지, 완만한 상승",
            "시세상승률": "+10~15%", "분담금": "중간", "이주기간": "1년 지연",
            "잠실주공5단지": {"신축가": 52, "분담금": 3.5, "이주기간": 5.0},
            "장미1·2차": {"신축가": 42, "분담금": 2.5, "이주기간": 4.5},
            "장미3차": {"신축가": 40, "분담금": 2.5, "이주기간": 4.5},
            "우성1·2·3차": {"신축가": 40, "분담금": 2.5, "이주기간": 4.5},
            "우성4차": {"신축가": 35, "분담금": 2.0, "이주기간": 3.5},
        },
        "비관 (Bear)": {
            "전제": "금리상승, 규제강화, 경기침체",
            "시세상승률": "0~-5%", "분담금": "상한", "이주기간": "2년+ 지연",
            "잠실주공5단지": {"신축가": 48, "분담금": 5.0, "이주기간": 6.5},
            "장미1·2차": {"신축가": 38, "분담금": 4.0, "이주기간": 6.0},
            "장미3차": {"신축가": 36, "분담금": 4.0, "이주기간": 6.0},
            "우성1·2·3차": {"신축가": 35, "분담금": 4.0, "이주기간": 6.0},
            "우성4차": {"신축가": 30, "분담금": 3.0, "이주기간": 4.5},
        },
    }

    for scenario_name, sc in scenarios.items():
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        cell = ws.cell(row=r, column=1, value=f"시나리오: {scenario_name}")
        cell.font = Font(name="맑은 고딕", size=13, bold=True, color=WHITE)
        if "낙관" in scenario_name: cell.fill = PatternFill("solid", fgColor=GREEN)
        elif "중립" in scenario_name: cell.fill = PatternFill("solid", fgColor=BLUE)
        else: cell.fill = PatternFill("solid", fgColor=RED)
        ws.row_dimensions[r].height = 30
        r += 1

        ws.cell(row=r, column=1, value=f"전제: {sc['전제']}").font = small_font
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1

        headers = ["항목"] + COMPLEXES
        r = write_headers(ws, r, headers)

        items = [
            ("예상 신축가(억)", lambda cpx: sc[cpx]["신축가"]),
            ("분담금(억)", lambda cpx: sc[cpx]["분담금"]),
            ("이주기간(년)", lambda cpx: sc[cpx]["이주기간"]),
            ("총투자비(억)", lambda cpx: round(
                PRICE_DATA[cpx]["현재매매가"] +
                PRICE_DATA[cpx]["현재매매가"]*0.03 + # 취득세
                PRICE_DATA[cpx]["현재매매가"]*0.005 + # 중개
                sc[cpx]["분담금"] +
                (0.3+0.15+0.8)*sc[cpx]["이주기간"] + # 기회비용+보유세+이자
                sc[cpx]["신축가"]*0.0296, 1)),
            ("예상차익(억)", lambda cpx: round(
                sc[cpx]["신축가"] - (
                    PRICE_DATA[cpx]["현재매매가"] +
                    PRICE_DATA[cpx]["현재매매가"]*0.03 +
                    PRICE_DATA[cpx]["현재매매가"]*0.005 +
                    sc[cpx]["분담금"] +
                    (0.3+0.15+0.8)*sc[cpx]["이주기간"] +
                    sc[cpx]["신축가"]*0.0296), 1)),
            ("ROI(%)", lambda cpx: round(
                (sc[cpx]["신축가"] - (
                    PRICE_DATA[cpx]["현재매매가"] +
                    PRICE_DATA[cpx]["현재매매가"]*0.035 +
                    sc[cpx]["분담금"] +
                    1.25*sc[cpx]["이주기간"] +
                    sc[cpx]["신축가"]*0.0296)) /
                (PRICE_DATA[cpx]["현재매매가"]+PRICE_DATA[cpx]["현재매매가"]*0.035) * 100, 1)),
        ]

        for item_name, fn in items:
            row_data = [item_name]
            for cpx in COMPLEXES:
                row_data.append(fn(cpx))
            r = write_data_row(ws, r, row_data, fill=sub_fill if items.index((item_name, fn)) % 2 == 0 else white_fill)
            # ROI 색상
            if "ROI" in item_name or "차익" in item_name:
                for ci in range(len(COMPLEXES)):
                    cell = ws.cell(row=r-1, column=2+ci)
                    try:
                        v = float(cell.value)
                        if v > 0: cell.fill = green_fill; cell.font = Font(name="맑은 고딕", size=10, bold=True, color=GREEN)
                        elif v < 0: cell.fill = red_fill; cell.font = Font(name="맑은 고딕", size=10, bold=True, color=RED)
                    except: pass
        r += 1

    return ws


def create_sheet_cashflow(wb):
    """Sheet 10: 연도별 현금흐름"""
    ws = wb.create_sheet("현금흐름 분석")
    set_col_widths(ws, [16, 14, 14, 14, 14, 14, 14])
    r = write_title(ws, 1, "연도별 현금흐름 시뮬레이션 (잠실주공5단지 vs 우성4차)", 7)
    r += 1

    # 잠실주공5단지 현금흐름
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="[A] 잠실주공5단지 (전용 84m², 매입가 45억)").font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
    r += 1
    cf_headers = ["연도", "구분", "지출(억)", "수입(억)", "누적투자(억)", "시점자산가치(억)", "순자산(억)"]
    r = write_headers(ws, r, cf_headers)
    cf5 = [
        (2026, "매입+취득세+중개", -46.6, 0, -46.6, 45.0, -1.6),
        (2026, "보유세+대출이자", -1.1, 0, -47.7, 45.0, -2.7),
        (2027, "이주+전세기회비용+보유세+이자", -1.4, 0, -49.1, 46.0, -3.1),
        (2028, "분담금(1차)+보유비용", -3.0, 0, -52.1, 47.0, -5.1),
        (2029, "보유비용+분담금(잔금)", -1.9, 0, -54.0, 48.0, -6.0),
        (2030, "보유비용", -1.4, 0, -55.4, 50.0, -5.4),
        (2031, "준공+신축취득세", -1.5, 0, -56.9, 52.0, -4.9),
        (2031, "입주 (자산가치 확정)", 0, 52.0, -56.9, 52.0, -4.9),
    ]
    for row_data in cf5:
        r = write_data_row(ws, r, row_data)
        # 순자산 색상
        cell = ws.cell(row=r-1, column=7)
        if row_data[6] < 0: cell.fill = red_fill; cell.font = Font(name="맑은 고딕", size=10, bold=True, color=RED)
        else: cell.fill = green_fill

    # 우성4차 현금흐름
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="[B] 우성4차 (전용 84m², 매입가 22억)").font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
    r += 1
    r = write_headers(ws, r, cf_headers)
    cf4 = [
        (2026, "매입+취득세+중개", -22.8, 0, -22.8, 22.0, -0.8),
        (2026, "이주+보유비용", -0.9, 0, -23.7, 22.0, -1.7),
        (2027, "분담금(1차)+보유비용", -2.4, 0, -26.1, 24.0, -2.1),
        (2028, "보유비용+분담금(잔금)", -1.2, 0, -27.3, 28.0, 0.7),
        (2029, "준공+신축취득세", -1.0, 0, -28.3, 35.0, 6.7),
        (2029, "입주 (자산가치 확정)", 0, 35.0, -28.3, 35.0, 6.7),
    ]
    for row_data in cf4:
        r = write_data_row(ws, r, row_data)
        cell = ws.cell(row=r-1, column=7)
        if row_data[6] < 0: cell.fill = red_fill; cell.font = Font(name="맑은 고딕", size=10, bold=True, color=RED)
        else: cell.fill = green_fill; cell.font = Font(name="맑은 고딕", size=10, bold=True, color=GREEN)

    # 비교 요약
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="[비교 요약]").font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
    r += 1
    comp_headers = ["항목", "잠실주공5단지", "우성4차", "차이", "유리한 단지", "", ""]
    r = write_headers(ws, r, comp_headers)
    comparisons = [
        ("총투자금(억)", 56.9, 28.3, -28.6, "우성4차"),
        ("예상자산가치(억)", 52.0, 35.0, -17.0, "잠실주공5단지"),
        ("순자산(억)", -4.9, 6.7, 11.6, "우성4차"),
        ("투자기간(년)", 5.5, 3.5, -2.0, "우성4차"),
        ("ROI(%)", -8.6, 23.7, 32.3, "우성4차"),
        ("연환산ROI(%)", -1.6, 6.8, 8.4, "우성4차"),
        ("리스크수준", "높음", "중간", "-", "우성4차"),
        ("미래프리미엄", "매우높음", "중간", "-", "잠실주공5단지"),
    ]
    for comp in comparisons:
        r = write_data_row(ws, r, list(comp) + ["", ""])

    return ws


def create_sheet_policy(wb):
    """Sheet 11: 정책/규제 영향"""
    ws = wb.create_sheet("정책규제 영향")
    set_col_widths(ws, [22, 28, 28, 12, 30])
    r = write_title(ws, 1, "정책/규제 영향 분석", 5)
    r += 1

    headers = ["정책/규제", "현황", "투자 영향", "영향도", "잠실 적용"]
    r = write_headers(ws, r, headers)

    for pol in POLICIES:
        r = write_data_row(ws, r, pol, fill=sub_fill if POLICIES.index(pol) % 2 == 0 else white_fill)
        # 영향도 색상
        cell = ws.cell(row=r-1, column=4)
        if "매우" in str(pol[3]): cell.fill = red_fill
        elif "높음" in str(pol[3]): cell.fill = orange_fill
        elif "긍정" in str(pol[3]): cell.fill = green_fill

    return ws


def create_sheet_infra(wb):
    """Sheet 12: 입지/인프라"""
    ws = wb.create_sheet("입지 인프라 분석")
    set_col_widths(ws, [24, 18, 18, 18])
    r = write_title(ws, 1, "주요 단지별 입지/인프라 접근성 비교", 4)
    r += 1

    infra_cpx = ["잠실주공5단지", "장미1·2차", "우성4차"]
    headers = ["인프라 항목"] + infra_cpx
    r = write_headers(ws, r, headers)

    all_items = set()
    for cpx in infra_cpx:
        all_items.update(INFRA[cpx].keys())
    all_items = sorted(all_items)

    for item in all_items:
        row_data = [item]
        for cpx in infra_cpx:
            row_data.append(INFRA[cpx].get(item, "-"))
        fill_row = sub_fill if all_items.index(item) % 2 == 0 else white_fill
        r = write_data_row(ws, r, row_data, fill=fill_row)

    # 개발호재 점수 비교
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1, value="[미래 개발호재 영향도 점수 (10점 만점)]").font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
    r += 1
    dev_headers = ["개발호재", "잠실주공5단지", "장미1·2차", "우성4차"]
    r = write_headers(ws, r, dev_headers)
    developments = [
        ("GTX-A 수서역 개통", 8, 7, 7),
        ("위례신사선 개통", 9, 7, 8),
        ("MICE 복합단지", 10, 7, 8),
        ("GBC(현대차)", 7, 6, 6),
        ("탄천 수변공간 정비", 7, 8, 7),
        ("잠실종합운동장 리모델링", 9, 7, 8),
        ("롯데월드타워 상권", 9, 8, 8),
        ("영동대로 지하공간", 8, 6, 7),
        ("합계", 67, 56, 59),
    ]
    for dev in developments:
        r = write_data_row(ws, r, dev, fill=sub_fill if developments.index(dev) % 2 == 0 else white_fill)
        if dev[0] == "합계":
            style_range(ws, r-1, 1, 4, font=Font(name="맑은 고딕", size=11, bold=True), fill=gold_fill)

    return ws


def create_sheet_final(wb):
    """Sheet 13: 최종 투자 판단"""
    ws = wb.create_sheet("최종 투자 판단")
    set_col_widths(ws, [24, 14, 14, 14, 14, 14])
    r = write_title(ws, 1, "최종 투자 판단 종합 매트릭스", 6)
    r += 1

    # 투자자 유형별 추천
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value="[투자자 유형별 추천 단지]").font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
    r += 1
    type_headers = ["투자자 유형", "1순위 추천", "2순위 추천", "이유", "주의사항", ""]
    r = write_headers(ws, r, type_headers)
    investor_types = [
        ("공격적 투자자\n(높은 수익 추구)", "우성4차", "장미1·2차",
         "빠른 사업진행, 높은 ROI", "소규모 단지 리스크"),
        ("안정적 투자자\n(자산가치 보전)", "잠실주공5단지", "장미1·2차",
         "입지/브랜드 최상위", "높은 매입가, 낮은 ROI"),
        ("중장기 투자자\n(5년+ 보유)", "장미1·2차", "우성1·2·3차",
         "적정 매입가+한강조망", "사업 지연 가능성"),
        ("갭투자자\n(레버리지 활용)", "우성4차", "우성1·2·3차",
         "낮은 매입가, 빠른 회수", "토지거래허가 주의"),
        ("실거주+투자\n(자가거주 후 매도)", "잠실주공5단지", "장미1·2차",
         "최고의 거주환경 확보", "이주기간 전세비용"),
    ]
    for inv in investor_types:
        r = write_data_row(ws, r, list(inv) + [""])

    # 종합 점수 재표시
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value="[종합 평가 매트릭스]").font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
    r += 1
    eval_headers = ["평가 항목"] + COMPLEXES
    r = write_headers(ws, r, eval_headers)

    eval_items = [
        ("입지/환경 (30%)", [9.4, 8.6, 8.0, 7.4, 7.0]),
        ("재건축 사업성 (25%)", [8.8, 6.2, 5.8, 6.0, 6.6]),
        ("투자 수익성 (25%)", [3.6, 5.8, 5.8, 5.6, 7.0]),
        ("안정성/리스크 (10%)", [6.5, 6.0, 5.5, 5.8, 7.3]),
        ("미래 가치 (10%)", [10.0, 8.0, 7.0, 6.0, 5.0]),
    ]
    for item_name, scores in eval_items:
        row_data = [item_name] + scores
        r = write_data_row(ws, r, row_data, fill=sub_fill if eval_items.index((item_name, scores)) % 2 == 0 else white_fill)
        for ci in range(len(scores)):
            cell = ws.cell(row=r-1, column=2+ci)
            cell.fill = score_fill(scores[ci])
            cell.font = Font(name="맑은 고딕", size=10, bold=True)

    # 가중 종합
    weights = [0.30, 0.25, 0.25, 0.10, 0.10]
    r_total = r
    ws.cell(row=r, column=1, value="가중 종합점수 (100점환산)").font = Font(name="맑은 고딕", size=12, bold=True)
    ws.cell(row=r, column=1).fill = gold_fill
    for ci in range(len(COMPLEXES)):
        weighted = sum(eval_items[ei][1][ci] * weights[ei] for ei in range(len(eval_items)))
        cell = ws.cell(row=r, column=2+ci, value=round(weighted*10, 1))
        cell.font = Font(name="맑은 고딕", size=14, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = center_align
    r += 1

    # 최종 순위
    ws.cell(row=r, column=1, value="최종 투자 추천 순위").font = Font(name="맑은 고딕", size=12, bold=True)
    ws.cell(row=r, column=1).fill = gold_fill
    final_scores = []
    for ci in range(len(COMPLEXES)):
        weighted = sum(eval_items[ei][1][ci] * weights[ei] for ei in range(len(eval_items)))
        final_scores.append((COMPLEXES[ci], weighted))
    ranked = sorted(final_scores, key=lambda x: x[1], reverse=True)
    for ci, cpx in enumerate(COMPLEXES):
        rank = [i+1 for i, (c, _) in enumerate(ranked) if c == cpx][0]
        cell = ws.cell(row=r, column=2+ci, value=f"{rank}위")
        cell.font = Font(name="맑은 고딕", size=14, bold=True)
        cell.alignment = center_align
        if rank == 1: cell.fill = PatternFill("solid", fgColor="FFD700")
        elif rank == 2: cell.fill = PatternFill("solid", fgColor="C0C0C0")
        elif rank == 3: cell.fill = PatternFill("solid", fgColor="CD7F32")
        else: cell.fill = sub_fill
    r += 2

    # 핵심 결론
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value="[핵심 결론 및 전략적 제언]").font = Font(name="맑은 고딕", size=13, bold=True, color=WHITE)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws.row_dimensions[r].height = 30
    r += 1

    conclusions = [
        "1. [최고 ROI] 우성4차: 가장 빠른 사업진행(2029년 준공)+낮은 매입가→단기 수익률 최고",
        "   → 22억 매입 후 3년 내 35억 달성 가능, 연환산 ROI 약 6.8%",
        "",
        "2. [밸런스] 장미1·2차: 한강조망+적정가+중기 수익률의 균형 잡힌 선택지",
        "   → 26.5억 매입, 한강변 프리미엄으로 42억+ 기대, ROI 5~8%",
        "",
        "3. [트로피 자산] 잠실주공5단지: 최고 입지/브랜드이나 현재 매입가에 프리미엄 과반영",
        "   → 45억 매입가 대비 순수익 제한적, 자산가치 보전+α 전략에 적합",
        "",
        "4. [전략] 투자 가용 자금이 충분하다면 우성4차+장미1·2차 분산투자 고려",
        "   → 총 48.5억으로 두 단지 분산, 리스크 분산+수익 극대화",
        "",
        "5. [주의] 토지거래허가제(2026.06까지)+재초환+보유세 강화는 모든 단지 공통 리스크",
        "   → 매입 전 실거주 계획 및 세금 시뮬레이션 필수",
        "",
        "6. [타이밍] 우성4차는 관리처분인가 임박→매입 타이밍 중요 (지금이 적기)",
        "   장미1·2차/우성1·2·3차는 시공사 선정 전후가 매입 적기",
    ]
    for conc in conclusions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value=conc).font = Font(name="맑은 고딕", size=10)
        ws.cell(row=r, column=1).alignment = left_align
        r += 1

    # 면책 조항
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    cell = ws.cell(row=r, column=1, value="※ 본 분석은 2026년 2월 기준 공개 데이터 및 시장 정보를 바탕으로 작성되었으며, 투자 판단의 참고자료입니다. 실제 투자 시 전문가 상담을 권장합니다.")
    cell.font = Font(name="맑은 고딕", size=9, color="999999", italic=True)
    cell.alignment = left_align

    return ws


# ========== 메인 실행 ==========
def main():
    wb = openpyxl.Workbook()

    print("1/13 종합 스코어카드 생성...")
    create_sheet_scorecard(wb)
    print("2/13 단지별 기본정보 생성...")
    create_sheet_basic_info(wb)
    print("3/13 재건축 진행현황 생성...")
    create_sheet_progress(wb)
    print("4/13 시세/가격분석 생성...")
    create_sheet_price(wb)
    print("5/13 수익성 시뮬레이션 생성...")
    create_sheet_roi(wb)
    print("6/13 비용구조 분석 생성...")
    create_sheet_cost(wb)
    print("7/13 리스크 매트릭스 생성...")
    create_sheet_risk(wb)
    print("8/13 벤치마크 비교 생성...")
    create_sheet_benchmark(wb)
    print("9/13 시나리오 분석 생성...")
    create_sheet_scenario(wb)
    print("10/13 현금흐름 분석 생성...")
    create_sheet_cashflow(wb)
    print("11/13 정책/규제 영향 생성...")
    create_sheet_policy(wb)
    print("12/13 입지/인프라 생성...")
    create_sheet_infra(wb)
    print("13/13 최종 투자 판단 생성...")
    create_sheet_final(wb)

    output = "C:/Devs/KRE/잠실_재건축_다차원_투자분석.xlsx"
    wb.save(output)
    print(f"\n완료! 파일 저장: {output}")
    print(f"총 {len(wb.sheetnames)}개 시트: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
