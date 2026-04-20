#!/usr/bin/env python3
"""서울 25개 자치구 + 구내 상위단지 심층분석 생성기 (v6)

핵심 기능
- 국토부 실거래 API(아파트 매매/전월세) 월별 수집
- 서울 25개 구 병렬 수집 + 병렬 다중전문가 스코어링
- 확률(분위수) 기반 낙관/중립/비관 시나리오
- 실거주/투자자/법인 3개 프로필 별도 엑셀 생성

실행 예시
python generate_seoul_investment_v6.py --months 12 --all-profiles
"""

from __future__ import annotations

import argparse
import concurrent.futures as cf
import datetime as dt
import json
import math
import os
import random
import statistics
import time
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DEFAULT_DATA_GO_KR_API_KEY = (
    "1deaf1b26dc45dd6f26d5b481d28d2fe784327eb82cc8ccdb727a51baf462419"
)

APT_TRADE_URL = "https://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade"
APT_RENT_URL = "https://apis.data.go.kr/1613000/RTMSDataSvcAptRent/getRTMSDataSvcAptRent"

DEFAULT_AREA_MIN = 59.0
DEFAULT_AREA_MAX = 99.0
DEFAULT_DEPOSIT_YIELD_PCT = 3.0
DEFAULT_TOP_COMPLEX_PER_DISTRICT = 7
DEFAULT_COMPLEX_MIN_TOTAL_TRADE = 18
DEFAULT_COMPLEX_MIN_MONTH_COVERAGE = 6
DEFAULT_COMPLEX_SHRINK_TRADE_REF = 120
DEFAULT_SCENARIO_RATE_SHOCK = 0.01
DEFAULT_COMPLETE_MONTH_SHIFT = -1
DEFAULT_BACKTEST_LOOKBACK = 4
DEFAULT_BACKTEST_HORIZON = 3
DEFAULT_BACKTEST_TOP_FRACTION = 0.20

BACKTEST_SIGNAL_CANDIDATES: list[dict[str, float]] = [
    {"mom": 0.20, "liq": 0.25, "fresh": 0.15, "val": 0.20, "yld": 0.20},
    {"mom": 0.25, "liq": 0.25, "fresh": 0.20, "yld": 0.30},
    {"mom": 0.25, "liq": 0.25, "fresh": 0.20, "val": 0.15, "lowvol": 0.15},
    {"mom": 0.30, "liq": 0.25, "fresh": 0.20, "val": 0.25},
    {"mom": 0.30, "liq": 0.30, "fresh": 0.20, "lowvol": 0.20},
    {"mom": 0.35, "liq": 0.35, "fresh": 0.30},
]

BACKTEST_TOP_FRACTION_CANDIDATES = (0.16, 0.20, 0.24, 0.30)

HIGH_INTENSITY_LTPS = {"강남구", "서초구", "송파구", "용산구"}

PROFILE_CONFIGS: dict[str, dict[str, Any]] = {
    "실거주": {
        "weights": {
            "trend": 17,
            "value": 23,
            "liquidity": 18,
            "stability": 27,
            "policy": 15,
        },
        "ltv": 0.40,
        "loan_rate": 0.037,
        "buy_tax": 0.015,
        "annual_hold_cost": 0.004,
        "capital_gain_tax": 0.15,
        "horizon_years": 5,
        "display": "실거주 안정성 + 생활비/현금흐름 방어",
    },
    "투자자": {
        "weights": {
            "trend": 29,
            "value": 24,
            "liquidity": 22,
            "stability": 10,
            "policy": 15,
        },
        "ltv": 0.50,
        "loan_rate": 0.041,
        "buy_tax": 0.025,
        "annual_hold_cost": 0.006,
        "capital_gain_tax": 0.30,
        "horizon_years": 3,
        "display": "모멘텀 + 밸류 + 유동성 중심",
    },
    "법인": {
        "weights": {
            "trend": 24,
            "value": 17,
            "liquidity": 20,
            "stability": 9,
            "policy": 30,
        },
        "ltv": 0.55,
        "loan_rate": 0.044,
        "buy_tax": 0.120,
        "annual_hold_cost": 0.009,
        "capital_gain_tax": 0.45,
        "horizon_years": 3,
        "display": "세후수익/규제리스크 중심",
    },
}

# 분위수 기반 z-score (현실적 범위: P20/P50/P80)
SCENARIO_Z = {
    "bull": 0.8416,
    "base": 0.0,
    "bear": -0.8416,
}

SOURCES: list[dict[str, str]] = [
    {
        "category": "official",
        "title": "서울시 토지거래허가구역 지정현황",
        "url": "https://land.seoul.go.kr/land/other/appointStatusSeoul.do",
        "as_of": "2026-01-07",
        "usage": "정책/규제 리스크(서울 전체 아파트 LTPS, 강남·서초·송파·용산 집중규제) 반영",
    },
    {
        "category": "official",
        "title": "한국은행 통화정책방향(2026-01-15, 기준금리 2.50%)",
        "url": "https://www.bok.or.kr/portal/singl/newsData/view.do?menuNo=200643&nttId=10095737",
        "as_of": "2026-01-15",
        "usage": "기준금리 2.50% 동결을 금리 가정의 기준점으로 사용",
    },
    {
        "category": "official",
        "title": "금융위 3단계 스트레스 DSR 시행방안(2025-05-20)",
        "url": "https://www.fsc.go.kr/no010101/84617",
        "as_of": "2025-05-20",
        "usage": "대출한도 제약 시나리오(레버리지 민감도) 반영",
    },
    {
        "category": "official",
        "title": "금융위 2026년 하반기 스트레스 DSR 운영방향",
        "url": "https://www.fsc.go.kr/no010101/85903",
        "as_of": "2026-02-20",
        "usage": "스트레스금리 체계 변경을 금리충격 시나리오에 반영",
    },
    {
        "category": "official",
        "title": "한국부동산원 R-ONE 월간 주택가격동향",
        "url": "https://www.reb.or.kr/r-one/portal/main/indexPage.do",
        "as_of": "2026-01-15",
        "usage": "서울 주택가격 사이클/벤치마크 비교 검증",
    },
    {
        "category": "official",
        "title": "IMF Article IV Consultation with Korea (2025-02-06)",
        "url": "https://www.imf.org/en/news/articles/2025/02/06/pr-2530-korea-imf-concludes-2024-art-iv-consultation-with-republic-of-korea",
        "as_of": "2025-02-06",
        "usage": "부동산 PF/비은행권 리스크를 하방 시나리오 리스크 프리미엄으로 반영",
    },
    {
        "category": "paper",
        "title": "Housing Prices and Future Returns (Case-Shiller, 1990, NBER w3368)",
        "url": "https://www.nber.org/papers/w3368",
        "as_of": "1990-06-01",
        "usage": "가격-임대료 비율(전세가율 유사)의 예측력 및 밸류 신호 근거",
    },
    {
        "category": "paper",
        "title": "House prices and rents: New evidence from OECD countries (JIMF, 2015)",
        "url": "https://www.sciencedirect.com/science/article/pii/S0261560615001332",
        "as_of": "2015-10-01",
        "usage": "주택가격-임대료 간 장기균형/괴리(밸류에이션 리스크) 근거",
    },
    {
        "category": "paper",
        "title": "Micro-Analysis of Price Spillover Effect (Land, 2021)",
        "url": "https://www.mdpi.com/2073-445X/10/8/879",
        "as_of": "2021-08-21",
        "usage": "서울권역 내 가격 파급효과(상위가격지 주도 전이) 해석 근거",
    },
    {
        "category": "paper",
        "title": "Aging and Real Estate Prices (BIS WP 2020)",
        "url": "https://www.bis.org/publ/work903.htm",
        "as_of": "2020-11-01",
        "usage": "인구구조 변화가 지역 부동산 수요/가격 민감도에 미치는 영향 근거",
    },
    {
        "category": "paper",
        "title": "Household lending, interest rates and housing price bubbles in Korea (Economic Modelling, 2011)",
        "url": "https://www.sciencedirect.com/science/article/abs/pii/S0264999311000198",
        "as_of": "2011-05-01",
        "usage": "금리/가계대출 변화가 버블 국면 변동성에 미치는 영향 근거",
    },
    {
        "category": "paper",
        "title": "Real estate illiquidity and expected returns in private and public real estate (Int. J. Finance & Economics, 2023)",
        "url": "https://onlinelibrary.wiley.com/doi/10.1002/ijfe.2780",
        "as_of": "2023-04-01",
        "usage": "유동성·비유동성 프리미엄을 랭킹/리스크 해석에 반영",
    },
    {
        "category": "community",
        "title": "Reddit r/korea: Housing market sentiment thread",
        "url": "https://www.reddit.com/r/korea/comments/17ir2l6/could_someone_shed_some_lights_on_the_housing/",
        "as_of": "2023-10-27",
        "usage": "외국인/실수요 관점의 체감 유동성·주거비 스트레스 정성검증",
    },
    {
        "category": "community",
        "title": "BiggerPockets: Seoul real estate discussion",
        "url": "https://www.biggerpockets.com/forums/79/topics/11793-seoul-korean-real-estate",
        "as_of": "2026-02-20",
        "usage": "글로벌 투자자 관점의 매수/보유/환금성 이슈 정성검증",
    },
]


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

DARK = "1B2A4A"
BLUE = "2C3E6B"
GOLD = "C9A84C"
WHITE = "FFFFFF"
LIGHT = "F3F5F9"
LINE = "D9D9D9"
GREEN = "27AE60"
RED = "E74C3C"
ORANGE = "F39C12"

FONT_TITLE = Font(name="맑은 고딕", size=14, bold=True, color=WHITE)
FONT_HEADER = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
FONT_CELL = Font(name="맑은 고딕", size=10)
FONT_SMALL = Font(name="맑은 고딕", size=9, color="666666")

FILL_TITLE = PatternFill("solid", fgColor=DARK)
FILL_HEADER = PatternFill("solid", fgColor=BLUE)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT)
FILL_GOLD = PatternFill("solid", fgColor=GOLD)
FILL_WHITE = PatternFill("solid", fgColor=WHITE)

BORDER = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class DistrictInfo:
    code: str
    name: str


@dataclass
class MonthSummary:
    year_month: str
    trade_count: int
    rent_count: int
    median_trade_10k: int | None
    median_rent_10k: int | None
    median_monthly_rent_10k: int | None
    jeonse_ratio_pct: float | None
    rent_yield_proxy_pct: float | None


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------


def clamp(v: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, v))


def pct(v: float | None) -> str:
    if v is None:
        return "-"
    return f"{v * 100:.1f}%"


def eok_from_10k(v: int | None) -> float | None:
    if v is None:
        return None
    return v / 10000.0


def now_yyyymm() -> str:
    t = dt.date.today()
    return f"{t.year:04d}{t.month:02d}"


def latest_complete_yyyymm(today: dt.date | None = None) -> str:
    # 실거래 API 당월 데이터는 통상 미완성(부분월)인 경우가 많아 직전월을 기본으로 사용.
    t = today or dt.date.today()
    cur = f"{t.year:04d}{t.month:02d}"
    return shift_yyyymm(cur, DEFAULT_COMPLETE_MONTH_SHIFT)


def shift_yyyymm(yyyymm: str, delta_months: int) -> str:
    y = int(yyyymm[:4])
    m = int(yyyymm[4:])
    total = y * 12 + (m - 1) + delta_months
    ny = total // 12
    nm = total % 12 + 1
    return f"{ny:04d}{nm:02d}"


def iter_months(end_yyyymm: str, months: int) -> list[str]:
    start_delta = -(months - 1)
    return [shift_yyyymm(end_yyyymm, i) for i in range(start_delta, 1)]


def month_diff(newer_yyyymm: str, older_yyyymm: str) -> int:
    y1, m1 = int(newer_yyyymm[:4]), int(newer_yyyymm[4:])
    y0, m0 = int(older_yyyymm[:4]), int(older_yyyymm[4:])
    return (y1 - y0) * 12 + (m1 - m0)


def median_int(values: list[int]) -> int | None:
    if not values:
        return None
    return int(statistics.median(values))


def stdev(values: list[float]) -> float:
    if len(values) <= 1:
        return 0.0
    return statistics.stdev(values)


def winsorize(values: list[float], alpha: float = 0.10) -> list[float]:
    if not values:
        return []
    if len(values) < 5:
        return values[:]
    sorted_vals = sorted(values)
    n = len(sorted_vals)
    lo_idx = int((n - 1) * alpha)
    hi_idx = int((n - 1) * (1.0 - alpha))
    lo = sorted_vals[lo_idx]
    hi = sorted_vals[hi_idx]
    return [clamp(v, lo, hi) for v in values]


def robust_mean_std(values: list[float]) -> tuple[float, float]:
    vals = [v for v in values if v is not None and not math.isnan(v)]
    if not vals:
        return 0.0, 0.0
    w = winsorize(vals, 0.10)
    mu = statistics.mean(w)
    sigma = statistics.stdev(w) if len(w) > 1 else 0.0
    return mu, sigma


def expected_shortfall(returns: list[float], tail_q: float = 0.20) -> float:
    vals = [float(x) for x in returns if x is not None and not math.isnan(x)]
    if not vals:
        return 0.0
    n = len(vals)
    k = max(1, int(math.ceil(n * tail_q)))
    worst = sorted(vals)[:k]
    return statistics.mean(worst)


def safe_div(a: float, b: float, default: float = 0.0) -> float:
    if not b:
        return default
    return a / b


def percentile_rank(value: float | None, arr: list[float], reverse: bool = False) -> float:
    valid = [x for x in arr if x is not None and not math.isnan(x)]
    if value is None or not valid:
        return 50.0
    sorted_vals = sorted(valid)
    n = len(sorted_vals)
    less = sum(1 for x in sorted_vals if x < value)
    equal = sum(1 for x in sorted_vals if x == value)
    rank = (less + 0.5 * equal) / n * 100.0
    if reverse:
        rank = 100.0 - rank
    return clamp(rank, 0.0, 100.0)


def shrink_toward_anchor(value: float | None, anchor: float | None, weight: float) -> float | None:
    if value is None and anchor is None:
        return None
    if value is None:
        return anchor
    if anchor is None:
        return value
    w = clamp(weight, 0.0, 1.0)
    return w * value + (1.0 - w) * anchor


def confidence_bucket(data_quality: float, total_trade_samples: int, coverage_trade: float) -> str:
    if data_quality >= 75 and total_trade_samples >= 120 and coverage_trade >= 0.75:
        return "high"
    if data_quality >= 58 and total_trade_samples >= 45 and coverage_trade >= 0.50:
        return "medium"
    return "low"


def confidence_penalty(bucket: str) -> float:
    if bucket == "high":
        return 0.00
    if bucket == "medium":
        return 0.05
    return 0.12


def style_cell(cell, *, font=None, fill=None, align=None):
    cell.font = font or FONT_CELL
    cell.fill = fill or FILL_WHITE
    cell.alignment = align or ALIGN_CENTER
    cell.border = BORDER


def write_title(ws, row: int, title: str, col_end: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=1, value=title)
    c.font = FONT_TITLE
    c.fill = FILL_TITLE
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = 30
    return row + 1


def write_headers(ws, row: int, headers: list[str]) -> int:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        style_cell(c, font=FONT_HEADER, fill=FILL_HEADER)
    ws.row_dimensions[row].height = 22
    return row + 1


def write_row(ws, row: int, values: list[Any], fill: PatternFill | None = None, left: bool = False) -> int:
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        style_cell(c, fill=fill or FILL_WHITE, align=ALIGN_LEFT if left else ALIGN_CENTER)
    return row + 1


def set_widths(ws, widths: list[float]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def score_band_fill(score: float) -> PatternFill:
    if score >= 75:
        return PatternFill("solid", fgColor="D5F5E3")
    if score >= 55:
        return PatternFill("solid", fgColor="FCF3CF")
    return PatternFill("solid", fgColor="FADBD8")


# ---------------------------------------------------------------------------
# Region / API helpers
# ---------------------------------------------------------------------------


def get_api_key(explicit: str | None = None) -> str:
    if explicit:
        return explicit.strip()
    env = os.getenv("DATA_GO_KR_API_KEY", "").strip()
    if env:
        return env
    return DEFAULT_DATA_GO_KR_API_KEY


def load_seoul_gu_list(region_code_file: Path) -> list[DistrictInfo]:
    text = region_code_file.read_text(encoding="utf-8")
    rows: list[DistrictInfo] = []
    prefix = "\uc11c\uc6b8\ud2b9\ubcc4\uc2dc "
    for line in text.splitlines():
        parts = line.split("\t")
        if len(parts) < 2:
            continue
        code10 = parts[0].strip()
        name = parts[1].strip()
        if not (code10.isdigit() and len(code10) == 10):
            continue
        if code10 == "1100000000":
            continue
        if not code10.startswith("11"):
            continue
        if not code10.endswith("00000"):
            continue
        if not name.startswith(prefix):
            continue
        gu_name = name.replace(prefix, "")
        rows.append(DistrictInfo(code=code10[:5], name=gu_name))

    rows.sort(key=lambda x: x.code)
    if len(rows) != 25:
        raise RuntimeError(f"서울 자치구 코드 25개를 찾지 못했습니다. 현재: {len(rows)}")
    return rows


def build_url(base_url: str, api_key: str, region_code: str, yyyymm: str, num_rows: int, page_no: int) -> str:
    params = {
        "serviceKey": api_key,
        "LAWD_CD": region_code,
        "DEAL_YMD": yyyymm,
        "numOfRows": str(num_rows),
        "pageNo": str(page_no),
    }
    return f"{base_url}?{urllib.parse.urlencode(params)}"


def fetch_xml(url: str, timeout_sec: int = 25, retries: int = 4) -> str:
    backoff = 0.6
    last_err: Exception | None = None
    for _ in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "KRE-Seoul-Analyzer/1.0"})
            with urllib.request.urlopen(req, timeout=timeout_sec) as resp:
                return resp.read().decode("utf-8", errors="ignore")
        except Exception as e:  # noqa: BLE001
            last_err = e
            time.sleep(backoff + random.random() * 0.3)
            backoff *= 1.9
    raise RuntimeError(f"API 요청 실패: {last_err}")


def parse_xml_items_trade(xml_text: str) -> tuple[list[dict[str, Any]], int, str | None]:
    root = ET.fromstring(xml_text)
    code = (root.findtext(".//resultCode") or "").strip()
    if code != "000":
        return [], 0, code

    total_count = int((root.findtext(".//totalCount") or "0").strip() or "0")
    out: list[dict[str, Any]] = []
    for item in root.findall(".//item"):
        if ((item.findtext("cdealType") or "").strip()) == "O":
            continue
        raw_price = (item.findtext("dealAmount") or "").replace(",", "").strip()
        if not raw_price.isdigit():
            continue
        area_raw = (item.findtext("excluUseAr") or "0").strip()
        try:
            area = float(area_raw)
        except ValueError:
            area = 0.0
        out.append(
            {
                "apt_name": (item.findtext("aptNm") or "").strip(),
                "dong": (item.findtext("umdNm") or "").strip(),
                "area_sqm": area,
                "price_10k": int(raw_price),
                "year": int((item.findtext("dealYear") or "0").strip() or 0),
                "month": int((item.findtext("dealMonth") or "0").strip() or 0),
                "day": int((item.findtext("dealDay") or "0").strip() or 0),
            }
        )
    return out, total_count, None


def parse_xml_items_rent(xml_text: str) -> tuple[list[dict[str, Any]], int, str | None]:
    root = ET.fromstring(xml_text)
    code = (root.findtext(".//resultCode") or "").strip()
    if code != "000":
        return [], 0, code

    total_count = int((root.findtext(".//totalCount") or "0").strip() or "0")
    out: list[dict[str, Any]] = []
    for item in root.findall(".//item"):
        if ((item.findtext("cdealType") or "").strip()) == "O":
            continue
        raw_dep = (item.findtext("deposit") or "").replace(",", "").strip()
        raw_mon = (item.findtext("monthlyRent") or "0").replace(",", "").strip() or "0"
        if not raw_dep.isdigit():
            continue
        area_raw = (item.findtext("excluUseAr") or "0").strip()
        try:
            area = float(area_raw)
        except ValueError:
            area = 0.0
        out.append(
            {
                "apt_name": (item.findtext("aptNm") or "").strip(),
                "dong": (item.findtext("umdNm") or "").strip(),
                "area_sqm": area,
                "deposit_10k": int(raw_dep),
                "monthly_rent_10k": int(raw_mon) if raw_mon.isdigit() else 0,
                "contract_type": (item.findtext("contractType") or "").strip(),
                "year": int((item.findtext("dealYear") or "0").strip() or 0),
                "month": int((item.findtext("dealMonth") or "0").strip() or 0),
                "day": int((item.findtext("dealDay") or "0").strip() or 0),
            }
        )
    return out, total_count, None


def fetch_all_pages(
    *,
    base_url: str,
    api_key: str,
    region_code: str,
    yyyymm: str,
    num_rows: int,
    parse_kind: str,
    max_pages: int = 80,
) -> tuple[list[dict[str, Any]], int, str | None]:
    parser = parse_xml_items_trade if parse_kind == "trade" else parse_xml_items_rent
    all_items: list[dict[str, Any]] = []
    total_count = 0

    for page in range(1, max_pages + 1):
        url = build_url(base_url, api_key, region_code, yyyymm, num_rows, page)
        xml_text = fetch_xml(url)
        items, tcount, err = parser(xml_text)
        if err is not None:
            return [], 0, err

        total_count = max(total_count, tcount)
        all_items.extend(items)

        if not items:
            break
        if len(items) < num_rows:
            break
        if total_count and len(all_items) >= total_count:
            break

    return all_items, total_count, None


def filter_area(items: list[dict[str, Any]], area_min: float, area_max: float) -> list[dict[str, Any]]:
    return [x for x in items if area_min <= float(x.get("area_sqm", 0.0)) <= area_max]


def normalize_apt_name(name: str) -> str:
    t = (name or "").strip()
    t = " ".join(t.split())
    if not t:
        return "(미상단지)"
    return t


def build_complex_month_snapshot(
    *,
    district_name: str,
    yyyymm: str,
    trade_items: list[dict[str, Any]],
    rent_items: list[dict[str, Any]],
    deposit_yield_pct: float,
) -> dict[str, dict[str, Any]]:
    trade_map: dict[str, list[int]] = {}
    rent_dep_map: dict[str, list[int]] = {}
    rent_mon_map: dict[str, list[int]] = {}

    for it in trade_items:
        nm = normalize_apt_name(str(it.get("apt_name") or ""))
        p = int(it.get("price_10k") or 0)
        if p <= 0:
            continue
        trade_map.setdefault(nm, []).append(p)

    for it in rent_items:
        nm = normalize_apt_name(str(it.get("apt_name") or ""))
        d = int(it.get("deposit_10k") or 0)
        m = int(it.get("monthly_rent_10k") or 0)
        if d <= 0:
            continue
        rent_dep_map.setdefault(nm, []).append(d)
        if m > 0:
            rent_mon_map.setdefault(nm, []).append(m)

    names = sorted(set(trade_map) | set(rent_dep_map))
    out: dict[str, dict[str, Any]] = {}
    for nm in names:
        prices = trade_map.get(nm, [])
        deps = rent_dep_map.get(nm, [])
        mons = rent_mon_map.get(nm, [])

        med_price = median_int(prices)
        med_dep = median_int(deps)
        med_mon = median_int(mons)
        jeonse_ratio = None
        if med_price and med_dep:
            jeonse_ratio = med_dep / med_price * 100.0

        rent_yield_proxy = None
        if med_price and (med_dep or med_mon):
            annual_income_10k = (med_mon or 0) * 12 + int((med_dep or 0) * (deposit_yield_pct / 100.0))
            rent_yield_proxy = annual_income_10k / med_price * 100.0

        out[nm] = {
            "district": district_name,
            "year_month": yyyymm,
            "complex": nm,
            "trade_count": len(prices),
            "rent_count": len(deps),
            "median_trade_10k": med_price,
            "median_rent_10k": med_dep,
            "median_monthly_rent_10k": med_mon,
            "jeonse_ratio_pct": jeonse_ratio,
            "rent_yield_proxy_pct": rent_yield_proxy,
        }

    return out


def summarize_month(
    *,
    yyyymm: str,
    trade_items: list[dict[str, Any]],
    rent_items: list[dict[str, Any]],
    deposit_yield_pct: float,
) -> MonthSummary:
    prices = [int(x["price_10k"]) for x in trade_items if int(x.get("price_10k", 0)) > 0]
    deposits = [int(x["deposit_10k"]) for x in rent_items if int(x.get("deposit_10k", 0)) > 0]
    monthly_rents = [
        int(x["monthly_rent_10k"])
        for x in rent_items
        if int(x.get("monthly_rent_10k", 0)) > 0
    ]

    med_price = median_int(prices)
    med_deposit = median_int(deposits)
    med_monthly = median_int(monthly_rents)

    jeonse_ratio = None
    if med_price and med_deposit:
        jeonse_ratio = med_deposit / med_price * 100.0

    rent_yield_proxy = None
    if med_price and (med_deposit or med_monthly):
        annual_income_10k = (med_monthly or 0) * 12 + int((med_deposit or 0) * (deposit_yield_pct / 100.0))
        rent_yield_proxy = annual_income_10k / med_price * 100.0

    return MonthSummary(
        year_month=yyyymm,
        trade_count=len(prices),
        rent_count=len(deposits),
        median_trade_10k=med_price,
        median_rent_10k=med_deposit,
        median_monthly_rent_10k=med_monthly,
        jeonse_ratio_pct=jeonse_ratio,
        rent_yield_proxy_pct=rent_yield_proxy,
    )

# ---------------------------------------------------------------------------
# Collection and analytics
# ---------------------------------------------------------------------------


def collect_one_district_month(
    district: DistrictInfo,
    yyyymm: str,
    api_key: str,
    num_rows: int,
    area_min: float,
    area_max: float,
    deposit_yield_pct: float,
) -> tuple[str, str, dict[str, Any]]:
    t0 = time.time()

    trade_items, _, trade_err = fetch_all_pages(
        base_url=APT_TRADE_URL,
        api_key=api_key,
        region_code=district.code,
        yyyymm=yyyymm,
        num_rows=num_rows,
        parse_kind="trade",
    )
    if trade_err:
        return district.name, yyyymm, {
            "error": f"trade_api_error:{trade_err}",
            "elapsed": round(time.time() - t0, 2),
        }

    rent_items, _, rent_err = fetch_all_pages(
        base_url=APT_RENT_URL,
        api_key=api_key,
        region_code=district.code,
        yyyymm=yyyymm,
        num_rows=num_rows,
        parse_kind="rent",
    )
    if rent_err:
        return district.name, yyyymm, {
            "error": f"rent_api_error:{rent_err}",
            "elapsed": round(time.time() - t0, 2),
        }

    trade_f = filter_area(trade_items, area_min, area_max)
    rent_f = filter_area(rent_items, area_min, area_max)
    summary = summarize_month(
        yyyymm=yyyymm,
        trade_items=trade_f,
        rent_items=rent_f,
        deposit_yield_pct=deposit_yield_pct,
    )
    complex_snapshot = build_complex_month_snapshot(
        district_name=district.name,
        yyyymm=yyyymm,
        trade_items=trade_f,
        rent_items=rent_f,
        deposit_yield_pct=deposit_yield_pct,
    )

    return district.name, yyyymm, {
        "error": None,
        "elapsed": round(time.time() - t0, 2),
        "summary": summary.__dict__,
        "complexes": complex_snapshot,
        "raw_counts": {
            "trade_total": len(trade_items),
            "trade_filtered": len(trade_f),
            "rent_total": len(rent_items),
            "rent_filtered": len(rent_f),
        },
    }


def collect_market_panel(
    *,
    districts: list[DistrictInfo],
    months: list[str],
    api_key: str,
    num_rows: int,
    max_workers: int,
    area_min: float,
    area_max: float,
    deposit_yield_pct: float,
) -> tuple[dict[str, dict[str, dict[str, Any]]], list[dict[str, Any]]]:
    panel: dict[str, dict[str, dict[str, Any]]] = {d.name: {} for d in districts}
    logs: list[dict[str, Any]] = []

    with cf.ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {}
        for d in districts:
            for ym in months:
                fut = ex.submit(
                    collect_one_district_month,
                    d,
                    ym,
                    api_key,
                    num_rows,
                    area_min,
                    area_max,
                    deposit_yield_pct,
                )
                futures[fut] = (d.name, ym)

        total = len(futures)
        done = 0
        for fut in cf.as_completed(futures):
            name, ym = futures[fut]
            done += 1
            try:
                n, y, payload = fut.result()
            except Exception as e:  # noqa: BLE001
                payload = {"error": f"exception:{e}", "elapsed": None}
                n, y = name, ym

            panel[n][y] = payload
            logs.append({"district": n, "year_month": y, **payload})
            if done % 40 == 0 or done == total:
                print(f"[collect] {done}/{total} 완료")

    return panel, logs


def ffill(values: list[float | None]) -> list[float | None]:
    out: list[float | None] = []
    last = None
    for v in values:
        if v is None:
            out.append(last)
        else:
            out.append(v)
            last = v
    return out


def compute_max_drawdown(series: list[float]) -> float:
    if not series:
        return 0.0
    peak = series[0]
    mdd = 0.0
    for x in series:
        peak = max(peak, x)
        if peak > 0:
            dd = (x / peak) - 1.0
            mdd = min(mdd, dd)
    return abs(mdd)


def annualized_vol_from_prices(prices: list[float]) -> float:
    rets: list[float] = []
    for i in range(1, len(prices)):
        p0 = prices[i - 1]
        p1 = prices[i]
        if p0 > 0 and p1 > 0:
            rets.append(math.log(p1 / p0))
    if len(rets) <= 1:
        return 0.0
    return stdev(rets) * math.sqrt(12.0)


def growth(prices: list[float | None], lag: int) -> float | None:
    if len(prices) <= lag or lag <= 0:
        return None
    p1 = prices[-1]
    p0 = prices[-1 - lag]
    if p1 is None or p0 is None:
        return None
    if p0 <= 0:
        return None
    return p1 / p0 - 1.0


def build_district_metrics(
    districts: list[DistrictInfo],
    months: list[str],
    panel: dict[str, dict[str, dict[str, Any]]],
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []

    for d in districts:
        month_rows = []
        for ym in months:
            payload = panel.get(d.name, {}).get(ym, {})
            if payload.get("error"):
                month_rows.append({"year_month": ym, "error": payload.get("error")})
                continue
            summary = payload.get("summary") or {}
            raw_counts = payload.get("raw_counts") or {}
            month_rows.append(
                {
                    "year_month": ym,
                    **summary,
                    **raw_counts,
                }
            )

        trade_series: list[float | None] = [
            (r.get("median_trade_10k") / 10000.0) if r.get("median_trade_10k") else None
            for r in month_rows
        ]
        rent_series: list[float | None] = [
            (r.get("median_rent_10k") / 10000.0) if r.get("median_rent_10k") else None
            for r in month_rows
        ]
        jeonse_series: list[float | None] = [r.get("jeonse_ratio_pct") for r in month_rows]
        yield_series: list[float | None] = [r.get("rent_yield_proxy_pct") for r in month_rows]

        trade_counts = [int(r.get("trade_count") or 0) for r in month_rows]
        rent_counts = [int(r.get("rent_count") or 0) for r in month_rows]

        trade_ffill = ffill(trade_series)
        rent_ffill = ffill(rent_series)

        valid_price = [x for x in trade_ffill if x is not None and x > 0]
        latest_price = valid_price[-1] if valid_price else None

        valid_rent = [x for x in rent_ffill if x is not None and x > 0]
        latest_rent = valid_rent[-1] if valid_rent else None

        lag_3m = min(3, max(1, len(trade_ffill) - 1))
        lag_6m = min(6, max(1, len(trade_ffill) - 1))
        # 12개월 데이터가 있으면 끝-처음(11간격)으로 YoY에 준하는 성장률을 계산한다.
        lag_12m = min(11, max(1, len(trade_ffill) - 1))
        g3 = growth(trade_ffill, lag_3m)
        g6 = growth(trade_ffill, lag_6m)
        g12 = growth(trade_ffill, lag_12m)

        vols = annualized_vol_from_prices([x for x in trade_ffill if x is not None])
        mdd = compute_max_drawdown([x for x in trade_ffill if x is not None])

        latest_jeonse = next((x for x in reversed(jeonse_series) if x is not None), None)
        latest_yield = next((x for x in reversed(yield_series) if x is not None), None)

        latest_trade_count = trade_counts[-1] if trade_counts else 0
        latest_rent_count = rent_counts[-1] if rent_counts else 0
        avg_trade_3m = statistics.mean(trade_counts[-3:]) if len(trade_counts) >= 3 else statistics.mean(trade_counts) if trade_counts else 0.0
        last_trade_idx = max((i for i, c in enumerate(trade_counts) if c > 0), default=None)
        last_trade_yyyymm = months[last_trade_idx] if last_trade_idx is not None else None
        trade_stale_months = (len(months) - 1 - last_trade_idx) if last_trade_idx is not None else len(months)
        market_freshness = clamp(1.0 - safe_div(trade_stale_months, 4.0), 0.0, 1.0)

        monthly_returns: list[float] = []
        for i in range(1, len(trade_ffill)):
            p0 = trade_ffill[i - 1]
            p1 = trade_ffill[i]
            if p0 is None or p1 is None or p0 <= 0:
                continue
            monthly_returns.append(clamp((p1 / p0) - 1.0, -0.35, 0.35))
        tail_es = expected_shortfall(monthly_returns, 0.20)
        tail_es_abs = abs(min(0.0, tail_es))

        months_with_trade = sum(1 for x in trade_counts if x > 0)
        months_with_rent = sum(1 for x in rent_counts if x > 0)
        coverage_trade = months_with_trade / len(months) if months else 0.0
        coverage_rent = months_with_rent / len(months) if months else 0.0

        total_trade_samples = sum(trade_counts)
        total_rent_samples = sum(rent_counts)

        quality = (
            40.0 * clamp(coverage_trade, 0.0, 1.0)
            + 25.0 * clamp(coverage_rent, 0.0, 1.0)
            + 20.0 * clamp(safe_div(latest_trade_count, 60.0), 0.0, 1.0)
            + 15.0 * clamp(safe_div(latest_rent_count, 120.0), 0.0, 1.0)
        )
        quality -= 12.0 * clamp(safe_div(trade_stale_months, 6.0), 0.0, 1.0)
        quality = clamp(quality, 0.0, 100.0)

        out.append(
            {
                "district": d.name,
                "code": d.code,
                "months": month_rows,
                "latest_price_eok": latest_price,
                "latest_rent_eok": latest_rent,
                "growth_3m": g3,
                "growth_6m": g6,
                "growth_12m": g12,
                "vol_ann": vols,
                "max_drawdown": mdd,
                "latest_jeonse_ratio": latest_jeonse,
                "latest_rent_yield": latest_yield,
                "latest_trade_count": latest_trade_count,
                "latest_rent_count": latest_rent_count,
                "avg_trade_3m": float(avg_trade_3m),
                "coverage_trade": coverage_trade,
                "coverage_rent": coverage_rent,
                "total_trade_samples": total_trade_samples,
                "total_rent_samples": total_rent_samples,
                "last_trade_yyyymm": last_trade_yyyymm,
                "trade_stale_months": trade_stale_months,
                "market_freshness": round(market_freshness, 3),
                "monthly_returns": monthly_returns,
                "tail_es_m20": tail_es,
                "tail_es_abs": tail_es_abs,
                "data_quality": round(quality, 1),
            }
        )

    # 서울 전체 상대가치
    latest_prices = [m["latest_price_eok"] for m in out if m["latest_price_eok"]]
    seoul_median_price = statistics.median(latest_prices) if latest_prices else 0.0

    for m in out:
        p = m["latest_price_eok"]
        if p and seoul_median_price > 0:
            m["price_vs_seoul"] = p / seoul_median_price
        else:
            m["price_vs_seoul"] = None

    return out


def build_complex_metrics(
    *,
    months: list[str],
    panel: dict[str, dict[str, dict[str, Any]]],
    district_metrics: list[dict[str, Any]],
    min_total_trade: int,
    min_month_coverage: int,
    shrink_trade_ref: int,
) -> list[dict[str, Any]]:
    district_lookup = {d["district"]: d for d in district_metrics}
    series_map: dict[tuple[str, str], dict[str, dict[str, Any]]] = {}

    for district_name, month_map in panel.items():
        for ym in months:
            payload = month_map.get(ym, {})
            for cname, cval in (payload.get("complexes") or {}).items():
                key = (district_name, cname)
                series_map.setdefault(key, {})
                series_map[key][ym] = cval

    out: list[dict[str, Any]] = []
    for (district_name, cname), ym_map in series_map.items():
        month_rows = []
        for ym in months:
            c = ym_map.get(ym)
            if c is None:
                month_rows.append(
                    {
                        "year_month": ym,
                        "trade_count": 0,
                        "rent_count": 0,
                        "median_trade_10k": None,
                        "median_rent_10k": None,
                        "jeonse_ratio_pct": None,
                        "rent_yield_proxy_pct": None,
                    }
                )
            else:
                month_rows.append(c)

        trade_series: list[float | None] = [
            (r.get("median_trade_10k") / 10000.0) if r.get("median_trade_10k") else None
            for r in month_rows
        ]
        rent_series: list[float | None] = [
            (r.get("median_rent_10k") / 10000.0) if r.get("median_rent_10k") else None
            for r in month_rows
        ]
        jeonse_series: list[float | None] = [r.get("jeonse_ratio_pct") for r in month_rows]
        yield_series: list[float | None] = [r.get("rent_yield_proxy_pct") for r in month_rows]

        trade_counts = [int(r.get("trade_count") or 0) for r in month_rows]
        rent_counts = [int(r.get("rent_count") or 0) for r in month_rows]
        total_trade_samples = sum(trade_counts)
        total_rent_samples = sum(rent_counts)

        if total_trade_samples < min_total_trade:
            continue

        trade_ffill = ffill(trade_series)
        rent_ffill = ffill(rent_series)
        valid_price = [x for x in trade_ffill if x is not None and x > 0]
        valid_rent = [x for x in rent_ffill if x is not None and x > 0]

        latest_price = valid_price[-1] if valid_price else None
        latest_rent = valid_rent[-1] if valid_rent else None

        lag_3m = min(3, max(1, len(trade_ffill) - 1))
        lag_6m = min(6, max(1, len(trade_ffill) - 1))
        lag_12m = min(11, max(1, len(trade_ffill) - 1))
        g3_raw = growth(trade_ffill, lag_3m)
        g6_raw = growth(trade_ffill, lag_6m)
        g12_raw = growth(trade_ffill, lag_12m)

        vols_raw = annualized_vol_from_prices([x for x in trade_ffill if x is not None])
        mdd = compute_max_drawdown([x for x in trade_ffill if x is not None])

        latest_jeonse = next((x for x in reversed(jeonse_series) if x is not None), None)
        latest_yield = next((x for x in reversed(yield_series) if x is not None), None)

        latest_trade_count = trade_counts[-1] if trade_counts else 0
        latest_rent_count = rent_counts[-1] if rent_counts else 0
        avg_trade_3m = (
            statistics.mean(trade_counts[-3:])
            if len(trade_counts) >= 3
            else statistics.mean(trade_counts)
            if trade_counts
            else 0.0
        )
        last_trade_idx = max((i for i, c in enumerate(trade_counts) if c > 0), default=None)
        last_trade_yyyymm = months[last_trade_idx] if last_trade_idx is not None else None
        trade_stale_months = (len(months) - 1 - last_trade_idx) if last_trade_idx is not None else len(months)
        market_freshness = clamp(1.0 - safe_div(trade_stale_months, 4.0), 0.0, 1.0)

        monthly_returns: list[float] = []
        for i in range(1, len(trade_ffill)):
            p0 = trade_ffill[i - 1]
            p1 = trade_ffill[i]
            if p0 is None or p1 is None or p0 <= 0:
                continue
            monthly_returns.append(clamp((p1 / p0) - 1.0, -0.35, 0.35))
        tail_es = expected_shortfall(monthly_returns, 0.20)
        tail_es_abs = abs(min(0.0, tail_es))

        months_with_trade = sum(1 for x in trade_counts if x > 0)
        months_with_rent = sum(1 for x in rent_counts if x > 0)
        if months_with_trade < min_month_coverage:
            continue
        coverage_trade = months_with_trade / len(months) if months else 0.0
        coverage_rent = months_with_rent / len(months) if months else 0.0

        quality = (
            42.0 * clamp(coverage_trade, 0.0, 1.0)
            + 18.0 * clamp(coverage_rent, 0.0, 1.0)
            + 22.0 * clamp(safe_div(total_trade_samples, 120.0), 0.0, 1.0)
            + 18.0 * clamp(safe_div(latest_trade_count, 12.0), 0.0, 1.0)
        )
        quality -= 18.0 * clamp(safe_div(trade_stale_months, 6.0), 0.0, 1.0)
        quality = clamp(quality, 0.0, 100.0)

        parent = district_lookup.get(district_name, {})
        sample_weight = clamp(safe_div(total_trade_samples, float(shrink_trade_ref)), 0.0, 1.0) * clamp(coverage_trade, 0.0, 1.0)
        g3 = shrink_toward_anchor(g3_raw, parent.get("growth_3m"), sample_weight)
        g6 = shrink_toward_anchor(g6_raw, parent.get("growth_6m"), sample_weight)
        g12 = shrink_toward_anchor(g12_raw, parent.get("growth_12m"), sample_weight)

        conf_bucket = confidence_bucket(quality, total_trade_samples, coverage_trade)
        unc_pen = confidence_penalty(conf_bucket)
        vols = vols_raw * (1.0 + unc_pen * 1.2)

        district_price = parent.get("latest_price_eok")
        price_vs_district = safe_div(latest_price, district_price) if latest_price and district_price else None
        rel_growth_12m = (
            (g12 - parent.get("growth_12m"))
            if g12 is not None and parent.get("growth_12m") is not None
            else None
        )

        out.append(
            {
                "district": district_name,
                "complex": cname,
                "complex_key": f"{district_name}::{cname}",
                "months": month_rows,
                "latest_price_eok": latest_price,
                "latest_rent_eok": latest_rent,
                "growth_3m": g3,
                "growth_6m": g6,
                "growth_12m": g12,
                "rel_growth_12m": rel_growth_12m,
                "vol_ann": vols,
                "max_drawdown": mdd,
                "latest_jeonse_ratio": latest_jeonse,
                "latest_rent_yield": latest_yield,
                "latest_trade_count": latest_trade_count,
                "latest_rent_count": latest_rent_count,
                "avg_trade_3m": float(avg_trade_3m),
                "coverage_trade": coverage_trade,
                "coverage_rent": coverage_rent,
                "total_trade_samples": total_trade_samples,
                "total_rent_samples": total_rent_samples,
                "last_trade_yyyymm": last_trade_yyyymm,
                "trade_stale_months": trade_stale_months,
                "market_freshness": round(market_freshness, 3),
                "monthly_returns": monthly_returns,
                "tail_es_m20": tail_es,
                "tail_es_abs": tail_es_abs,
                "data_quality": round(quality, 1),
                "confidence_bucket": conf_bucket,
                "uncertainty_penalty": unc_pen,
                "sample_weight": round(sample_weight, 3),
                "price_vs_district": price_vs_district,
                "parent_district_score": parent.get("total_score"),
                "parent_policy_score": parent.get("policy_score"),
            }
        )

    latest_prices = [m["latest_price_eok"] for m in out if m["latest_price_eok"]]
    seoul_complex_median = statistics.median(latest_prices) if latest_prices else 0.0
    for m in out:
        p = m["latest_price_eok"]
        m["price_vs_seoul"] = p / seoul_complex_median if p and seoul_complex_median > 0 else None

    return out


def run_multi_expert_scoring(metrics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    # 병렬 전문가 에이전트 입력 벡터
    growth12_all = [m.get("growth_12m") for m in metrics if m.get("growth_12m") is not None]
    growth3_all = [m.get("growth_3m") for m in metrics if m.get("growth_3m") is not None]
    ext12_all = [abs(float(m.get("growth_12m"))) for m in metrics if m.get("growth_12m") is not None]
    jeonse_all = [m.get("latest_jeonse_ratio") for m in metrics if m.get("latest_jeonse_ratio") is not None]
    yield_all = [m.get("latest_rent_yield") for m in metrics if m.get("latest_rent_yield") is not None]
    pvs_all = [m.get("price_vs_seoul") for m in metrics if m.get("price_vs_seoul") is not None]
    liq_all = [m.get("avg_trade_3m") for m in metrics if m.get("avg_trade_3m") is not None]
    vol_all = [m.get("vol_ann") for m in metrics if m.get("vol_ann") is not None]
    dd_all = [m.get("max_drawdown") for m in metrics if m.get("max_drawdown") is not None]
    es_all = [m.get("tail_es_abs") for m in metrics if m.get("tail_es_abs") is not None]
    fresh_all = [m.get("market_freshness") for m in metrics if m.get("market_freshness") is not None]
    stale_all = [m.get("trade_stale_months") for m in metrics if m.get("trade_stale_months") is not None]

    def gap_risk_penalty(jeonse_ratio: float | None) -> float:
        if jeonse_ratio is None:
            return 0.0
        j = float(jeonse_ratio)
        penalty = 0.0
        if j >= 80.0:
            penalty += 20.0
        elif j >= 75.0:
            penalty += 12.0
        elif j >= 70.0:
            penalty += 6.0
        if j <= 30.0:
            penalty += 6.0
        elif j <= 35.0:
            penalty += 3.0
        return penalty

    def trend_expert(m: dict[str, Any]) -> float:
        s12 = percentile_rank(m.get("growth_12m"), growth12_all)
        s3 = percentile_rank(m.get("growth_3m"), growth3_all)
        s_ext = percentile_rank(abs(float(m.get("growth_12m") or 0.0)), ext12_all, reverse=True)
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        return 0.50 * s12 + 0.22 * s3 + 0.18 * s_ext + 0.10 * s_fresh

    def value_expert(m: dict[str, Any]) -> float:
        s_jeonse = percentile_rank(m.get("latest_jeonse_ratio"), jeonse_all)
        s_yield = percentile_rank(m.get("latest_rent_yield"), yield_all)
        s_pvs = percentile_rank(m.get("price_vs_seoul"), pvs_all, reverse=True)
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        score = 0.34 * s_jeonse + 0.30 * s_yield + 0.24 * s_pvs + 0.12 * s_fresh
        score -= gap_risk_penalty(m.get("latest_jeonse_ratio"))
        return clamp(score, 0.0, 100.0)

    def liquidity_expert(m: dict[str, Any]) -> float:
        s_liq = percentile_rank(m.get("avg_trade_3m"), liq_all)
        s_latest = percentile_rank(float(m.get("latest_trade_count") or 0), liq_all)
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        s_stale = percentile_rank(m.get("trade_stale_months"), stale_all, reverse=True)
        return 0.40 * s_liq + 0.25 * s_latest + 0.20 * s_fresh + 0.15 * s_stale

    def stability_expert(m: dict[str, Any]) -> float:
        s_vol = percentile_rank(m.get("vol_ann"), vol_all, reverse=True)
        s_dd = percentile_rank(m.get("max_drawdown"), dd_all, reverse=True)
        dq = m.get("data_quality") or 0.0
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        s_es = percentile_rank(m.get("tail_es_abs"), es_all, reverse=True)
        return 0.28 * s_vol + 0.24 * s_dd + 0.22 * dq + 0.14 * s_fresh + 0.12 * s_es

    def policy_expert(m: dict[str, Any]) -> float:
        base = 45.0  # 서울 전체 LTPS 가정
        if m["district"] in HIGH_INTENSITY_LTPS:
            base -= 15.0
        # 거래가 충분하면 정책 충격 흡수력이 상대적으로 높다고 가정
        base += clamp((m.get("avg_trade_3m") or 0) / 12.0, 0.0, 10.0)
        base += clamp((m.get("data_quality") or 0) / 20.0, 0.0, 5.0)
        base += clamp((m.get("market_freshness") or 0.0) * 8.0, 0.0, 8.0)
        base -= clamp(safe_div(float(m.get("trade_stale_months") or 0), 6.0), 0.0, 1.0) * 10.0
        base += clamp((1.0 - clamp(safe_div(float(m.get("tail_es_abs") or 0.0), 0.10), 0.0, 1.0)) * 4.0, 0.0, 4.0)
        base -= gap_risk_penalty(m.get("latest_jeonse_ratio")) * 0.35
        return clamp(base, 0.0, 100.0)

    with cf.ThreadPoolExecutor(max_workers=5) as ex:
        f_trend = {ex.submit(trend_expert, m): m for m in metrics}
        f_value = {ex.submit(value_expert, m): m for m in metrics}
        f_liq = {ex.submit(liquidity_expert, m): m for m in metrics}
        f_stb = {ex.submit(stability_expert, m): m for m in metrics}
        f_pol = {ex.submit(policy_expert, m): m for m in metrics}

        for fut_map, key in [
            (f_trend, "trend_score"),
            (f_value, "value_score"),
            (f_liq, "liquidity_score"),
            (f_stb, "stability_score"),
            (f_pol, "policy_score"),
        ]:
            for fut in cf.as_completed(fut_map):
                m = fut_map[fut]
                try:
                    m[key] = round(float(fut.result()), 2)
                except Exception:  # noqa: BLE001
                    m[key] = 50.0

    return metrics


def run_complex_expert_scoring(metrics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    growth12_all = [m.get("growth_12m") for m in metrics if m.get("growth_12m") is not None]
    ext12_all = [abs(float(m.get("growth_12m"))) for m in metrics if m.get("growth_12m") is not None]
    rel_growth_all = [m.get("rel_growth_12m") for m in metrics if m.get("rel_growth_12m") is not None]
    jeonse_all = [m.get("latest_jeonse_ratio") for m in metrics if m.get("latest_jeonse_ratio") is not None]
    yield_all = [m.get("latest_rent_yield") for m in metrics if m.get("latest_rent_yield") is not None]
    pvd_all = [m.get("price_vs_district") for m in metrics if m.get("price_vs_district") is not None]
    pvs_all = [m.get("price_vs_seoul") for m in metrics if m.get("price_vs_seoul") is not None]
    liq_all = [m.get("avg_trade_3m") for m in metrics if m.get("avg_trade_3m") is not None]
    vol_all = [m.get("vol_ann") for m in metrics if m.get("vol_ann") is not None]
    dd_all = [m.get("max_drawdown") for m in metrics if m.get("max_drawdown") is not None]
    es_all = [m.get("tail_es_abs") for m in metrics if m.get("tail_es_abs") is not None]
    fresh_all = [m.get("market_freshness") for m in metrics if m.get("market_freshness") is not None]
    stale_all = [m.get("trade_stale_months") for m in metrics if m.get("trade_stale_months") is not None]

    def confidence_score(m: dict[str, Any]) -> float:
        b = m.get("confidence_bucket")
        if b == "high":
            return 100.0
        if b == "medium":
            return 70.0
        return 40.0

    def gap_risk_penalty(jeonse_ratio: float | None) -> float:
        if jeonse_ratio is None:
            return 0.0
        j = float(jeonse_ratio)
        penalty = 0.0
        if j >= 80.0:
            penalty += 20.0
        elif j >= 75.0:
            penalty += 12.0
        elif j >= 70.0:
            penalty += 6.0
        if j <= 30.0:
            penalty += 6.0
        elif j <= 35.0:
            penalty += 3.0
        return penalty

    def trend_expert(m: dict[str, Any]) -> float:
        s12 = percentile_rank(m.get("growth_12m"), growth12_all)
        srel = percentile_rank(m.get("rel_growth_12m"), rel_growth_all)
        sw = (m.get("sample_weight") or 0.0) * 100.0
        s_ext = percentile_rank(abs(float(m.get("growth_12m") or 0.0)), ext12_all, reverse=True)
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        return 0.42 * s12 + 0.26 * srel + 0.14 * sw + 0.10 * s_ext + 0.08 * s_fresh

    def value_expert(m: dict[str, Any]) -> float:
        s_jeonse = percentile_rank(m.get("latest_jeonse_ratio"), jeonse_all)
        s_yield = percentile_rank(m.get("latest_rent_yield"), yield_all)
        s_pvd = percentile_rank(m.get("price_vs_district"), pvd_all, reverse=True)
        s_pvs = percentile_rank(m.get("price_vs_seoul"), pvs_all, reverse=True)
        conf = confidence_score(m)
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        score = 0.28 * s_jeonse + 0.22 * s_yield + 0.18 * s_pvd + 0.10 * s_pvs + 0.12 * conf + 0.10 * s_fresh
        score -= gap_risk_penalty(m.get("latest_jeonse_ratio")) * 0.55
        return clamp(score, 0.0, 100.0)

    def liquidity_expert(m: dict[str, Any]) -> float:
        s_liq = percentile_rank(m.get("avg_trade_3m"), liq_all)
        s_latest = percentile_rank(float(m.get("latest_trade_count") or 0), liq_all)
        conf = confidence_score(m)
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        s_stale = percentile_rank(m.get("trade_stale_months"), stale_all, reverse=True)
        return 0.38 * s_liq + 0.24 * s_latest + 0.18 * s_fresh + 0.08 * s_stale + 0.12 * conf

    def stability_expert(m: dict[str, Any]) -> float:
        s_vol = percentile_rank(m.get("vol_ann"), vol_all, reverse=True)
        s_dd = percentile_rank(m.get("max_drawdown"), dd_all, reverse=True)
        dq = m.get("data_quality") or 0.0
        conf = confidence_score(m)
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        s_es = percentile_rank(m.get("tail_es_abs"), es_all, reverse=True)
        return 0.23 * s_vol + 0.21 * s_dd + 0.18 * dq + 0.14 * conf + 0.12 * s_fresh + 0.12 * s_es

    def policy_expert(m: dict[str, Any]) -> float:
        base = 45.0
        if m["district"] in HIGH_INTENSITY_LTPS:
            base -= 15.0
        base += clamp((m.get("avg_trade_3m") or 0) / 2.0, 0.0, 15.0)
        base += clamp((m.get("data_quality") or 0) / 25.0, 0.0, 5.0)
        base += clamp((m.get("market_freshness") or 0.0) * 12.0, 0.0, 12.0)
        base -= clamp(safe_div(float(m.get("trade_stale_months") or 0), 6.0), 0.0, 1.0) * 12.0
        base += clamp((1.0 - clamp(safe_div(float(m.get("tail_es_abs") or 0.0), 0.10), 0.0, 1.0)) * 5.0, 0.0, 5.0)
        base -= gap_risk_penalty(m.get("latest_jeonse_ratio")) * 0.40
        base -= (m.get("uncertainty_penalty") or 0.0) * 40.0
        return clamp(base, 0.0, 100.0)

    with cf.ThreadPoolExecutor(max_workers=5) as ex:
        f_trend = {ex.submit(trend_expert, m): m for m in metrics}
        f_value = {ex.submit(value_expert, m): m for m in metrics}
        f_liq = {ex.submit(liquidity_expert, m): m for m in metrics}
        f_stb = {ex.submit(stability_expert, m): m for m in metrics}
        f_pol = {ex.submit(policy_expert, m): m for m in metrics}

        for fut_map, key in [
            (f_trend, "trend_score"),
            (f_value, "value_score"),
            (f_liq, "liquidity_score"),
            (f_stb, "stability_score"),
            (f_pol, "policy_score"),
        ]:
            for fut in cf.as_completed(fut_map):
                m = fut_map[fut]
                try:
                    m[key] = round(float(fut.result()), 2)
                except Exception:  # noqa: BLE001
                    m[key] = 50.0

    return metrics


def score_by_profile(metrics: list[dict[str, Any]], profile: str) -> list[dict[str, Any]]:
    conf = PROFILE_CONFIGS[profile]
    w = conf["weights"]

    rows = []
    for m in metrics:
        total = (
            m["trend_score"] * w["trend"]
            + m["value_score"] * w["value"]
            + m["liquidity_score"] * w["liquidity"]
            + m["stability_score"] * w["stability"]
            + m["policy_score"] * w["policy"]
        ) / 100.0

        x = dict(m)
        x["profile"] = profile
        x["total_score"] = round(total, 2)
        rows.append(x)

    rows.sort(key=lambda r: r["total_score"], reverse=True)
    for i, r in enumerate(rows, 1):
        r["rank"] = i
    return rows


def score_complex_by_profile(metrics: list[dict[str, Any]], profile: str) -> list[dict[str, Any]]:
    conf = PROFILE_CONFIGS[profile]
    w = conf["weights"]

    rows = []
    for m in metrics:
        core = (
            m["trend_score"] * w["trend"]
            + m["value_score"] * w["value"]
            + m["liquidity_score"] * w["liquidity"]
            + m["stability_score"] * w["stability"]
            + m["policy_score"] * w["policy"]
        ) / 100.0

        # 단지 평점은 구(상위 레벨) 컨텍스트를 20% 섞어 과적합을 완화한다.
        district_tail = float(m.get("parent_district_score") or 50.0)
        total = 0.80 * core + 0.20 * district_tail
        total = total * (1.0 - (m.get("uncertainty_penalty") or 0.0))
        stale_months = float(m.get("trade_stale_months") or 0.0)
        recency_penalty = clamp(stale_months, 0.0, 6.0) * 1.2
        tail_penalty = clamp(safe_div(float(m.get("tail_es_abs") or 0.0), 0.10), 0.0, 1.0) * 2.5
        total -= recency_penalty + tail_penalty

        x = dict(m)
        x["profile"] = profile
        x["total_score"] = round(total, 2)
        x["core_score"] = round(core, 2)
        x["recency_penalty"] = round(recency_penalty, 2)
        x["tail_penalty"] = round(tail_penalty, 2)
        rows.append(x)

    rows.sort(key=lambda r: r["total_score"], reverse=True)
    for i, r in enumerate(rows, 1):
        r["rank"] = i

    # 구 내 순위 부여
    group: dict[str, int] = {}
    for r in rows:
        d = r["district"]
        group[d] = group.get(d, 0) + 1
        r["district_rank"] = group[d]
    return rows


def scenario_for_district(m: dict[str, Any], profile: str) -> dict[str, Any]:
    cfg = PROFILE_CONFIGS[profile]

    price = m.get("latest_price_eok") or 0.0
    if price <= 0:
        return {
            "bull_cum": None,
            "base_cum": None,
            "bear_cum": None,
            "bull_roi": None,
            "base_roi": None,
            "bear_roi": None,
            "horizon_years": cfg["horizon_years"],
        }

    h = int(cfg["horizon_years"])
    t_months = h * 12

    g12 = m.get("growth_12m")
    g3 = m.get("growth_3m")
    g12 = 0.02 if g12 is None else g12
    g3 = 0.005 if g3 is None else g3
    anchor_annual = clamp(0.60 * g12 + 0.40 * g3, -0.05, 0.09)
    anchor_monthly = (1.0 + anchor_annual) ** (1.0 / 12.0) - 1.0

    unc_pen = float(m.get("uncertainty_penalty") or 0.0)

    monthly_returns = [
        clamp(float(r), -0.25, 0.25)
        for r in (m.get("monthly_returns") or [])
        if r is not None
    ]
    mu_hist, sigma_hist = robust_mean_std(monthly_returns)
    vol_ann = m.get("vol_ann") or 0.08
    vol_from_ann = clamp(vol_ann, 0.04, 0.30) / math.sqrt(12.0)

    if monthly_returns:
        mu_m = 0.55 * mu_hist + 0.45 * anchor_monthly
        sigma_m = 0.65 * sigma_hist + 0.35 * vol_from_ann
    else:
        mu_m = anchor_monthly
        sigma_m = vol_from_ann

    sigma_m = sigma_m * (1.0 + unc_pen * 0.8)
    sigma_m = clamp(sigma_m, 0.008, 0.09)
    stale_m = float(m.get("trade_stale_months") or 0.0)
    tail_es_abs = float(m.get("tail_es_abs") or 0.0)
    jeonse = float(m.get("latest_jeonse_ratio") or 0.0)
    gap_risk = 0.0
    if jeonse >= 80.0:
        gap_risk = 1.0
    elif jeonse >= 75.0:
        gap_risk = 0.7
    elif jeonse >= 70.0:
        gap_risk = 0.4
    stale_pen = clamp(safe_div(stale_m, 6.0), 0.0, 1.0) * 0.0012

    if m["district"] in HIGH_INTENSITY_LTPS:
        regime_bias = -0.0008
        bear_extra = -0.0010
    else:
        regime_bias = -0.0003
        bear_extra = -0.0005

    # 과도한 낙관/비관 왜곡을 줄이기 위해 월수익률 분포를 robust하게 보정.
    bull_m = mu_m + 0.90 * sigma_m - regime_bias * 0.2
    base_m = mu_m - 0.10 * sigma_m + regime_bias - stale_pen * 0.5 - tail_es_abs * 0.02 - gap_risk * 0.0002
    bear_m = mu_m - 1.10 * sigma_m + regime_bias + bear_extra - stale_pen - unc_pen * 0.0015 - tail_es_abs * 0.08 - gap_risk * 0.0006

    def annual_from_monthly(r_m: float) -> float:
        return (1.0 + r_m) ** 12 - 1.0

    bull_a = clamp(annual_from_monthly(bull_m), -0.08, 0.18)
    base_a = clamp(annual_from_monthly(base_m), -0.07, 0.12)
    bear_a = clamp(annual_from_monthly(bear_m), -0.18, 0.06)

    bull_c = (1.0 + bull_a) ** h - 1.0
    base_c = (1.0 + base_a) ** h - 1.0
    bear_c = (1.0 + bear_a) ** h - 1.0

    buy_tax = cfg["buy_tax"]
    hold_cost = cfg["annual_hold_cost"]
    cg_tax = cfg["capital_gain_tax"]
    ltv = cfg["ltv"]
    loan_rate = cfg["loan_rate"]

    debt = price * ltv
    equity = price - debt
    carry_total = price * hold_cost * h

    rent_yield = clamp((m.get("latest_rent_yield") or 1.8) / 100.0, 0.008, 0.05)
    if profile == "실거주":
        annual_utility_yield = max(0.012, rent_yield * 0.60)
    elif profile == "투자자":
        annual_utility_yield = rent_yield * 0.85
    else:
        annual_utility_yield = rent_yield * 0.80
    utility_income_total = price * annual_utility_yield * h

    fresh = float(m.get("market_freshness") or 0.0)
    vol_norm = clamp((m.get("vol_ann") or 0.10) / 0.25, 0.0, 1.0)
    liq_norm = clamp((m.get("avg_trade_3m") or 0.0) / 12.0, 0.0, 1.0)
    yield_norm = clamp(rent_yield / 0.04, 0.0, 1.0)
    es_norm = clamp(safe_div(tail_es_abs, 0.10), 0.0, 1.0)
    interest_sensitivity = 0.75 + 0.30 * vol_norm + 0.25 * (1.0 - liq_norm) + 0.18 * (1.0 - yield_norm) + 0.18 * (1.0 - fresh) + 0.15 * unc_pen + 0.20 * es_norm + 0.10 * gap_risk
    interest_sensitivity = clamp(interest_sensitivity, 0.70, 1.90)

    def leveraged_roi(cum_ret: float, loan_rate_override: float | None = None) -> float:
        exit_price = price * (1.0 + cum_ret)
        gross_gain = exit_price - price
        tx = max(gross_gain, 0.0) * cg_tax
        r_loan = loan_rate if loan_rate_override is None else loan_rate_override
        interest_total = debt * r_loan * h * interest_sensitivity
        freshness_cost = price * (1.0 - fresh) * 0.0015 * h
        net_gain = (
            gross_gain
            + utility_income_total
            - tx
            - (price * buy_tax)
            - carry_total
            - interest_total
            - freshness_cost
        )
        if equity <= 0:
            return 0.0
        return clamp(net_gain / equity, -0.95, 3.00)

    shock_scale = 0.75 + 0.55 * interest_sensitivity + 0.35 * unc_pen + 0.35 * clamp(stale_m / 6.0, 0.0, 1.0)
    shock_scale += 0.25 * es_norm + 0.20 * gap_risk
    effective_rate_shock = DEFAULT_SCENARIO_RATE_SHOCK * clamp(shock_scale, 0.60, 2.20)
    stress_up_rate = loan_rate + effective_rate_shock
    stress_dn_rate = max(0.0, loan_rate - effective_rate_shock)

    return {
        "bull_cum": bull_c,
        "base_cum": base_c,
        "bear_cum": bear_c,
        "bull_roi": leveraged_roi(bull_c),
        "base_roi": leveraged_roi(base_c),
        "bear_roi": leveraged_roi(bear_c),
        "horizon_years": h,
        "annual_base": base_a,
        "annual_vol": sigma_m * math.sqrt(12.0),
        "annual_utility_yield": annual_utility_yield,
        "base_roi_rate_up": leveraged_roi(base_c, stress_up_rate),
        "base_roi_rate_down": leveraged_roi(base_c, stress_dn_rate),
        "confidence_bucket": m.get("confidence_bucket", "medium"),
        "assumption": {
            "buy_tax": buy_tax,
            "hold_cost": hold_cost,
            "capital_gain_tax": cg_tax,
            "ltv": ltv,
            "loan_rate": loan_rate,
            "loan_rate_up": stress_up_rate,
            "loan_rate_down": stress_dn_rate,
            "effective_rate_shock": effective_rate_shock,
            "interest_sensitivity": interest_sensitivity,
            "trade_stale_months": stale_m,
            "tail_es_abs": tail_es_abs,
            "gap_risk": gap_risk,
            "utility_income_total": utility_income_total,
            "t_months": t_months,
        },
    }


def attach_scenarios(rows: list[dict[str, Any]], profile: str) -> list[dict[str, Any]]:
    for r in rows:
        r["scenario"] = scenario_for_district(r, profile)
    return rows

# ---------------------------------------------------------------------------
# Workbook builders
# ---------------------------------------------------------------------------


def add_common_header(ws, title: str, subtitle: str, cols: int) -> int:
    r = write_title(ws, 1, title, cols)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    c = ws.cell(row=r, column=1, value=subtitle)
    c.font = FONT_SMALL
    c.alignment = ALIGN_LEFT
    c.fill = FILL_LIGHT
    c.border = BORDER
    ws.row_dimensions[r].height = 20
    return r + 1


def add_sheet_dashboard(wb, rows: list[dict[str, Any]], profile: str, months: list[str]):
    ws = wb.create_sheet("종합 대시보드")
    r = add_common_header(
        ws,
        f"서울 25개구 아파트 심층분석 대시보드 ({profile})",
        f"기준: {months[0]}~{months[-1]} / 면적필터 59~99m² / 생성시각 {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}",
        10,
    )

    top5 = rows[:5]
    bot5 = rows[-5:]

    r = write_row(ws, r, ["[핵심 가정]"] + [""] * 9, fill=FILL_GOLD, left=True)
    assumptions = [
        "1) 월별 중앙값 기반(거래 outlier 완화), 구/단지 2단계 비교",
        "2) 낙관/중립/비관은 P80/P50/P20 분위수 기반 확률 시나리오",
        "3) 레버리지 ROI는 프로필별 LTV/금리/세율/보유비용 반영",
        "4) 서울 LTPS 규제와 4개 집중규제구(강남·서초·송파·용산) 정책 패널티 반영",
        "5) 데이터 품질점수(표본수+커버리지) 낮으면 점수 자동 보수화",
    ]
    for line in assumptions:
        r = write_row(ws, r, [line] + [""] * 9, left=True)

    r += 1
    r = write_row(ws, r, ["[TOP 5 추천]"] + [""] * 9, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["순위", "구", "총점", "12M성장", "전세가율", "변동성", "거래량(최근)", "낙관ROI", "중립ROI", "비관ROI"])
    for x in top5:
        sc = x["scenario"]
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["total_score"],
                pct(x.get("growth_12m")),
                f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                x.get("latest_trade_count") or 0,
                pct(sc.get("bull_roi")),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
            ],
        )

    r += 1
    r = write_row(ws, r, ["[BOTTOM 5 (관망/주의)]"] + [""] * 9, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["순위", "구", "총점", "12M성장", "전세가율", "변동성", "거래량(최근)", "낙관ROI", "중립ROI", "비관ROI"])
    for x in bot5:
        sc = x["scenario"]
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["total_score"],
                pct(x.get("growth_12m")),
                f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                x.get("latest_trade_count") or 0,
                pct(sc.get("bull_roi")),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
            ],
        )

    set_widths(ws, [8, 13, 9, 11, 11, 10, 12, 11, 11, 11])


def add_sheet_scorecard(wb, rows: list[dict[str, Any]], profile: str, months: list[str]):
    ws = wb.create_sheet("서울25구 스코어카드")
    r = add_common_header(ws, f"서울 25개구 스코어카드 ({profile})", f"분석기간: {months[0]}~{months[-1]}", 16)
    headers = [
        "순위", "구", "총점", "Trend", "Value", "Liquidity", "Stability", "Policy", "DataQ",
        "최근매매(억)", "최근전세(억)", "12M성장", "전세가율", "변동성", "최근거래수", "거래공백월",
    ]
    r = write_headers(ws, r, headers)
    for x in rows:
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["total_score"],
                x["trend_score"],
                x["value_score"],
                x["liquidity_score"],
                x["stability_score"],
                x["policy_score"],
                x["data_quality"],
                round(x["latest_price_eok"], 2) if x.get("latest_price_eok") else "-",
                round(x["latest_rent_eok"], 2) if x.get("latest_rent_eok") else "-",
                pct(x.get("growth_12m")),
                f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                x.get("latest_trade_count") or 0,
                x.get("trade_stale_months") or 0,
            ],
        )
        ws.cell(row=r - 1, column=3).fill = score_band_fill(float(x["total_score"]))

    set_widths(ws, [6, 12, 8, 8, 8, 9, 9, 8, 8, 10, 10, 10, 9, 9, 9, 10])


def add_sheet_coverage(wb, rows: list[dict[str, Any]], months: list[str]):
    ws = wb.create_sheet("데이터 커버리지")
    r = add_common_header(ws, "데이터 품질/커버리지", "표본수와 월별 커버리지를 점수화하여 신뢰도 반영", 13)
    headers = [
        "구", "분석월수", "매매커버리지", "전월세커버리지", "총매매표본", "총전월세표본", "최근매매표본", "최근전월세표본", "평균거래(3M)", "최종거래월", "거래공백월", "DataQ", "비고"
    ]
    r = write_headers(ws, r, headers)
    for x in rows:
        note = "양호"
        if x["data_quality"] < 55:
            note = "주의(표본 부족)"
        elif x["data_quality"] < 70:
            note = "보통"
        r = write_row(
            ws,
            r,
            [
                x["district"],
                len(months),
                f"{x['coverage_trade']*100:.1f}%",
                f"{x['coverage_rent']*100:.1f}%",
                x["total_trade_samples"],
                x["total_rent_samples"],
                x["latest_trade_count"],
                x["latest_rent_count"],
                round(x["avg_trade_3m"], 1),
                x.get("last_trade_yyyymm") or "-",
                x.get("trade_stale_months") or 0,
                x["data_quality"],
                note,
            ],
        )
        ws.cell(row=r - 1, column=12).fill = score_band_fill(float(x["data_quality"]))
    set_widths(ws, [12, 9, 11, 12, 11, 12, 11, 12, 11, 10, 10, 8, 16])


def add_sheet_trade_trend(wb, rows: list[dict[str, Any]], months: list[str]):
    ws = wb.create_sheet("매매 추세")
    ncol = 1 + len(months)
    r = add_common_header(ws, "구별 매매 중앙값 추세(억)", "면적필터 59~99m², 월별 중앙값", ncol)
    r = write_headers(ws, r, ["구"] + months)
    for x in rows:
        row_vals = [x["district"]]
        month_map = {m["year_month"]: m for m in x["months"]}
        for ym in months:
            m = month_map.get(ym, {})
            if m.get("median_trade_10k"):
                row_vals.append(round(m["median_trade_10k"] / 10000.0, 2))
            else:
                row_vals.append(None)
        r = write_row(ws, r, row_vals)
    set_widths(ws, [12] + [9] * len(months))


def add_sheet_rent_trend(wb, rows: list[dict[str, Any]], months: list[str]):
    ws = wb.create_sheet("전세 추세")
    ncol = 1 + len(months)
    r = add_common_header(ws, "구별 전세 중앙값 추세(억)", "면적필터 59~99m², 월별 중앙값", ncol)
    r = write_headers(ws, r, ["구"] + months)
    for x in rows:
        row_vals = [x["district"]]
        month_map = {m["year_month"]: m for m in x["months"]}
        for ym in months:
            m = month_map.get(ym, {})
            if m.get("median_rent_10k"):
                row_vals.append(round(m["median_rent_10k"] / 10000.0, 2))
            else:
                row_vals.append(None)
        r = write_row(ws, r, row_vals)

    r += 1
    r = write_row(ws, r, ["[최근 전세가율/환산수익률]"] + [""] * (ncol - 1), fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["구", "전세가율(%)", "전세환산 수익률(%)"])
    for x in rows:
        r = write_row(
            ws,
            r,
            [
                x["district"],
                round(x["latest_jeonse_ratio"], 2) if x.get("latest_jeonse_ratio") else "-",
                round(x["latest_rent_yield"], 2) if x.get("latest_rent_yield") else "-",
            ],
        )
    set_widths(ws, [12] + [9] * len(months))


def add_sheet_value(wb, rows: list[dict[str, Any]]):
    ws = wb.create_sheet("밸류에이션")
    r = add_common_header(ws, "상대가치 분석", "서울 중앙값 대비 가격수준/전세가율/전세환산수익률", 10)
    headers = ["구", "최근매매(억)", "최근전세(억)", "서울대비가격배수", "전세가율(%)", "전세환산수익률(%)", "Value점수", "해석", "위험", "메모"]
    r = write_headers(ws, r, headers)

    for x in rows:
        pvs = x.get("price_vs_seoul")
        label = "중립"
        if pvs is not None and pvs < 0.9:
            label = "상대저평가"
        elif pvs is not None and pvs > 1.15:
            label = "상대고평가"

        risk = "보통"
        if (x.get("latest_jeonse_ratio") or 0) > 65:
            risk = "갭리스크"
        if (x.get("latest_jeonse_ratio") or 0) < 30:
            risk = "현금부담"

        r = write_row(
            ws,
            r,
            [
                x["district"],
                round(x["latest_price_eok"], 2) if x.get("latest_price_eok") else "-",
                round(x["latest_rent_eok"], 2) if x.get("latest_rent_eok") else "-",
                round(pvs, 3) if pvs else "-",
                round(x["latest_jeonse_ratio"], 2) if x.get("latest_jeonse_ratio") else "-",
                round(x["latest_rent_yield"], 2) if x.get("latest_rent_yield") else "-",
                x["value_score"],
                label,
                risk,
                "",
            ],
        )
        ws.cell(row=r - 1, column=7).fill = score_band_fill(float(x["value_score"]))

    set_widths(ws, [12, 10, 10, 12, 11, 14, 9, 12, 10, 12])


def add_sheet_liquidity(wb, rows: list[dict[str, Any]]):
    ws = wb.create_sheet("유동성")
    r = add_common_header(ws, "유동성 분석", "최근 거래량/3개월 평균 거래량 중심", 9)
    r = write_headers(ws, r, ["구", "최근매매표본", "3M평균표본", "최근전월세표본", "Liquidity점수", "정책충격흡수", "체결난이도", "권고", "비고"])

    for x in rows:
        liq_score = x["liquidity_score"]
        absorb = "높음" if liq_score >= 70 else "보통" if liq_score >= 55 else "낮음"
        exec_cost = "낮음" if liq_score >= 70 else "보통" if liq_score >= 50 else "높음"
        rec = "분할진입" if liq_score >= 70 else "선별진입" if liq_score >= 55 else "관망"

        r = write_row(
            ws,
            r,
            [
                x["district"],
                x["latest_trade_count"],
                round(x["avg_trade_3m"], 1),
                x["latest_rent_count"],
                liq_score,
                absorb,
                exec_cost,
                rec,
                "",
            ],
        )
        ws.cell(row=r - 1, column=5).fill = score_band_fill(float(liq_score))

    set_widths(ws, [12, 11, 11, 12, 10, 12, 10, 10, 12])


def add_sheet_risk(wb, rows: list[dict[str, Any]]):
    ws = wb.create_sheet("변동성리스크")
    r = add_common_header(ws, "변동성/낙폭 리스크", "변동성·낙폭·꼬리위험(ES)으로 하방위험 측정", 11)
    r = write_headers(ws, r, ["구", "12M성장", "3M성장", "변동성(연)", "최대낙폭", "월ES(20%)", "Stability점수", "Policy점수", "종합리스크", "LTPS강도", "코멘트"])

    for x in rows:
        risk_index = 100.0 - (0.65 * x["stability_score"] + 0.35 * x["policy_score"])
        if risk_index <= 35:
            bucket = "낮음"
        elif risk_index <= 55:
            bucket = "보통"
        else:
            bucket = "높음"
        intensity = "집중규제" if x["district"] in HIGH_INTENSITY_LTPS else "일반규제"

        r = write_row(
            ws,
            r,
            [
                x["district"],
                pct(x.get("growth_12m")),
                pct(x.get("growth_3m")),
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                f"{(x.get('max_drawdown') or 0)*100:.1f}%",
                f"{(x.get('tail_es_m20') or 0)*100:.1f}%",
                x["stability_score"],
                x["policy_score"],
                round(risk_index, 2),
                intensity,
                bucket,
            ],
        )
        ws.cell(row=r - 1, column=9).fill = score_band_fill(100 - risk_index)

    set_widths(ws, [12, 10, 10, 10, 10, 10, 11, 10, 10, 11, 10])


def add_sheet_benchmark(wb, rows: list[dict[str, Any]]):
    ws = wb.create_sheet("벤치마크")
    r = add_common_header(ws, "벤치마크 및 상대비교", "서울 중앙값/상하위 20% 비교", 10)

    # 상위/하위 20%
    n = len(rows)
    k = max(1, int(round(n * 0.2)))
    top = rows[:k]
    bottom = rows[-k:]

    def avg(items: list[dict[str, Any]], key: str) -> float:
        vals = [float(x[key]) for x in items if x.get(key) is not None]
        return statistics.mean(vals) if vals else 0.0

    r = write_row(ws, r, ["[요약]"] + [""] * 9, fill=FILL_GOLD, left=True)
    summary_rows = [
        ["그룹", "총점", "12M성장", "전세가율", "변동성", "거래량", "Trend", "Value", "Liquidity", "Stability"],
        ["상위20%", round(avg(top, "total_score"), 2), pct(avg(top, "growth_12m")), f"{avg(top, 'latest_jeonse_ratio'):.1f}%", f"{avg(top, 'vol_ann')*100:.1f}%", round(avg(top, "avg_trade_3m"), 1), round(avg(top, "trend_score"), 1), round(avg(top, "value_score"), 1), round(avg(top, "liquidity_score"), 1), round(avg(top, "stability_score"), 1)],
        ["하위20%", round(avg(bottom, "total_score"), 2), pct(avg(bottom, "growth_12m")), f"{avg(bottom, 'latest_jeonse_ratio'):.1f}%", f"{avg(bottom, 'vol_ann')*100:.1f}%", round(avg(bottom, "avg_trade_3m"), 1), round(avg(bottom, "trend_score"), 1), round(avg(bottom, "value_score"), 1), round(avg(bottom, "liquidity_score"), 1), round(avg(bottom, "stability_score"), 1)],
    ]
    r = write_headers(ws, r, summary_rows[0])
    r = write_row(ws, r, summary_rows[1])
    r = write_row(ws, r, summary_rows[2])

    r += 1
    r = write_row(ws, r, ["[강남3구+용산 vs 기타구]"] + [""] * 9, fill=FILL_GOLD, left=True)
    core_names = {"강남구", "서초구", "송파구", "용산구"}
    core = [x for x in rows if x["district"] in core_names]
    non_core = [x for x in rows if x["district"] not in core_names]
    r = write_headers(ws, r, ["그룹", "평균총점", "평균매매(억)", "12M성장", "전세가율", "변동성", "평균거래량", "정책점수", "리스크해석", "비고"])
    for nm, grp in [("강남3구+용산", core), ("기타21구", non_core)]:
        risk = "정책민감" if nm == "강남3구+용산" else "분산형"
        r = write_row(
            ws,
            r,
            [
                nm,
                round(avg(grp, "total_score"), 2),
                round(avg(grp, "latest_price_eok"), 2),
                pct(avg(grp, "growth_12m")),
                f"{avg(grp, 'latest_jeonse_ratio'):.1f}%",
                f"{avg(grp, 'vol_ann')*100:.1f}%",
                round(avg(grp, "avg_trade_3m"), 1),
                round(avg(grp, "policy_score"), 1),
                risk,
                "",
            ],
        )

    set_widths(ws, [14, 11, 12, 10, 10, 10, 10, 10, 12, 10])


def compute_walkforward_backtest(
    rows: list[dict[str, Any]],
    months: list[str],
    lookback: int = DEFAULT_BACKTEST_LOOKBACK,
    horizon: int = DEFAULT_BACKTEST_HORIZON,
    top_fraction: float = DEFAULT_BACKTEST_TOP_FRACTION,
) -> dict[str, Any]:
    if len(months) < lookback + horizon:
        return {"records": [], "config": None}

    slices: list[dict[str, Any]] = []
    for t in range(lookback - 1, len(months) - horizon):
        signal_rows: list[dict[str, Any]] = []
        all_forward: list[float] = []

        for r in rows:
            month_map = {m["year_month"]: m for m in r.get("months", [])}
            prices: list[float | None] = []
            counts: list[int] = []
            jeonse: list[float] = []
            yields: list[float] = []
            for ym in months:
                m = month_map.get(ym, {})
                p10k = m.get("median_trade_10k")
                prices.append((p10k / 10000.0) if p10k else None)
                counts.append(int(m.get("trade_count") or 0))
                jeonse.append(float(m.get("jeonse_ratio_pct") or 0.0))
                yields.append(float(m.get("rent_yield_proxy_pct") or 0.0))

            p_now = prices[t]
            p_ref = prices[t - lookback + 1]
            p_fwd = prices[t + horizon]
            if p_now is None or p_ref is None or p_fwd is None or p_ref <= 0 or p_now <= 0:
                continue

            momentum = p_now / p_ref - 1.0
            liq = statistics.mean(counts[max(0, t - lookback + 1): t + 1]) if counts else 0.0
            freshness = 1.0 if counts[t] > 0 else 0.0
            val = jeonse[t]
            yld = yields[t]
            rets: list[float] = []
            for i in range(max(1, t - lookback + 1), t + 1):
                p0 = prices[i - 1]
                p1 = prices[i]
                if p0 is None or p1 is None or p0 <= 0:
                    continue
                rets.append(math.log(p1 / p0))
            lowvol = stdev(rets) if len(rets) > 1 else 0.0
            fwd_ret = p_fwd / p_now - 1.0
            all_forward.append(fwd_ret)
            signal_rows.append(
                {
                    "district": r["district"],
                    "momentum": momentum,
                    "liq": liq,
                    "freshness": freshness,
                    "val": val,
                    "yld": yld,
                    "lowvol": lowvol,
                    "fwd_ret": fwd_ret,
                }
            )

        if len(signal_rows) < 8 or not all_forward:
            continue
        slices.append({"month": months[t], "rows": signal_rows, "all_forward": all_forward})

    if not slices:
        return {"records": [], "config": None}

    def eval_cfg(weights: dict[str, float], frac: float) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        recs: list[dict[str, Any]] = []
        cum_s = 1.0
        cum_b = 1.0
        for s in slices:
            rows_i = [dict(x) for x in s["rows"]]
            moms = [x["momentum"] for x in rows_i]
            liqs = [x["liq"] for x in rows_i]
            freshes = [x["freshness"] for x in rows_i]
            vals = [x["val"] for x in rows_i]
            ylds = [x["yld"] for x in rows_i]
            vols = [x["lowvol"] for x in rows_i]

            for x in rows_i:
                signal = 0.0
                if weights.get("mom"):
                    signal += weights["mom"] * percentile_rank(x["momentum"], moms)
                if weights.get("liq"):
                    signal += weights["liq"] * percentile_rank(x["liq"], liqs)
                if weights.get("fresh"):
                    signal += weights["fresh"] * percentile_rank(x["freshness"], freshes)
                if weights.get("val"):
                    signal += weights["val"] * percentile_rank(x["val"], vals)
                if weights.get("yld"):
                    signal += weights["yld"] * percentile_rank(x["yld"], ylds)
                if weights.get("lowvol"):
                    signal += weights["lowvol"] * percentile_rank(x["lowvol"], vols, reverse=True)
                x["signal"] = signal

            rows_i.sort(key=lambda x: x["signal"], reverse=True)
            k = max(3, int(round(len(rows_i) * frac)))
            k = min(k, len(rows_i))
            top = rows_i[:k]
            strategy = statistics.mean([x["fwd_ret"] for x in top]) if top else 0.0
            bench = statistics.mean(s["all_forward"]) if s["all_forward"] else 0.0
            excess = strategy - bench
            cum_s *= 1.0 + strategy
            cum_b *= 1.0 + bench
            recs.append(
                {
                    "rebalance_month": s["month"],
                    "strategy_ret": strategy,
                    "benchmark_ret": bench,
                    "excess_ret": excess,
                    "hit": 1 if excess > 0 else 0,
                    "sample_count": len(top),
                    "top_districts": ", ".join(x["district"] for x in top[:5]),
                }
            )
        avg_ex = statistics.mean([x["excess_ret"] for x in recs]) if recs else -1.0
        hit = safe_div(sum(x["hit"] for x in recs), float(len(recs))) if recs else 0.0
        return recs, {
            "avg_excess": avg_ex,
            "hit_ratio": hit,
            "cum_excess": cum_s - cum_b,
            "top_fraction": frac,
            "weights": weights,
            "samples": len(recs),
        }

    best_records: list[dict[str, Any]] = []
    best_cfg: dict[str, Any] | None = None
    candidates_tested = 0
    for weights in BACKTEST_SIGNAL_CANDIDATES:
        for frac in BACKTEST_TOP_FRACTION_CANDIDATES:
            candidates_tested += 1
            recs, cfg = eval_cfg(weights, frac)
            if not recs:
                continue
            if best_cfg is None:
                best_cfg = cfg
                best_records = recs
                continue
            lhs = (cfg["avg_excess"], cfg["hit_ratio"], cfg["cum_excess"])
            rhs = (best_cfg["avg_excess"], best_cfg["hit_ratio"], best_cfg["cum_excess"])
            if lhs > rhs:
                best_cfg = cfg
                best_records = recs

    if best_cfg is None:
        return {"records": [], "config": None}

    return {
        "records": best_records,
        "config": {
            **best_cfg,
            "lookback": lookback,
            "horizon": horizon,
            "candidates_tested": candidates_tested,
        },
    }


def add_sheet_backtest(
    wb,
    rows: list[dict[str, Any]],
    months: list[str],
    lookback: int = DEFAULT_BACKTEST_LOOKBACK,
    horizon: int = DEFAULT_BACKTEST_HORIZON,
    top_fraction: float = DEFAULT_BACKTEST_TOP_FRACTION,
):
    ws = wb.create_sheet("전략 백테스트")
    r = add_common_header(
        ws,
        "워크포워드 백테스트(구 단위)",
        f"lookback {lookback}M / forward {horizon}M / 다중신호 자동튜닝 전략",
        10,
    )
    r = write_headers(
        ws,
        r,
        ["리밸런스월", "전략수익", "벤치마크수익", "초과수익", "적중", "표본수", "누적전략", "누적벤치", "누적초과", "Top구(일부)"],
    )

    bt_obj = compute_walkforward_backtest(rows, months, lookback=lookback, horizon=horizon, top_fraction=top_fraction)
    bt = bt_obj.get("records", [])
    cfg = bt_obj.get("config")
    if not bt:
        write_row(ws, r, ["백테스트 표본이 부족합니다."] + [""] * 9, left=True)
        set_widths(ws, [10, 10, 12, 10, 6, 7, 10, 10, 10, 36])
        return

    cum_s = 1.0
    cum_b = 1.0
    for x in bt:
        cum_s *= 1.0 + x["strategy_ret"]
        cum_b *= 1.0 + x["benchmark_ret"]
        r = write_row(
            ws,
            r,
            [
                x["rebalance_month"],
                pct(x["strategy_ret"]),
                pct(x["benchmark_ret"]),
                pct(x["excess_ret"]),
                "Y" if x["hit"] else "N",
                x["sample_count"],
                pct(cum_s - 1.0),
                pct(cum_b - 1.0),
                pct((cum_s - cum_b)),
                x["top_districts"],
            ],
            left=True,
        )

    hit_ratio = safe_div(sum(x["hit"] for x in bt), float(len(bt)))
    avg_excess = statistics.mean([x["excess_ret"] for x in bt]) if bt else 0.0
    r += 1
    r = write_row(ws, r, ["[요약]"] + [""] * 9, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["평가구간", "횟수", "적중률", "평균초과수익", "누적전략", "누적벤치", "누적초과", "", "", ""])
    r = write_row(
        ws,
        r,
        [
            f"{bt[0]['rebalance_month']}~{bt[-1]['rebalance_month']}",
            len(bt),
            pct(hit_ratio),
            pct(avg_excess),
            pct(cum_s - 1.0),
            pct(cum_b - 1.0),
            pct(cum_s - cum_b),
            "",
            "",
            "",
        ],
    )
    if cfg:
        r += 1
        r = write_row(ws, r, ["[튜닝 설정]"] + [""] * 9, fill=FILL_GOLD, left=True)
        r = write_headers(ws, r, ["lookback", "horizon", "Top비중", "가중치", "탐색조합", "평균초과", "적중률", "누적초과", "", ""])
        write_row(
            ws,
            r,
            [
                cfg.get("lookback"),
                cfg.get("horizon"),
                f"{(cfg.get('top_fraction') or 0)*100:.0f}%",
                ", ".join(f"{k}:{v:.2f}" for k, v in (cfg.get("weights") or {}).items()),
                cfg.get("candidates_tested"),
                pct(cfg.get("avg_excess")),
                pct(cfg.get("hit_ratio")),
                pct(cfg.get("cum_excess")),
                "",
                "",
            ],
            left=True,
        )
    set_widths(ws, [10, 10, 12, 10, 6, 7, 10, 10, 10, 36])


def add_sheet_scenario(wb, rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("시나리오 분석")
    r = add_common_header(ws, f"확률기반 시나리오 분석 ({profile})", "P80/P50/P20 분위수 기반 누적수익률 + 레버리지 ROI", 12)
    r = write_headers(ws, r, ["순위", "구", "현재매매(억)", "기간(년)", "낙관누적", "중립누적", "비관누적", "낙관ROI", "중립ROI", "비관ROI", "기준연변동성", "비고"])

    for x in rows:
        sc = x["scenario"]
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                round(x["latest_price_eok"], 2) if x.get("latest_price_eok") else "-",
                sc.get("horizon_years"),
                pct(sc.get("bull_cum")),
                pct(sc.get("base_cum")),
                pct(sc.get("bear_cum")),
                pct(sc.get("bull_roi")),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
                f"{(sc.get('annual_vol') or 0)*100:.1f}%",
                f"보유수익률 {((sc.get('annual_utility_yield') or 0)*100):.2f}%",
            ],
        )

    set_widths(ws, [7, 12, 11, 8, 10, 10, 10, 10, 10, 10, 11, 12])


def add_sheet_scenario_method(wb, profile: str):
    ws = wb.create_sheet("시나리오 방법론")
    r = add_common_header(ws, "시나리오 산정 방법론", "현실반영형 확률모형 + 정책/레버리지/세금 보정", 6)

    steps = [
        "1) 월별 매매중앙값 시계열로 12M/3M 성장률 및 연환산 변동성 산출",
        "2) 기초성장률 = 0.60*12M + 0.40*3M (상·하한 clipping으로 과최적화 방지)",
        "3) 월수익률은 winsorize(10%) + robust 평균/표준편차로 추정",
        "4) 월 하방꼬리위험(ES 20%)을 stability/시나리오 하방에 반영",
        "5) 정책/신선도 레짐(집중규제·거래공백월) 보정으로 상하방 비대칭 반영",
        "6) 단지 성장률은 구 성장률로 shrinkage(표본가중)해 노이즈 축소",
        "7) 단지 레벨 총점은 구 점수(거시) 20% + 단지 점수(미시) 80% 결합",
        "8) 전세가율 과열구간(갭리스크) 패널티를 value/policy/시나리오에 반영",
        "9) 프로필별 LTV/금리/세금/보유비용으로 레버리지 ROI 계산",
        "10) 금리충격은 고정 ±1%p가 아니라 단지별 가변충격폭으로 계산",
        "11) 워크포워드(lookback/forward 자동튜닝)로 전략 초과수익을 점검",
    ]

    r = write_row(ws, r, ["[방법론 단계]"] + [""] * 5, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["단계", "설명", "수식/로직", "현실보정", "한계", "완화장치"])

    rows = [
        ["Step1", "기초통계", "median(price_t), median(rent_t)", "이상치 완화", "표본부족", "DataQ 경고"],
        ["Step2", "추세 산출", "g_base=0.6*g12+0.4*g3", "최근/중기 균형", "최근왜곡", "clipping"],
        ["Step3", "강건분포", "winsorize + robust(μ,σ)", "꼬리값 완화", "구조변화", "주기 재추정"],
        ["Step4", "꼬리위험", "ES20 = avg(bottom20% monthly return)", "하방리스크 반영", "표본 민감", "최소표본 필터"],
        ["Step5", "레짐보정", "R'=R+regime(stale/policy)", "현실 충격 반영", "정책변화", "월별 업데이트"],
        ["Step6", "노이즈축소", "g_adj=w*g+(1-w)*g_district", "표본작은 단지 안정화", "과도한 평활", "w 상한/하한"],
        ["Step7", "재무반영", "ROI=(NetGain-이자-세금)/자기자본", "투자자별 현실성", "세법변경", "가정표 분리"],
        ["Step8", "민감도", "ROI(r±shock_i)", "단지별 금리충격", "단순 모수", "충격폭 가변화"],
    ]
    for x in rows:
        r = write_row(ws, r, x, left=True)

    r += 1
    r = write_row(ws, r, ["[프로필 설정]"] + [""] * 5, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["프로필", "핵심목표", "LTV", "대출금리", "취득세", "양도세"])
    for p, cfg in PROFILE_CONFIGS.items():
        r = write_row(
            ws,
            r,
            [
                p,
                cfg["display"],
                f"{cfg['ltv']*100:.0f}%",
                f"{cfg['loan_rate']*100:.2f}%",
                f"{cfg['buy_tax']*100:.2f}%",
                f"{cfg['capital_gain_tax']*100:.1f}%",
            ],
            left=True,
        )

    r += 1
    r = write_row(ws, r, ["[현재 생성 프로필]", profile, "", "", "", ""], fill=FILL_LIGHT, left=True)

    set_widths(ws, [11, 37, 20, 16, 14, 14])


def add_sheet_profile_weights(wb):
    ws = wb.create_sheet("프로필가중치")
    r = add_common_header(ws, "프로필별 멀티전문가 가중치", "실거주/투자자/법인 전략차이를 수치화", 8)
    r = write_headers(ws, r, ["프로필", "Trend", "Value", "Liquidity", "Stability", "Policy", "합계", "전략요약"])

    for p, cfg in PROFILE_CONFIGS.items():
        w = cfg["weights"]
        r = write_row(
            ws,
            r,
            [
                p,
                w["trend"],
                w["value"],
                w["liquidity"],
                w["stability"],
                w["policy"],
                sum(w.values()),
                cfg["display"],
            ],
            left=True,
        )

    set_widths(ws, [10, 8, 8, 10, 10, 8, 8, 32])


def add_sheet_tax_leverage(wb):
    ws = wb.create_sheet("세금레버리지가정")
    r = add_common_header(ws, "세금/레버리지 가정", "프로필별 보수적 가정치", 9)
    r = write_headers(ws, r, ["프로필", "LTV", "대출금리", "취득세", "연보유비용", "양도세", "투자기간(년)", "의미", "주의"])

    notes = {
        "실거주": "거주효용 포함, 세후수익보다 안정성 우선",
        "투자자": "현금흐름·회전율·세후ROI 균형",
        "법인": "세율/취득세 부담이 커 고수익·고확실성 지역 선호",
    }
    warns = {
        "실거주": "거주의무/이주비용 누락 금지",
        "투자자": "DSR/금리변동에 민감",
        "법인": "취득세 12%와 세후수익 급감 리스크",
    }

    for p, cfg in PROFILE_CONFIGS.items():
        r = write_row(
            ws,
            r,
            [
                p,
                f"{cfg['ltv']*100:.0f}%",
                f"{cfg['loan_rate']*100:.2f}%",
                f"{cfg['buy_tax']*100:.2f}%",
                f"{cfg['annual_hold_cost']*100:.2f}%",
                f"{cfg['capital_gain_tax']*100:.1f}%",
                cfg["horizon_years"],
                notes[p],
                warns[p],
            ],
            left=True,
        )

    set_widths(ws, [10, 8, 10, 9, 11, 8, 10, 28, 24])


def add_sheet_policy(wb, rows: list[dict[str, Any]]):
    ws = wb.create_sheet("정책규제체크")
    r = add_common_header(ws, "정책/규제 체크", "LTPS·금리·DSR 최신 공개자료 반영", 9)

    r = write_row(ws, r, ["[정책 이벤트]"] + [""] * 8, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["구분", "내용", "기준일", "시장영향", "점수반영", "리스크", "출처", "링크", "비고"])

    policy_rows = [
        ["LTPS", "서울시 전체 아파트 토지거래허가구역", "2025-10-20~2026-12-31", "유동성/레버리지 제약", "Policy score base=45", "중", "서울시", "https://land.seoul.go.kr/land/other/appointStatusSeoul.do", ""],
        ["LTPS", "강남·서초·송파·용산 집중규제", "2026-01-07 고시", "상방축소/하방확대", "집중규제구 -15", "중~상", "서울시", "https://land.seoul.go.kr/land/other/appointStatusSeoul.do", ""],
        ["금리", "한국은행 기준금리 2.50% 동결", "2026-01-15", "차입비용 기준선", "loan_rate anchor", "중", "한국은행", "https://www.bok.or.kr/portal/bbs/B0000169/view.do?menuNo=200059&nttId=10095737", ""],
        ["대출", "3단계 스트레스 DSR", "2025-07 시행", "대출가능액 축소", "LTV/금리 보수화", "중", "금융위원회", "https://www.fsc.go.kr/no010101/84617", ""],
    ]
    for p in policy_rows:
        r = write_row(ws, r, p, left=True)

    r += 1
    r = write_row(ws, r, ["[구별 규제강도 매핑]"] + [""] * 8, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["구", "정책점수", "강도", "집중규제", "유동성보정", "데이터품질", "권고", "메모", ""])
    for x in rows:
        intensity = "집중" if x["district"] in HIGH_INTENSITY_LTPS else "일반"
        rec = "보수진입" if intensity == "집중" else "분할검토"
        r = write_row(
            ws,
            r,
            [
                x["district"],
                x["policy_score"],
                intensity,
                "Y" if intensity == "집중" else "N",
                round(x["avg_trade_3m"], 1),
                x["data_quality"],
                rec,
                "",
                "",
            ],
        )

    set_widths(ws, [10, 9, 8, 9, 10, 10, 12, 15, 8])


def add_sheet_timing(wb, rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("매수타이밍신호")
    r = add_common_header(ws, f"매수 타이밍 신호 ({profile})", "추세·유동성·안정성 결합신호", 11)
    r = write_headers(ws, r, ["순위", "구", "타이밍점수", "Trend", "Liquidity", "Stability", "최근3M", "12M", "신호", "전략", "주의"])

    for x in rows:
        fresh_bonus = (x.get("market_freshness") or 0.0) * 8.0
        stale_penalty = (x.get("trade_stale_months") or 0.0) * 1.2
        timing = 0.50 * x["trend_score"] + 0.30 * x["liquidity_score"] + 0.20 * x["stability_score"] + fresh_bonus - stale_penalty
        if timing >= 70:
            sig = "진입 우호"
            strat = "3회 분할매수"
        elif timing >= 55:
            sig = "중립"
            strat = "조건부 진입"
        else:
            sig = "관망"
            strat = "지표 개선 대기"

        caution = "정책민감" if x["district"] in HIGH_INTENSITY_LTPS else "일반"
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                round(timing, 2),
                x["trend_score"],
                x["liquidity_score"],
                x["stability_score"],
                pct(x.get("growth_3m")),
                pct(x.get("growth_12m")),
                sig,
                strat,
                caution,
            ],
        )
        ws.cell(row=r - 1, column=3).fill = score_band_fill(float(timing))

    set_widths(ws, [7, 11, 10, 8, 9, 9, 9, 9, 10, 12, 9])


def add_sheet_complex_overview(
    wb,
    complex_rows: list[dict[str, Any]],
    district_rows: list[dict[str, Any]],
    top_per_district: int,
):
    ws = wb.create_sheet("구내상위단지 요약")
    r = add_common_header(
        ws,
        "구 내부 상위단지 요약",
        f"구별 상위 {top_per_district}개 단지의 평균 점수/수익률 비교",
        11,
    )
    r = write_headers(
        ws,
        r,
        ["구", "구점수", "단지수(분석)", f"Top{top_per_district} 평균점수", f"Top{top_per_district} 평균중립ROI", f"Top{top_per_district} 평균비관ROI", "Top 단지명", "Top 단지점수", "Top단지 중립ROI", "전세가율", "비고"],
    )

    dmap = {d["district"]: d for d in district_rows}
    group: dict[str, list[dict[str, Any]]] = {}
    for c in complex_rows:
        group.setdefault(c["district"], []).append(c)

    for district, items in sorted(group.items(), key=lambda x: dmap.get(x[0], {}).get("rank", 999)):
        items_sorted = sorted(items, key=lambda x: x["total_score"], reverse=True)
        topn = items_sorted[:top_per_district]
        top1 = items_sorted[0]
        avg_score = statistics.mean([x["total_score"] for x in topn]) if topn else 0.0
        avg_base_roi = statistics.mean([x["scenario"]["base_roi"] for x in topn if x["scenario"].get("base_roi") is not None]) if topn else 0.0
        avg_bear_roi = statistics.mean([x["scenario"]["bear_roi"] for x in topn if x["scenario"].get("bear_roi") is not None]) if topn else 0.0
        top1_base = top1["scenario"].get("base_roi")
        note = "표본주의" if top1.get("data_quality", 0) < 55 else ""
        r = write_row(
            ws,
            r,
            [
                district,
                round((dmap.get(district, {}).get("total_score") or 0), 2),
                len(items_sorted),
                round(avg_score, 2),
                pct(avg_base_roi),
                pct(avg_bear_roi),
                top1["complex"],
                round(top1["total_score"], 2),
                pct(top1_base),
                f"{(top1.get('latest_jeonse_ratio') or 0):.1f}%" if top1.get("latest_jeonse_ratio") else "-",
                note,
            ],
            left=True,
        )

    set_widths(ws, [11, 9, 10, 11, 12, 12, 26, 10, 12, 9, 10])


def add_sheet_complex_rank(wb, complex_rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("서울단지 랭킹")
    r = add_common_header(
        ws,
        f"서울 단지 통합 랭킹 ({profile})",
        "구 단위 점수 + 단지 미시지표 결합",
        19,
    )
    r = write_headers(
        ws,
        r,
        ["전체순위", "구내순위", "구", "단지", "총점", "Core점수", "최근매매(억)", "12M성장", "전세가율", "변동성", "최근거래", "거래공백월", "중립ROI", "비관ROI", "DataQ", "신뢰도", "매매표본합", "신선도패널티", "꼬리패널티"],
    )

    for x in complex_rows:
        sc = x["scenario"]
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district_rank"],
                x["district"],
                x["complex"],
                x["total_score"],
                x.get("core_score"),
                round(x["latest_price_eok"], 2) if x.get("latest_price_eok") else "-",
                pct(x.get("growth_12m")),
                f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                x.get("latest_trade_count") or 0,
                x.get("trade_stale_months") or 0,
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
                x.get("data_quality"),
                x.get("confidence_bucket"),
                x.get("total_trade_samples"),
                x.get("recency_penalty"),
                x.get("tail_penalty"),
            ],
            left=True,
        )
        ws.cell(row=r - 1, column=5).fill = score_band_fill(float(x["total_score"]))

    set_widths(ws, [8, 8, 11, 26, 9, 9, 10, 10, 9, 9, 9, 10, 10, 10, 8, 8, 9, 11, 11])


def add_sheet_complex_district_top(wb, complex_rows: list[dict[str, Any]], top_per_district: int):
    ws = wb.create_sheet("구별단지 TopN")
    r = add_common_header(
        ws,
        "구별 상위 단지 리스트",
        f"각 구당 상위 {top_per_district}개 단지 표시",
        12,
    )
    r = write_headers(
        ws,
        r,
        ["구", "구내순위", "단지", "총점", "최근매매(억)", "12M성장", "전세가율", "거래량(최근)", "중립ROI", "비관ROI", "DataQ", "비고"],
    )

    by_d: dict[str, list[dict[str, Any]]] = {}
    for c in complex_rows:
        by_d.setdefault(c["district"], []).append(c)

    for district in sorted(by_d):
        items = sorted(by_d[district], key=lambda x: x["total_score"], reverse=True)[:top_per_district]
        for x in items:
            sc = x["scenario"]
            note = "집중규제" if district in HIGH_INTENSITY_LTPS else ""
            r = write_row(
                ws,
                r,
                [
                    district,
                    x["district_rank"],
                    x["complex"],
                    x["total_score"],
                    round(x["latest_price_eok"], 2) if x.get("latest_price_eok") else "-",
                    pct(x.get("growth_12m")),
                    f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                    x.get("latest_trade_count") or 0,
                    pct(sc.get("base_roi")),
                    pct(sc.get("bear_roi")),
                    x.get("data_quality"),
                    note,
                ],
                left=True,
            )
        r = write_row(ws, r, [""] * 12)

    set_widths(ws, [10, 8, 26, 9, 10, 10, 9, 10, 10, 10, 8, 10])


def add_sheet_complex_scenario(wb, complex_rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("단지 시나리오")
    r = add_common_header(
        ws,
        f"단지별 시나리오 상세 ({profile})",
        "상위 단지 중심 낙관/중립/비관 수익률",
        12,
    )
    r = write_headers(
        ws,
        r,
        ["순위", "구", "단지", "현재가(억)", "기간(년)", "낙관누적", "중립누적", "비관누적", "낙관ROI", "중립ROI", "비관ROI", "가정(연보유수익률)"],
    )

    # 시나리오 시트는 상위 300개까지만 표기해 파일 크기/가독성 균형
    for x in complex_rows[:300]:
        sc = x["scenario"]
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["complex"],
                round(x["latest_price_eok"], 2) if x.get("latest_price_eok") else "-",
                sc.get("horizon_years"),
                pct(sc.get("bull_cum")),
                pct(sc.get("base_cum")),
                pct(sc.get("bear_cum")),
                pct(sc.get("bull_roi")),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
                f"{((sc.get('annual_utility_yield') or 0)*100):.2f}%",
            ],
            left=True,
        )

    set_widths(ws, [7, 10, 26, 10, 8, 10, 10, 10, 10, 10, 10, 13])


def add_sheet_complex_confidence(wb, complex_rows: list[dict[str, Any]]):
    ws = wb.create_sheet("단지 신뢰도")
    r = add_common_header(
        ws,
        "단지 데이터 신뢰도",
        "표본수/커버리지/품질점수 기반 high·medium·low 등급",
        14,
    )
    r = write_headers(
        ws,
        r,
        ["순위", "구", "단지", "신뢰도", "DataQ", "매매표본합", "최근매매표본", "매매커버리지", "최종거래월", "거래공백월", "불확실성패널티", "샘플가중치", "중립ROI", "비고"],
    )
    for x in complex_rows:
        sc = x["scenario"]
        bucket = x.get("confidence_bucket", "medium")
        note = "주의" if bucket == "low" else ""
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["complex"],
                bucket,
                x.get("data_quality"),
                x.get("total_trade_samples"),
                x.get("latest_trade_count"),
                f"{(x.get('coverage_trade') or 0)*100:.1f}%",
                x.get("last_trade_yyyymm") or "-",
                x.get("trade_stale_months") or 0,
                round(x.get("uncertainty_penalty") or 0.0, 3),
                x.get("sample_weight"),
                pct(sc.get("base_roi")),
                note,
            ],
            left=True,
        )
        score_col = 4
        bfill = (
            PatternFill("solid", fgColor="D5F5E3")
            if bucket == "high"
            else PatternFill("solid", fgColor="FCF3CF")
            if bucket == "medium"
            else PatternFill("solid", fgColor="FADBD8")
        )
        ws.cell(row=r - 1, column=score_col).fill = bfill
    set_widths(ws, [7, 10, 26, 9, 8, 10, 10, 10, 10, 10, 12, 10, 10, 10])


def add_sheet_complex_sensitivity(wb, complex_rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("단지 민감도")
    r = add_common_header(
        ws,
        f"단지 금리 민감도 ({profile})",
        "중립 시나리오 기준 단지별 금리충격(가변) 영향",
        12,
    )
    r = write_headers(
        ws,
        r,
        ["순위", "구", "단지", "중립ROI", "금리상승 ROI", "금리하락 ROI", "민감도폭", "적용충격폭", "LTV", "기준금리", "신뢰도", "비고"],
    )
    for x in complex_rows[:400]:
        sc = x["scenario"]
        roi_base = sc.get("base_roi")
        roi_up = sc.get("base_roi_rate_up")
        roi_dn = sc.get("base_roi_rate_down")
        band = None
        if roi_up is not None and roi_dn is not None:
            band = roi_dn - roi_up
        asmp = sc.get("assumption", {})
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["complex"],
                pct(roi_base),
                pct(roi_up),
                pct(roi_dn),
                pct(band),
                f"{(asmp.get('effective_rate_shock') or 0)*100:.2f}%p",
                f"{(asmp.get('ltv') or 0)*100:.0f}%",
                f"{(asmp.get('loan_rate') or 0)*100:.2f}%",
                sc.get("confidence_bucket"),
                "",
            ],
            left=True,
        )
    set_widths(ws, [7, 10, 26, 10, 11, 11, 10, 10, 8, 9, 8, 10])


def add_sheet_universe_meta(
    wb,
    complex_rows: list[dict[str, Any]],
    universe_label: str,
    original_complex_count: int | None,
):
    ws = wb.create_sheet("유니버스 설정")
    r = add_common_header(
        ws,
        "실전매수 유니버스 설정",
        "신뢰도 필터 적용 내역",
        8,
    )
    r = write_headers(ws, r, ["항목", "값", "비고", "", "", "", "", ""])
    total_now = len(complex_rows)
    total_org = original_complex_count if original_complex_count is not None else total_now
    excluded = max(0, total_org - total_now)
    cnt_high = sum(1 for x in complex_rows if x.get("confidence_bucket") == "high")
    cnt_med = sum(1 for x in complex_rows if x.get("confidence_bucket") == "medium")
    cnt_low = sum(1 for x in complex_rows if x.get("confidence_bucket") == "low")
    rows = [
        ["유니버스 타입", universe_label, "예: High+Medium", "", "", "", "", ""],
        ["원본 단지 수", total_org, "필터 전", "", "", "", "", ""],
        ["유니버스 단지 수", total_now, "필터 후", "", "", "", "", ""],
        ["제외 단지 수", excluded, "low 신뢰도 제외", "", "", "", "", ""],
        ["신뢰도 분포", f"high={cnt_high}, medium={cnt_med}, low={cnt_low}", "유니버스 기준", "", "", "", "", ""],
    ]
    for row in rows:
        r = write_row(ws, r, row, left=True)
    set_widths(ws, [18, 40, 22, 8, 8, 8, 8, 8])


def filter_complex_universe(
    complex_rows: list[dict[str, Any]],
    allowed_buckets: tuple[str, ...] = ("high", "medium"),
) -> list[dict[str, Any]]:
    allowed = set(allowed_buckets)
    filtered = [dict(x) for x in complex_rows if x.get("confidence_bucket") in allowed]
    filtered.sort(key=lambda r: r["total_score"], reverse=True)
    for i, r in enumerate(filtered, 1):
        r["rank"] = i
    group: dict[str, int] = {}
    for r in filtered:
        d = r["district"]
        group[d] = group.get(d, 0) + 1
        r["district_rank"] = group[d]
    return filtered


def add_sheet_profile_recommendation(wb, rows: list[dict[str, Any]], profile: str):
    ws_name = {
        "실거주": "실거주 추천",
        "투자자": "투자자 추천",
        "법인": "법인 추천",
    }[profile]
    ws = wb.create_sheet(ws_name)
    r = add_common_header(ws, f"{profile} 관점 Top 10", PROFILE_CONFIGS[profile]["display"], 11)
    r = write_headers(ws, r, ["순위", "구", "총점", "중립ROI", "비관ROI", "전세가율", "변동성", "정책", "추천전략", "핵심리스크", "한줄코멘트"])

    def rec_score(x: dict[str, Any]) -> float:
        sc = x["scenario"]
        base = float(sc.get("base_roi") or -1.0)
        bear = float(sc.get("bear_roi") or -1.0)
        fresh = clamp(1.0 - safe_div(float(x.get("trade_stale_months") or 0), 6.0), 0.0, 1.0)
        utility = float(sc.get("annual_utility_yield") or 0.0) * 100.0
        vol_pen = max(0.0, (x.get("vol_ann") or 0.0) - 0.22) * 100.0
        if profile == "실거주":
            return 0.25 * x["total_score"] + 0.40 * x["stability_score"] + 0.20 * x["value_score"] + 8.0 * fresh + 0.20 * utility - 0.90 * vol_pen
        if profile == "투자자":
            return 0.28 * x["total_score"] + 0.35 * x["trend_score"] + 0.22 * x["liquidity_score"] + 12.0 * base + 6.0 * fresh
        return 0.22 * x["total_score"] + 0.34 * x["policy_score"] + 0.20 * x["liquidity_score"] + 10.0 * base + 5.0 * fresh - 4.0 * max(0.0, -0.75 - bear)

    def hard_filter(x: dict[str, Any]) -> bool:
        stale = float(x.get("trade_stale_months") or 99)
        dq = float(x.get("data_quality") or 0)
        vol = float(x.get("vol_ann") or 0)
        latest = float(x.get("latest_trade_count") or 0)
        if profile == "실거주":
            return dq >= 60 and stale <= 2 and vol <= 0.40
        if profile == "투자자":
            return dq >= 58 and stale <= 2 and latest > 0
        return dq >= 58 and stale <= 3 and x.get("policy_score", 0) >= 38

    ranked = sorted(rows, key=rec_score, reverse=True)
    filtered = [x for x in ranked if hard_filter(x)]
    if len(filtered) < 10:
        seen = {x["district"] for x in filtered}
        for x in ranked:
            if x["district"] in seen:
                continue
            filtered.append(x)
            seen.add(x["district"])
            if len(filtered) >= 10:
                break
    top = filtered[:10]

    for i, x in enumerate(top, 1):
        sc = x["scenario"]
        if profile == "실거주":
            strategy = "실거주+자산방어"
            risk = "생활권/거래공백"
            comment = "신선도·안정성 우선"
        elif profile == "투자자":
            strategy = "분할매수+회전관리"
            risk = "금리/레버리지"
            comment = "모멘텀·유동성 우선"
        else:
            strategy = "세후ROI 선별진입"
            risk = "취득세/법인세"
            comment = "정책·세후수익 우선"

        r = write_row(
            ws,
            r,
            [
                i,
                x["district"],
                x["total_score"],
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
                f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                x["policy_score"],
                strategy,
                risk,
                comment,
            ],
        )

    set_widths(ws, [7, 11, 9, 10, 10, 10, 9, 8, 14, 12, 12])


def add_sheet_references(wb):
    ws = wb.create_sheet("참고문헌")
    r = add_common_header(ws, "근거자료(공식/논문/커뮤니티)", "본 분석의 현실반영 근거", 6)
    r = write_headers(ws, r, ["분류", "자료명", "기준시점", "활용목적", "URL", "비고"])
    for s in SOURCES:
        r = write_row(
            ws,
            r,
            [
                s["category"],
                s["title"],
                s["as_of"],
                s["usage"],
                s["url"],
                "",
            ],
            left=True,
        )
    set_widths(ws, [10, 40, 12, 42, 65, 10])


def add_sheet_checklist(wb, profile: str):
    ws = wb.create_sheet("실행체크리스트")
    r = add_common_header(ws, f"실행 체크리스트 ({profile})", "실제 투자 실행 전 필수 검증", 6)
    r = write_headers(ws, r, ["No", "체크항목", "왜 중요한가", "데이터/문서", "판단기준", "완료"])

    checklist = [
        [1, "자금계획", "LTV/DSR 충족 여부", "대출사전심사", "DSR 허용범위", ""],
        [2, "세금시뮬레이션", "세후수익 급변 방지", "세무사 검토", "보수세율 적용", ""],
        [3, "정책확인", "허가구역/규제 변경", "서울시/금융위 공지", "최근 공고 기준", ""],
        [4, "표본신뢰도", "왜곡 방지", "DataQ/표본수", "DataQ 60 이상 권장", ""],
        [5, "시나리오점검", "하방손실 관리", "Bear ROI", "손실허용범위 내", ""],
        [6, "유동성", "매도/갈아타기 가능성", "최근 거래량", "상위 50% 권장", ""],
        [7, "현장검증", "입지/소음/학군 체감", "현장방문", "핵심 하자 무", ""],
        [8, "출구전략", "매도 타이밍", "목표수익률/기간", "사전 규칙화", ""],
    ]
    for row in checklist:
        r = write_row(ws, r, row, left=True)

    set_widths(ws, [6, 14, 24, 22, 20, 8])


def build_workbook(
    profile_rows: list[dict[str, Any]],
    complex_rows: list[dict[str, Any]],
    profile: str,
    months: list[str],
    rows_by_profile: dict[str, list[dict[str, Any]]] | None = None,
    complex_rows_by_profile: dict[str, list[dict[str, Any]]] | None = None,
    top_per_district: int = DEFAULT_TOP_COMPLEX_PER_DISTRICT,
    universe_label: str = "전체(ALL)",
    original_complex_count: int | None = None,
    backtest_lookback: int = DEFAULT_BACKTEST_LOOKBACK,
    backtest_horizon: int = DEFAULT_BACKTEST_HORIZON,
    backtest_top_fraction: float = DEFAULT_BACKTEST_TOP_FRACTION,
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    add_sheet_dashboard(wb, profile_rows, profile, months)
    add_sheet_scorecard(wb, profile_rows, profile, months)
    add_sheet_coverage(wb, profile_rows, months)
    add_sheet_trade_trend(wb, profile_rows, months)
    add_sheet_rent_trend(wb, profile_rows, months)
    add_sheet_value(wb, profile_rows)
    add_sheet_liquidity(wb, profile_rows)
    add_sheet_risk(wb, profile_rows)
    add_sheet_benchmark(wb, profile_rows)
    add_sheet_backtest(
        wb,
        profile_rows,
        months,
        lookback=backtest_lookback,
        horizon=backtest_horizon,
        top_fraction=backtest_top_fraction,
    )
    add_sheet_scenario(wb, profile_rows, profile)
    add_sheet_scenario_method(wb, profile)
    add_sheet_profile_weights(wb)
    add_sheet_tax_leverage(wb)
    add_sheet_policy(wb, profile_rows)
    add_sheet_timing(wb, profile_rows, profile)
    add_sheet_complex_overview(wb, complex_rows, profile_rows, top_per_district)
    add_sheet_complex_rank(wb, complex_rows, profile)
    add_sheet_complex_district_top(wb, complex_rows, top_per_district)
    add_sheet_complex_scenario(wb, complex_rows, profile)
    add_sheet_complex_confidence(wb, complex_rows)
    add_sheet_complex_sensitivity(wb, complex_rows, profile)
    add_sheet_universe_meta(wb, complex_rows, universe_label, original_complex_count)

    # 프로필별 추천 시트 3개를 모두 포함(비교 가능)
    rows_by_profile = rows_by_profile or {profile: profile_rows}
    complex_rows_by_profile = complex_rows_by_profile or {profile: complex_rows}
    for p in ["실거주", "투자자", "법인"]:
        src = rows_by_profile.get(p) or profile_rows
        add_sheet_profile_recommendation(wb, src, p)

    add_sheet_references(wb)
    add_sheet_checklist(wb, profile)

    return wb


# ---------------------------------------------------------------------------
# Cache and main flow
# ---------------------------------------------------------------------------


def save_cache(path: Path, payload: dict[str, Any]):
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def save_workbook_safe(wb: openpyxl.Workbook, path: Path) -> Path:
    try:
        wb.save(path)
        return path
    except PermissionError:
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = path.with_name(f"{path.stem}_locked_{ts}{path.suffix}")
        wb.save(alt)
        return alt


def load_cache(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def serialize_panel(panel: dict[str, dict[str, dict[str, Any]]]) -> dict[str, Any]:
    return panel


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="서울 전체 아파트 심층분석 엑셀 생성기")
    p.add_argument("--months", type=int, default=12, help="수집 월 수 (기본 12)")
    p.add_argument("--end-yyyymm", default="", help="종료월 YYYYMM (미지정 시 직전 완료월)")
    p.add_argument("--include-current-month", action="store_true", help="당월(부분월 가능)까지 포함")
    p.add_argument("--max-workers", type=int, default=16, help="병렬 작업 수")
    p.add_argument("--num-rows", type=int, default=1000, help="API 페이지 행수")
    p.add_argument("--area-min", type=float, default=DEFAULT_AREA_MIN)
    p.add_argument("--area-max", type=float, default=DEFAULT_AREA_MAX)
    p.add_argument("--deposit-yield-pct", type=float, default=DEFAULT_DEPOSIT_YIELD_PCT)
    p.add_argument("--top-complex-per-district", type=int, default=DEFAULT_TOP_COMPLEX_PER_DISTRICT, help="구별 상위 단지 표시 개수")
    p.add_argument("--complex-min-total-trade", type=int, default=DEFAULT_COMPLEX_MIN_TOTAL_TRADE, help="단지 분석 최소 누적 매매표본")
    p.add_argument("--complex-min-month-coverage", type=int, default=DEFAULT_COMPLEX_MIN_MONTH_COVERAGE, help="단지 분석 최소 거래월 수")
    p.add_argument("--complex-shrink-trade-ref", type=int, default=DEFAULT_COMPLEX_SHRINK_TRADE_REF, help="단지 성장률 shrinkage 기준 표본수")
    p.add_argument("--no-universe", action="store_true", help="High+Medium 실전유니버스 파일 생성 비활성화")
    p.add_argument("--api-key", default="", help="data.go.kr API key (없으면 env/내장키 사용)")
    p.add_argument("--refresh-cache", action="store_true", help="캐시 무시하고 재수집")
    p.add_argument("--profile", choices=list(PROFILE_CONFIGS.keys()), default="투자자")
    p.add_argument("--all-profiles", action="store_true", help="3개 프로필 모두 생성")
    p.add_argument("--backtest-lookback", type=int, default=DEFAULT_BACKTEST_LOOKBACK, help="백테스트 시그널 lookback 월수")
    p.add_argument("--backtest-horizon", type=int, default=DEFAULT_BACKTEST_HORIZON, help="백테스트 forward 월수")
    p.add_argument("--backtest-top-frac", type=float, default=DEFAULT_BACKTEST_TOP_FRACTION, help="백테스트 Top 선택 비중(0~1)")
    p.add_argument("--output-dir", default="C:/Devs/KRE")
    p.add_argument(
        "--region-code-file",
        default="C:/Devs/KRE/real-estate-mcp/src/real_estate/resources/region_codes.txt",
        help="region_codes.txt 경로",
    )
    return p.parse_args()


def main() -> int:
    args = parse_args()

    if args.months < 6:
        raise SystemExit("months는 최소 6 이상을 권장합니다.")
    if args.area_min >= args.area_max:
        raise SystemExit("area-min은 area-max보다 작아야 합니다.")
    if args.top_complex_per_district < 1:
        raise SystemExit("top-complex-per-district는 1 이상이어야 합니다.")
    if args.complex_min_total_trade < 1:
        raise SystemExit("complex-min-total-trade는 1 이상이어야 합니다.")
    if args.complex_min_month_coverage < 1:
        raise SystemExit("complex-min-month-coverage는 1 이상이어야 합니다.")
    if args.complex_shrink_trade_ref < 10:
        raise SystemExit("complex-shrink-trade-ref는 10 이상을 권장합니다.")
    if args.backtest_lookback < 3:
        raise SystemExit("backtest-lookback는 3 이상이어야 합니다.")
    if args.backtest_horizon < 1:
        raise SystemExit("backtest-horizon은 1 이상이어야 합니다.")
    if not (0.05 <= args.backtest_top_frac <= 0.50):
        raise SystemExit("backtest-top-frac는 0.05~0.50 범위여야 합니다.")

    api_key = get_api_key(args.api_key)
    if not api_key:
        raise SystemExit("API key를 찾지 못했습니다.")

    end_yyyymm = (args.end_yyyymm or "").strip()
    if not end_yyyymm:
        end_yyyymm = now_yyyymm() if args.include_current_month else latest_complete_yyyymm()
    if len(end_yyyymm) != 6 or not end_yyyymm.isdigit():
        raise SystemExit("end-yyyymm 형식은 YYYYMM 이어야 합니다.")
    if (not args.include_current_month) and end_yyyymm == now_yyyymm():
        print("[warn] 당월은 부분월일 수 있습니다. 직전 완료월 사용을 권장합니다.")

    months = iter_months(end_yyyymm, args.months)
    if args.months < args.backtest_lookback + args.backtest_horizon:
        print("[warn] 백테스트 표본이 부족할 수 있습니다. months 증가를 권장합니다.")
    region_code_file = Path(args.region_code_file)
    districts = load_seoul_gu_list(region_code_file)

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"[period] {months[0]}~{months[-1]} / include_current_month={args.include_current_month}")
    cache_path = out_dir / f"seoul_market_cache_v6_{months[0]}_{months[-1]}_{int(args.area_min)}_{int(args.area_max)}.json"

    if cache_path.exists() and not args.refresh_cache:
        print(f"[cache] 로드: {cache_path}")
        cache = load_cache(cache_path)
        panel = cache["panel"]
        logs = cache.get("logs", [])
    else:
        print(f"[collect] 서울 25개구 x {len(months)}개월 수집 시작")
        panel, logs = collect_market_panel(
            districts=districts,
            months=months,
            api_key=api_key,
            num_rows=args.num_rows,
            max_workers=args.max_workers,
            area_min=args.area_min,
            area_max=args.area_max,
            deposit_yield_pct=args.deposit_yield_pct,
        )
        cache = {
            "generated_at": dt.datetime.now().isoformat(),
            "months": months,
            "area_filter": [args.area_min, args.area_max],
            "panel": serialize_panel(panel),
            "logs": logs,
        }
        save_cache(cache_path, cache)
        print(f"[cache] 저장: {cache_path}")

    print("[analyze] 지표 계산")
    metrics = build_district_metrics(districts, months, panel)
    metrics = run_multi_expert_scoring(metrics)
    complex_metrics = build_complex_metrics(
        months=months,
        panel=panel,
        district_metrics=metrics,
        min_total_trade=args.complex_min_total_trade,
        min_month_coverage=min(args.complex_min_month_coverage, len(months)),
        shrink_trade_ref=args.complex_shrink_trade_ref,
    )
    complex_metrics = run_complex_expert_scoring(complex_metrics)
    if not complex_metrics:
        raise SystemExit("단지 메트릭이 비어 있습니다. area/min-trade 조건을 완화해 주세요.")

    profiles = ["실거주", "투자자", "법인"] if args.all_profiles else [args.profile]
    generated: list[Path] = []

    rows_by_profile: dict[str, list[dict[str, Any]]] = {}
    complex_rows_by_profile: dict[str, list[dict[str, Any]]] = {}
    for p in ["실거주", "투자자", "법인"]:
        rows_p = score_by_profile(metrics, p)
        rows_by_profile[p] = attach_scenarios(rows_p, p)
        crows_p = score_complex_by_profile(complex_metrics, p)
        complex_rows_by_profile[p] = attach_scenarios(crows_p, p)

    for profile in profiles:
        rows = rows_by_profile[profile]
        crows = complex_rows_by_profile[profile]
        wb = build_workbook(
            rows,
            crows,
            profile,
            months,
            rows_by_profile=rows_by_profile,
            complex_rows_by_profile=complex_rows_by_profile,
            top_per_district=args.top_complex_per_district,
            universe_label="전체(ALL)",
            original_complex_count=len(crows),
            backtest_lookback=args.backtest_lookback,
            backtest_horizon=args.backtest_horizon,
            backtest_top_fraction=args.backtest_top_frac,
        )
        out_path = out_dir / f"서울_전체_아파트_심층분석_v6_{profile}.xlsx"
        actual = save_workbook_safe(wb, out_path)
        generated.append(actual)
        print(f"[done] {profile}: {actual}")

    # 통합본(투자자 기준)도 추가 생성
    rows_i = rows_by_profile["투자자"]
    crows_i = complex_rows_by_profile["투자자"]
    wb_i = build_workbook(
        rows_i,
        crows_i,
        "투자자",
        months,
        rows_by_profile=rows_by_profile,
        complex_rows_by_profile=complex_rows_by_profile,
        top_per_district=args.top_complex_per_district,
        universe_label="전체(ALL)",
        original_complex_count=len(crows_i),
        backtest_lookback=args.backtest_lookback,
        backtest_horizon=args.backtest_horizon,
        backtest_top_fraction=args.backtest_top_frac,
    )
    master_path = out_dir / "서울_전체_아파트_심층분석_v6_통합.xlsx"
    actual_master = save_workbook_safe(wb_i, master_path)
    generated.append(actual_master)
    print(f"[done] 통합본: {actual_master}")

    if not args.no_universe:
        print("[universe] High+Medium 실전유니버스 생성")
        for profile in profiles:
            rows = rows_by_profile[profile]
            crows_all = complex_rows_by_profile[profile]
            crows_uni = filter_complex_universe(crows_all, ("high", "medium"))
            if not crows_uni:
                print(f"[skip] {profile} 유니버스가 비어 생성 생략")
                continue
            wb_u = build_workbook(
                rows,
                crows_uni,
                profile,
                months,
                rows_by_profile=rows_by_profile,
                complex_rows_by_profile=complex_rows_by_profile,
                top_per_district=args.top_complex_per_district,
                universe_label="High+Medium",
                original_complex_count=len(crows_all),
                backtest_lookback=args.backtest_lookback,
                backtest_horizon=args.backtest_horizon,
                backtest_top_fraction=args.backtest_top_frac,
            )
            out_u = out_dir / f"서울_전체_아파트_실전유니버스_v6_{profile}.xlsx"
            actual_u = save_workbook_safe(wb_u, out_u)
            generated.append(actual_u)
            print(f"[done] 유니버스 {profile}: {actual_u}")

        crows_all_i = complex_rows_by_profile["투자자"]
        crows_uni_i = filter_complex_universe(crows_all_i, ("high", "medium"))
        if crows_uni_i:
            wb_ui = build_workbook(
                rows_i,
                crows_uni_i,
                "투자자",
                months,
                rows_by_profile=rows_by_profile,
                complex_rows_by_profile=complex_rows_by_profile,
                top_per_district=args.top_complex_per_district,
                universe_label="High+Medium",
                original_complex_count=len(crows_all_i),
                backtest_lookback=args.backtest_lookback,
                backtest_horizon=args.backtest_horizon,
                backtest_top_fraction=args.backtest_top_frac,
            )
            out_ui = out_dir / "서울_전체_아파트_실전유니버스_v6_통합.xlsx"
            actual_ui = save_workbook_safe(wb_ui, out_ui)
            generated.append(actual_ui)
            print(f"[done] 유니버스 통합본: {actual_ui}")

    print("\n생성 완료 파일:")
    for p in generated:
        print(f"- {p}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
