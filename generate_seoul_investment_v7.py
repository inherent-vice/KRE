#!/usr/bin/env python3
"""서울 25개 자치구 + 구내 상위단지 심층분석 생성기 (v7)

핵심 기능
- 국토부 실거래 API(아파트 매매/전월세) 월별 수집
- 서울 25개 구 병렬 수집 + 병렬 다중전문가 스코어링
- 확률(분위수) 기반 낙관/중립/비관 시나리오
- 비정량 오버레이(거버넌스/생활환경/수요지속/관리품질/심리지표)
- 실거주/투자자/법인 3개 프로필 별도 엑셀 생성

실행 예시
python generate_seoul_investment_v7.py --months 12 --all-profiles
"""

from __future__ import annotations

import argparse
import concurrent.futures as cf
import datetime as dt
import json
import math
import os
import random
import statistics
import time
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DEFAULT_DATA_GO_KR_API_KEY = (
    "1deaf1b26dc45dd6f26d5b481d28d2fe784327eb82cc8ccdb727a51baf462419"
)

APT_TRADE_URL = "https://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade"
APT_RENT_URL = "https://apis.data.go.kr/1613000/RTMSDataSvcAptRent/getRTMSDataSvcAptRent"

DEFAULT_AREA_MIN = 59.0
DEFAULT_AREA_MAX = 99.0
DEFAULT_DEPOSIT_YIELD_PCT = 3.0
DEFAULT_TOP_COMPLEX_PER_DISTRICT = 7
DEFAULT_COMPLEX_MIN_TOTAL_TRADE = 18
DEFAULT_COMPLEX_MIN_MONTH_COVERAGE = 6
DEFAULT_COMPLEX_SHRINK_TRADE_REF = 120
DEFAULT_SCENARIO_RATE_SHOCK = 0.01
DEFAULT_COMPLETE_MONTH_SHIFT = -1
DEFAULT_BACKTEST_LOOKBACK = 4
DEFAULT_BACKTEST_HORIZON = 3
DEFAULT_BACKTEST_TOP_FRACTION = 0.20
DEFAULT_NQ_PRIOR_WEIGHT = 0.58
DEFAULT_NQ_OVERLAY_CAP = 12.0
DEFAULT_NQ_COMPLEX_OVERLAY_CAP = 9.0
DEFAULT_NONQUANT_PRIOR_FILE = "C:/Devs/KRE/seoul_nonquant_priors_v7.json"
DEFAULT_ADV_RISK_DISTRICT_CAP = 6.0
DEFAULT_ADV_RISK_COMPLEX_CAP = 5.0
DEFAULT_ADV_SIGNAL_FILE = "C:/Devs/KRE/seoul_advanced_signals_v7.json"

NQ_DIM_KEYS = (
    "nq_governance",
    "nq_liveability",
    "nq_demand_persistence",
    "nq_management_quality",
    "nq_sentiment",
)

NQ_PROFILE_WEIGHTS: dict[str, dict[str, float]] = {
    "실거주": {
        "nq_governance": 0.18,
        "nq_liveability": 0.34,
        "nq_demand_persistence": 0.18,
        "nq_management_quality": 0.20,
        "nq_sentiment": 0.10,
    },
    "투자자": {
        "nq_governance": 0.20,
        "nq_liveability": 0.14,
        "nq_demand_persistence": 0.34,
        "nq_management_quality": 0.12,
        "nq_sentiment": 0.20,
    },
    "법인": {
        "nq_governance": 0.34,
        "nq_liveability": 0.08,
        "nq_demand_persistence": 0.26,
        "nq_management_quality": 0.20,
        "nq_sentiment": 0.12,
    },
}

ADV_RISK_KEYS = (
    "adv_liquidity_trap",
    "adv_crowding",
    "adv_policy_flip",
    "adv_execution",
)

ADV_RISK_PROFILE_WEIGHTS: dict[str, dict[str, float]] = {
    "실거주": {
        "adv_liquidity_trap": 0.26,
        "adv_crowding": 0.14,
        "adv_policy_flip": 0.22,
        "adv_execution": 0.38,
    },
    "투자자": {
        "adv_liquidity_trap": 0.34,
        "adv_crowding": 0.30,
        "adv_policy_flip": 0.18,
        "adv_execution": 0.18,
    },
    "법인": {
        "adv_liquidity_trap": 0.20,
        "adv_crowding": 0.10,
        "adv_policy_flip": 0.42,
        "adv_execution": 0.28,
    },
}

DEFAULT_NQ_PRIOR_PAYLOAD: dict[str, Any] = {
    "as_of": "2026-02-20",
    "version": "v7",
    "method": "district prior + market proxy blend",
    "district_priors": {
        "강남구": {"governance": 55, "liveability": 86, "demand_persistence": 92, "management_quality": 80, "sentiment": 76, "flags": ["집중규제", "가격부담"]},
        "강동구": {"governance": 58, "liveability": 72, "demand_persistence": 70, "management_quality": 68, "sentiment": 63, "flags": ["공급변동"]},
        "강북구": {"governance": 60, "liveability": 58, "demand_persistence": 56, "management_quality": 57, "sentiment": 52, "flags": ["노후주거비중"]},
        "강서구": {"governance": 61, "liveability": 63, "demand_persistence": 66, "management_quality": 62, "sentiment": 58, "flags": ["공항소음권역"]},
        "관악구": {"governance": 59, "liveability": 62, "demand_persistence": 67, "management_quality": 60, "sentiment": 59, "flags": ["대학수요편중"]},
        "광진구": {"governance": 62, "liveability": 74, "demand_persistence": 76, "management_quality": 66, "sentiment": 68, "flags": []},
        "구로구": {"governance": 60, "liveability": 60, "demand_persistence": 64, "management_quality": 61, "sentiment": 57, "flags": ["노후산업지혼재"]},
        "금천구": {"governance": 61, "liveability": 57, "demand_persistence": 58, "management_quality": 59, "sentiment": 54, "flags": ["업무핵심거리"]},
        "노원구": {"governance": 62, "liveability": 65, "demand_persistence": 63, "management_quality": 63, "sentiment": 58, "flags": ["노후단지비중"]},
        "도봉구": {"governance": 61, "liveability": 59, "demand_persistence": 56, "management_quality": 60, "sentiment": 53, "flags": ["수요탄성낮음"]},
        "동대문구": {"governance": 61, "liveability": 64, "demand_persistence": 65, "management_quality": 63, "sentiment": 60, "flags": []},
        "동작구": {"governance": 63, "liveability": 76, "demand_persistence": 78, "management_quality": 68, "sentiment": 70, "flags": []},
        "마포구": {"governance": 64, "liveability": 82, "demand_persistence": 85, "management_quality": 72, "sentiment": 74, "flags": []},
        "서대문구": {"governance": 63, "liveability": 73, "demand_persistence": 72, "management_quality": 68, "sentiment": 67, "flags": []},
        "서초구": {"governance": 56, "liveability": 87, "demand_persistence": 90, "management_quality": 80, "sentiment": 77, "flags": ["집중규제", "가격부담"]},
        "성동구": {"governance": 64, "liveability": 80, "demand_persistence": 83, "management_quality": 73, "sentiment": 75, "flags": []},
        "성북구": {"governance": 62, "liveability": 68, "demand_persistence": 66, "management_quality": 65, "sentiment": 61, "flags": []},
        "송파구": {"governance": 55, "liveability": 86, "demand_persistence": 91, "management_quality": 79, "sentiment": 76, "flags": ["집중규제", "가격부담"]},
        "양천구": {"governance": 60, "liveability": 78, "demand_persistence": 74, "management_quality": 71, "sentiment": 69, "flags": ["목동재건축정책민감"]},
        "영등포구": {"governance": 62, "liveability": 81, "demand_persistence": 84, "management_quality": 71, "sentiment": 73, "flags": []},
        "용산구": {"governance": 54, "liveability": 85, "demand_persistence": 89, "management_quality": 77, "sentiment": 75, "flags": ["집중규제", "개발이벤트변동"]},
        "은평구": {"governance": 62, "liveability": 67, "demand_persistence": 64, "management_quality": 64, "sentiment": 60, "flags": ["공급변동"]},
        "종로구": {"governance": 59, "liveability": 72, "demand_persistence": 66, "management_quality": 67, "sentiment": 62, "flags": ["도심업무수요의존"]},
        "중구": {"governance": 58, "liveability": 74, "demand_persistence": 68, "management_quality": 66, "sentiment": 63, "flags": ["도심업무수요의존"]},
        "중랑구": {"governance": 61, "liveability": 60, "demand_persistence": 59, "management_quality": 61, "sentiment": 56, "flags": ["상대저밀고용"]},
    },
    "complex_overrides": {},
}

DEFAULT_ADV_SIGNAL_PAYLOAD: dict[str, Any] = {
    "as_of": "2026-02-20",
    "version": "v7",
    "method": "external supplemental risk signals (supply/guarantee/management DD) + model blend",
    "guide": {
        "value_range": "0~100 (또는 0~1 비율 입력 가능, 내부에서 자동 변환)",
        "district_signals": "key=자치구명, value={supply_pressure, guarantee_risk, management_dd_risk, confidence, source, flags}",
        "complex_signals": "key=구::단지명, district 미입력 시 complex에는 구 시그널 fallback 적용",
        "confidence": "high|medium|low",
    },
    "district_signals": {
        d: {
            "supply_pressure": None,
            "guarantee_risk": None,
            "management_dd_risk": None,
            "confidence": "medium",
            "source": "fill_with_official_data",
            "flags": [],
        }
        for d in (DEFAULT_NQ_PRIOR_PAYLOAD.get("district_priors") or {}).keys()
    },
    "complex_signals": {},
}

BACKTEST_SIGNAL_CANDIDATES: list[dict[str, float]] = [
    {"mom": 0.20, "liq": 0.25, "fresh": 0.15, "val": 0.20, "yld": 0.20},
    {"mom": 0.25, "liq": 0.25, "fresh": 0.20, "yld": 0.30},
    {"mom": 0.25, "liq": 0.25, "fresh": 0.20, "val": 0.15, "lowvol": 0.15},
    {"mom": 0.30, "liq": 0.25, "fresh": 0.20, "val": 0.25},
    {"mom": 0.30, "liq": 0.30, "fresh": 0.20, "lowvol": 0.20},
    {"mom": 0.35, "liq": 0.35, "fresh": 0.30},
]

BACKTEST_TOP_FRACTION_CANDIDATES = (0.16, 0.20, 0.24, 0.30)

HIGH_INTENSITY_LTPS = {"강남구", "서초구", "송파구", "용산구"}

PROFILE_CONFIGS: dict[str, dict[str, Any]] = {
    "실거주": {
        "weights": {
            "trend": 17,
            "value": 23,
            "liquidity": 18,
            "stability": 27,
            "policy": 15,
        },
        "ltv": 0.40,
        "loan_rate": 0.037,
        "buy_tax": 0.015,
        "annual_hold_cost": 0.004,
        "capital_gain_tax": 0.15,
        "horizon_years": 5,
        "display": "실거주 안정성 + 생활비/현금흐름 방어",
    },
    "투자자": {
        "weights": {
            "trend": 29,
            "value": 24,
            "liquidity": 22,
            "stability": 10,
            "policy": 15,
        },
        "ltv": 0.50,
        "loan_rate": 0.041,
        "buy_tax": 0.025,
        "annual_hold_cost": 0.006,
        "capital_gain_tax": 0.30,
        "horizon_years": 3,
        "display": "모멘텀 + 밸류 + 유동성 중심",
    },
    "법인": {
        "weights": {
            "trend": 24,
            "value": 17,
            "liquidity": 20,
            "stability": 9,
            "policy": 30,
        },
        "ltv": 0.55,
        "loan_rate": 0.044,
        "buy_tax": 0.120,
        "annual_hold_cost": 0.009,
        "capital_gain_tax": 0.45,
        "horizon_years": 3,
        "display": "세후수익/규제리스크 중심",
    },
}

THEME_CATALOG: list[dict[str, Any]] = [
    {
        "key": "rebuild",
        "name": "재건축·정비사업",
        "desc": "노후축 정비와 사업단계 진전에 따른 가치 재평가",
        "opp_weights": {"prior": 0.26, "policy": 0.16, "premium": 0.14, "trend": 0.12, "liq": 0.12, "demand": 0.10, "fresh": 0.10},
        "risk_weights": {"adv_policy_risk": 0.30, "adv_supply_risk": 0.24, "adv_mgmt_risk": 0.20, "data_low": 0.14, "fresh_low": 0.12},
    },
    {
        "key": "redevelop",
        "name": "재개발·도시정비수혜",
        "desc": "정비사업 인접 수혜/도심 재편 수혜",
        "opp_weights": {"prior": 0.24, "value": 0.16, "trend": 0.16, "liq": 0.14, "policy": 0.12, "demand": 0.10, "fresh": 0.08},
        "risk_weights": {"adv_supply_risk": 0.28, "adv_policy_risk": 0.26, "adv_mgmt_risk": 0.20, "data_low": 0.14, "vol_risk": 0.12},
    },
    {
        "key": "transport",
        "name": "교통개선·역세권확장",
        "desc": "광역교통/환승개선에 따른 접근성 프리미엄",
        "opp_weights": {"prior": 0.22, "trend": 0.18, "liq": 0.16, "demand": 0.14, "live": 0.12, "premium": 0.10, "policy": 0.08},
        "risk_weights": {"adv_policy_risk": 0.28, "adv_supply_risk": 0.22, "vol_risk": 0.20, "data_low": 0.15, "fresh_low": 0.15},
    },
    {
        "key": "job_core",
        "name": "업무지구 직주근접",
        "desc": "고용핵심지 접근성 기반 실수요/임차수요",
        "opp_weights": {"prior": 0.20, "demand": 0.20, "liq": 0.18, "live": 0.12, "premium": 0.12, "trend": 0.10, "policy": 0.08},
        "risk_weights": {"adv_policy_risk": 0.22, "adv_guarantee_risk": 0.20, "vol_risk": 0.18, "gap_risk": 0.20, "fresh_low": 0.20},
    },
    {
        "key": "education",
        "name": "학군·교육프리미엄",
        "desc": "교육 선호에 기반한 방어적 수요",
        "opp_weights": {"prior": 0.30, "live": 0.22, "demand": 0.14, "stability": 0.12, "premium": 0.12, "liq": 0.10},
        "risk_weights": {"premium_risk": 0.30, "adv_policy_risk": 0.22, "vol_risk": 0.18, "fresh_low": 0.15, "data_low": 0.15},
    },
    {
        "key": "new_big",
        "name": "신축희소·대단지유동성",
        "desc": "거래회전이 빠른 대단지 중심의 유동성 프리미엄",
        "opp_weights": {"prior": 0.18, "liq": 0.22, "sample_depth": 0.18, "fresh": 0.14, "stability": 0.12, "demand": 0.10, "live": 0.06},
        "risk_weights": {"adv_supply_risk": 0.26, "adv_guarantee_risk": 0.22, "vol_risk": 0.16, "gap_risk": 0.18, "data_low": 0.18},
    },
    {
        "key": "value_re",
        "name": "저평가 리레이팅",
        "desc": "가격괴리 축소와 회복국면 수혜",
        "opp_weights": {"value": 0.30, "trend": 0.20, "liq": 0.16, "policy": 0.10, "demand": 0.10, "fresh": 0.08, "prior": 0.06},
        "risk_weights": {"adv_guarantee_risk": 0.26, "adv_supply_risk": 0.22, "vol_risk": 0.18, "data_low": 0.18, "fresh_low": 0.16},
    },
    {
        "key": "cashflow",
        "name": "임대수익·현금흐름",
        "desc": "전월세 수익률과 현금흐름 안정성",
        "opp_weights": {"yield": 0.30, "value": 0.16, "liq": 0.12, "demand": 0.10, "stability": 0.10, "fresh": 0.08, "prior": 0.14},
        "risk_weights": {"gap_risk": 0.30, "adv_guarantee_risk": 0.28, "adv_supply_risk": 0.18, "data_low": 0.14, "fresh_low": 0.10},
    },
    {
        "key": "policy_mo",
        "name": "정책모멘텀",
        "desc": "규제완화/정책전환에 따른 단기 모멘텀",
        "opp_weights": {"policy": 0.30, "trend": 0.18, "liq": 0.16, "demand": 0.10, "prior": 0.14, "fresh": 0.12},
        "risk_weights": {"adv_policy_risk": 0.35, "vol_risk": 0.20, "adv_supply_risk": 0.20, "data_low": 0.15, "fresh_low": 0.10},
    },
    {
        "key": "supply_cycle",
        "name": "공급사이클",
        "desc": "입주/공급 사이클 저점 구간 선별",
        "opp_weights": {"adv_supply_low": 0.34, "demand": 0.18, "liq": 0.16, "policy": 0.10, "trend": 0.10, "prior": 0.12},
        "risk_weights": {"adv_supply_risk": 0.40, "adv_policy_risk": 0.20, "gap_risk": 0.16, "data_low": 0.14, "fresh_low": 0.10},
    },
    {
        "key": "defensive",
        "name": "방어형 실거주",
        "desc": "변동성 완충과 거주만족 중심 장기보유",
        "opp_weights": {"stability": 0.24, "live": 0.18, "gov": 0.14, "mgmt": 0.14, "adv_low": 0.12, "dataq": 0.10, "fresh": 0.08},
        "risk_weights": {"vol_risk": 0.24, "adv_guarantee_risk": 0.20, "adv_mgmt_risk": 0.20, "data_low": 0.18, "gap_risk": 0.18},
    },
]

THEME_FEATURE_LABELS: dict[str, str] = {
    "prior": "구조적선호",
    "policy": "정책점수",
    "trend": "가격모멘텀",
    "liq": "유동성",
    "premium": "프리미엄",
    "value": "상대저평가",
    "yield": "임대수익률",
    "demand": "수요지속성",
    "live": "생활환경",
    "stability": "안정성",
    "gov": "거버넌스",
    "mgmt": "관리품질",
    "fresh": "거래신선도",
    "sample_depth": "표본깊이",
    "adv_low": "ADV저위험",
    "adv_supply_low": "저공급압력",
    "adv_policy_risk": "정책플립위험",
    "adv_supply_risk": "공급압력위험",
    "adv_guarantee_risk": "보증리스크",
    "adv_mgmt_risk": "관리DD리스크",
    "gap_risk": "전세구조위험",
    "premium_risk": "고평가부담",
    "vol_risk": "고변동성",
    "data_low": "데이터취약",
    "fresh_low": "거래공백",
    "dataq": "데이터품질",
}

THEME_PROFILE_BIAS: dict[str, dict[str, float]] = {
    "실거주": {
        "defensive": 1.12,
        "education": 1.10,
        "job_core": 1.06,
        "cashflow": 0.92,
        "policy_mo": 0.90,
    },
    "투자자": {
        "rebuild": 1.08,
        "value_re": 1.10,
        "cashflow": 1.08,
        "policy_mo": 1.06,
        "defensive": 0.92,
    },
    "법인": {
        "job_core": 1.10,
        "cashflow": 1.10,
        "policy_mo": 1.08,
        "transport": 1.05,
        "education": 0.90,
    },
}

THEME_PRIOR_HIGH: dict[str, set[str]] = {
    "rebuild": {"강남구", "서초구", "송파구", "양천구", "강동구", "노원구", "영등포구", "용산구"},
    "redevelop": {"성동구", "마포구", "동작구", "용산구", "영등포구", "성북구", "은평구", "동대문구", "중랑구"},
    "transport": {"용산구", "송파구", "강동구", "강남구", "영등포구", "동작구", "마포구", "성동구"},
    "job_core": {"강남구", "서초구", "송파구", "영등포구", "용산구", "마포구", "종로구", "중구", "성동구"},
    "education": {"강남구", "서초구", "양천구", "송파구", "노원구", "광진구"},
    "new_big": {"송파구", "강동구", "노원구", "은평구", "성동구", "마포구", "영등포구"},
    "value_re": {"금천구", "구로구", "중랑구", "강북구", "도봉구", "은평구", "동대문구"},
    "cashflow": {"강서구", "구로구", "금천구", "영등포구", "동작구", "관악구", "중랑구"},
    "policy_mo": {"강남구", "서초구", "송파구", "용산구", "양천구", "강동구", "영등포구"},
    "supply_cycle": {"강남구", "서초구", "송파구", "마포구", "성동구", "동작구", "서대문구"},
    "defensive": {"강남구", "서초구", "송파구", "성동구", "마포구", "양천구", "광진구", "동작구"},
}

THEME_PRIOR_MED: dict[str, set[str]] = {
    "rebuild": {"성동구", "마포구", "동작구", "양천구", "영등포구", "광진구"},
    "redevelop": {"강북구", "도봉구", "관악구", "구로구", "금천구", "서대문구"},
    "transport": {"서초구", "서대문구", "강서구", "광진구", "중구", "종로구"},
    "job_core": {"동작구", "광진구", "강동구", "서대문구", "동대문구"},
    "education": {"성동구", "마포구", "동작구", "서대문구", "강동구"},
    "new_big": {"강서구", "강남구", "서초구", "동작구", "관악구"},
    "value_re": {"강서구", "관악구", "동대문구", "서대문구", "성북구"},
    "cashflow": {"은평구", "동대문구", "강북구", "도봉구", "성북구"},
    "policy_mo": {"마포구", "성동구", "강서구", "동작구", "은평구"},
    "supply_cycle": {"강동구", "영등포구", "은평구", "강서구", "동대문구"},
    "defensive": {"서대문구", "성북구", "강동구", "강서구", "영등포구"},
}

THEME_PLAYBOOK: dict[str, dict[str, str]] = {
    "rebuild": {
        "real": "실거주 장기보유 전제에서 사업단계 확정 구간만 분할 진입",
        "invest": "초기 과열구간 추격보다 조정구간 분할매수",
        "corp": "세후자금계획 확정 후 단계별 포지션 빌드업",
        "risk": "정책변경·사업지연·추가분담금 상승",
        "monitor": "정비계획 인가·조합일정·규제변경 공고 월별 점검",
    },
    "redevelop": {
        "real": "생활권 개선 체감 가능한 구역 중심으로 선별",
        "invest": "인접 수혜권 저평가 구간 위주로 분산",
        "corp": "사업속도·토지이슈 확인 후 제한적 진입",
        "risk": "인허가 지연·사업비 급등·수요 이동",
        "monitor": "도시정비 고시·구역별 진척도·거래 공백 추이 확인",
    },
    "transport": {
        "real": "실제 통행시간 단축이 확인되는 생활권 중심",
        "invest": "개통 임박/환승개선 가시화 구간에 집중",
        "corp": "업무권 접근성 개선 지역을 임대수요와 함께 검증",
        "risk": "착공·개통 지연·정책 변경",
        "monitor": "착공/개통 일정 업데이트, 사업비 증액 여부 확인",
    },
    "job_core": {
        "real": "출퇴근 편의와 생활인프라 동시 충족 단지 선호",
        "invest": "직주근접 수요가 유지되는 역세권 코어 선별",
        "corp": "임차회전율과 공실리스크를 동시 점검",
        "risk": "업무권 경기둔화·임대수요 약화",
        "monitor": "고용지표·공실률·환승편의 지표 점검",
    },
    "education": {
        "real": "학군/통학 동선 우선, 과도한 프리미엄 추격 금지",
        "invest": "학군 수요 견조 지역 중 상대가치 우위 선별",
        "corp": "학군 프리미엄 과열 구간은 비중 축소",
        "risk": "고평가 구간 진입, 정책·배정 변화",
        "monitor": "실거래 갭 확대 여부, 학교배정/학군 수요 변화 확인",
    },
    "new_big": {
        "real": "대단지 유동성과 커뮤니티 품질 중심 장기보유",
        "invest": "거래회전 빠른 신축 핵심단지 회전 전략",
        "corp": "유동성 높은 단지 우선, 고점추격 방지",
        "risk": "공급확대 구간, 고평가 피로",
        "monitor": "입주물량·거래회전율·관리비 추이 점검",
    },
    "value_re": {
        "real": "거주편의 저하 없는 저평가 지역부터 접근",
        "invest": "밸류+모멘텀 동시 개선 구간에서 분할매수",
        "corp": "세후 수익성 양호한 저평가 지역만 선별",
        "risk": "가치함정·유동성 저하·하락추세 지속",
        "monitor": "거래량 회복, 전세수요, 가격 추세 반전 확인",
    },
    "cashflow": {
        "real": "보유비용 대비 현금흐름 방어력 중심",
        "invest": "월세전환·임대회전 안정 단지 우선",
        "corp": "세후 임대수익률과 공실리스크 동시 관리",
        "risk": "보증사고·공실·전세구조 리스크",
        "monitor": "전세가율/보증리스크/체납·공실 지표 점검",
    },
    "policy_mo": {
        "real": "정책 급변 구간은 공격적 추격보다 방어적 접근",
        "invest": "정책 이벤트 이후 거래확인 구간에서 단계 진입",
        "corp": "세제·규제 변화에 맞춘 단기 대응 포지션",
        "risk": "정책 플립·거래 위축·규제 재강화",
        "monitor": "정부/서울시 공고, 허가구역 및 세제 변경 즉시 반영",
    },
    "supply_cycle": {
        "real": "입주물량 저점 국면의 생활권 중심 선별",
        "invest": "공급압력 완화 구간에서 선제 분할",
        "corp": "공급절벽/과잉 구간을 구분해 보수적 비중조절",
        "risk": "입주물량 급증·임대수요 둔화",
        "monitor": "2년 누적 입주물량·미분양·전월세 약세 점검",
    },
    "defensive": {
        "real": "변동성 낮고 관리품질 좋은 단지 위주 장기보유",
        "invest": "하방 방어 포지션으로 비중관리",
        "corp": "현금흐름 안정성을 최우선으로 편입",
        "risk": "수익률 저하·기회비용 증가",
        "monitor": "변동성·관리품질·유동성 유지 여부 점검",
    },
}

QUANT_FACTOR_CATALOG: list[dict[str, str]] = [
    {
        "key": "mom",
        "name": "모멘텀",
        "direction": "+",
        "formula": "0.65*12M수익률 + 0.35*(3M수익률*4)",
        "evidence": "Jegadeesh & Titman (1993), Guren (2018)",
    },
    {
        "key": "value",
        "name": "밸류(상대저평가)",
        "direction": "+",
        "formula": "-ln(상대가격지수)",
        "evidence": "Case-Shiller (1990), Campbell et al. (2009), Asness et al. (2013)",
    },
    {
        "key": "carry",
        "name": "캐리(임대수익)",
        "direction": "+",
        "formula": "최근 환산 임대수익률",
        "evidence": "Housing returns decomposition (Campbell et al. 2009)",
    },
    {
        "key": "liq",
        "name": "유동성",
        "direction": "+",
        "formula": "log(최근거래량)+신선도-거래공백",
        "evidence": "Pastor-Stambaugh (2001/2003), Amihud (2002)",
    },
    {
        "key": "quality",
        "name": "품질/실행가능성",
        "direction": "+",
        "formula": "DataQ + 거버넌스 + 관리품질 + 신뢰도",
        "evidence": "Taming the Factor Zoo (Feng et al. 2019) 보수적 선택",
    },
    {
        "key": "lowrisk",
        "name": "저위험",
        "direction": "+",
        "formula": "-(변동성 + 꼬리위험 + ADV리스크)",
        "evidence": "Frazzini-Pedersen (2010), Liquidity & Volatility (Drechsler et al. 2020)",
    },
    {
        "key": "policy",
        "name": "정책/규제체력",
        "direction": "+",
        "formula": "정책점수 + (100-정책플립위험)",
        "evidence": "Monetary Momentum (Neuhierl & Weber 2018)",
    },
    {
        "key": "sent",
        "name": "수요/심리",
        "direction": "+",
        "formula": "수요지속성 + 심리 + 생활환경",
        "evidence": "News Sentiment and Asset Prices (Engelberg et al. 2020)",
    },
    {
        "key": "supply",
        "name": "공급사이클",
        "direction": "+",
        "formula": "100-공급압력위험",
        "evidence": "Housing supply and cycles literature",
    },
]
QUANT_FACTOR_NAME_MAP: dict[str, str] = {x["key"]: x["name"] for x in QUANT_FACTOR_CATALOG}

QUANT_PROFILE_WEIGHTS: dict[str, dict[str, float]] = {
    "실거주": {
        "mom": 0.10,
        "value": 0.12,
        "carry": 0.08,
        "liq": 0.10,
        "quality": 0.20,
        "lowrisk": 0.18,
        "policy": 0.10,
        "sent": 0.08,
        "supply": 0.04,
    },
    "투자자": {
        "mom": 0.21,
        "value": 0.17,
        "carry": 0.12,
        "liq": 0.15,
        "quality": 0.08,
        "lowrisk": 0.10,
        "policy": 0.08,
        "sent": 0.05,
        "supply": 0.04,
    },
    "법인": {
        "mom": 0.15,
        "value": 0.12,
        "carry": 0.14,
        "liq": 0.14,
        "quality": 0.13,
        "lowrisk": 0.10,
        "policy": 0.16,
        "sent": 0.03,
        "supply": 0.03,
    },
}

QUANT_FACTOR_BLEND: dict[str, float] = {
    "실거주": 0.55,
    "투자자": 0.65,
    "법인": 0.60,
}

QUANT_CORR_SHRINK_THRESHOLD = 0.55
QUANT_CORR_SHRINK_INTENSITY = 0.35

# 분위수 기반 z-score (현실적 범위: P20/P50/P80)
SCENARIO_Z = {
    "bull": 0.8416,
    "base": 0.0,
    "bear": -0.8416,
}

SOURCES: list[dict[str, str]] = [
    {
        "category": "official",
        "title": "서울시 토지거래허가구역 지정현황",
        "url": "https://land.seoul.go.kr/land/other/appointStatusSeoul.do",
        "as_of": "2026-01-07",
        "usage": "정책/규제 리스크(서울 전체 아파트 LTPS, 강남·서초·송파·용산 집중규제) 반영",
    },
    {
        "category": "official",
        "title": "한국은행 통화정책방향(2026-01-15, 기준금리 2.50%)",
        "url": "https://www.bok.or.kr/portal/singl/newsData/view.do?menuNo=200643&nttId=10095737",
        "as_of": "2026-01-15",
        "usage": "기준금리 2.50% 동결을 금리 가정의 기준점으로 사용",
    },
    {
        "category": "official",
        "title": "금융위 3단계 스트레스 DSR 시행방안(2025-05-20)",
        "url": "https://www.fsc.go.kr/no010101/84617",
        "as_of": "2025-05-20",
        "usage": "대출한도 제약 시나리오(레버리지 민감도) 반영",
    },
    {
        "category": "official",
        "title": "금융위 2026년 하반기 스트레스 DSR 운영방향",
        "url": "https://www.fsc.go.kr/no010101/85903",
        "as_of": "2026-02-20",
        "usage": "스트레스금리 체계 변경을 금리충격 시나리오에 반영",
    },
    {
        "category": "official",
        "title": "한국부동산원 R-ONE 월간 주택가격동향",
        "url": "https://www.reb.or.kr/r-one/portal/main/indexPage.do",
        "as_of": "2026-01-15",
        "usage": "서울 주택가격 사이클/벤치마크 비교 검증",
    },
    {
        "category": "official",
        "title": "IMF Article IV Consultation with Korea (2025-02-06)",
        "url": "https://www.imf.org/en/news/articles/2025/02/06/pr-2530-korea-imf-concludes-2024-art-iv-consultation-with-republic-of-korea",
        "as_of": "2025-02-06",
        "usage": "부동산 PF/비은행권 리스크를 하방 시나리오 리스크 프리미엄으로 반영",
    },
    {
        "category": "official",
        "title": "한국은행 금융안정보고서 (2025년 6월)",
        "url": "https://www.bok.or.kr/portal/singl/newsData/view.do?menuNo=200285&nttId=10094737",
        "as_of": "2025-06-25",
        "usage": "가계부채·부동산PF·비은행권 리스크를 ADV 하방요인(정책플립/실행리스크)으로 반영",
    },
    {
        "category": "official",
        "title": "한국은행 이슈노트: 수도권 주택시장 수급 상황 평가 및 시사점",
        "url": "https://www.bok.or.kr/portal/singl/newsData/view.do?menuNo=200066&nttId=10090506",
        "as_of": "2025-02-17",
        "usage": "수급불균형·지역별 탄력성 차이를 군집쏠림/유동성함정 해석에 반영",
    },
    {
        "category": "official",
        "title": "서울 열린데이터광장: 서울시 도시침수 통계(자치구별)",
        "url": "https://data.seoul.go.kr/dataList/OA-15636/S/1/datasetView.do",
        "as_of": "2026-02-20",
        "usage": "침수 취약도 비정량 지표(환경 리스크)의 구 단위 사전값 근거",
    },
    {
        "category": "official",
        "title": "서울 열린데이터광장: 지하철 호선별 역별 승하차 인원",
        "url": "https://data.seoul.go.kr/dataList/OA-12914/S/1/datasetView.do",
        "as_of": "2026-02-20",
        "usage": "대중교통 접근성/수요지속성 비정량 지표 설계 근거",
    },
    {
        "category": "official",
        "title": "행안부 생활안전지도(SafeMap) API",
        "url": "https://www.safemap.go.kr/opna/data/dataView.do?objtId=3",
        "as_of": "2026-02-20",
        "usage": "치안/안전 체감 지표 연계 후보(비정량 리스크 계층화) 근거",
    },
    {
        "category": "official",
        "title": "K-apt 공동주택관리정보시스템",
        "url": "https://www.k-apt.go.kr/",
        "as_of": "2026-02-20",
        "usage": "관리품질(하자·장충금·관리비 구조) 정성검증 데이터 소스",
    },
    {
        "category": "official",
        "title": "AirKorea 대기오염정보 OpenAPI",
        "url": "https://www.data.go.kr/data/15073861/openapi.do",
        "as_of": "2026-02-20",
        "usage": "생활환경(대기질) 비정량 지표 연계 후보",
    },
    {
        "category": "official",
        "title": "HUG 공지/보도자료(보증사고·미분양안심환매 등)",
        "url": "https://khug.or.kr/hug/web/cs/no/csno000002.jsp",
        "as_of": "2026-02-20",
        "usage": "전세보증/분양리스크 이벤트를 ADV(보증리스크·공급압력) 해석 보조근거로 사용",
    },
    {
        "category": "official",
        "title": "한국주택금융공사(HF) 정책금융/보증 제도",
        "url": "https://www.hf.go.kr/index.do",
        "as_of": "2026-02-20",
        "usage": "주택금융·보증제도 변동의 레버리지 민감도/실행리스크 해석 보조근거",
    },
    {
        "category": "paper",
        "title": "Housing Prices and Future Returns (Case-Shiller, 1990, NBER w3368)",
        "url": "https://www.nber.org/papers/w3368",
        "as_of": "1990-06-01",
        "usage": "가격-임대료 비율(전세가율 유사)의 예측력 및 밸류 신호 근거",
    },
    {
        "category": "paper",
        "title": "House prices and rents: New evidence from OECD countries (JIMF, 2015)",
        "url": "https://www.sciencedirect.com/science/article/pii/S0261560615001332",
        "as_of": "2015-10-01",
        "usage": "주택가격-임대료 간 장기균형/괴리(밸류에이션 리스크) 근거",
    },
    {
        "category": "paper",
        "title": "Micro-Analysis of Price Spillover Effect (Land, 2021)",
        "url": "https://www.mdpi.com/2073-445X/10/8/879",
        "as_of": "2021-08-21",
        "usage": "서울권역 내 가격 파급효과(상위가격지 주도 전이) 해석 근거",
    },
    {
        "category": "paper",
        "title": "Aging and Real Estate Prices (BIS WP 2020)",
        "url": "https://www.bis.org/publ/work903.htm",
        "as_of": "2020-11-01",
        "usage": "인구구조 변화가 지역 부동산 수요/가격 민감도에 미치는 영향 근거",
    },
    {
        "category": "paper",
        "title": "Household lending, interest rates and housing price bubbles in Korea (Economic Modelling, 2011)",
        "url": "https://www.sciencedirect.com/science/article/abs/pii/S0264999311000198",
        "as_of": "2011-05-01",
        "usage": "금리/가계대출 변화가 버블 국면 변동성에 미치는 영향 근거",
    },
    {
        "category": "paper",
        "title": "Real estate illiquidity and expected returns in private and public real estate (Int. J. Finance & Economics, 2023)",
        "url": "https://onlinelibrary.wiley.com/doi/10.1002/ijfe.2780",
        "as_of": "2023-04-01",
        "usage": "유동성·비유동성 프리미엄을 랭킹/리스크 해석에 반영",
    },
    {
        "category": "paper",
        "title": "Urban rail transit and nearby property values (Cities, 2019)",
        "url": "https://www.sciencedirect.com/science/article/pii/S026427511831490X",
        "as_of": "2019-04-01",
        "usage": "역세권 접근성이 주택가치에 미치는 영향(수요지속성) 근거",
    },
    {
        "category": "paper",
        "title": "School choice, school quality and housing prices (Cities, 2021)",
        "url": "https://www.sciencedirect.com/science/article/pii/S0264275121000138",
        "as_of": "2021-02-01",
        "usage": "학군·교육품질 비정량 점수의 가격 반영 근거",
    },
    {
        "category": "paper",
        "title": "Railway noise and residential property values (Transportation Research D, 2013)",
        "url": "https://www.sciencedirect.com/science/article/pii/S136192091200136X",
        "as_of": "2013-01-01",
        "usage": "소음 리스크가 거주선호/가격에 미치는 음(-)의 영향 근거",
    },
    {
        "category": "paper",
        "title": "Identifying urban flood risk and its effect on property values (Sustainability, 2018)",
        "url": "https://www.mdpi.com/2071-1050/10/11/4008",
        "as_of": "2018-11-01",
        "usage": "침수 취약지역의 가치 할인 및 하방 리스크 보정 근거",
    },
    {
        "category": "paper",
        "title": "Urban park accessibility and housing prices (Sustainability, 2017)",
        "url": "https://www.mdpi.com/2071-1050/9/2/185",
        "as_of": "2017-02-01",
        "usage": "공원 접근성/환경 쾌적성이 거주수요에 주는 프리미엄 근거",
    },
    {
        "category": "paper",
        "title": "News Sentiment and Asset Prices (Review of Financial Studies, 2020)",
        "url": "https://academic.oup.com/rfs/article/33/9/3799/5248705",
        "as_of": "2020-09-01",
        "usage": "심리지표(커뮤니티/뉴스)와 가격 변동성 연결의 방법론 근거",
    },
    {
        "category": "paper",
        "title": "CAViaR: Conditional Autoregressive Value at Risk by Regression Quantiles (J. Business & Economic Statistics, 2004)",
        "url": "https://www.sciencedirect.com/science/article/pii/S0304407604002200",
        "as_of": "2004-10-01",
        "usage": "시나리오 drift를 분위수(P20/P50/P80) 중심으로 구성하는 위험관리 방법론 근거",
    },
    {
        "category": "paper",
        "title": "Common risk factors in the returns on stocks and bonds (Fama & French, 1993)",
        "url": "https://www.sciencedirect.com/science/article/pii/0304405X93900235",
        "as_of": "1993-02-01",
        "usage": "멀티팩터(시장·가치·규모) 프레임의 기초 근거로 팩터형 랭킹 구조 설계",
    },
    {
        "category": "paper",
        "title": "Returns to Buying Winners and Selling Losers: Implications for Stock Market Efficiency (Jegadeesh & Titman, 1993)",
        "url": "https://www.jstor.org/stable/2328882",
        "as_of": "1993-03-01",
        "usage": "모멘텀 팩터(12M·단기 추세)의 정량 근거",
    },
    {
        "category": "paper",
        "title": "Value and Momentum Everywhere (Asness, Moskowitz, Pedersen, 2013)",
        "url": "https://www.sciencedirect.com/science/article/pii/S0304405X13002675",
        "as_of": "2013-12-01",
        "usage": "가치+모멘텀 결합 신호의 일관성 근거(테마/구/단지 동시 적용)",
    },
    {
        "category": "paper",
        "title": "Liquidity Risk and Expected Stock Returns (Pastor & Stambaugh, 2003; NBER w8462)",
        "url": "https://www.nber.org/papers/w8462",
        "as_of": "2001-08-01",
        "usage": "유동성 프리미엄/리스크를 별도 팩터로 반영하는 근거",
    },
    {
        "category": "paper",
        "title": "Illiquidity and stock returns: cross-section and time-series effects (Amihud, 2002)",
        "url": "https://www.sciencedirect.com/science/article/pii/S1386418101000246",
        "as_of": "2002-03-01",
        "usage": "거래량·유동성 기반 팩터 점수(로그 거래량/신선도) 설계 근거",
    },
    {
        "category": "paper",
        "title": "Taming the Factor Zoo: A Test of New Factors (Feng, Giglio, Xiu, 2020)",
        "url": "https://academic.oup.com/rfs/article/33/5/2227/5758276",
        "as_of": "2020-05-01",
        "usage": "다중 팩터 과최적화 방지를 위한 상관축소/보수적 가중 설계 근거",
    },
    {
        "category": "paper",
        "title": "A Rent-Price Ratio for the Housing Market (Campbell, Davis, Gallin, Martin, NBER w14819)",
        "url": "https://www.nber.org/papers/w14819",
        "as_of": "2009-03-01",
        "usage": "임대수익/가격 괴리 기반 밸류·캐리 팩터의 주택시장 적용 근거",
    },
    {
        "category": "paper",
        "title": "The Housing Wealth Effect: The Crucial Roles of Demographics, Wealth Distribution and Social Security (NBER w17740)",
        "url": "https://www.nber.org/papers/w17740",
        "as_of": "2012-01-01",
        "usage": "수요측 구조요인(인구/자산분포)과 가격 민감도의 장기 변화 해석 근거",
    },
    {
        "category": "paper",
        "title": "House Price Momentum and Strategic Complementarity (Guren, NBER w23087)",
        "url": "https://www.nber.org/papers/w23087",
        "as_of": "2018-04-01",
        "usage": "주택시장 모멘텀 지속/반전 메커니즘을 모멘텀 팩터 해석에 반영",
    },
    {
        "category": "community",
        "title": "Reddit r/korea: Housing market sentiment thread",
        "url": "https://www.reddit.com/r/korea/comments/17ir2l6/could_someone_shed_some_lights_on_the_housing/",
        "as_of": "2023-10-27",
        "usage": "외국인/실수요 관점의 체감 유동성·주거비 스트레스 정성검증",
    },
    {
        "category": "community",
        "title": "Reddit r/Living_in_Korea: Local guide to Seoul housing by district",
        "url": "https://www.reddit.com/r/Living_in_Korea/comments/1p7oaut/im_a_local_korean_who_moved_6_times_in_seoul/",
        "as_of": "2025-11-27",
        "usage": "거주 체감(습기/반지하/동네별 체감안전/생활편의) 비정량 리스크 검증",
    },
    {
        "category": "community",
        "title": "Reddit r/Living_in_Korea: Buying condo in Seoul (tax/residency concerns)",
        "url": "https://www.reddit.com/r/Living_in_Korea/comments/1p1d3n1/want_to_buy_a_condo_in_seoul_as_a_koreanamerican/",
        "as_of": "2025-11-20",
        "usage": "실수요·투자자 체감 규제/세금 부담의 현장형 정성검증",
    },
    {
        "category": "community",
        "title": "BiggerPockets: Seoul real estate discussion",
        "url": "https://www.biggerpockets.com/forums/79/topics/11793-seoul-korean-real-estate",
        "as_of": "2026-02-20",
        "usage": "글로벌 투자자 관점의 매수/보유/환금성 이슈 정성검증",
    },
]


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

DARK = "1B2A4A"
BLUE = "2C3E6B"
GOLD = "C9A84C"
WHITE = "FFFFFF"
LIGHT = "F3F5F9"
LINE = "D9D9D9"
GREEN = "27AE60"
RED = "E74C3C"
ORANGE = "F39C12"

FONT_TITLE = Font(name="맑은 고딕", size=14, bold=True, color=WHITE)
FONT_HEADER = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
FONT_CELL = Font(name="맑은 고딕", size=10)
FONT_SMALL = Font(name="맑은 고딕", size=9, color="666666")

FILL_TITLE = PatternFill("solid", fgColor=DARK)
FILL_HEADER = PatternFill("solid", fgColor=BLUE)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT)
FILL_GOLD = PatternFill("solid", fgColor=GOLD)
FILL_WHITE = PatternFill("solid", fgColor=WHITE)

BORDER = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class DistrictInfo:
    code: str
    name: str


@dataclass
class MonthSummary:
    year_month: str
    trade_count: int
    rent_count: int
    median_trade_10k: int | None
    median_rent_10k: int | None
    median_monthly_rent_10k: int | None
    jeonse_ratio_pct: float | None
    rent_yield_proxy_pct: float | None


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------


def clamp(v: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, v))


def pct(v: float | None) -> str:
    if v is None:
        return "-"
    return f"{v * 100:.1f}%"


def eok_from_10k(v: int | None) -> float | None:
    if v is None:
        return None
    return v / 10000.0


def now_yyyymm() -> str:
    t = dt.date.today()
    return f"{t.year:04d}{t.month:02d}"


def latest_complete_yyyymm(today: dt.date | None = None) -> str:
    # 실거래 API 당월 데이터는 통상 미완성(부분월)인 경우가 많아 직전월을 기본으로 사용.
    t = today or dt.date.today()
    cur = f"{t.year:04d}{t.month:02d}"
    return shift_yyyymm(cur, DEFAULT_COMPLETE_MONTH_SHIFT)


def shift_yyyymm(yyyymm: str, delta_months: int) -> str:
    y = int(yyyymm[:4])
    m = int(yyyymm[4:])
    total = y * 12 + (m - 1) + delta_months
    ny = total // 12
    nm = total % 12 + 1
    return f"{ny:04d}{nm:02d}"


def iter_months(end_yyyymm: str, months: int) -> list[str]:
    start_delta = -(months - 1)
    return [shift_yyyymm(end_yyyymm, i) for i in range(start_delta, 1)]


def month_diff(newer_yyyymm: str, older_yyyymm: str) -> int:
    y1, m1 = int(newer_yyyymm[:4]), int(newer_yyyymm[4:])
    y0, m0 = int(older_yyyymm[:4]), int(older_yyyymm[4:])
    return (y1 - y0) * 12 + (m1 - m0)


def median_int(values: list[int]) -> int | None:
    if not values:
        return None
    return int(statistics.median(values))


def stdev(values: list[float]) -> float:
    if len(values) <= 1:
        return 0.0
    return statistics.stdev(values)


def winsorize(values: list[float], alpha: float = 0.10) -> list[float]:
    if not values:
        return []
    if len(values) < 5:
        return values[:]
    sorted_vals = sorted(values)
    n = len(sorted_vals)
    lo_idx = int((n - 1) * alpha)
    hi_idx = int((n - 1) * (1.0 - alpha))
    lo = sorted_vals[lo_idx]
    hi = sorted_vals[hi_idx]
    return [clamp(v, lo, hi) for v in values]


def robust_mean_std(values: list[float]) -> tuple[float, float]:
    vals = [v for v in values if v is not None and not math.isnan(v)]
    if not vals:
        return 0.0, 0.0
    w = winsorize(vals, 0.10)
    mu = statistics.mean(w)
    sigma = statistics.stdev(w) if len(w) > 1 else 0.0
    return mu, sigma


def percentile_value(values: list[float], q: float, default: float = 0.0) -> float:
    vals: list[float] = []
    for v in values:
        try:
            if v is None:
                continue
            fv = float(v)
            if math.isnan(fv):
                continue
            vals.append(fv)
        except Exception:  # noqa: BLE001
            continue
    vals.sort()
    if not vals:
        return default
    qq = clamp(q, 0.0, 1.0)
    n = len(vals)
    if n == 1:
        return vals[0]
    pos = qq * (n - 1)
    lo = int(math.floor(pos))
    hi = int(math.ceil(pos))
    if lo == hi:
        return vals[lo]
    w = pos - lo
    return vals[lo] * (1.0 - w) + vals[hi] * w


def expected_shortfall(returns: list[float], tail_q: float = 0.20) -> float:
    vals = [float(x) for x in returns if x is not None and not math.isnan(x)]
    if not vals:
        return 0.0
    n = len(vals)
    k = max(1, int(math.ceil(n * tail_q)))
    worst = sorted(vals)[:k]
    return statistics.mean(worst)


def safe_div(a: float, b: float, default: float = 0.0) -> float:
    if not b:
        return default
    return a / b


def percentile_rank(value: float | None, arr: list[float], reverse: bool = False) -> float:
    valid = [x for x in arr if x is not None and not math.isnan(x)]
    if value is None or not valid:
        return 50.0
    sorted_vals = sorted(valid)
    n = len(sorted_vals)
    less = sum(1 for x in sorted_vals if x < value)
    equal = sum(1 for x in sorted_vals if x == value)
    rank = (less + 0.5 * equal) / n * 100.0
    if reverse:
        rank = 100.0 - rank
    return clamp(rank, 0.0, 100.0)


def shrink_toward_anchor(value: float | None, anchor: float | None, weight: float) -> float | None:
    if value is None and anchor is None:
        return None
    if value is None:
        return anchor
    if anchor is None:
        return value
    w = clamp(weight, 0.0, 1.0)
    return w * value + (1.0 - w) * anchor


def confidence_bucket(data_quality: float, total_trade_samples: int, coverage_trade: float) -> str:
    if data_quality >= 75 and total_trade_samples >= 120 and coverage_trade >= 0.75:
        return "high"
    if data_quality >= 58 and total_trade_samples >= 45 and coverage_trade >= 0.50:
        return "medium"
    return "low"


def confidence_penalty(bucket: str) -> float:
    if bucket == "high":
        return 0.00
    if bucket == "medium":
        return 0.05
    return 0.12


def norm01(value: float | None, lo: float, hi: float, default: float = 0.5) -> float:
    if value is None:
        return default
    if hi <= lo:
        return default
    try:
        v = float(value)
    except Exception:  # noqa: BLE001
        return default
    return clamp((v - lo) / (hi - lo), 0.0, 1.0)


def safe_float(v: Any, default: float = 0.0) -> float:
    try:
        if v is None:
            return default
        return float(v)
    except Exception:  # noqa: BLE001
        return default


def nq_gap_risk_norm(jeonse_ratio: float | None) -> float:
    j = safe_float(jeonse_ratio, 0.0)
    if j >= 80.0:
        return 1.0
    if j >= 75.0:
        return 0.75
    if j >= 70.0:
        return 0.45
    if 0 < j <= 30.0:
        return 0.70
    if 0 < j <= 35.0:
        return 0.40
    return 0.10


def _parse_nq_dim_record(raw: dict[str, Any] | None) -> tuple[dict[str, float], list[str], str | None]:
    if not raw:
        return {}, [], None
    key_map = {
        "governance": "nq_governance",
        "liveability": "nq_liveability",
        "demand_persistence": "nq_demand_persistence",
        "management_quality": "nq_management_quality",
        "sentiment": "nq_sentiment",
        "nq_governance": "nq_governance",
        "nq_liveability": "nq_liveability",
        "nq_demand_persistence": "nq_demand_persistence",
        "nq_management_quality": "nq_management_quality",
        "nq_sentiment": "nq_sentiment",
    }
    dims: dict[str, float] = {}
    for src, dst in key_map.items():
        if src not in raw:
            continue
        try:
            dims[dst] = clamp(float(raw[src]), 0.0, 100.0)
        except Exception:  # noqa: BLE001
            continue

    raw_flags = raw.get("flags", [])
    flags: list[str] = []
    if isinstance(raw_flags, str):
        flags = [x.strip() for x in raw_flags.split(",") if x.strip()]
    elif isinstance(raw_flags, list):
        flags = [str(x).strip() for x in raw_flags if str(x).strip()]

    conf = raw.get("confidence")
    if conf is not None:
        conf = str(conf).strip().lower()
    return dims, flags, conf


def ensure_nonquant_prior_file(path: Path):
    if path.exists():
        return
    path.write_text(
        json.dumps(DEFAULT_NQ_PRIOR_PAYLOAD, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_nonquant_prior_payload(path: Path | None) -> dict[str, Any]:
    # 내장 기본값을 기반으로, 외부 JSON 파일이 있으면 덮어쓴다.
    base_district = dict(DEFAULT_NQ_PRIOR_PAYLOAD.get("district_priors", {}))
    base_complex = dict(DEFAULT_NQ_PRIOR_PAYLOAD.get("complex_overrides", {}))
    meta = {
        "as_of": DEFAULT_NQ_PRIOR_PAYLOAD.get("as_of", ""),
        "version": DEFAULT_NQ_PRIOR_PAYLOAD.get("version", "v7"),
        "method": DEFAULT_NQ_PRIOR_PAYLOAD.get("method", ""),
        "source": "default",
    }
    if not path:
        return {"district_priors": base_district, "complex_overrides": base_complex, "meta": meta}

    if not path.exists():
        return {"district_priors": base_district, "complex_overrides": base_complex, "meta": meta}

    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001
        return {"district_priors": base_district, "complex_overrides": base_complex, "meta": meta}

    dist_obj = raw.get("district_priors") or raw.get("districts") or {}
    if isinstance(dist_obj, dict):
        for district, vals in dist_obj.items():
            if not isinstance(vals, dict):
                continue
            base_district[str(district).strip()] = vals

    comp_obj = raw.get("complex_overrides") or raw.get("complexes") or {}
    if isinstance(comp_obj, dict):
        for ckey, vals in comp_obj.items():
            if not isinstance(vals, dict):
                continue
            base_complex[str(ckey).strip()] = vals

    meta = dict(meta)
    meta["as_of"] = str(raw.get("as_of") or meta.get("as_of", ""))
    meta["version"] = str(raw.get("version") or meta.get("version", "v7"))
    meta["method"] = str(raw.get("method") or meta.get("method", ""))
    meta["source"] = str(path)
    return {"district_priors": base_district, "complex_overrides": base_complex, "meta": meta}


def _adv_signal_to_pct(v: Any) -> float | None:
    if v is None:
        return None
    try:
        x = float(v)
    except Exception:  # noqa: BLE001
        return None
    if not math.isfinite(x):
        return None
    if x <= 1.0:
        x *= 100.0
    return clamp(x, 0.0, 100.0)


def _parse_adv_signal_record(raw: dict[str, Any] | None) -> tuple[dict[str, float], list[str], str | None, str]:
    if not raw:
        return {}, [], None, ""

    key_map = {
        "supply_pressure": "adv_supply_pressure",
        "supply_risk": "adv_supply_pressure",
        "adv_supply_pressure": "adv_supply_pressure",
        "guarantee_risk": "adv_guarantee_risk",
        "guarantee": "adv_guarantee_risk",
        "adv_guarantee_risk": "adv_guarantee_risk",
        "management_dd_risk": "adv_management_dd_risk",
        "management_risk": "adv_management_dd_risk",
        "adv_management_dd_risk": "adv_management_dd_risk",
    }

    dims: dict[str, float] = {}
    for src, dst in key_map.items():
        if src not in raw:
            continue
        pct_v = _adv_signal_to_pct(raw.get(src))
        if pct_v is not None:
            dims[dst] = pct_v

    raw_flags = raw.get("flags", [])
    flags: list[str] = []
    if isinstance(raw_flags, str):
        flags = [x.strip() for x in raw_flags.split(",") if x.strip()]
    elif isinstance(raw_flags, list):
        flags = [str(x).strip() for x in raw_flags if str(x).strip()]

    conf = raw.get("confidence")
    if conf is None:
        conf = raw.get("confidence_bucket")
    if conf is not None:
        conf = str(conf).strip().lower()

    source = str(raw.get("source") or raw.get("data_source") or "").strip()
    return dims, flags, conf, source


def ensure_advanced_signal_file(path: Path):
    if path.exists():
        return
    path.write_text(
        json.dumps(DEFAULT_ADV_SIGNAL_PAYLOAD, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_advanced_signal_payload(path: Path | None) -> dict[str, Any]:
    base_district = dict(DEFAULT_ADV_SIGNAL_PAYLOAD.get("district_signals", {}))
    base_complex = dict(DEFAULT_ADV_SIGNAL_PAYLOAD.get("complex_signals", {}))
    base_guide = dict(DEFAULT_ADV_SIGNAL_PAYLOAD.get("guide", {}))
    meta = {
        "as_of": DEFAULT_ADV_SIGNAL_PAYLOAD.get("as_of", ""),
        "version": DEFAULT_ADV_SIGNAL_PAYLOAD.get("version", "v7"),
        "method": DEFAULT_ADV_SIGNAL_PAYLOAD.get("method", ""),
        "source": "default",
    }
    if not path:
        return {"district_signals": base_district, "complex_signals": base_complex, "guide": base_guide, "meta": meta}

    if not path.exists():
        return {"district_signals": base_district, "complex_signals": base_complex, "guide": base_guide, "meta": meta}

    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001
        return {"district_signals": base_district, "complex_signals": base_complex, "guide": base_guide, "meta": meta}

    raw_meta = raw.get("meta") or {}
    if isinstance(raw_meta, dict):
        for mk, mv in raw_meta.items():
            meta[str(mk)] = mv

    guide_obj = raw.get("guide") or {}
    if isinstance(guide_obj, dict):
        for k, v in guide_obj.items():
            base_guide[str(k)] = v

    dist_obj = raw.get("district_signals") or raw.get("districts") or {}
    if isinstance(dist_obj, dict):
        for district, vals in dist_obj.items():
            if not isinstance(vals, dict):
                continue
            base_district[str(district).strip()] = vals

    comp_obj = raw.get("complex_signals") or raw.get("complexes") or {}
    if isinstance(comp_obj, dict):
        for ckey, vals in comp_obj.items():
            if not isinstance(vals, dict):
                continue
            base_complex[str(ckey).strip()] = vals

    meta = {
        "as_of": str(raw.get("as_of") or meta["as_of"]),
        "version": str(raw.get("version") or meta["version"]),
        "method": str(raw.get("method") or meta["method"]),
        "source": str(path),
    }
    last_changes = raw.get("autofill_last_changes")
    if not isinstance(last_changes, dict):
        last_changes = {}
    run_history = raw.get("autofill_run_history")
    if not isinstance(run_history, list):
        run_history = []

    return {
        "district_signals": base_district,
        "complex_signals": base_complex,
        "guide": base_guide,
        "meta": meta,
        "autofill_last_changes": last_changes,
        "autofill_run_history": run_history,
    }


def save_advanced_signal_payload(path: Path, payload: dict[str, Any]):
    meta = payload.get("meta") if isinstance(payload, dict) else {}
    if not isinstance(meta, dict):
        meta = {}
    raw = {
        "as_of": str(meta.get("as_of") or dt.date.today().isoformat()),
        "version": str(meta.get("version") or DEFAULT_ADV_SIGNAL_PAYLOAD.get("version", "v7")),
        "method": str(meta.get("method") or DEFAULT_ADV_SIGNAL_PAYLOAD.get("method", "")),
        "meta": meta,
        "guide": payload.get("guide") if isinstance(payload.get("guide"), dict) else dict(DEFAULT_ADV_SIGNAL_PAYLOAD.get("guide", {})),
        "district_signals": payload.get("district_signals") if isinstance(payload.get("district_signals"), dict) else {},
        "complex_signals": payload.get("complex_signals") if isinstance(payload.get("complex_signals"), dict) else {},
        "autofill_last_changes": payload.get("autofill_last_changes") if isinstance(payload.get("autofill_last_changes"), dict) else {},
        "autofill_run_history": payload.get("autofill_run_history") if isinstance(payload.get("autofill_run_history"), list) else [],
    }
    path.write_text(json.dumps(raw, ensure_ascii=False, indent=2), encoding="utf-8")


def _adv_autofill_conf_bucket(conf_score: float) -> str:
    if conf_score >= 0.76:
        return "high"
    if conf_score >= 0.56:
        return "medium"
    return "low"


def _adv_merge_flags(a: list[str], b: list[str]) -> list[str]:
    out: list[str] = []
    for arr in (a, b):
        for x in arr:
            sx = str(x).strip()
            if sx and sx not in out:
                out.append(sx)
    return out


def _adv_signal_row_snapshot(row: dict[str, Any]) -> dict[str, Any]:
    flags: list[str] = []
    raw_flags = row.get("flags", [])
    if isinstance(raw_flags, str):
        flags = [x.strip() for x in raw_flags.split(",") if x.strip()]
    elif isinstance(raw_flags, list):
        flags = [str(x).strip() for x in raw_flags if str(x).strip()]
    flags = _adv_merge_flags(flags, [])

    conf = str(row.get("confidence") or row.get("confidence_bucket") or "").strip().lower()
    if conf not in {"high", "medium", "low"}:
        conf = ""

    return {
        "supply_pressure": _adv_signal_to_pct(row.get("supply_pressure")),
        "guarantee_risk": _adv_signal_to_pct(row.get("guarantee_risk")),
        "management_dd_risk": _adv_signal_to_pct(row.get("management_dd_risk")),
        "confidence": conf,
        "source": str(row.get("source") or "").strip(),
        "flags": flags,
    }


def _adv_changed_fields(before: dict[str, Any], after: dict[str, Any]) -> list[str]:
    fields = ["supply_pressure", "guarantee_risk", "management_dd_risk", "confidence", "source", "flags"]
    changed: list[str] = []
    for f in fields:
        if before.get(f) != after.get(f):
            changed.append(f)
    return changed


def _adv_pct_eq(a: float | None, b: float | None) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return abs(float(a) - float(b)) <= 1e-9


def _adv_autofill_dimensions(
    m: dict[str, Any],
    *,
    arr_growth_neg: list[float],
    arr_liq: list[float],
    arr_stale: list[float],
    arr_yield: list[float],
    arr_premium: list[float],
    arr_jeonse: list[float],
    arr_tail: list[float],
    arr_dq: list[float],
    arr_nq_gov: list[float],
    arr_nq_mgt: list[float],
    arr_unc: list[float],
) -> tuple[dict[str, float], str, list[str]]:
    g12 = safe_float(m.get("growth_12m"), 0.0)
    liq = safe_float(m.get("avg_trade_3m"), 0.0)
    stale = safe_float(m.get("trade_stale_months"), 0.0)
    yld = safe_float(m.get("latest_rent_yield"), 2.0)
    pvs = safe_float(m.get("price_vs_seoul"), 1.0)
    jeonse = safe_float(m.get("latest_jeonse_ratio"), 0.0)
    tail = safe_float(m.get("tail_es_abs"), 0.0)
    dq = safe_float(m.get("data_quality"), 60.0)
    nq_gov = safe_float(m.get("nq_governance"), 50.0)
    nq_mgt = safe_float(m.get("nq_management_quality"), 50.0)
    unc = safe_float(m.get("uncertainty_penalty"), 0.0)

    decel_rank = percentile_rank(-g12, arr_growth_neg)
    illiq_rank = percentile_rank(liq, arr_liq, reverse=True)
    stale_rank = percentile_rank(stale, arr_stale)
    low_yield_rank = percentile_rank(yld, arr_yield, reverse=True)
    premium_rank = percentile_rank(pvs, arr_premium)
    gap_rank = percentile_rank(jeonse, arr_jeonse)
    tail_rank = percentile_rank(tail, arr_tail)
    dq_low_rank = percentile_rank(dq, arr_dq, reverse=True)
    gov_low_rank = percentile_rank(nq_gov, arr_nq_gov, reverse=True)
    mgt_low_rank = percentile_rank(nq_mgt, arr_nq_mgt, reverse=True)
    unc_rank = percentile_rank(unc, arr_unc)

    supply = clamp(
        0.35 * decel_rank
        + 0.23 * illiq_rank
        + 0.16 * low_yield_rank
        + 0.16 * premium_rank
        + 0.10 * stale_rank,
        0.0,
        100.0,
    )
    guarantee = clamp(
        0.44 * gap_rank
        + 0.22 * tail_rank
        + 0.20 * stale_rank
        + 0.14 * illiq_rank,
        0.0,
        100.0,
    )
    mgmt = clamp(
        0.38 * dq_low_rank
        + 0.24 * gov_low_rank
        + 0.20 * mgt_low_rank
        + 0.18 * unc_rank,
        0.0,
        100.0,
    )

    conf_score = clamp(
        0.45 * norm01(dq, 40.0, 100.0, default=0.5)
        + 0.30 * clamp(safe_float(m.get("coverage_trade"), 0.0), 0.0, 1.0)
        + 0.15 * clamp(safe_float(m.get("market_freshness"), 0.0), 0.0, 1.0)
        + 0.10 * (1.0 - clamp(safe_div(stale, 6.0), 0.0, 1.0)),
        0.0,
        1.0,
    )
    conf = _adv_autofill_conf_bucket(conf_score)

    flags: list[str] = []
    if supply >= 64.0:
        flags.append("공급압력")
    if guarantee >= 64.0:
        flags.append("보증회수리스크")
    if mgmt >= 64.0:
        flags.append("관리DD리스크")
    if jeonse >= 78.0:
        flags.append("고전세가율")
    if stale >= 3.0:
        flags.append("거래공백")
    if dq <= 55.0:
        flags.append("데이터취약")

    dims = {
        "adv_supply_pressure": round(supply, 2),
        "adv_guarantee_risk": round(guarantee, 2),
        "adv_management_dd_risk": round(mgmt, 2),
    }
    return dims, conf, flags


def autofill_advanced_signal_districts(
    metrics: list[dict[str, Any]],
    payload: dict[str, Any],
    *,
    months: list[str],
    overwrite: bool,
    run_id: str | None = None,
) -> dict[str, Any]:
    district_signals = payload.get("district_signals") if isinstance(payload.get("district_signals"), dict) else {}
    district_signals = dict(district_signals)

    arr_growth_neg = [-safe_float(x.get("growth_12m"), 0.0) for x in metrics]
    arr_liq = [safe_float(x.get("avg_trade_3m"), 0.0) for x in metrics]
    arr_stale = [safe_float(x.get("trade_stale_months"), 0.0) for x in metrics]
    arr_yield = [safe_float(x.get("latest_rent_yield"), 2.0) for x in metrics]
    arr_premium = [safe_float(x.get("price_vs_seoul"), 1.0) for x in metrics]
    arr_jeonse = [safe_float(x.get("latest_jeonse_ratio"), 0.0) for x in metrics]
    arr_tail = [safe_float(x.get("tail_es_abs"), 0.0) for x in metrics]
    arr_dq = [safe_float(x.get("data_quality"), 60.0) for x in metrics]
    arr_nq_gov = [safe_float(x.get("nq_governance"), 50.0) for x in metrics]
    arr_nq_mgt = [safe_float(x.get("nq_management_quality"), 50.0) for x in metrics]
    arr_unc = [safe_float(x.get("uncertainty_penalty"), 0.0) for x in metrics]
    source = f"autofill:molit_rtms:{months[0]}-{months[-1]}"

    updated_rows = 0
    updated_dims = 0
    changed_dims = 0
    assigned_rows = 0
    run_id_eff = run_id or dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_at = dt.datetime.now().isoformat(timespec="seconds")
    month_range = f"{months[0]}~{months[-1]}"

    prev_last_changes = payload.get("autofill_last_changes") if isinstance(payload.get("autofill_last_changes"), dict) else {}
    prev_complex_changes = prev_last_changes.get("complex") if isinstance(prev_last_changes.get("complex"), dict) else {}
    district_changes: dict[str, Any] = {}

    for m in metrics:
        district = str(m.get("district") or "").strip()
        if not district:
            continue
        auto_dims, auto_conf, auto_flags = _adv_autofill_dimensions(
            m,
            arr_growth_neg=arr_growth_neg,
            arr_liq=arr_liq,
            arr_stale=arr_stale,
            arr_yield=arr_yield,
            arr_premium=arr_premium,
            arr_jeonse=arr_jeonse,
            arr_tail=arr_tail,
            arr_dq=arr_dq,
            arr_nq_gov=arr_nq_gov,
            arr_nq_mgt=arr_nq_mgt,
            arr_unc=arr_unc,
        )

        cur = district_signals.get(district)
        row = dict(cur) if isinstance(cur, dict) else {}
        before_snap = _adv_signal_row_snapshot(row)
        _, cur_flags, _, _ = _parse_adv_signal_record(row)
        cur_source_txt = str(row.get("source") or "").strip().lower()
        placeholder_source = cur_source_txt in {"", "fill_with_official_data", "manual_pending", "todo", "tbd"}
        assigned_this_row = 0

        key_map = {
            "supply_pressure": "adv_supply_pressure",
            "guarantee_risk": "adv_guarantee_risk",
            "management_dd_risk": "adv_management_dd_risk",
        }
        for rk, ak in key_map.items():
            cur_val = _adv_signal_to_pct(row.get(rk))
            if overwrite or cur_val is None:
                row[rk] = auto_dims[ak]
                updated_dims += 1
                assigned_this_row += 1
                if not _adv_pct_eq(cur_val, auto_dims[ak]):
                    changed_dims += 1

        if overwrite or placeholder_source or str(row.get("confidence") or "").lower() not in {"high", "medium", "low"}:
            assigned_this_row += 1
            if str(before_snap.get("confidence") or "") != auto_conf:
                changed_dims += 1
            row["confidence"] = auto_conf
        if overwrite or placeholder_source:
            assigned_this_row += 1
            if str(before_snap.get("source") or "") != source:
                changed_dims += 1
            row["source"] = source

        row["flags"] = _adv_merge_flags(cur_flags, auto_flags)
        district_signals[district] = row
        if assigned_this_row > 0:
            assigned_rows += 1

        after_snap = _adv_signal_row_snapshot(row)
        changed_fields = _adv_changed_fields(before_snap, after_snap)
        if changed_fields:
            updated_rows += 1
        if changed_fields or (overwrite and assigned_this_row > 0):
            district_changes[district] = {
                "level": "district",
                "key": district,
                "changed_fields": changed_fields,
                "touched": assigned_this_row > 0,
                "changed": bool(changed_fields),
                "before": before_snap,
                "after": after_snap,
                "run_id": run_id_eff,
                "run_at": run_at,
                "month_range": month_range,
            }

    meta = payload.get("meta") if isinstance(payload.get("meta"), dict) else {}
    meta = dict(meta)
    meta["as_of"] = dt.date.today().isoformat()
    meta["version"] = str(meta.get("version") or DEFAULT_ADV_SIGNAL_PAYLOAD.get("version", "v7"))
    meta["method"] = (
        "external supplemental risk signals (supply/guarantee/management DD) + "
        "autofill from MOLIT trade/rent public panel"
    )
    meta["autofill_enabled"] = True
    meta["autofill_overwrite"] = bool(overwrite)
    meta["autofill_month_range"] = month_range
    meta["autofill_run_id"] = run_id_eff
    meta["autofill_run_at"] = run_at
    meta["autofill_total_rows_district"] = int(len(metrics))
    meta["autofill_signal_rows_district"] = int(len(district_signals))
    meta["autofill_assigned_rows_district"] = int(assigned_rows)
    meta["autofill_updated_rows_district"] = int(updated_rows)
    meta["autofill_updated_dims_district"] = int(updated_dims)
    meta["autofill_changed_rows_district"] = int(updated_rows)
    meta["autofill_changed_dims_district"] = int(changed_dims)

    payload["district_signals"] = district_signals
    payload["meta"] = meta
    payload["autofill_last_changes"] = {
        "run_id": run_id_eff,
        "run_at": run_at,
        "month_range": month_range,
        "district": district_changes,
        "complex": prev_complex_changes,
    }
    return payload


def autofill_advanced_signal_complexes(
    complex_metrics: list[dict[str, Any]],
    payload: dict[str, Any],
    *,
    months: list[str],
    overwrite: bool,
    run_id: str | None = None,
) -> dict[str, Any]:
    complex_signals = payload.get("complex_signals") if isinstance(payload.get("complex_signals"), dict) else {}
    complex_signals = dict(complex_signals)

    arr_growth_neg = [-safe_float(x.get("growth_12m"), 0.0) for x in complex_metrics]
    arr_liq = [safe_float(x.get("avg_trade_3m"), 0.0) for x in complex_metrics]
    arr_stale = [safe_float(x.get("trade_stale_months"), 0.0) for x in complex_metrics]
    arr_yield = [safe_float(x.get("latest_rent_yield"), 2.0) for x in complex_metrics]
    arr_premium = [safe_float(x.get("price_vs_seoul"), 1.0) for x in complex_metrics]
    arr_jeonse = [safe_float(x.get("latest_jeonse_ratio"), 0.0) for x in complex_metrics]
    arr_tail = [safe_float(x.get("tail_es_abs"), 0.0) for x in complex_metrics]
    arr_dq = [safe_float(x.get("data_quality"), 60.0) for x in complex_metrics]
    arr_nq_gov = [safe_float(x.get("nq_governance"), 50.0) for x in complex_metrics]
    arr_nq_mgt = [safe_float(x.get("nq_management_quality"), 50.0) for x in complex_metrics]
    arr_unc = [safe_float(x.get("uncertainty_penalty"), 0.0) for x in complex_metrics]
    source = f"autofill:molit_rtms_complex:{months[0]}-{months[-1]}"

    updated_rows = 0
    updated_dims = 0
    changed_dims = 0
    assigned_rows = 0
    run_id_eff = run_id or dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_at = dt.datetime.now().isoformat(timespec="seconds")
    month_range = f"{months[0]}~{months[-1]}"

    prev_last_changes = payload.get("autofill_last_changes") if isinstance(payload.get("autofill_last_changes"), dict) else {}
    prev_district_changes = prev_last_changes.get("district") if isinstance(prev_last_changes.get("district"), dict) else {}
    complex_changes: dict[str, Any] = {}

    for m in complex_metrics:
        ckey = str(m.get("complex_key") or "").strip()
        if not ckey:
            district = str(m.get("district") or "").strip()
            cname = str(m.get("complex") or "").strip()
            if district and cname:
                ckey = f"{district}::{cname}"
            else:
                continue
        auto_dims, auto_conf, auto_flags = _adv_autofill_dimensions(
            m,
            arr_growth_neg=arr_growth_neg,
            arr_liq=arr_liq,
            arr_stale=arr_stale,
            arr_yield=arr_yield,
            arr_premium=arr_premium,
            arr_jeonse=arr_jeonse,
            arr_tail=arr_tail,
            arr_dq=arr_dq,
            arr_nq_gov=arr_nq_gov,
            arr_nq_mgt=arr_nq_mgt,
            arr_unc=arr_unc,
        )

        cur = complex_signals.get(ckey)
        row = dict(cur) if isinstance(cur, dict) else {}
        before_snap = _adv_signal_row_snapshot(row)
        _, cur_flags, _, _ = _parse_adv_signal_record(row)
        cur_source_txt = str(row.get("source") or "").strip().lower()
        placeholder_source = cur_source_txt in {"", "fill_with_official_data", "manual_pending", "todo", "tbd"}
        assigned_this_row = 0

        key_map = {
            "supply_pressure": "adv_supply_pressure",
            "guarantee_risk": "adv_guarantee_risk",
            "management_dd_risk": "adv_management_dd_risk",
        }
        for rk, ak in key_map.items():
            cur_val = _adv_signal_to_pct(row.get(rk))
            if overwrite or cur_val is None:
                row[rk] = auto_dims[ak]
                updated_dims += 1
                assigned_this_row += 1
                if not _adv_pct_eq(cur_val, auto_dims[ak]):
                    changed_dims += 1

        if overwrite or placeholder_source or str(row.get("confidence") or "").lower() not in {"high", "medium", "low"}:
            assigned_this_row += 1
            if str(before_snap.get("confidence") or "") != auto_conf:
                changed_dims += 1
            row["confidence"] = auto_conf
        if overwrite or placeholder_source:
            assigned_this_row += 1
            if str(before_snap.get("source") or "") != source:
                changed_dims += 1
            row["source"] = source

        row["flags"] = _adv_merge_flags(cur_flags, auto_flags)
        complex_signals[ckey] = row
        if assigned_this_row > 0:
            assigned_rows += 1

        after_snap = _adv_signal_row_snapshot(row)
        changed_fields = _adv_changed_fields(before_snap, after_snap)
        if changed_fields:
            updated_rows += 1
        if changed_fields or (overwrite and assigned_this_row > 0):
            complex_changes[ckey] = {
                "level": "complex",
                "key": ckey,
                "changed_fields": changed_fields,
                "touched": assigned_this_row > 0,
                "changed": bool(changed_fields),
                "before": before_snap,
                "after": after_snap,
                "run_id": run_id_eff,
                "run_at": run_at,
                "month_range": month_range,
            }

    meta = payload.get("meta") if isinstance(payload.get("meta"), dict) else {}
    meta = dict(meta)
    meta["as_of"] = dt.date.today().isoformat()
    meta["version"] = str(meta.get("version") or DEFAULT_ADV_SIGNAL_PAYLOAD.get("version", "v7"))
    meta["method"] = (
        "external supplemental risk signals (supply/guarantee/management DD) + "
        "autofill from MOLIT trade/rent public panel"
    )
    meta["autofill_enabled"] = True
    meta["autofill_overwrite"] = bool(overwrite)
    meta["autofill_month_range"] = month_range
    meta["autofill_run_id"] = run_id_eff
    meta["autofill_run_at"] = run_at
    meta["autofill_total_rows_complex"] = int(len(complex_metrics))
    meta["autofill_signal_rows_complex"] = int(len(complex_signals))
    meta["autofill_assigned_rows_complex"] = int(assigned_rows)
    meta["autofill_updated_rows_complex"] = int(updated_rows)
    meta["autofill_updated_dims_complex"] = int(updated_dims)
    meta["autofill_changed_rows_complex"] = int(updated_rows)
    meta["autofill_changed_dims_complex"] = int(changed_dims)

    payload["complex_signals"] = complex_signals
    payload["meta"] = meta
    payload["autofill_last_changes"] = {
        "run_id": run_id_eff,
        "run_at": run_at,
        "month_range": month_range,
        "district": prev_district_changes,
        "complex": complex_changes,
    }

    run_history = payload.get("autofill_run_history") if isinstance(payload.get("autofill_run_history"), list) else []
    run_summary = {
        "run_id": run_id_eff,
        "run_at": run_at,
        "month_range": month_range,
        "overwrite": bool(overwrite),
        "district_total": int(meta.get("autofill_total_rows_district", 0)),
        "district_signal_rows": int(meta.get("autofill_signal_rows_district", 0)),
        "district_assigned_rows": int(meta.get("autofill_assigned_rows_district", 0)),
        "district_changed_rows": int(meta.get("autofill_changed_rows_district", 0)),
        "district_changed_dims": int(meta.get("autofill_changed_dims_district", 0)),
        "complex_total": int(meta.get("autofill_total_rows_complex", 0)),
        "complex_signal_rows": int(meta.get("autofill_signal_rows_complex", 0)),
        "complex_assigned_rows": int(meta.get("autofill_assigned_rows_complex", 0)),
        "complex_changed_rows": int(meta.get("autofill_changed_rows_complex", 0)),
        "complex_changed_dims": int(meta.get("autofill_changed_dims_complex", 0)),
    }
    run_history = [x for x in run_history if not (isinstance(x, dict) and x.get("run_id") == run_id_eff)]
    run_history.append(run_summary)
    payload["autofill_run_history"] = run_history[-40:]
    return payload


def _adv_conf_score_from_bucket(conf: str | None) -> float:
    c = str(conf or "").strip().lower()
    if c == "high":
        return 0.86
    if c == "medium":
        return 0.66
    if c == "low":
        return 0.46
    return 0.58


def _model_nonquant_dimensions(m: dict[str, Any], district_name: str | None = None) -> tuple[dict[str, float], list[str], float]:
    district_name = district_name or str(m.get("district") or "")
    fresh = clamp(safe_float(m.get("market_freshness"), 0.0), 0.0, 1.0)
    stale_norm = clamp(safe_div(safe_float(m.get("trade_stale_months"), 0.0), 6.0), 0.0, 1.0)
    liq_norm = norm01(m.get("avg_trade_3m"), 0.0, 15.0, default=0.30)
    dq_norm = norm01(m.get("data_quality"), 40.0, 100.0, default=0.50)
    cov_trade = clamp(safe_float(m.get("coverage_trade"), 0.0), 0.0, 1.0)
    cov_rent = clamp(safe_float(m.get("coverage_rent"), 0.0), 0.0, 1.0)
    trend12 = norm01(m.get("growth_12m"), -0.10, 0.15, default=0.50)
    trend3 = norm01(m.get("growth_3m"), -0.04, 0.06, default=0.50)
    vol_norm = norm01(m.get("vol_ann"), 0.05, 0.35, default=0.45)
    yield_norm = norm01(m.get("latest_rent_yield"), 1.0, 4.0, default=0.50)
    tail_norm = norm01(m.get("tail_es_abs"), 0.0, 0.10, default=0.25)
    gap_norm = nq_gap_risk_norm(m.get("latest_jeonse_ratio"))
    pvs = safe_float(m.get("price_vs_seoul"), 1.0)
    if pvs <= 0:
        pvs = 1.0
    price_premium = norm01(pvs, 0.80, 1.60, default=0.50)
    affordability = clamp(1.0 - abs(math.log(pvs) / math.log(2.0)), 0.0, 1.0)

    governance = 45.0 + 22.0 * fresh + 10.0 * liq_norm + 8.0 * dq_norm + 8.0 * cov_trade
    if district_name in HIGH_INTENSITY_LTPS:
        governance -= 15.0
    governance -= 10.0 * gap_norm + 8.0 * stale_norm + 6.0 * tail_norm

    liveability = 46.0 + 20.0 * price_premium + 16.0 * (1.0 - vol_norm) + 12.0 * cov_rent + 10.0 * yield_norm + 8.0 * affordability
    liveability -= 8.0 * tail_norm + 6.0 * stale_norm

    demand = 40.0 + 28.0 * liq_norm + 18.0 * fresh + 16.0 * trend12 + 8.0 * trend3 + 6.0 * cov_trade
    demand -= 8.0 * tail_norm + 6.0 * gap_norm

    management = 42.0 + 32.0 * dq_norm + 16.0 * cov_trade + 12.0 * cov_rent + 8.0 * fresh
    management -= 10.0 * stale_norm + 8.0 * tail_norm

    sentiment = 38.0 + 28.0 * trend12 + 14.0 * trend3 + 16.0 * liq_norm + 10.0 * fresh
    sentiment -= 14.0 * tail_norm + 10.0 * gap_norm + 6.0 * stale_norm

    dims = {
        "nq_governance": clamp(governance, 0.0, 100.0),
        "nq_liveability": clamp(liveability, 0.0, 100.0),
        "nq_demand_persistence": clamp(demand, 0.0, 100.0),
        "nq_management_quality": clamp(management, 0.0, 100.0),
        "nq_sentiment": clamp(sentiment, 0.0, 100.0),
    }

    flags: list[str] = []
    if district_name in HIGH_INTENSITY_LTPS:
        flags.append("집중규제")
    if stale_norm >= 0.50:
        flags.append("거래공백")
    if gap_norm >= 0.75:
        flags.append("갭과열")
    if vol_norm >= 0.85:
        flags.append("고변동성")
    if dq_norm <= 0.35:
        flags.append("표본취약")

    model_cov = (dq_norm + fresh + cov_trade) / 3.0
    return dims, flags, clamp(model_cov, 0.0, 1.0)


def _compose_nonquant_dims(
    *,
    model_dims: dict[str, float],
    model_flags: list[str],
    model_cov: float,
    prior_dims: dict[str, float] | None,
    prior_flags: list[str] | None,
    prior_conf: str | None,
    prior_weight: float,
    override_dims: dict[str, float] | None = None,
    override_flags: list[str] | None = None,
    override_conf: str | None = None,
    uncertainty_penalty: float = 0.0,
) -> dict[str, Any]:
    prior_dims = prior_dims or {}
    prior_flags = prior_flags or []
    override_dims = override_dims or {}
    override_flags = override_flags or []

    pw = clamp(prior_weight, 0.0, 1.0)
    dims: dict[str, float] = {}
    for k in NQ_DIM_KEYS:
        mv = model_dims.get(k)
        pv = prior_dims.get(k)
        ov = override_dims.get(k)
        if ov is not None:
            if mv is None and pv is None:
                dims[k] = ov
            elif pv is None:
                dims[k] = clamp(0.70 * ov + 0.30 * mv, 0.0, 100.0)
            elif mv is None:
                dims[k] = clamp(0.80 * ov + 0.20 * pv, 0.0, 100.0)
            else:
                dims[k] = clamp(0.65 * ov + 0.20 * pv + 0.15 * mv, 0.0, 100.0)
            continue
        if pv is not None and mv is not None:
            dims[k] = clamp(pw * pv + (1.0 - pw) * mv, 0.0, 100.0)
        elif pv is not None:
            dims[k] = clamp(pv, 0.0, 100.0)
        elif mv is not None:
            dims[k] = clamp(mv, 0.0, 100.0)
        else:
            dims[k] = 50.0

    flags: list[str] = []
    for arr in [prior_flags, model_flags, override_flags]:
        for f in arr:
            sf = str(f).strip()
            if sf and sf not in flags:
                flags.append(sf)

    raw = (
        0.24 * dims["nq_governance"]
        + 0.24 * dims["nq_liveability"]
        + 0.22 * dims["nq_demand_persistence"]
        + 0.18 * dims["nq_management_quality"]
        + 0.12 * dims["nq_sentiment"]
    )
    flag_pen = min(14.0, len(flags) * 1.8 + clamp(uncertainty_penalty, 0.0, 0.20) * 25.0)
    comp = clamp(raw - flag_pen, 0.0, 100.0)
    risk_index = 100.0 - comp

    prior_cov = 1.0 if prior_dims else 0.0
    conf_score = clamp(0.55 * prior_cov + 0.45 * model_cov, 0.0, 1.0)
    conf_hint = (override_conf or prior_conf or "").lower()
    if conf_hint == "high":
        conf_score = max(conf_score, 0.82)
    elif conf_hint == "low":
        conf_score = min(conf_score, 0.45)
    elif conf_hint == "medium":
        conf_score = max(conf_score, 0.58)

    if conf_score >= 0.75:
        conf_bucket = "high"
    elif conf_score >= 0.55:
        conf_bucket = "medium"
    else:
        conf_bucket = "low"

    return {
        **dims,
        "nq_composite": round(comp, 2),
        "nq_risk_index": round(risk_index, 2),
        "nq_flag_count": len(flags),
        "nq_flags": ", ".join(flags) if flags else "",
        "nq_confidence_score": round(conf_score, 3),
        "nq_confidence_bucket": conf_bucket,
    }


def apply_nonquant_to_district_metrics(
    metrics: list[dict[str, Any]],
    prior_payload: dict[str, Any],
    *,
    prior_weight: float,
    enabled: bool,
) -> list[dict[str, Any]]:
    priors_raw = prior_payload.get("district_priors", {}) if prior_payload else {}
    for m in metrics:
        if not enabled:
            m.update(
                {
                    "nq_governance": 50.0,
                    "nq_liveability": 50.0,
                    "nq_demand_persistence": 50.0,
                    "nq_management_quality": 50.0,
                    "nq_sentiment": 50.0,
                    "nq_composite": 50.0,
                    "nq_risk_index": 50.0,
                    "nq_flag_count": 0,
                    "nq_flags": "비정량 비활성",
                    "nq_confidence_score": 0.0,
                    "nq_confidence_bucket": "low",
                    "nq_source": "disabled",
                }
            )
            continue

        prior_raw = priors_raw.get(m.get("district"), {}) if isinstance(priors_raw, dict) else {}
        prior_dims, prior_flags, prior_conf = _parse_nq_dim_record(prior_raw if isinstance(prior_raw, dict) else {})
        model_dims, model_flags, model_cov = _model_nonquant_dimensions(m, m.get("district"))
        nq = _compose_nonquant_dims(
            model_dims=model_dims,
            model_flags=model_flags,
            model_cov=model_cov,
            prior_dims=prior_dims,
            prior_flags=prior_flags,
            prior_conf=prior_conf,
            prior_weight=prior_weight,
            uncertainty_penalty=0.0,
        )
        nq["nq_source"] = "prior+model" if prior_dims else "model"
        m.update(nq)
    return metrics


def apply_nonquant_to_complex_metrics(
    complex_metrics: list[dict[str, Any]],
    district_metrics: list[dict[str, Any]],
    prior_payload: dict[str, Any],
    *,
    prior_weight: float,
    enabled: bool,
) -> list[dict[str, Any]]:
    district_nq_map = {d["district"]: d for d in district_metrics}
    c_overrides_raw = prior_payload.get("complex_overrides", {}) if prior_payload else {}
    for m in complex_metrics:
        if not enabled:
            m.update(
                {
                    "nq_governance": 50.0,
                    "nq_liveability": 50.0,
                    "nq_demand_persistence": 50.0,
                    "nq_management_quality": 50.0,
                    "nq_sentiment": 50.0,
                    "nq_composite": 50.0,
                    "nq_risk_index": 50.0,
                    "nq_flag_count": 0,
                    "nq_flags": "비정량 비활성",
                    "nq_confidence_score": 0.0,
                    "nq_confidence_bucket": "low",
                    "nq_source": "disabled",
                }
            )
            continue

        parent = district_nq_map.get(m.get("district"), {})
        parent_dims = {k: safe_float(parent.get(k), 50.0) for k in NQ_DIM_KEYS}
        parent_flags = [x.strip() for x in str(parent.get("nq_flags") or "").split(",") if x.strip()]
        parent_conf = str(parent.get("nq_confidence_bucket") or "medium")

        model_dims, model_flags, model_cov = _model_nonquant_dimensions(m, m.get("district"))
        # 단지 모델치는 구 비정량 특성을 65% 앵커링해 노이즈를 완화한다.
        anchored_model: dict[str, float] = {}
        for k in NQ_DIM_KEYS:
            anchored_model[k] = clamp(0.65 * parent_dims.get(k, 50.0) + 0.35 * model_dims.get(k, 50.0), 0.0, 100.0)

        ckey = str(m.get("complex_key") or "")
        override_raw = c_overrides_raw.get(ckey, {}) if isinstance(c_overrides_raw, dict) else {}
        ov_dims, ov_flags, ov_conf = _parse_nq_dim_record(override_raw if isinstance(override_raw, dict) else {})

        nq = _compose_nonquant_dims(
            model_dims=anchored_model,
            model_flags=model_flags + parent_flags,
            model_cov=clamp((model_cov + safe_float(parent.get("nq_confidence_score"), 0.5)) / 2.0, 0.0, 1.0),
            prior_dims=parent_dims,
            prior_flags=parent_flags,
            prior_conf=parent_conf,
            prior_weight=prior_weight,
            override_dims=ov_dims,
            override_flags=ov_flags,
            override_conf=ov_conf,
            uncertainty_penalty=safe_float(m.get("uncertainty_penalty"), 0.0),
        )
        nq["nq_source"] = "complex_override" if ov_dims else "district_anchor+model"
        m.update(nq)
    return complex_metrics


def nonquant_overlay_for_profile(m: dict[str, Any], profile: str, cap: float) -> float:
    if cap <= 0:
        return 0.0
    w = NQ_PROFILE_WEIGHTS.get(profile) or NQ_PROFILE_WEIGHTS["투자자"]
    signal = 0.0
    for k, wk in w.items():
        v = safe_float(m.get(k), 50.0)
        signal += ((v - 50.0) / 50.0) * wk

    conf_bucket = str(m.get("nq_confidence_bucket") or "medium").lower()
    conf_mult = 1.0 if conf_bucket == "high" else 0.85 if conf_bucket == "medium" else 0.65
    flag_count = int(safe_float(m.get("nq_flag_count"), 0.0))
    flag_drag = min(0.55, flag_count * 0.08)
    overlay = cap * signal * conf_mult * (1.0 - flag_drag)
    overlay -= min(cap * 0.30, flag_count * 0.35)
    return clamp(overlay, -cap, cap)


def _bucket_from_score(score: float, hi: float = 70.0, mid: float = 55.0) -> str:
    if score >= hi:
        return "high"
    if score >= mid:
        return "medium"
    return "low"


def model_advanced_risk(m: dict[str, Any]) -> dict[str, Any]:
    district = str(m.get("district") or "")
    fresh = clamp(safe_float(m.get("market_freshness"), 0.0), 0.0, 1.0)
    fresh_risk = 1.0 - fresh
    stale_norm = clamp(safe_div(safe_float(m.get("trade_stale_months"), 0.0), 6.0), 0.0, 1.0)
    illiq_norm = clamp(1.0 - safe_div(safe_float(m.get("avg_trade_3m"), 0.0), 12.0), 0.0, 1.0)
    dq_norm = norm01(m.get("data_quality"), 40.0, 100.0, default=0.50)
    dq_risk = 1.0 - dq_norm
    cov_trade = clamp(safe_float(m.get("coverage_trade"), 0.0), 0.0, 1.0)
    vol_norm = norm01(m.get("vol_ann"), 0.05, 0.35, default=0.45)
    tail_norm = norm01(m.get("tail_es_abs"), 0.0, 0.10, default=0.25)
    gap_norm = nq_gap_risk_norm(m.get("latest_jeonse_ratio"))
    trend12_norm = norm01(m.get("growth_12m"), -0.05, 0.15, default=0.50)
    trend3_norm = norm01(m.get("growth_3m"), -0.03, 0.06, default=0.50)
    yield_low = 1.0 - norm01(m.get("latest_rent_yield"), 1.0, 4.0, default=0.50)
    pvs = safe_float(m.get("price_vs_seoul"), 1.0)
    if pvs <= 0:
        pvs = 1.0
    premium_norm = norm01(pvs, 0.85, 1.80, default=0.45)
    policy_score = safe_float(m.get("policy_score"), 45.0)
    policy_risk = 1.0 - norm01(policy_score, 30.0, 72.0, default=0.40)
    ltps_intense = 1.0 if district in HIGH_INTENSITY_LTPS else 0.0
    sent_mania = norm01(m.get("nq_sentiment"), 58.0, 90.0, default=0.35)
    gov_risk = clamp(1.0 - safe_div(safe_float(m.get("nq_governance"), 50.0), 100.0), 0.0, 1.0)
    unc_norm = norm01(m.get("uncertainty_penalty"), 0.0, 0.20, default=0.22)
    nq_flag_norm = clamp(safe_div(safe_float(m.get("nq_flag_count"), 0.0), 8.0), 0.0, 1.0)
    mgt_risk = clamp(1.0 - safe_div(safe_float(m.get("nq_management_quality"), 50.0), 100.0), 0.0, 1.0)

    # 실전 DD 관점의 보조 리스크 축 (공급/보증/관리)
    supply_pressure = clamp(
        0.38 * (1.0 - trend12_norm)
        + 0.22 * illiq_norm
        + 0.20 * yield_low
        + 0.20 * premium_norm,
        0.0,
        1.0,
    )
    guarantee_risk = clamp(
        0.45 * gap_norm
        + 0.20 * tail_norm
        + 0.20 * stale_norm
        + 0.15 * illiq_norm,
        0.0,
        1.0,
    )
    management_dd_risk = clamp(
        0.42 * dq_risk
        + 0.28 * gov_risk
        + 0.15 * unc_norm
        + 0.15 * max(mgt_risk, nq_flag_norm),
        0.0,
        1.0,
    )

    liq_trap = (
        18.0
        + 34.0 * illiq_norm
        + 22.0 * stale_norm
        + 12.0 * fresh_risk
        + 10.0 * dq_risk
        + 8.0 * tail_norm
        + 8.0 * gap_norm
        + 6.0 * premium_norm
        + 9.0 * supply_pressure
    )
    crowding = (
        14.0
        + 24.0 * trend12_norm
        + 19.0 * trend3_norm
        + 15.0 * premium_norm
        + 14.0 * gap_norm
        + 12.0 * yield_low
        + 10.0 * sent_mania
        + 8.0 * vol_norm
    )
    policy_flip = (
        16.0
        + 30.0 * ltps_intense
        + 16.0 * gap_norm
        + 14.0 * policy_risk
        + 12.0 * premium_norm
        + 8.0 * illiq_norm
        + 8.0 * sent_mania
        + 6.0 * trend12_norm
        + 7.0 * supply_pressure
    )
    execution = (
        16.0
        + 26.0 * dq_risk
        + 21.0 * stale_norm
        + 16.0 * tail_norm
        + 13.0 * unc_norm
        + 12.0 * gov_risk
        + 8.0 * nq_flag_norm
        + 8.0 * fresh_risk
        + 11.0 * guarantee_risk
        + 12.0 * management_dd_risk
    )

    liq_trap = clamp(liq_trap, 0.0, 100.0)
    crowding = clamp(crowding, 0.0, 100.0)
    policy_flip = clamp(policy_flip, 0.0, 100.0)
    execution = clamp(execution, 0.0, 100.0)

    flags: list[str] = []
    if liq_trap >= 68.0:
        flags.append("유동성함정")
    if crowding >= 70.0:
        flags.append("군집쏠림")
    if policy_flip >= 68.0:
        flags.append("정책플립")
    if execution >= 66.0:
        flags.append("실행리스크")
    if supply_pressure >= 0.62:
        flags.append("공급압력")
    if guarantee_risk >= 0.60:
        flags.append("보증회수리스크")
    if management_dd_risk >= 0.60:
        flags.append("관리DD리스크")
    if ltps_intense > 0 and premium_norm >= 0.60:
        flags.append("고가규제중첩")
    if stale_norm >= 0.50 and illiq_norm >= 0.60:
        flags.append("거래절벽")

    comp_raw = 0.32 * liq_trap + 0.23 * crowding + 0.27 * policy_flip + 0.18 * execution
    comp = clamp(comp_raw + min(8.0, len(flags) * 1.1), 0.0, 100.0)

    conf_score = clamp(0.45 * dq_norm + 0.35 * (1.0 - stale_norm) + 0.20 * cov_trade, 0.0, 1.0)
    if conf_score >= 0.75:
        conf_bucket = "high"
    elif conf_score >= 0.55:
        conf_bucket = "medium"
    else:
        conf_bucket = "low"

    return {
        "adv_liquidity_trap": round(liq_trap, 2),
        "adv_crowding": round(crowding, 2),
        "adv_policy_flip": round(policy_flip, 2),
        "adv_execution": round(execution, 2),
        "adv_supply_pressure": round(supply_pressure * 100.0, 2),
        "adv_guarantee_risk": round(guarantee_risk * 100.0, 2),
        "adv_management_dd_risk": round(management_dd_risk * 100.0, 2),
        "adv_risk_composite": round(comp, 2),
        "adv_flag_count": len(flags),
        "adv_flags": ", ".join(flags) if flags else "",
        "adv_confidence_score": round(conf_score, 3),
        "adv_confidence_bucket": conf_bucket,
        "adv_risk_bucket": _bucket_from_score(comp, hi=70.0, mid=55.0),
    }


def _resolve_adv_signal_raw(
    m: dict[str, Any],
    *,
    level: str,
    district_signals: dict[str, Any],
    complex_signals: dict[str, Any],
) -> tuple[dict[str, Any], str]:
    district = str(m.get("district") or "").strip()
    if level == "complex":
        ckey = str(m.get("complex_key") or "").strip()
        if ckey and isinstance(complex_signals.get(ckey), dict):
            return complex_signals[ckey], "complex"
        cname = str(m.get("complex") or "").strip()
        if district and cname:
            ckey2 = f"{district}::{cname}"
            if isinstance(complex_signals.get(ckey2), dict):
                return complex_signals[ckey2], "complex"
        if district and isinstance(district_signals.get(district), dict):
            return district_signals[district], "district_fallback"
        return {}, "none"

    if district and isinstance(district_signals.get(district), dict):
        return district_signals[district], "district"
    return {}, "none"


def _compose_advanced_risk(
    *,
    model_adv: dict[str, Any],
    signal_dims: dict[str, float],
    signal_flags: list[str],
    signal_conf: str | None,
    signal_source: str,
    signal_scope: str,
    level: str,
) -> dict[str, Any]:
    out = dict(model_adv)

    model_flags = [x.strip() for x in str(model_adv.get("adv_flags") or "").split(",") if x.strip()]
    m_supply = safe_float(model_adv.get("adv_supply_pressure"), 50.0)
    m_guar = safe_float(model_adv.get("adv_guarantee_risk"), 50.0)
    m_mgt = safe_float(model_adv.get("adv_management_dd_risk"), 50.0)
    m_liq = safe_float(model_adv.get("adv_liquidity_trap"), 50.0)
    m_crowd = safe_float(model_adv.get("adv_crowding"), 50.0)
    m_policy = safe_float(model_adv.get("adv_policy_flip"), 50.0)
    m_exec = safe_float(model_adv.get("adv_execution"), 50.0)
    out["adv_model_supply_pressure"] = round(m_supply, 2)
    out["adv_model_guarantee_risk"] = round(m_guar, 2)
    out["adv_model_management_dd_risk"] = round(m_mgt, 2)
    out["adv_signal_supply_pressure"] = round(safe_float(signal_dims.get("adv_supply_pressure"), float("nan")), 2) if "adv_supply_pressure" in signal_dims else None
    out["adv_signal_guarantee_risk"] = round(safe_float(signal_dims.get("adv_guarantee_risk"), float("nan")), 2) if "adv_guarantee_risk" in signal_dims else None
    out["adv_signal_management_dd_risk"] = round(safe_float(signal_dims.get("adv_management_dd_risk"), float("nan")), 2) if "adv_management_dd_risk" in signal_dims else None
    out["adv_signal_confidence"] = str(signal_conf or "").lower()
    out["adv_signal_flags"] = ", ".join([x for x in signal_flags if str(x).strip()]) if signal_flags else ""

    coverage = safe_div(len(signal_dims), 3.0)
    if coverage <= 0:
        out["adv_source"] = "model"
        out["adv_signal_scope"] = "none"
        out["adv_signal_weight"] = 0.0
        out["adv_signal_source"] = ""
        return out

    conf_mult = {"high": 0.85, "medium": 0.60, "low": 0.38}.get(str(signal_conf or "").lower(), 0.52)
    base_weight = 0.56 if level == "district" else 0.48
    blend_w = clamp(base_weight * conf_mult * (0.70 + 0.30 * coverage), 0.15, 0.80)

    def blended(model_v: float, k: str) -> float:
        sv = signal_dims.get(k)
        if sv is None:
            return model_v
        return clamp((1.0 - blend_w) * model_v + blend_w * sv, 0.0, 100.0)

    supply = blended(m_supply, "adv_supply_pressure")
    guar = blended(m_guar, "adv_guarantee_risk")
    mgt = blended(m_mgt, "adv_management_dd_risk")

    d_supply = supply - m_supply
    d_guar = guar - m_guar
    d_mgt = mgt - m_mgt

    liq = clamp(m_liq + 0.20 * d_supply + 0.08 * d_guar, 0.0, 100.0)
    crowd = clamp(m_crowd + 0.08 * d_supply, 0.0, 100.0)
    policy = clamp(m_policy + 0.16 * d_supply + 0.10 * d_guar, 0.0, 100.0)
    execution = clamp(m_exec + 0.08 * d_supply + 0.16 * d_guar + 0.20 * d_mgt, 0.0, 100.0)

    flags: list[str] = []
    for f in model_flags + signal_flags:
        sf = str(f).strip()
        if sf and sf not in flags:
            flags.append(sf)

    if supply >= 62.0 and "공급압력" not in flags:
        flags.append("공급압력")
    if guar >= 60.0 and "보증회수리스크" not in flags:
        flags.append("보증회수리스크")
    if mgt >= 60.0 and "관리DD리스크" not in flags:
        flags.append("관리DD리스크")
    if liq >= 68.0 and "유동성함정" not in flags:
        flags.append("유동성함정")
    if crowd >= 70.0 and "군집쏠림" not in flags:
        flags.append("군집쏠림")
    if policy >= 68.0 and "정책플립" not in flags:
        flags.append("정책플립")
    if execution >= 66.0 and "실행리스크" not in flags:
        flags.append("실행리스크")
    if signal_scope == "district_fallback" and level == "complex" and "구시그널대체" not in flags:
        flags.append("구시그널대체")

    comp_raw = 0.32 * liq + 0.23 * crowd + 0.27 * policy + 0.18 * execution
    comp = clamp(comp_raw + min(8.0, len(flags) * 1.1), 0.0, 100.0)

    model_conf = clamp(safe_float(model_adv.get("adv_confidence_score"), 0.5), 0.0, 1.0)
    signal_conf = _adv_conf_score_from_bucket(signal_conf)
    conf_score = clamp(0.70 * model_conf + 0.30 * signal_conf, 0.0, 1.0)
    if signal_scope == "district_fallback" and level == "complex":
        conf_score = clamp(conf_score * 0.92, 0.0, 1.0)
    if conf_score >= 0.75:
        conf_bucket = "high"
    elif conf_score >= 0.55:
        conf_bucket = "medium"
    else:
        conf_bucket = "low"

    out.update(
        {
            "adv_liquidity_trap": round(liq, 2),
            "adv_crowding": round(crowd, 2),
            "adv_policy_flip": round(policy, 2),
            "adv_execution": round(execution, 2),
            "adv_supply_pressure": round(supply, 2),
            "adv_guarantee_risk": round(guar, 2),
            "adv_management_dd_risk": round(mgt, 2),
            "adv_risk_composite": round(comp, 2),
            "adv_flag_count": len(flags),
            "adv_flags": ", ".join(flags) if flags else "",
            "adv_confidence_score": round(conf_score, 3),
            "adv_confidence_bucket": conf_bucket,
            "adv_risk_bucket": _bucket_from_score(comp, hi=70.0, mid=55.0),
            "adv_source": "model+signal",
            "adv_signal_scope": signal_scope,
            "adv_signal_weight": round(blend_w, 3),
            "adv_signal_source": signal_source or "external",
        }
    )
    return out


def apply_advanced_risk_to_metrics(
    metrics: list[dict[str, Any]],
    *,
    enabled: bool,
    signal_payload: dict[str, Any] | None = None,
    level: str = "district",
) -> list[dict[str, Any]]:
    signal_payload = signal_payload or {}
    district_signals = signal_payload.get("district_signals") if isinstance(signal_payload, dict) else {}
    complex_signals = signal_payload.get("complex_signals") if isinstance(signal_payload, dict) else {}
    if not isinstance(district_signals, dict):
        district_signals = {}
    if not isinstance(complex_signals, dict):
        complex_signals = {}

    for m in metrics:
        if not enabled:
            m.update(
                {
                    "adv_liquidity_trap": 50.0,
                    "adv_crowding": 50.0,
                    "adv_policy_flip": 50.0,
                    "adv_execution": 50.0,
                    "adv_supply_pressure": 50.0,
                    "adv_guarantee_risk": 50.0,
                    "adv_management_dd_risk": 50.0,
                    "adv_risk_composite": 50.0,
                    "adv_flag_count": 0,
                    "adv_flags": "고급리스크 비활성",
                    "adv_confidence_score": 0.0,
                    "adv_confidence_bucket": "low",
                    "adv_risk_bucket": "medium",
                    "adv_source": "disabled",
                    "adv_signal_scope": "disabled",
                    "adv_signal_weight": 0.0,
                    "adv_signal_source": "",
                    "adv_model_supply_pressure": 50.0,
                    "adv_model_guarantee_risk": 50.0,
                    "adv_model_management_dd_risk": 50.0,
                    "adv_signal_supply_pressure": None,
                    "adv_signal_guarantee_risk": None,
                    "adv_signal_management_dd_risk": None,
                    "adv_signal_confidence": "",
                    "adv_signal_flags": "",
                }
            )
            continue
        model_adv = model_advanced_risk(m)
        raw_signal, scope = _resolve_adv_signal_raw(
            m,
            level=level,
            district_signals=district_signals,
            complex_signals=complex_signals,
        )
        signal_dims, signal_flags, signal_conf, signal_source = _parse_adv_signal_record(raw_signal if isinstance(raw_signal, dict) else {})
        adv = _compose_advanced_risk(
            model_adv=model_adv,
            signal_dims=signal_dims,
            signal_flags=signal_flags,
            signal_conf=signal_conf,
            signal_source=signal_source,
            signal_scope=scope,
            level=level,
        )
        m.update(adv)
    return metrics


def advanced_risk_penalty_for_profile(m: dict[str, Any], profile: str, cap: float) -> float:
    if cap <= 0:
        return 0.0
    w = ADV_RISK_PROFILE_WEIGHTS.get(profile) or ADV_RISK_PROFILE_WEIGHTS["투자자"]
    signal = 0.0
    for k, wk in w.items():
        signal += clamp(safe_float(m.get(k), 50.0) / 100.0, 0.0, 1.0) * wk

    conf_bucket = str(m.get("adv_confidence_bucket") or m.get("nq_confidence_bucket") or "medium").lower()
    conf_mult = 1.0 if conf_bucket == "high" else 0.90 if conf_bucket == "medium" else 0.78
    flag_count = int(safe_float(m.get("adv_flag_count"), 0.0))
    flag_mult = 1.0 + min(0.35, flag_count * 0.06)
    penalty = cap * signal * conf_mult * flag_mult

    comp = safe_float(m.get("adv_risk_composite"), 50.0)
    if comp >= 80.0:
        penalty = max(penalty, cap * 0.62)
    elif comp >= 70.0:
        penalty = max(penalty, cap * 0.45)
    return clamp(penalty, 0.0, cap)


def style_cell(cell, *, font=None, fill=None, align=None):
    cell.font = font or FONT_CELL
    cell.fill = fill or FILL_WHITE
    cell.alignment = align or ALIGN_CENTER
    cell.border = BORDER


def write_title(ws, row: int, title: str, col_end: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=1, value=title)
    c.font = FONT_TITLE
    c.fill = FILL_TITLE
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = 30
    return row + 1


def write_headers(ws, row: int, headers: list[str]) -> int:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        style_cell(c, font=FONT_HEADER, fill=FILL_HEADER)
    ws.row_dimensions[row].height = 22
    return row + 1


def write_row(ws, row: int, values: list[Any], fill: PatternFill | None = None, left: bool = False) -> int:
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        style_cell(c, fill=fill or FILL_WHITE, align=ALIGN_LEFT if left else ALIGN_CENTER)
    return row + 1


def set_widths(ws, widths: list[float]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def score_band_fill(score: float) -> PatternFill:
    if score >= 75:
        return PatternFill("solid", fgColor="D5F5E3")
    if score >= 55:
        return PatternFill("solid", fgColor="FCF3CF")
    return PatternFill("solid", fgColor="FADBD8")


# ---------------------------------------------------------------------------
# Region / API helpers
# ---------------------------------------------------------------------------


def get_api_key(explicit: str | None = None) -> str:
    if explicit:
        return explicit.strip()
    env = os.getenv("DATA_GO_KR_API_KEY", "").strip()
    if env:
        return env
    return DEFAULT_DATA_GO_KR_API_KEY


def load_seoul_gu_list(region_code_file: Path) -> list[DistrictInfo]:
    text = region_code_file.read_text(encoding="utf-8")
    rows: list[DistrictInfo] = []
    prefix = "\uc11c\uc6b8\ud2b9\ubcc4\uc2dc "
    for line in text.splitlines():
        parts = line.split("\t")
        if len(parts) < 2:
            continue
        code10 = parts[0].strip()
        name = parts[1].strip()
        if not (code10.isdigit() and len(code10) == 10):
            continue
        if code10 == "1100000000":
            continue
        if not code10.startswith("11"):
            continue
        if not code10.endswith("00000"):
            continue
        if not name.startswith(prefix):
            continue
        gu_name = name.replace(prefix, "")
        rows.append(DistrictInfo(code=code10[:5], name=gu_name))

    rows.sort(key=lambda x: x.code)
    if len(rows) != 25:
        raise RuntimeError(f"서울 자치구 코드 25개를 찾지 못했습니다. 현재: {len(rows)}")
    return rows


def build_url(base_url: str, api_key: str, region_code: str, yyyymm: str, num_rows: int, page_no: int) -> str:
    params = {
        "serviceKey": api_key,
        "LAWD_CD": region_code,
        "DEAL_YMD": yyyymm,
        "numOfRows": str(num_rows),
        "pageNo": str(page_no),
    }
    return f"{base_url}?{urllib.parse.urlencode(params)}"


def fetch_xml(url: str, timeout_sec: int = 25, retries: int = 4) -> str:
    backoff = 0.6
    last_err: Exception | None = None
    for _ in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "KRE-Seoul-Analyzer/1.0"})
            with urllib.request.urlopen(req, timeout=timeout_sec) as resp:
                return resp.read().decode("utf-8", errors="ignore")
        except Exception as e:  # noqa: BLE001
            last_err = e
            time.sleep(backoff + random.random() * 0.3)
            backoff *= 1.9
    raise RuntimeError(f"API 요청 실패: {last_err}")


def parse_xml_items_trade(xml_text: str) -> tuple[list[dict[str, Any]], int, str | None]:
    root = ET.fromstring(xml_text)
    code = (root.findtext(".//resultCode") or "").strip()
    if code != "000":
        return [], 0, code

    total_count = int((root.findtext(".//totalCount") or "0").strip() or "0")
    out: list[dict[str, Any]] = []
    for item in root.findall(".//item"):
        if ((item.findtext("cdealType") or "").strip()) == "O":
            continue
        raw_price = (item.findtext("dealAmount") or "").replace(",", "").strip()
        if not raw_price.isdigit():
            continue
        area_raw = (item.findtext("excluUseAr") or "0").strip()
        try:
            area = float(area_raw)
        except ValueError:
            area = 0.0
        out.append(
            {
                "apt_name": (item.findtext("aptNm") or "").strip(),
                "dong": (item.findtext("umdNm") or "").strip(),
                "area_sqm": area,
                "price_10k": int(raw_price),
                "year": int((item.findtext("dealYear") or "0").strip() or 0),
                "month": int((item.findtext("dealMonth") or "0").strip() or 0),
                "day": int((item.findtext("dealDay") or "0").strip() or 0),
            }
        )
    return out, total_count, None


def parse_xml_items_rent(xml_text: str) -> tuple[list[dict[str, Any]], int, str | None]:
    root = ET.fromstring(xml_text)
    code = (root.findtext(".//resultCode") or "").strip()
    if code != "000":
        return [], 0, code

    total_count = int((root.findtext(".//totalCount") or "0").strip() or "0")
    out: list[dict[str, Any]] = []
    for item in root.findall(".//item"):
        if ((item.findtext("cdealType") or "").strip()) == "O":
            continue
        raw_dep = (item.findtext("deposit") or "").replace(",", "").strip()
        raw_mon = (item.findtext("monthlyRent") or "0").replace(",", "").strip() or "0"
        if not raw_dep.isdigit():
            continue
        area_raw = (item.findtext("excluUseAr") or "0").strip()
        try:
            area = float(area_raw)
        except ValueError:
            area = 0.0
        out.append(
            {
                "apt_name": (item.findtext("aptNm") or "").strip(),
                "dong": (item.findtext("umdNm") or "").strip(),
                "area_sqm": area,
                "deposit_10k": int(raw_dep),
                "monthly_rent_10k": int(raw_mon) if raw_mon.isdigit() else 0,
                "contract_type": (item.findtext("contractType") or "").strip(),
                "year": int((item.findtext("dealYear") or "0").strip() or 0),
                "month": int((item.findtext("dealMonth") or "0").strip() or 0),
                "day": int((item.findtext("dealDay") or "0").strip() or 0),
            }
        )
    return out, total_count, None


def fetch_all_pages(
    *,
    base_url: str,
    api_key: str,
    region_code: str,
    yyyymm: str,
    num_rows: int,
    parse_kind: str,
    max_pages: int = 80,
) -> tuple[list[dict[str, Any]], int, str | None]:
    parser = parse_xml_items_trade if parse_kind == "trade" else parse_xml_items_rent
    all_items: list[dict[str, Any]] = []
    total_count = 0

    for page in range(1, max_pages + 1):
        url = build_url(base_url, api_key, region_code, yyyymm, num_rows, page)
        xml_text = fetch_xml(url)
        items, tcount, err = parser(xml_text)
        if err is not None:
            return [], 0, err

        total_count = max(total_count, tcount)
        all_items.extend(items)

        if not items:
            break
        if len(items) < num_rows:
            break
        if total_count and len(all_items) >= total_count:
            break

    return all_items, total_count, None


def filter_area(items: list[dict[str, Any]], area_min: float, area_max: float) -> list[dict[str, Any]]:
    return [x for x in items if area_min <= float(x.get("area_sqm", 0.0)) <= area_max]


def normalize_apt_name(name: str) -> str:
    t = (name or "").strip()
    t = " ".join(t.split())
    if not t:
        return "(미상단지)"
    return t


def build_complex_month_snapshot(
    *,
    district_name: str,
    yyyymm: str,
    trade_items: list[dict[str, Any]],
    rent_items: list[dict[str, Any]],
    deposit_yield_pct: float,
) -> dict[str, dict[str, Any]]:
    trade_map: dict[str, list[int]] = {}
    rent_dep_map: dict[str, list[int]] = {}
    rent_mon_map: dict[str, list[int]] = {}

    for it in trade_items:
        nm = normalize_apt_name(str(it.get("apt_name") or ""))
        p = int(it.get("price_10k") or 0)
        if p <= 0:
            continue
        trade_map.setdefault(nm, []).append(p)

    for it in rent_items:
        nm = normalize_apt_name(str(it.get("apt_name") or ""))
        d = int(it.get("deposit_10k") or 0)
        m = int(it.get("monthly_rent_10k") or 0)
        if d <= 0:
            continue
        rent_dep_map.setdefault(nm, []).append(d)
        if m > 0:
            rent_mon_map.setdefault(nm, []).append(m)

    names = sorted(set(trade_map) | set(rent_dep_map))
    out: dict[str, dict[str, Any]] = {}
    for nm in names:
        prices = trade_map.get(nm, [])
        deps = rent_dep_map.get(nm, [])
        mons = rent_mon_map.get(nm, [])

        med_price = median_int(prices)
        med_dep = median_int(deps)
        med_mon = median_int(mons)
        jeonse_ratio = None
        if med_price and med_dep:
            jeonse_ratio = med_dep / med_price * 100.0

        rent_yield_proxy = None
        if med_price and (med_dep or med_mon):
            annual_income_10k = (med_mon or 0) * 12 + int((med_dep or 0) * (deposit_yield_pct / 100.0))
            rent_yield_proxy = annual_income_10k / med_price * 100.0

        out[nm] = {
            "district": district_name,
            "year_month": yyyymm,
            "complex": nm,
            "trade_count": len(prices),
            "rent_count": len(deps),
            "median_trade_10k": med_price,
            "median_rent_10k": med_dep,
            "median_monthly_rent_10k": med_mon,
            "jeonse_ratio_pct": jeonse_ratio,
            "rent_yield_proxy_pct": rent_yield_proxy,
        }

    return out


def summarize_month(
    *,
    yyyymm: str,
    trade_items: list[dict[str, Any]],
    rent_items: list[dict[str, Any]],
    deposit_yield_pct: float,
) -> MonthSummary:
    prices = [int(x["price_10k"]) for x in trade_items if int(x.get("price_10k", 0)) > 0]
    deposits = [int(x["deposit_10k"]) for x in rent_items if int(x.get("deposit_10k", 0)) > 0]
    monthly_rents = [
        int(x["monthly_rent_10k"])
        for x in rent_items
        if int(x.get("monthly_rent_10k", 0)) > 0
    ]

    med_price = median_int(prices)
    med_deposit = median_int(deposits)
    med_monthly = median_int(monthly_rents)

    jeonse_ratio = None
    if med_price and med_deposit:
        jeonse_ratio = med_deposit / med_price * 100.0

    rent_yield_proxy = None
    if med_price and (med_deposit or med_monthly):
        annual_income_10k = (med_monthly or 0) * 12 + int((med_deposit or 0) * (deposit_yield_pct / 100.0))
        rent_yield_proxy = annual_income_10k / med_price * 100.0

    return MonthSummary(
        year_month=yyyymm,
        trade_count=len(prices),
        rent_count=len(deposits),
        median_trade_10k=med_price,
        median_rent_10k=med_deposit,
        median_monthly_rent_10k=med_monthly,
        jeonse_ratio_pct=jeonse_ratio,
        rent_yield_proxy_pct=rent_yield_proxy,
    )

# ---------------------------------------------------------------------------
# Collection and analytics
# ---------------------------------------------------------------------------


def collect_one_district_month(
    district: DistrictInfo,
    yyyymm: str,
    api_key: str,
    num_rows: int,
    area_min: float,
    area_max: float,
    deposit_yield_pct: float,
) -> tuple[str, str, dict[str, Any]]:
    t0 = time.time()

    trade_items, _, trade_err = fetch_all_pages(
        base_url=APT_TRADE_URL,
        api_key=api_key,
        region_code=district.code,
        yyyymm=yyyymm,
        num_rows=num_rows,
        parse_kind="trade",
    )
    if trade_err:
        return district.name, yyyymm, {
            "error": f"trade_api_error:{trade_err}",
            "elapsed": round(time.time() - t0, 2),
        }

    rent_items, _, rent_err = fetch_all_pages(
        base_url=APT_RENT_URL,
        api_key=api_key,
        region_code=district.code,
        yyyymm=yyyymm,
        num_rows=num_rows,
        parse_kind="rent",
    )
    if rent_err:
        return district.name, yyyymm, {
            "error": f"rent_api_error:{rent_err}",
            "elapsed": round(time.time() - t0, 2),
        }

    trade_f = filter_area(trade_items, area_min, area_max)
    rent_f = filter_area(rent_items, area_min, area_max)
    summary = summarize_month(
        yyyymm=yyyymm,
        trade_items=trade_f,
        rent_items=rent_f,
        deposit_yield_pct=deposit_yield_pct,
    )
    complex_snapshot = build_complex_month_snapshot(
        district_name=district.name,
        yyyymm=yyyymm,
        trade_items=trade_f,
        rent_items=rent_f,
        deposit_yield_pct=deposit_yield_pct,
    )

    return district.name, yyyymm, {
        "error": None,
        "elapsed": round(time.time() - t0, 2),
        "summary": summary.__dict__,
        "complexes": complex_snapshot,
        "raw_counts": {
            "trade_total": len(trade_items),
            "trade_filtered": len(trade_f),
            "rent_total": len(rent_items),
            "rent_filtered": len(rent_f),
        },
    }


def collect_market_panel(
    *,
    districts: list[DistrictInfo],
    months: list[str],
    api_key: str,
    num_rows: int,
    max_workers: int,
    area_min: float,
    area_max: float,
    deposit_yield_pct: float,
) -> tuple[dict[str, dict[str, dict[str, Any]]], list[dict[str, Any]]]:
    panel: dict[str, dict[str, dict[str, Any]]] = {d.name: {} for d in districts}
    logs: list[dict[str, Any]] = []

    with cf.ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {}
        for d in districts:
            for ym in months:
                fut = ex.submit(
                    collect_one_district_month,
                    d,
                    ym,
                    api_key,
                    num_rows,
                    area_min,
                    area_max,
                    deposit_yield_pct,
                )
                futures[fut] = (d.name, ym)

        total = len(futures)
        done = 0
        for fut in cf.as_completed(futures):
            name, ym = futures[fut]
            done += 1
            try:
                n, y, payload = fut.result()
            except Exception as e:  # noqa: BLE001
                payload = {"error": f"exception:{e}", "elapsed": None}
                n, y = name, ym

            panel[n][y] = payload
            logs.append({"district": n, "year_month": y, **payload})
            if done % 40 == 0 or done == total:
                print(f"[collect] {done}/{total} 완료")

    return panel, logs


def ffill(values: list[float | None]) -> list[float | None]:
    out: list[float | None] = []
    last = None
    for v in values:
        if v is None:
            out.append(last)
        else:
            out.append(v)
            last = v
    return out


def compute_max_drawdown(series: list[float]) -> float:
    if not series:
        return 0.0
    peak = series[0]
    mdd = 0.0
    for x in series:
        peak = max(peak, x)
        if peak > 0:
            dd = (x / peak) - 1.0
            mdd = min(mdd, dd)
    return abs(mdd)


def annualized_vol_from_prices(prices: list[float]) -> float:
    rets: list[float] = []
    for i in range(1, len(prices)):
        p0 = prices[i - 1]
        p1 = prices[i]
        if p0 > 0 and p1 > 0:
            rets.append(math.log(p1 / p0))
    if len(rets) <= 1:
        return 0.0
    return stdev(rets) * math.sqrt(12.0)


def growth(prices: list[float | None], lag: int) -> float | None:
    if len(prices) <= lag or lag <= 0:
        return None
    p1 = prices[-1]
    p0 = prices[-1 - lag]
    if p1 is None or p0 is None:
        return None
    if p0 <= 0:
        return None
    return p1 / p0 - 1.0


def build_district_metrics(
    districts: list[DistrictInfo],
    months: list[str],
    panel: dict[str, dict[str, dict[str, Any]]],
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []

    for d in districts:
        month_rows = []
        for ym in months:
            payload = panel.get(d.name, {}).get(ym, {})
            if payload.get("error"):
                month_rows.append({"year_month": ym, "error": payload.get("error")})
                continue
            summary = payload.get("summary") or {}
            raw_counts = payload.get("raw_counts") or {}
            month_rows.append(
                {
                    "year_month": ym,
                    **summary,
                    **raw_counts,
                }
            )

        trade_series: list[float | None] = [
            (r.get("median_trade_10k") / 10000.0) if r.get("median_trade_10k") else None
            for r in month_rows
        ]
        rent_series: list[float | None] = [
            (r.get("median_rent_10k") / 10000.0) if r.get("median_rent_10k") else None
            for r in month_rows
        ]
        jeonse_series: list[float | None] = [r.get("jeonse_ratio_pct") for r in month_rows]
        yield_series: list[float | None] = [r.get("rent_yield_proxy_pct") for r in month_rows]

        trade_counts = [int(r.get("trade_count") or 0) for r in month_rows]
        rent_counts = [int(r.get("rent_count") or 0) for r in month_rows]

        trade_ffill = ffill(trade_series)
        rent_ffill = ffill(rent_series)

        valid_price = [x for x in trade_ffill if x is not None and x > 0]
        latest_price = valid_price[-1] if valid_price else None

        valid_rent = [x for x in rent_ffill if x is not None and x > 0]
        latest_rent = valid_rent[-1] if valid_rent else None

        lag_3m = min(3, max(1, len(trade_ffill) - 1))
        lag_6m = min(6, max(1, len(trade_ffill) - 1))
        # 12개월 데이터가 있으면 끝-처음(11간격)으로 YoY에 준하는 성장률을 계산한다.
        lag_12m = min(11, max(1, len(trade_ffill) - 1))
        g3 = growth(trade_ffill, lag_3m)
        g6 = growth(trade_ffill, lag_6m)
        g12 = growth(trade_ffill, lag_12m)

        vols = annualized_vol_from_prices([x for x in trade_ffill if x is not None])
        mdd = compute_max_drawdown([x for x in trade_ffill if x is not None])

        latest_jeonse = next((x for x in reversed(jeonse_series) if x is not None), None)
        latest_yield = next((x for x in reversed(yield_series) if x is not None), None)

        latest_trade_count = trade_counts[-1] if trade_counts else 0
        latest_rent_count = rent_counts[-1] if rent_counts else 0
        avg_trade_3m = statistics.mean(trade_counts[-3:]) if len(trade_counts) >= 3 else statistics.mean(trade_counts) if trade_counts else 0.0
        last_trade_idx = max((i for i, c in enumerate(trade_counts) if c > 0), default=None)
        last_trade_yyyymm = months[last_trade_idx] if last_trade_idx is not None else None
        trade_stale_months = (len(months) - 1 - last_trade_idx) if last_trade_idx is not None else len(months)
        market_freshness = clamp(1.0 - safe_div(trade_stale_months, 4.0), 0.0, 1.0)

        monthly_returns: list[float] = []
        for i in range(1, len(trade_ffill)):
            p0 = trade_ffill[i - 1]
            p1 = trade_ffill[i]
            if p0 is None or p1 is None or p0 <= 0:
                continue
            monthly_returns.append(clamp((p1 / p0) - 1.0, -0.35, 0.35))
        tail_es = expected_shortfall(monthly_returns, 0.20)
        tail_es_abs = abs(min(0.0, tail_es))

        months_with_trade = sum(1 for x in trade_counts if x > 0)
        months_with_rent = sum(1 for x in rent_counts if x > 0)
        coverage_trade = months_with_trade / len(months) if months else 0.0
        coverage_rent = months_with_rent / len(months) if months else 0.0

        total_trade_samples = sum(trade_counts)
        total_rent_samples = sum(rent_counts)

        quality = (
            40.0 * clamp(coverage_trade, 0.0, 1.0)
            + 25.0 * clamp(coverage_rent, 0.0, 1.0)
            + 20.0 * clamp(safe_div(latest_trade_count, 60.0), 0.0, 1.0)
            + 15.0 * clamp(safe_div(latest_rent_count, 120.0), 0.0, 1.0)
        )
        quality -= 12.0 * clamp(safe_div(trade_stale_months, 6.0), 0.0, 1.0)
        quality = clamp(quality, 0.0, 100.0)

        out.append(
            {
                "district": d.name,
                "code": d.code,
                "months": month_rows,
                "latest_price_eok": latest_price,
                "latest_rent_eok": latest_rent,
                "growth_3m": g3,
                "growth_6m": g6,
                "growth_12m": g12,
                "vol_ann": vols,
                "max_drawdown": mdd,
                "latest_jeonse_ratio": latest_jeonse,
                "latest_rent_yield": latest_yield,
                "latest_trade_count": latest_trade_count,
                "latest_rent_count": latest_rent_count,
                "avg_trade_3m": float(avg_trade_3m),
                "coverage_trade": coverage_trade,
                "coverage_rent": coverage_rent,
                "total_trade_samples": total_trade_samples,
                "total_rent_samples": total_rent_samples,
                "last_trade_yyyymm": last_trade_yyyymm,
                "trade_stale_months": trade_stale_months,
                "market_freshness": round(market_freshness, 3),
                "monthly_returns": monthly_returns,
                "tail_es_m20": tail_es,
                "tail_es_abs": tail_es_abs,
                "data_quality": round(quality, 1),
            }
        )

    # 서울 전체 상대가치
    latest_prices = [m["latest_price_eok"] for m in out if m["latest_price_eok"]]
    seoul_median_price = statistics.median(latest_prices) if latest_prices else 0.0

    for m in out:
        p = m["latest_price_eok"]
        if p and seoul_median_price > 0:
            m["price_vs_seoul"] = p / seoul_median_price
        else:
            m["price_vs_seoul"] = None

    return out


def build_complex_metrics(
    *,
    months: list[str],
    panel: dict[str, dict[str, dict[str, Any]]],
    district_metrics: list[dict[str, Any]],
    min_total_trade: int,
    min_month_coverage: int,
    shrink_trade_ref: int,
) -> list[dict[str, Any]]:
    district_lookup = {d["district"]: d for d in district_metrics}
    series_map: dict[tuple[str, str], dict[str, dict[str, Any]]] = {}

    for district_name, month_map in panel.items():
        for ym in months:
            payload = month_map.get(ym, {})
            for cname, cval in (payload.get("complexes") or {}).items():
                key = (district_name, cname)
                series_map.setdefault(key, {})
                series_map[key][ym] = cval

    out: list[dict[str, Any]] = []
    for (district_name, cname), ym_map in series_map.items():
        month_rows = []
        for ym in months:
            c = ym_map.get(ym)
            if c is None:
                month_rows.append(
                    {
                        "year_month": ym,
                        "trade_count": 0,
                        "rent_count": 0,
                        "median_trade_10k": None,
                        "median_rent_10k": None,
                        "jeonse_ratio_pct": None,
                        "rent_yield_proxy_pct": None,
                    }
                )
            else:
                month_rows.append(c)

        trade_series: list[float | None] = [
            (r.get("median_trade_10k") / 10000.0) if r.get("median_trade_10k") else None
            for r in month_rows
        ]
        rent_series: list[float | None] = [
            (r.get("median_rent_10k") / 10000.0) if r.get("median_rent_10k") else None
            for r in month_rows
        ]
        jeonse_series: list[float | None] = [r.get("jeonse_ratio_pct") for r in month_rows]
        yield_series: list[float | None] = [r.get("rent_yield_proxy_pct") for r in month_rows]

        trade_counts = [int(r.get("trade_count") or 0) for r in month_rows]
        rent_counts = [int(r.get("rent_count") or 0) for r in month_rows]
        total_trade_samples = sum(trade_counts)
        total_rent_samples = sum(rent_counts)

        if total_trade_samples < min_total_trade:
            continue

        trade_ffill = ffill(trade_series)
        rent_ffill = ffill(rent_series)
        valid_price = [x for x in trade_ffill if x is not None and x > 0]
        valid_rent = [x for x in rent_ffill if x is not None and x > 0]

        latest_price = valid_price[-1] if valid_price else None
        latest_rent = valid_rent[-1] if valid_rent else None

        lag_3m = min(3, max(1, len(trade_ffill) - 1))
        lag_6m = min(6, max(1, len(trade_ffill) - 1))
        lag_12m = min(11, max(1, len(trade_ffill) - 1))
        g3_raw = growth(trade_ffill, lag_3m)
        g6_raw = growth(trade_ffill, lag_6m)
        g12_raw = growth(trade_ffill, lag_12m)

        vols_raw = annualized_vol_from_prices([x for x in trade_ffill if x is not None])
        mdd = compute_max_drawdown([x for x in trade_ffill if x is not None])

        latest_jeonse = next((x for x in reversed(jeonse_series) if x is not None), None)
        latest_yield = next((x for x in reversed(yield_series) if x is not None), None)

        latest_trade_count = trade_counts[-1] if trade_counts else 0
        latest_rent_count = rent_counts[-1] if rent_counts else 0
        avg_trade_3m = (
            statistics.mean(trade_counts[-3:])
            if len(trade_counts) >= 3
            else statistics.mean(trade_counts)
            if trade_counts
            else 0.0
        )
        last_trade_idx = max((i for i, c in enumerate(trade_counts) if c > 0), default=None)
        last_trade_yyyymm = months[last_trade_idx] if last_trade_idx is not None else None
        trade_stale_months = (len(months) - 1 - last_trade_idx) if last_trade_idx is not None else len(months)
        market_freshness = clamp(1.0 - safe_div(trade_stale_months, 4.0), 0.0, 1.0)

        monthly_returns: list[float] = []
        for i in range(1, len(trade_ffill)):
            p0 = trade_ffill[i - 1]
            p1 = trade_ffill[i]
            if p0 is None or p1 is None or p0 <= 0:
                continue
            monthly_returns.append(clamp((p1 / p0) - 1.0, -0.35, 0.35))
        tail_es = expected_shortfall(monthly_returns, 0.20)
        tail_es_abs = abs(min(0.0, tail_es))

        months_with_trade = sum(1 for x in trade_counts if x > 0)
        months_with_rent = sum(1 for x in rent_counts if x > 0)
        if months_with_trade < min_month_coverage:
            continue
        coverage_trade = months_with_trade / len(months) if months else 0.0
        coverage_rent = months_with_rent / len(months) if months else 0.0

        quality = (
            42.0 * clamp(coverage_trade, 0.0, 1.0)
            + 18.0 * clamp(coverage_rent, 0.0, 1.0)
            + 22.0 * clamp(safe_div(total_trade_samples, 120.0), 0.0, 1.0)
            + 18.0 * clamp(safe_div(latest_trade_count, 12.0), 0.0, 1.0)
        )
        quality -= 18.0 * clamp(safe_div(trade_stale_months, 6.0), 0.0, 1.0)
        quality = clamp(quality, 0.0, 100.0)

        parent = district_lookup.get(district_name, {})
        sample_weight = clamp(safe_div(total_trade_samples, float(shrink_trade_ref)), 0.0, 1.0) * clamp(coverage_trade, 0.0, 1.0)
        g3 = shrink_toward_anchor(g3_raw, parent.get("growth_3m"), sample_weight)
        g6 = shrink_toward_anchor(g6_raw, parent.get("growth_6m"), sample_weight)
        g12 = shrink_toward_anchor(g12_raw, parent.get("growth_12m"), sample_weight)

        conf_bucket = confidence_bucket(quality, total_trade_samples, coverage_trade)
        unc_pen = confidence_penalty(conf_bucket)
        vols = vols_raw * (1.0 + unc_pen * 1.2)

        district_price = parent.get("latest_price_eok")
        price_vs_district = safe_div(latest_price, district_price) if latest_price and district_price else None
        rel_growth_12m = (
            (g12 - parent.get("growth_12m"))
            if g12 is not None and parent.get("growth_12m") is not None
            else None
        )
        parent_base_score = (
            0.24 * safe_float(parent.get("trend_score"), 50.0)
            + 0.22 * safe_float(parent.get("value_score"), 50.0)
            + 0.20 * safe_float(parent.get("liquidity_score"), 50.0)
            + 0.20 * safe_float(parent.get("stability_score"), 50.0)
            + 0.14 * safe_float(parent.get("policy_score"), 45.0)
        )

        out.append(
            {
                "district": district_name,
                "complex": cname,
                "complex_key": f"{district_name}::{cname}",
                "months": month_rows,
                "latest_price_eok": latest_price,
                "latest_rent_eok": latest_rent,
                "growth_3m": g3,
                "growth_6m": g6,
                "growth_12m": g12,
                "rel_growth_12m": rel_growth_12m,
                "vol_ann": vols,
                "max_drawdown": mdd,
                "latest_jeonse_ratio": latest_jeonse,
                "latest_rent_yield": latest_yield,
                "latest_trade_count": latest_trade_count,
                "latest_rent_count": latest_rent_count,
                "avg_trade_3m": float(avg_trade_3m),
                "coverage_trade": coverage_trade,
                "coverage_rent": coverage_rent,
                "total_trade_samples": total_trade_samples,
                "total_rent_samples": total_rent_samples,
                "last_trade_yyyymm": last_trade_yyyymm,
                "trade_stale_months": trade_stale_months,
                "market_freshness": round(market_freshness, 3),
                "monthly_returns": monthly_returns,
                "tail_es_m20": tail_es,
                "tail_es_abs": tail_es_abs,
                "data_quality": round(quality, 1),
                "confidence_bucket": conf_bucket,
                "uncertainty_penalty": unc_pen,
                "sample_weight": round(sample_weight, 3),
                "price_vs_district": price_vs_district,
                "parent_district_score": round(clamp(parent_base_score, 0.0, 100.0), 2),
                "parent_policy_score": parent.get("policy_score"),
            }
        )

    latest_prices = [m["latest_price_eok"] for m in out if m["latest_price_eok"]]
    seoul_complex_median = statistics.median(latest_prices) if latest_prices else 0.0
    for m in out:
        p = m["latest_price_eok"]
        m["price_vs_seoul"] = p / seoul_complex_median if p and seoul_complex_median > 0 else None

    return out


def run_multi_expert_scoring(metrics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    # 병렬 전문가 에이전트 입력 벡터
    growth12_all = [m.get("growth_12m") for m in metrics if m.get("growth_12m") is not None]
    growth3_all = [m.get("growth_3m") for m in metrics if m.get("growth_3m") is not None]
    ext12_all = [abs(float(m.get("growth_12m"))) for m in metrics if m.get("growth_12m") is not None]
    jeonse_all = [m.get("latest_jeonse_ratio") for m in metrics if m.get("latest_jeonse_ratio") is not None]
    yield_all = [m.get("latest_rent_yield") for m in metrics if m.get("latest_rent_yield") is not None]
    pvs_all = [m.get("price_vs_seoul") for m in metrics if m.get("price_vs_seoul") is not None]
    liq_all = [m.get("avg_trade_3m") for m in metrics if m.get("avg_trade_3m") is not None]
    vol_all = [m.get("vol_ann") for m in metrics if m.get("vol_ann") is not None]
    dd_all = [m.get("max_drawdown") for m in metrics if m.get("max_drawdown") is not None]
    es_all = [m.get("tail_es_abs") for m in metrics if m.get("tail_es_abs") is not None]
    fresh_all = [m.get("market_freshness") for m in metrics if m.get("market_freshness") is not None]
    stale_all = [m.get("trade_stale_months") for m in metrics if m.get("trade_stale_months") is not None]

    def gap_risk_penalty(jeonse_ratio: float | None) -> float:
        if jeonse_ratio is None:
            return 0.0
        j = float(jeonse_ratio)
        penalty = 0.0
        if j >= 80.0:
            penalty += 20.0
        elif j >= 75.0:
            penalty += 12.0
        elif j >= 70.0:
            penalty += 6.0
        if j <= 30.0:
            penalty += 6.0
        elif j <= 35.0:
            penalty += 3.0
        return penalty

    def trend_expert(m: dict[str, Any]) -> float:
        s12 = percentile_rank(m.get("growth_12m"), growth12_all)
        s3 = percentile_rank(m.get("growth_3m"), growth3_all)
        s_ext = percentile_rank(abs(float(m.get("growth_12m") or 0.0)), ext12_all, reverse=True)
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        return 0.50 * s12 + 0.22 * s3 + 0.18 * s_ext + 0.10 * s_fresh

    def value_expert(m: dict[str, Any]) -> float:
        s_jeonse = percentile_rank(m.get("latest_jeonse_ratio"), jeonse_all)
        s_yield = percentile_rank(m.get("latest_rent_yield"), yield_all)
        s_pvs = percentile_rank(m.get("price_vs_seoul"), pvs_all, reverse=True)
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        score = 0.34 * s_jeonse + 0.30 * s_yield + 0.24 * s_pvs + 0.12 * s_fresh
        score -= gap_risk_penalty(m.get("latest_jeonse_ratio"))
        return clamp(score, 0.0, 100.0)

    def liquidity_expert(m: dict[str, Any]) -> float:
        s_liq = percentile_rank(m.get("avg_trade_3m"), liq_all)
        s_latest = percentile_rank(float(m.get("latest_trade_count") or 0), liq_all)
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        s_stale = percentile_rank(m.get("trade_stale_months"), stale_all, reverse=True)
        return 0.40 * s_liq + 0.25 * s_latest + 0.20 * s_fresh + 0.15 * s_stale

    def stability_expert(m: dict[str, Any]) -> float:
        s_vol = percentile_rank(m.get("vol_ann"), vol_all, reverse=True)
        s_dd = percentile_rank(m.get("max_drawdown"), dd_all, reverse=True)
        dq = m.get("data_quality") or 0.0
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        s_es = percentile_rank(m.get("tail_es_abs"), es_all, reverse=True)
        return 0.28 * s_vol + 0.24 * s_dd + 0.22 * dq + 0.14 * s_fresh + 0.12 * s_es

    def policy_expert(m: dict[str, Any]) -> float:
        base = 45.0  # 서울 전체 LTPS 가정
        if m["district"] in HIGH_INTENSITY_LTPS:
            base -= 15.0
        # 거래가 충분하면 정책 충격 흡수력이 상대적으로 높다고 가정
        base += clamp((m.get("avg_trade_3m") or 0) / 12.0, 0.0, 10.0)
        base += clamp((m.get("data_quality") or 0) / 20.0, 0.0, 5.0)
        base += clamp((m.get("market_freshness") or 0.0) * 8.0, 0.0, 8.0)
        base -= clamp(safe_div(float(m.get("trade_stale_months") or 0), 6.0), 0.0, 1.0) * 10.0
        base += clamp((1.0 - clamp(safe_div(float(m.get("tail_es_abs") or 0.0), 0.10), 0.0, 1.0)) * 4.0, 0.0, 4.0)
        base -= gap_risk_penalty(m.get("latest_jeonse_ratio")) * 0.35
        return clamp(base, 0.0, 100.0)

    with cf.ThreadPoolExecutor(max_workers=5) as ex:
        f_trend = {ex.submit(trend_expert, m): m for m in metrics}
        f_value = {ex.submit(value_expert, m): m for m in metrics}
        f_liq = {ex.submit(liquidity_expert, m): m for m in metrics}
        f_stb = {ex.submit(stability_expert, m): m for m in metrics}
        f_pol = {ex.submit(policy_expert, m): m for m in metrics}

        for fut_map, key in [
            (f_trend, "trend_score"),
            (f_value, "value_score"),
            (f_liq, "liquidity_score"),
            (f_stb, "stability_score"),
            (f_pol, "policy_score"),
        ]:
            for fut in cf.as_completed(fut_map):
                m = fut_map[fut]
                try:
                    m[key] = round(float(fut.result()), 2)
                except Exception:  # noqa: BLE001
                    m[key] = 50.0

    return metrics


def run_complex_expert_scoring(metrics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    growth12_all = [m.get("growth_12m") for m in metrics if m.get("growth_12m") is not None]
    ext12_all = [abs(float(m.get("growth_12m"))) for m in metrics if m.get("growth_12m") is not None]
    rel_growth_all = [m.get("rel_growth_12m") for m in metrics if m.get("rel_growth_12m") is not None]
    jeonse_all = [m.get("latest_jeonse_ratio") for m in metrics if m.get("latest_jeonse_ratio") is not None]
    yield_all = [m.get("latest_rent_yield") for m in metrics if m.get("latest_rent_yield") is not None]
    pvd_all = [m.get("price_vs_district") for m in metrics if m.get("price_vs_district") is not None]
    pvs_all = [m.get("price_vs_seoul") for m in metrics if m.get("price_vs_seoul") is not None]
    liq_all = [m.get("avg_trade_3m") for m in metrics if m.get("avg_trade_3m") is not None]
    vol_all = [m.get("vol_ann") for m in metrics if m.get("vol_ann") is not None]
    dd_all = [m.get("max_drawdown") for m in metrics if m.get("max_drawdown") is not None]
    es_all = [m.get("tail_es_abs") for m in metrics if m.get("tail_es_abs") is not None]
    fresh_all = [m.get("market_freshness") for m in metrics if m.get("market_freshness") is not None]
    stale_all = [m.get("trade_stale_months") for m in metrics if m.get("trade_stale_months") is not None]

    def confidence_score(m: dict[str, Any]) -> float:
        b = m.get("confidence_bucket")
        if b == "high":
            return 100.0
        if b == "medium":
            return 70.0
        return 40.0

    def gap_risk_penalty(jeonse_ratio: float | None) -> float:
        if jeonse_ratio is None:
            return 0.0
        j = float(jeonse_ratio)
        penalty = 0.0
        if j >= 80.0:
            penalty += 20.0
        elif j >= 75.0:
            penalty += 12.0
        elif j >= 70.0:
            penalty += 6.0
        if j <= 30.0:
            penalty += 6.0
        elif j <= 35.0:
            penalty += 3.0
        return penalty

    def trend_expert(m: dict[str, Any]) -> float:
        s12 = percentile_rank(m.get("growth_12m"), growth12_all)
        srel = percentile_rank(m.get("rel_growth_12m"), rel_growth_all)
        sw = (m.get("sample_weight") or 0.0) * 100.0
        s_ext = percentile_rank(abs(float(m.get("growth_12m") or 0.0)), ext12_all, reverse=True)
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        return 0.42 * s12 + 0.26 * srel + 0.14 * sw + 0.10 * s_ext + 0.08 * s_fresh

    def value_expert(m: dict[str, Any]) -> float:
        s_jeonse = percentile_rank(m.get("latest_jeonse_ratio"), jeonse_all)
        s_yield = percentile_rank(m.get("latest_rent_yield"), yield_all)
        s_pvd = percentile_rank(m.get("price_vs_district"), pvd_all, reverse=True)
        s_pvs = percentile_rank(m.get("price_vs_seoul"), pvs_all, reverse=True)
        conf = confidence_score(m)
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        score = 0.28 * s_jeonse + 0.22 * s_yield + 0.18 * s_pvd + 0.10 * s_pvs + 0.12 * conf + 0.10 * s_fresh
        score -= gap_risk_penalty(m.get("latest_jeonse_ratio")) * 0.55
        return clamp(score, 0.0, 100.0)

    def liquidity_expert(m: dict[str, Any]) -> float:
        s_liq = percentile_rank(m.get("avg_trade_3m"), liq_all)
        s_latest = percentile_rank(float(m.get("latest_trade_count") or 0), liq_all)
        conf = confidence_score(m)
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        s_stale = percentile_rank(m.get("trade_stale_months"), stale_all, reverse=True)
        return 0.38 * s_liq + 0.24 * s_latest + 0.18 * s_fresh + 0.08 * s_stale + 0.12 * conf

    def stability_expert(m: dict[str, Any]) -> float:
        s_vol = percentile_rank(m.get("vol_ann"), vol_all, reverse=True)
        s_dd = percentile_rank(m.get("max_drawdown"), dd_all, reverse=True)
        dq = m.get("data_quality") or 0.0
        conf = confidence_score(m)
        s_fresh = percentile_rank(m.get("market_freshness"), fresh_all)
        s_es = percentile_rank(m.get("tail_es_abs"), es_all, reverse=True)
        return 0.23 * s_vol + 0.21 * s_dd + 0.18 * dq + 0.14 * conf + 0.12 * s_fresh + 0.12 * s_es

    def policy_expert(m: dict[str, Any]) -> float:
        base = 45.0
        if m["district"] in HIGH_INTENSITY_LTPS:
            base -= 15.0
        base += clamp((m.get("avg_trade_3m") or 0) / 2.0, 0.0, 15.0)
        base += clamp((m.get("data_quality") or 0) / 25.0, 0.0, 5.0)
        base += clamp((m.get("market_freshness") or 0.0) * 12.0, 0.0, 12.0)
        base -= clamp(safe_div(float(m.get("trade_stale_months") or 0), 6.0), 0.0, 1.0) * 12.0
        base += clamp((1.0 - clamp(safe_div(float(m.get("tail_es_abs") or 0.0), 0.10), 0.0, 1.0)) * 5.0, 0.0, 5.0)
        base -= gap_risk_penalty(m.get("latest_jeonse_ratio")) * 0.40
        base -= (m.get("uncertainty_penalty") or 0.0) * 40.0
        return clamp(base, 0.0, 100.0)

    with cf.ThreadPoolExecutor(max_workers=5) as ex:
        f_trend = {ex.submit(trend_expert, m): m for m in metrics}
        f_value = {ex.submit(value_expert, m): m for m in metrics}
        f_liq = {ex.submit(liquidity_expert, m): m for m in metrics}
        f_stb = {ex.submit(stability_expert, m): m for m in metrics}
        f_pol = {ex.submit(policy_expert, m): m for m in metrics}

        for fut_map, key in [
            (f_trend, "trend_score"),
            (f_value, "value_score"),
            (f_liq, "liquidity_score"),
            (f_stb, "stability_score"),
            (f_pol, "policy_score"),
        ]:
            for fut in cf.as_completed(fut_map):
                m = fut_map[fut]
                try:
                    m[key] = round(float(fut.result()), 2)
                except Exception:  # noqa: BLE001
                    m[key] = 50.0

    return metrics


def _theme_prior_score(district: str, theme_key: str) -> float:
    d = str(district or "").strip()
    if d and d in (THEME_PRIOR_HIGH.get(theme_key) or set()):
        return 85.0
    if d and d in (THEME_PRIOR_MED.get(theme_key) or set()):
        return 68.0
    return 50.0


def _theme_confidence_score(bucket: str | None) -> float:
    b = str(bucket or "").strip().lower()
    if b == "high":
        return 100.0
    if b == "medium":
        return 70.0
    if b == "low":
        return 40.0
    return 60.0


def _theme_rel_price(m: dict[str, Any]) -> float:
    if m.get("complex") is not None or m.get("complex_key") is not None:
        p = m.get("price_vs_district")
        if p is not None:
            return safe_float(p, 1.0)
    return safe_float(m.get("price_vs_seoul"), 1.0)


def _build_theme_context(rows: list[dict[str, Any]]) -> dict[str, list[float]]:
    return {
        "arr_growth12": [safe_float(x.get("growth_12m"), 0.0) for x in rows],
        "arr_liq": [safe_float(x.get("avg_trade_3m"), 0.0) for x in rows],
        "arr_price": [_theme_rel_price(x) for x in rows],
        "arr_yield": [safe_float(x.get("latest_rent_yield"), 2.0) for x in rows],
        "arr_jeonse": [safe_float(x.get("latest_jeonse_ratio"), 0.0) for x in rows],
        "arr_vol": [safe_float(x.get("vol_ann"), 0.12) for x in rows],
        "arr_trade_samples": [safe_float(x.get("total_trade_samples"), 0.0) for x in rows],
    }


def _theme_features_for_row(
    m: dict[str, Any],
    ctx: dict[str, list[float]],
    theme_key: str,
) -> dict[str, float]:
    district = str(m.get("district") or "").strip()
    rel_price = _theme_rel_price(m)
    market_fresh = clamp(safe_float(m.get("market_freshness"), 0.0), 0.0, 1.0) * 100.0
    dataq = clamp(safe_float(m.get("data_quality"), 60.0), 0.0, 100.0)
    adv_risk = clamp(safe_float(m.get("adv_risk_composite"), 50.0), 0.0, 100.0)
    adv_policy = clamp(safe_float(m.get("adv_policy_flip"), 50.0), 0.0, 100.0)
    adv_supply = clamp(safe_float(m.get("adv_supply_pressure"), 50.0), 0.0, 100.0)
    adv_guar = clamp(safe_float(m.get("adv_guarantee_risk"), 50.0), 0.0, 100.0)
    adv_mgt = clamp(safe_float(m.get("adv_management_dd_risk"), 50.0), 0.0, 100.0)
    conf = _theme_confidence_score(m.get("confidence_bucket") or m.get("adv_confidence_bucket"))

    vol_low = percentile_rank(safe_float(m.get("vol_ann"), 0.12), ctx["arr_vol"], reverse=True)
    feats = {
        "prior": _theme_prior_score(district, theme_key),
        "policy": clamp(safe_float(m.get("policy_score"), 45.0), 0.0, 100.0),
        "trend": clamp(safe_float(m.get("trend_score"), 50.0), 0.0, 100.0),
        "liq": percentile_rank(safe_float(m.get("avg_trade_3m"), 0.0), ctx["arr_liq"]),
        "premium": percentile_rank(rel_price, ctx["arr_price"]),
        "value": percentile_rank(rel_price, ctx["arr_price"], reverse=True),
        "yield": percentile_rank(safe_float(m.get("latest_rent_yield"), 2.0), ctx["arr_yield"]),
        "demand": clamp(safe_float(m.get("nq_demand_persistence"), 50.0), 0.0, 100.0),
        "live": clamp(safe_float(m.get("nq_liveability"), 50.0), 0.0, 100.0),
        "stability": clamp(safe_float(m.get("stability_score"), 50.0), 0.0, 100.0),
        "gov": clamp(safe_float(m.get("nq_governance"), 50.0), 0.0, 100.0),
        "mgmt": clamp(safe_float(m.get("nq_management_quality"), 50.0), 0.0, 100.0),
        "fresh": market_fresh,
        "sample_depth": percentile_rank(safe_float(m.get("total_trade_samples"), 0.0), ctx["arr_trade_samples"]),
        "adv_low": 100.0 - adv_risk,
        "adv_supply_low": 100.0 - adv_supply,
        "adv_policy_risk": adv_policy,
        "adv_supply_risk": adv_supply,
        "adv_guarantee_risk": adv_guar,
        "adv_mgmt_risk": adv_mgt,
        "gap_risk": percentile_rank(safe_float(m.get("latest_jeonse_ratio"), 0.0), ctx["arr_jeonse"]),
        "premium_risk": percentile_rank(rel_price, ctx["arr_price"]),
        "vol_risk": 100.0 - vol_low,
        "data_low": 100.0 - dataq,
        "fresh_low": 100.0 - market_fresh,
        "dataq": dataq,
        "conf": conf,
    }
    for k in list(feats.keys()):
        feats[k] = clamp(safe_float(feats[k], 50.0), 0.0, 100.0)
    return feats


def _theme_weighted_score(features: dict[str, float], weights: dict[str, float]) -> float:
    if not weights:
        return 50.0
    den = sum(max(0.0, safe_float(w, 0.0)) for w in weights.values())
    if den <= 0:
        return 50.0
    num = 0.0
    for k, w in weights.items():
        wk = max(0.0, safe_float(w, 0.0))
        num += wk * safe_float(features.get(k), 50.0)
    return clamp(num / den, 0.0, 100.0)


def _theme_driver_text(features: dict[str, float], weights: dict[str, float], top_n: int = 3) -> str:
    comps: list[tuple[str, float, float]] = []
    for k, w in weights.items():
        wk = max(0.0, safe_float(w, 0.0))
        fv = safe_float(features.get(k), 50.0)
        comps.append((k, wk * fv, fv))
    comps.sort(key=lambda x: x[1], reverse=True)
    items: list[str] = []
    for k, _, fv in comps[:top_n]:
        label = THEME_FEATURE_LABELS.get(k, k)
        items.append(f"{label}({fv:.0f})")
    return ", ".join(items)


def _theme_action_tag(conviction: float, risk: float, spread: float) -> str:
    conv = clamp(safe_float(conviction, 0.0), 0.0, 100.0)
    rk = clamp(safe_float(risk, 50.0), 0.0, 100.0)
    sp = safe_float(spread, 0.0)
    if rk >= 75.0:
        return "회피"
    if conv >= 82.0 and rk <= 46.0 and sp >= 4.0:
        return "집중매수"
    if conv >= 74.0 and rk <= 58.0:
        return "분할매수"
    if conv >= 66.0 and rk <= 66.0:
        return "관찰매수"
    return "관망"


def _theme_strategy_for_profile(theme_key: str | None, profile: str, action: str) -> tuple[str, str, str]:
    key = str(theme_key or "")
    pb = THEME_PLAYBOOK.get(key) or {}
    if profile == "실거주":
        tactic = str(pb.get("real") or pb.get("invest") or "실거주 안정성 중심으로 분할 접근")
    elif profile == "투자자":
        tactic = str(pb.get("invest") or pb.get("real") or "모멘텀/유동성 확인 후 분할 접근")
    else:
        tactic = str(pb.get("corp") or pb.get("invest") or "세후수익 기준으로 보수적 분할 접근")
    risk = str(pb.get("risk") or "정책/유동성/공급 리스크 동시 점검")
    monitor = str(pb.get("monitor") or "월별 거래량·정책·금리 업데이트 점검")
    act = str(action or "관찰")
    return f"{act} | {tactic}", risk, monitor


def apply_theme_classification(
    rows: list[dict[str, Any]],
    *,
    profile: str,
    level: str,
) -> list[dict[str, Any]]:
    if not rows:
        return rows
    ctx = _build_theme_context(rows)
    bias_map = THEME_PROFILE_BIAS.get(profile) or {}

    for r in rows:
        scored: list[dict[str, Any]] = []
        for td in THEME_CATALOG:
            key = str(td.get("key") or "")
            feats = _theme_features_for_row(r, ctx, key)
            opp = _theme_weighted_score(feats, td.get("opp_weights") or {})
            risk = _theme_weighted_score(feats, td.get("risk_weights") or {})
            bias = clamp(safe_float(bias_map.get(key), 1.0), 0.75, 1.25)
            conviction = clamp((0.70 * opp + 0.30 * (100.0 - risk)) * bias, 0.0, 100.0)
            edge = opp - risk
            scored.append(
                {
                    "key": key,
                    "name": str(td.get("name") or key),
                    "desc": str(td.get("desc") or ""),
                    "opportunity": round(opp, 2),
                    "risk": round(risk, 2),
                    "conviction": round(conviction, 2),
                    "edge": round(edge, 2),
                    "driver": _theme_driver_text(feats, td.get("opp_weights") or {}, top_n=3),
                    "risk_driver": _theme_driver_text(feats, td.get("risk_weights") or {}, top_n=2),
                }
            )

        scored.sort(key=lambda x: (safe_float(x.get("conviction"), 0.0), safe_float(x.get("edge"), 0.0)), reverse=True)
        top1 = scored[0]
        top2 = scored[1] if len(scored) > 1 else scored[0]
        spread = safe_float(top1.get("conviction"), 0.0) - safe_float(top2.get("conviction"), 0.0)
        caution = "주의" if safe_float(top1.get("risk"), 0.0) >= 66.0 else "보통"
        if safe_float(top1.get("risk"), 0.0) <= 45.0:
            caution = "양호"
        action = _theme_action_tag(
            safe_float(top1.get("conviction"), 0.0),
            safe_float(top1.get("risk"), 50.0),
            spread,
        )

        r["theme_primary_key"] = top1.get("key")
        r["theme_primary"] = top1.get("name")
        r["theme_primary_score"] = top1.get("conviction")
        r["theme_primary_risk"] = top1.get("risk")
        r["theme_primary_reason"] = f"호재:{top1.get('driver')} / 리스크:{top1.get('risk_driver')}"
        r["theme_secondary_key"] = top2.get("key")
        r["theme_secondary"] = top2.get("name")
        r["theme_secondary_score"] = top2.get("conviction")
        r["theme_secondary_risk"] = top2.get("risk")
        r["theme_spread"] = round(spread, 2)
        r["theme_caution"] = caution
        r["theme_action"] = action
        r["theme_candidates"] = ", ".join([f"{x['name']}({x['conviction']:.1f})" for x in scored[:3]])
        r["theme_level"] = level
        r["theme_score_map"] = {x["key"]: x for x in scored}
        r["theme_scores"] = scored

    return rows


def _safe_log1p(v: Any) -> float:
    return math.log1p(max(0.0, safe_float(v, 0.0)))


def _pearson_corr(xs: list[float], ys: list[float]) -> float:
    n = min(len(xs), len(ys))
    if n < 3:
        return 0.0
    x = [safe_float(v, 0.0) for v in xs[:n]]
    y = [safe_float(v, 0.0) for v in ys[:n]]
    mx = statistics.mean(x)
    my = statistics.mean(y)
    dx = [v - mx for v in x]
    dy = [v - my for v in y]
    sxx = sum(v * v for v in dx)
    syy = sum(v * v for v in dy)
    if sxx <= 1e-12 or syy <= 1e-12:
        return 0.0
    sxy = sum(a * b for a, b in zip(dx, dy))
    return clamp(sxy / math.sqrt(sxx * syy), -1.0, 1.0)


def _robust_z_series(values: list[float], clip: float = 3.0) -> list[float]:
    if not values:
        return []
    vals = [safe_float(v, 0.0) for v in values]
    win = winsorize(vals, alpha=0.10)
    med = statistics.median(win)
    mad = statistics.median([abs(v - med) for v in win])
    scale = mad * 1.4826
    center = med
    if scale <= 1e-9:
        center, sigma = robust_mean_std(win)
        scale = max(1e-9, sigma)
    z = [(v - center) / scale for v in win]
    return [clamp(v, -clip, clip) for v in z]


def _quant_factor_raw_row(m: dict[str, Any]) -> dict[str, float]:
    g12 = safe_float(m.get("growth_12m"), 0.0)
    g3 = safe_float(m.get("growth_3m"), 0.0)
    rel_price = max(0.30, safe_float(_theme_rel_price(m), 1.0))
    yld = clamp(safe_float(m.get("latest_rent_yield"), 2.0) / 100.0, 0.0, 0.10)
    liq = _safe_log1p(m.get("avg_trade_3m")) + 0.35 * clamp(safe_float(m.get("market_freshness"), 0.0), 0.0, 1.0) - 0.08 * clamp(safe_float(m.get("trade_stale_months"), 0.0), 0.0, 6.0)
    conf = _theme_confidence_score(m.get("confidence_bucket") or m.get("adv_confidence_bucket")) / 100.0
    dq = clamp(safe_float(m.get("data_quality"), 60.0) / 100.0, 0.0, 1.0)
    gov = clamp(safe_float(m.get("nq_governance"), 50.0) / 100.0, 0.0, 1.0)
    mgmt = clamp(safe_float(m.get("nq_management_quality"), 50.0) / 100.0, 0.0, 1.0)
    vol = clamp(safe_float(m.get("vol_ann"), 0.12), 0.0, 1.0)
    tail = clamp(safe_float(m.get("tail_es_abs"), 0.0), 0.0, 0.30)
    adv = clamp(safe_float(m.get("adv_risk_composite"), 50.0) / 100.0, 0.0, 1.0)
    policy_score = clamp(safe_float(m.get("policy_score"), 50.0) / 100.0, 0.0, 1.0)
    policy_flip = clamp(safe_float(m.get("adv_policy_flip"), 50.0) / 100.0, 0.0, 1.0)
    demand = clamp(safe_float(m.get("nq_demand_persistence"), 50.0) / 100.0, 0.0, 1.0)
    sent = clamp(safe_float(m.get("nq_sentiment"), 50.0) / 100.0, 0.0, 1.0)
    live = clamp(safe_float(m.get("nq_liveability"), 50.0) / 100.0, 0.0, 1.0)
    adv_supply = clamp(safe_float(m.get("adv_supply_pressure"), 50.0) / 100.0, 0.0, 1.0)

    return {
        "mom": 0.65 * g12 + 0.35 * (g3 * 4.0),
        "value": -math.log(rel_price),
        "carry": yld,
        "liq": liq,
        "quality": 0.45 * dq + 0.20 * gov + 0.20 * mgmt + 0.15 * conf,
        "lowrisk": -(0.65 * vol + 0.25 * (tail * 4.0) + 0.10 * adv),
        "policy": 0.65 * policy_score + 0.35 * (1.0 - policy_flip),
        "sent": 0.45 * demand + 0.30 * sent + 0.25 * live,
        "supply": 1.0 - adv_supply,
    }


def apply_quant_factor_model(
    rows: list[dict[str, Any]],
    *,
    profile: str,
    level: str,
) -> list[dict[str, Any]]:
    if not rows:
        return rows

    factor_keys = [str(x.get("key") or "") for x in QUANT_FACTOR_CATALOG]
    factor_raw_series: dict[str, list[float]] = {k: [] for k in factor_keys}
    for r in rows:
        raw = _quant_factor_raw_row(r)
        for k in factor_keys:
            factor_raw_series[k].append(safe_float(raw.get(k), 0.0))

    factor_z_series: dict[str, list[float]] = {}
    factor_score_series: dict[str, list[float]] = {}
    for k in factor_keys:
        z = _robust_z_series(factor_raw_series.get(k, []), clip=3.0)
        factor_z_series[k] = z
        factor_score_series[k] = [clamp(50.0 + (12.5 * v), 0.0, 100.0) for v in z]

    corr_penalty: dict[str, float] = {}
    for k in factor_keys:
        peers = [j for j in factor_keys if j != k]
        cors = [abs(_pearson_corr(factor_z_series.get(k, []), factor_z_series.get(j, []))) for j in peers]
        avg_abs_corr = statistics.mean(cors) if cors else 0.0
        over = max(0.0, avg_abs_corr - QUANT_CORR_SHRINK_THRESHOLD)
        penalty = 1.0 - QUANT_CORR_SHRINK_INTENSITY * safe_div(over, 1.0 - QUANT_CORR_SHRINK_THRESHOLD, 0.0)
        corr_penalty[k] = clamp(penalty, 0.72, 1.00)

    base_w = QUANT_PROFILE_WEIGHTS.get(profile) or {}
    w_raw = {k: max(0.0, safe_float(base_w.get(k), 0.0)) * corr_penalty.get(k, 1.0) for k in factor_keys}
    w_sum = sum(w_raw.values())
    if w_sum <= 1e-12:
        w_final = {k: 1.0 / max(1, len(factor_keys)) for k in factor_keys}
    else:
        w_final = {k: w_raw.get(k, 0.0) / w_sum for k in factor_keys}

    blend_alpha = clamp(safe_float(QUANT_FACTOR_BLEND.get(profile), 0.60), 0.35, 0.80)
    name_map = QUANT_FACTOR_NAME_MAP

    for i, r in enumerate(rows):
        factor_scores = {k: safe_float(factor_score_series.get(k, [50.0])[i], 50.0) for k in factor_keys}
        composite_raw = sum(w_final.get(k, 0.0) * factor_scores.get(k, 50.0) for k in factor_keys)

        dq = clamp(safe_float(r.get("data_quality"), 60.0) / 100.0, 0.0, 1.0)
        conf = clamp(_theme_confidence_score(r.get("confidence_bucket") or r.get("adv_confidence_bucket")) / 100.0, 0.40, 1.00)
        unc = clamp(safe_float(r.get("uncertainty_penalty"), 0.0) / 0.20, 0.0, 1.0)
        q_mult = clamp(0.80 + 0.12 * dq + 0.08 * conf - 0.12 * unc, 0.72, 1.05)
        factor_score = clamp(composite_raw * q_mult, 0.0, 100.0)

        legacy = safe_float(r.get("quant_total_score"), 50.0)
        blended = clamp((1.0 - blend_alpha) * legacy + blend_alpha * factor_score, 0.0, 100.0)

        contrib = [(k, w_final.get(k, 0.0) * factor_scores.get(k, 50.0)) for k in factor_keys]
        contrib.sort(key=lambda x: x[1], reverse=True)
        top_items = [f"{name_map.get(k, k)}({factor_scores.get(k, 50.0):.1f})" for k, _ in contrib[:3]]
        low_items = [f"{name_map.get(k, k)}({factor_scores.get(k, 50.0):.1f})" for k, _ in contrib[-2:]]

        r["qf_level"] = level
        r["qf_scores"] = {k: round(factor_scores.get(k, 50.0), 2) for k in factor_keys}
        r["qf_weight_map"] = {k: round(w_final.get(k, 0.0), 4) for k in factor_keys}
        r["qf_corr_penalty_map"] = {k: round(corr_penalty.get(k, 1.0), 4) for k in factor_keys}
        for k in factor_keys:
            r[f"qf_{k}"] = round(factor_scores.get(k, 50.0), 2)
            r[f"qfw_{k}"] = round(w_final.get(k, 0.0), 4)
        r["qf_composite_raw"] = round(composite_raw, 2)
        r["qf_quality_mult"] = round(q_mult, 3)
        r["quant_factor_score"] = round(factor_score, 2)
        r["quant_total_score_legacy"] = round(legacy, 2)
        r["quant_total_score"] = round(blended, 2)
        r["qf_driver"] = ", ".join(top_items)
        r["qf_risk_driver"] = ", ".join(low_items)

    return rows


def score_by_profile(
    metrics: list[dict[str, Any]],
    profile: str,
    *,
    nq_overlay_cap: float = DEFAULT_NQ_OVERLAY_CAP,
    adv_risk_cap: float = DEFAULT_ADV_RISK_DISTRICT_CAP,
) -> list[dict[str, Any]]:
    conf = PROFILE_CONFIGS[profile]
    w = conf["weights"]

    rows = []
    for m in metrics:
        quant_total_legacy = (
            m["trend_score"] * w["trend"]
            + m["value_score"] * w["value"]
            + m["liquidity_score"] * w["liquidity"]
            + m["stability_score"] * w["stability"]
            + m["policy_score"] * w["policy"]
        ) / 100.0
        nq_overlay = nonquant_overlay_for_profile(m, profile, nq_overlay_cap)
        adv_penalty = advanced_risk_penalty_for_profile(m, profile, adv_risk_cap)

        x = dict(m)
        x["profile"] = profile
        x["quant_total_score"] = round(quant_total_legacy, 2)
        x["nq_overlay"] = round(nq_overlay, 2)
        x["adv_risk_penalty"] = round(adv_penalty, 2)
        rows.append(x)

    rows = apply_quant_factor_model(rows, profile=profile, level="district")
    for x in rows:
        total = clamp(
            safe_float(x.get("quant_total_score"), 50.0)
            + safe_float(x.get("nq_overlay"), 0.0)
            - safe_float(x.get("adv_risk_penalty"), 0.0),
            0.0,
            100.0,
        )
        x["total_score"] = round(total, 2)

    rows.sort(key=lambda r: r["total_score"], reverse=True)
    for i, r in enumerate(rows, 1):
        r["rank"] = i
    return rows


def score_complex_by_profile(
    metrics: list[dict[str, Any]],
    profile: str,
    *,
    nq_overlay_cap: float = DEFAULT_NQ_COMPLEX_OVERLAY_CAP,
    adv_risk_cap: float = DEFAULT_ADV_RISK_COMPLEX_CAP,
) -> list[dict[str, Any]]:
    conf = PROFILE_CONFIGS[profile]
    w = conf["weights"]

    rows = []
    for m in metrics:
        core = (
            m["trend_score"] * w["trend"]
            + m["value_score"] * w["value"]
            + m["liquidity_score"] * w["liquidity"]
            + m["stability_score"] * w["stability"]
            + m["policy_score"] * w["policy"]
        ) / 100.0

        # 단지 평점은 구(상위 레벨) 컨텍스트를 20% 섞어 과적합을 완화한다.
        district_tail = float(m.get("parent_district_score") or 50.0)
        quant_total_legacy = 0.80 * core + 0.20 * district_tail
        nq_overlay = nonquant_overlay_for_profile(m, profile, nq_overlay_cap)
        adv_penalty = advanced_risk_penalty_for_profile(m, profile, adv_risk_cap)
        stale_months = float(m.get("trade_stale_months") or 0.0)
        recency_penalty = clamp(stale_months, 0.0, 6.0) * 1.2
        tail_penalty = clamp(safe_div(float(m.get("tail_es_abs") or 0.0), 0.10), 0.0, 1.0) * 2.5
        nq_flag_penalty = min(4.2, safe_float(m.get("nq_flag_count"), 0.0) * 0.7)

        x = dict(m)
        x["profile"] = profile
        x["quant_total_score"] = round(quant_total_legacy, 2)
        x["nq_overlay"] = round(nq_overlay, 2)
        x["adv_risk_penalty"] = round(adv_penalty, 2)
        x["core_score"] = round(core, 2)
        x["recency_penalty"] = round(recency_penalty, 2)
        x["tail_penalty"] = round(tail_penalty, 2)
        x["nq_flag_penalty"] = round(nq_flag_penalty, 2)
        rows.append(x)

    rows = apply_quant_factor_model(rows, profile=profile, level="complex")
    for x in rows:
        total = (
            safe_float(x.get("quant_total_score"), 50.0)
            + safe_float(x.get("nq_overlay"), 0.0)
            - safe_float(x.get("adv_risk_penalty"), 0.0)
        )
        total = total * (1.0 - safe_float(x.get("uncertainty_penalty"), 0.0))
        total -= (
            safe_float(x.get("recency_penalty"), 0.0)
            + safe_float(x.get("tail_penalty"), 0.0)
            + safe_float(x.get("nq_flag_penalty"), 0.0)
        )
        x["total_score"] = round(clamp(total, 0.0, 100.0), 2)

    rows.sort(key=lambda r: r["total_score"], reverse=True)
    for i, r in enumerate(rows, 1):
        r["rank"] = i

    # 구 내 순위 부여
    group: dict[str, int] = {}
    for r in rows:
        d = r["district"]
        group[d] = group.get(d, 0) + 1
        r["district_rank"] = group[d]
    return rows


def scenario_for_district(m: dict[str, Any], profile: str) -> dict[str, Any]:
    cfg = PROFILE_CONFIGS[profile]

    price = m.get("latest_price_eok") or 0.0
    if price <= 0:
        return {
            "bull_cum": None,
            "base_cum": None,
            "bear_cum": None,
            "bull_roi": None,
            "base_roi": None,
            "bear_roi": None,
            "horizon_years": cfg["horizon_years"],
        }

    h = int(cfg["horizon_years"])
    t_months = h * 12

    g12 = m.get("growth_12m")
    g3 = m.get("growth_3m")
    g12 = 0.02 if g12 is None else g12
    g3 = 0.005 if g3 is None else g3
    anchor_annual = clamp(0.60 * g12 + 0.40 * g3, -0.05, 0.09)
    anchor_monthly = (1.0 + anchor_annual) ** (1.0 / 12.0) - 1.0

    unc_pen = float(m.get("uncertainty_penalty") or 0.0)

    monthly_returns = [
        clamp(float(r), -0.25, 0.25)
        for r in (m.get("monthly_returns") or [])
        if r is not None
    ]
    mu_hist, sigma_hist = robust_mean_std(monthly_returns)
    vol_ann = m.get("vol_ann") or 0.08
    vol_from_ann = clamp(vol_ann, 0.04, 0.30) / math.sqrt(12.0)

    if monthly_returns:
        mu_m = 0.55 * mu_hist + 0.45 * anchor_monthly
        sigma_m = 0.65 * sigma_hist + 0.35 * vol_from_ann
    else:
        mu_m = anchor_monthly
        sigma_m = vol_from_ann

    q20_m = mu_m - 0.85 * sigma_m
    q50_m = mu_m
    q80_m = mu_m + 0.85 * sigma_m
    if len(monthly_returns) >= 6:
        wr = winsorize(monthly_returns, 0.10)
        q20_hist = percentile_value(wr, 0.20, default=q20_m)
        q50_hist = percentile_value(wr, 0.50, default=q50_m)
        q80_hist = percentile_value(wr, 0.80, default=q80_m)
        q20_m = 0.72 * q20_hist + 0.28 * (mu_m - 0.95 * sigma_m)
        q50_m = 0.75 * q50_hist + 0.25 * mu_m
        q80_m = 0.72 * q80_hist + 0.28 * (mu_m + 0.95 * sigma_m)

    sigma_m = sigma_m * (1.0 + unc_pen * 0.8)
    sigma_m = clamp(sigma_m, 0.008, 0.09)
    stale_m = float(m.get("trade_stale_months") or 0.0)
    tail_es_abs = float(m.get("tail_es_abs") or 0.0)
    jeonse = float(m.get("latest_jeonse_ratio") or 0.0)
    gap_risk = 0.0
    if jeonse >= 80.0:
        gap_risk = 1.0
    elif jeonse >= 75.0:
        gap_risk = 0.7
    elif jeonse >= 70.0:
        gap_risk = 0.4
    stale_pen = clamp(safe_div(stale_m, 6.0), 0.0, 1.0) * 0.0012
    nq_comp = safe_float(m.get("nq_composite"), 50.0)
    nq_risk = clamp(1.0 - safe_div(nq_comp, 100.0), 0.0, 1.0)
    nq_exec = clamp(1.0 - safe_div(safe_float(m.get("nq_governance"), 50.0), 100.0), 0.0, 1.0)
    nq_live = clamp(1.0 - safe_div(safe_float(m.get("nq_liveability"), 50.0), 100.0), 0.0, 1.0)
    nq_sent = clamp(1.0 - safe_div(safe_float(m.get("nq_sentiment"), 50.0), 100.0), 0.0, 1.0)
    nq_flag_ct = safe_float(m.get("nq_flag_count"), 0.0)
    nq_flag_pen = min(0.0018, nq_flag_ct * 0.00022)
    adv_comp = safe_float(m.get("adv_risk_composite"), 50.0)
    adv_risk = clamp(adv_comp / 100.0, 0.0, 1.0)
    adv_liq = clamp(safe_float(m.get("adv_liquidity_trap"), 50.0) / 100.0, 0.0, 1.0)
    adv_crowd = clamp(safe_float(m.get("adv_crowding"), 50.0) / 100.0, 0.0, 1.0)
    adv_policy = clamp(safe_float(m.get("adv_policy_flip"), 50.0) / 100.0, 0.0, 1.0)
    adv_exec = clamp(safe_float(m.get("adv_execution"), 50.0) / 100.0, 0.0, 1.0)
    adv_supply = clamp(safe_float(m.get("adv_supply_pressure"), 50.0) / 100.0, 0.0, 1.0)
    adv_guar = clamp(safe_float(m.get("adv_guarantee_risk"), 50.0) / 100.0, 0.0, 1.0)
    adv_mgt = clamp(safe_float(m.get("adv_management_dd_risk"), 50.0) / 100.0, 0.0, 1.0)
    adv_flag_ct = safe_float(m.get("adv_flag_count"), 0.0)
    adv_flag_pen = min(0.0020, adv_flag_ct * 0.00018)

    if m["district"] in HIGH_INTENSITY_LTPS:
        regime_bias = -0.0008
        bear_extra = -0.0010
    else:
        regime_bias = -0.0003
        bear_extra = -0.0005

    # 과거 분포 분위수(P20/P50/P80)와 모수(μ±σ)를 혼합해 시나리오 drift를 구성.
    bull_anchor = 0.72 * q80_m + 0.28 * (mu_m + 0.90 * sigma_m)
    base_anchor = 0.80 * q50_m + 0.20 * (mu_m - 0.10 * sigma_m)
    bear_anchor = 0.72 * q20_m + 0.28 * (mu_m - 1.00 * sigma_m)

    bull_m = (
        bull_anchor
        - regime_bias * 0.2
        - nq_risk * 0.00010
        - nq_exec * 0.00008
        - adv_risk * 0.00010
        - adv_policy * 0.00008
    )
    base_m = (
        base_anchor
        + regime_bias
        - stale_pen * 0.5
        - tail_es_abs * 0.02
        - gap_risk * 0.0002
        - nq_risk * 0.00025
        - nq_exec * 0.00015
        - nq_flag_pen
        - adv_risk * 0.00035
        - adv_liq * 0.00018
        - adv_policy * 0.00020
        - adv_exec * 0.00020
        - adv_supply * 0.00018
        - adv_guar * 0.00022
        - adv_flag_pen
    )
    bear_m = (
        bear_anchor
        + regime_bias
        + bear_extra
        - stale_pen
        - unc_pen * 0.0015
        - tail_es_abs * 0.08
        - gap_risk * 0.0006
        - nq_risk * 0.00055
        - nq_exec * 0.00025
        - nq_sent * 0.00020
        - nq_flag_pen * 1.5
        - adv_risk * 0.00070
        - adv_liq * 0.00030
        - adv_policy * 0.00035
        - adv_exec * 0.00035
        - adv_crowd * 0.00025
        - adv_supply * 0.00035
        - adv_guar * 0.00045
        - adv_mgt * 0.00030
        - adv_flag_pen * 1.4
    )

    def annual_from_monthly(r_m: float) -> float:
        return (1.0 + r_m) ** 12 - 1.0

    # 과도한 극단값을 피하기 위해 서울 장기 분포를 고려한 현실적 연환산 범위로 제한.
    bull_a = clamp(annual_from_monthly(bull_m), -0.06, 0.17)
    base_a = clamp(annual_from_monthly(base_m), -0.05, 0.10)
    bear_a = clamp(annual_from_monthly(bear_m), -0.12, 0.04)

    bull_c = (1.0 + bull_a) ** h - 1.0
    base_c = (1.0 + base_a) ** h - 1.0
    bear_c = (1.0 + bear_a) ** h - 1.0

    buy_tax = cfg["buy_tax"]
    hold_cost = cfg["annual_hold_cost"]
    cg_tax = cfg["capital_gain_tax"]
    ltv = cfg["ltv"]
    loan_rate = cfg["loan_rate"]

    debt = price * ltv
    equity = price - debt
    carry_total = price * hold_cost * h

    rent_yield = clamp((m.get("latest_rent_yield") or 1.8) / 100.0, 0.008, 0.05)
    if profile == "실거주":
        annual_utility_yield = max(0.012, rent_yield * 0.60)
    elif profile == "투자자":
        annual_utility_yield = rent_yield * 0.85
    else:
        annual_utility_yield = rent_yield * 0.80
    utility_income_total = price * annual_utility_yield * h

    fresh = float(m.get("market_freshness") or 0.0)
    vol_norm = clamp((m.get("vol_ann") or 0.10) / 0.25, 0.0, 1.0)
    liq_norm = clamp((m.get("avg_trade_3m") or 0.0) / 12.0, 0.0, 1.0)
    yield_norm = clamp(rent_yield / 0.04, 0.0, 1.0)
    es_norm = clamp(safe_div(tail_es_abs, 0.10), 0.0, 1.0)
    interest_sensitivity = (
        0.75
        + 0.30 * vol_norm
        + 0.25 * (1.0 - liq_norm)
        + 0.18 * (1.0 - yield_norm)
        + 0.18 * (1.0 - fresh)
        + 0.15 * unc_pen
        + 0.20 * es_norm
        + 0.10 * gap_risk
        + 0.20 * nq_risk
        + 0.08 * nq_exec
        + 0.05 * nq_live
        + 0.22 * adv_risk
        + 0.10 * adv_liq
        + 0.08 * adv_policy
        + 0.09 * adv_supply
        + 0.10 * adv_guar
    )
    interest_sensitivity = clamp(interest_sensitivity, 0.70, 1.90)

    def leveraged_roi(cum_ret: float, loan_rate_override: float | None = None) -> float:
        exit_price = price * (1.0 + cum_ret)
        gross_gain = exit_price - price
        tx = max(gross_gain, 0.0) * cg_tax
        r_loan = loan_rate if loan_rate_override is None else loan_rate_override
        interest_total = debt * r_loan * h * interest_sensitivity
        freshness_cost = price * (1.0 - fresh) * 0.0015 * h
        net_gain = (
            gross_gain
            + utility_income_total
            - tx
            - (price * buy_tax)
            - carry_total
            - interest_total
            - freshness_cost
        )
        if equity <= 0:
            return 0.0
        return clamp(net_gain / equity, -0.90, 3.00)

    shock_scale = 0.75 + 0.55 * interest_sensitivity + 0.35 * unc_pen + 0.35 * clamp(stale_m / 6.0, 0.0, 1.0)
    shock_scale += 0.25 * es_norm + 0.20 * gap_risk
    shock_scale += 0.25 * adv_risk + 0.12 * adv_liq
    effective_rate_shock = DEFAULT_SCENARIO_RATE_SHOCK * clamp(shock_scale, 0.60, 2.20)
    stress_up_rate = loan_rate + effective_rate_shock
    stress_dn_rate = max(0.0, loan_rate - effective_rate_shock)

    return {
        "bull_cum": bull_c,
        "base_cum": base_c,
        "bear_cum": bear_c,
        "bull_roi": leveraged_roi(bull_c),
        "base_roi": leveraged_roi(base_c),
        "bear_roi": leveraged_roi(bear_c),
        "horizon_years": h,
        "annual_base": base_a,
        "annual_vol": sigma_m * math.sqrt(12.0),
        "annual_utility_yield": annual_utility_yield,
        "base_roi_rate_up": leveraged_roi(base_c, stress_up_rate),
        "base_roi_rate_down": leveraged_roi(base_c, stress_dn_rate),
        "confidence_bucket": m.get("confidence_bucket", "medium"),
        "assumption": {
            "buy_tax": buy_tax,
            "hold_cost": hold_cost,
            "capital_gain_tax": cg_tax,
            "ltv": ltv,
            "loan_rate": loan_rate,
            "loan_rate_up": stress_up_rate,
            "loan_rate_down": stress_dn_rate,
            "effective_rate_shock": effective_rate_shock,
            "interest_sensitivity": interest_sensitivity,
            "q20_monthly": q20_m,
            "q50_monthly": q50_m,
            "q80_monthly": q80_m,
            "trade_stale_months": stale_m,
            "tail_es_abs": tail_es_abs,
            "gap_risk": gap_risk,
            "nq_composite": nq_comp,
            "nq_risk": nq_risk,
            "nq_exec": nq_exec,
            "nq_live": nq_live,
            "nq_sent": nq_sent,
            "nq_flag_count": nq_flag_ct,
            "adv_risk_composite": adv_comp,
            "adv_risk": adv_risk,
            "adv_liquidity_trap": adv_liq,
            "adv_crowding": adv_crowd,
            "adv_policy_flip": adv_policy,
            "adv_execution": adv_exec,
            "adv_supply_pressure": adv_supply,
            "adv_guarantee_risk": adv_guar,
            "adv_management_dd_risk": adv_mgt,
            "adv_flag_count": adv_flag_ct,
            "utility_income_total": utility_income_total,
            "t_months": t_months,
        },
    }


def attach_scenarios(rows: list[dict[str, Any]], profile: str) -> list[dict[str, Any]]:
    for r in rows:
        r["scenario"] = scenario_for_district(r, profile)
    return rows

# ---------------------------------------------------------------------------
# Workbook builders
# ---------------------------------------------------------------------------


def add_common_header(ws, title: str, subtitle: str, cols: int) -> int:
    r = write_title(ws, 1, title, cols)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    c = ws.cell(row=r, column=1, value=subtitle)
    c.font = FONT_SMALL
    c.alignment = ALIGN_LEFT
    c.fill = FILL_LIGHT
    c.border = BORDER
    ws.row_dimensions[r].height = 20
    return r + 1


def add_sheet_dashboard(wb, rows: list[dict[str, Any]], profile: str, months: list[str]):
    ws = wb.create_sheet("종합 대시보드")
    r = add_common_header(
        ws,
        f"서울 25개구 아파트 심층분석 대시보드 ({profile})",
        f"기준: {months[0]}~{months[-1]} / 면적필터 59~99m² / 생성시각 {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}",
        10,
    )

    top5 = rows[:5]
    bot5 = rows[-5:]

    r = write_row(ws, r, ["[핵심 가정]"] + [""] * 9, fill=FILL_GOLD, left=True)
    assumptions = [
        "1) 월별 중앙값 기반(거래 outlier 완화), 구/단지 2단계 비교",
        "2) 낙관/중립/비관은 P80/P50/P20 분위수 기반 확률 시나리오",
        "3) 레버리지 ROI는 프로필별 LTV/금리/세율/보유비용 반영",
        "4) 서울 LTPS 규제와 4개 집중규제구(강남·서초·송파·용산) 정책 패널티 반영",
        "5) 데이터 품질점수(표본수+커버리지) 낮으면 점수 자동 보수화",
        "6) 비정량 오버레이(거버넌스/생활환경/수요지속/관리품질/심리)로 총점·시나리오 보정",
        "7) 고급리스크(유동성함정·군집쏠림·정책플립·실행리스크) 패널티를 총점/하방드리프트에 반영",
    ]
    for line in assumptions:
        r = write_row(ws, r, [line] + [""] * 9, left=True)

    r += 1
    r = write_row(ws, r, ["[TOP 5 추천]"] + [""] * 9, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["순위", "구", "총점", "12M성장", "전세가율", "변동성", "거래량(최근)", "낙관ROI", "중립ROI", "비관ROI"])
    for x in top5:
        sc = x["scenario"]
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["total_score"],
                pct(x.get("growth_12m")),
                f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                x.get("latest_trade_count") or 0,
                pct(sc.get("bull_roi")),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
            ],
        )

    r += 1
    r = write_row(ws, r, ["[BOTTOM 5 (관망/주의)]"] + [""] * 9, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["순위", "구", "총점", "12M성장", "전세가율", "변동성", "거래량(최근)", "낙관ROI", "중립ROI", "비관ROI"])
    for x in bot5:
        sc = x["scenario"]
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["total_score"],
                pct(x.get("growth_12m")),
                f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                x.get("latest_trade_count") or 0,
                pct(sc.get("bull_roi")),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
            ],
        )

    set_widths(ws, [8, 13, 9, 11, 11, 10, 12, 11, 11, 11])


def add_sheet_scorecard(wb, rows: list[dict[str, Any]], profile: str, months: list[str]):
    ws = wb.create_sheet("서울25구 스코어카드")
    r = add_common_header(ws, f"서울 25개구 스코어카드 ({profile})", f"분석기간: {months[0]}~{months[-1]}", 27)
    headers = [
        "순위", "구", "총점", "1순위테마", "1순위점수", "테마액션", "퀀트최종", "퀀트레거시", "팩터순수", "NQ오버레이", "NQ합성", "NQ리스크", "ADV리스크", "ADV패널티", "Trend", "Value", "Liquidity", "Stability", "Policy", "DataQ",
        "최근매매(억)", "최근전세(억)", "12M성장", "전세가율", "변동성", "최근거래수", "거래공백월",
    ]
    r = write_headers(ws, r, headers)
    for x in rows:
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["total_score"],
                x.get("theme_primary"),
                x.get("theme_primary_score"),
                x.get("theme_action"),
                x.get("quant_total_score"),
                x.get("quant_total_score_legacy"),
                x.get("quant_factor_score"),
                x.get("nq_overlay"),
                x.get("nq_composite"),
                x.get("nq_risk_index"),
                x.get("adv_risk_composite"),
                x.get("adv_risk_penalty"),
                x["trend_score"],
                x["value_score"],
                x["liquidity_score"],
                x["stability_score"],
                x["policy_score"],
                x["data_quality"],
                round(x["latest_price_eok"], 2) if x.get("latest_price_eok") else "-",
                round(x["latest_rent_eok"], 2) if x.get("latest_rent_eok") else "-",
                pct(x.get("growth_12m")),
                f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                x.get("latest_trade_count") or 0,
                x.get("trade_stale_months") or 0,
            ],
        )
        ws.cell(row=r - 1, column=3).fill = score_band_fill(float(x["total_score"]))
        ws.cell(row=r - 1, column=5).fill = score_band_fill(float(x.get("theme_primary_score") or 0))
        ws.cell(row=r - 1, column=11).fill = score_band_fill(float(x.get("nq_composite") or 0))
        ws.cell(row=r - 1, column=13).fill = score_band_fill(100.0 - float(x.get("adv_risk_composite") or 50.0))

    set_widths(ws, [6, 12, 8, 18, 10, 10, 9, 9, 9, 9, 8, 8, 8, 9, 8, 8, 9, 9, 8, 8, 10, 10, 10, 9, 9, 9, 10])


def add_sheet_coverage(wb, rows: list[dict[str, Any]], months: list[str]):
    ws = wb.create_sheet("데이터 커버리지")
    r = add_common_header(ws, "데이터 품질/커버리지", "표본수와 월별 커버리지를 점수화하여 신뢰도 반영", 13)
    headers = [
        "구", "분석월수", "매매커버리지", "전월세커버리지", "총매매표본", "총전월세표본", "최근매매표본", "최근전월세표본", "평균거래(3M)", "최종거래월", "거래공백월", "DataQ", "비고"
    ]
    r = write_headers(ws, r, headers)
    for x in rows:
        note = "양호"
        if x["data_quality"] < 55:
            note = "주의(표본 부족)"
        elif x["data_quality"] < 70:
            note = "보통"
        r = write_row(
            ws,
            r,
            [
                x["district"],
                len(months),
                f"{x['coverage_trade']*100:.1f}%",
                f"{x['coverage_rent']*100:.1f}%",
                x["total_trade_samples"],
                x["total_rent_samples"],
                x["latest_trade_count"],
                x["latest_rent_count"],
                round(x["avg_trade_3m"], 1),
                x.get("last_trade_yyyymm") or "-",
                x.get("trade_stale_months") or 0,
                x["data_quality"],
                note,
            ],
        )
        ws.cell(row=r - 1, column=12).fill = score_band_fill(float(x["data_quality"]))
    set_widths(ws, [12, 9, 11, 12, 11, 12, 11, 12, 11, 10, 10, 8, 16])


def add_sheet_trade_trend(wb, rows: list[dict[str, Any]], months: list[str]):
    ws = wb.create_sheet("매매 추세")
    ncol = 1 + len(months)
    r = add_common_header(ws, "구별 매매 중앙값 추세(억)", "면적필터 59~99m², 월별 중앙값", ncol)
    r = write_headers(ws, r, ["구"] + months)
    for x in rows:
        row_vals = [x["district"]]
        month_map = {m["year_month"]: m for m in x["months"]}
        for ym in months:
            m = month_map.get(ym, {})
            if m.get("median_trade_10k"):
                row_vals.append(round(m["median_trade_10k"] / 10000.0, 2))
            else:
                row_vals.append(None)
        r = write_row(ws, r, row_vals)
    set_widths(ws, [12] + [9] * len(months))


def add_sheet_rent_trend(wb, rows: list[dict[str, Any]], months: list[str]):
    ws = wb.create_sheet("전세 추세")
    ncol = 1 + len(months)
    r = add_common_header(ws, "구별 전세 중앙값 추세(억)", "면적필터 59~99m², 월별 중앙값", ncol)
    r = write_headers(ws, r, ["구"] + months)
    for x in rows:
        row_vals = [x["district"]]
        month_map = {m["year_month"]: m for m in x["months"]}
        for ym in months:
            m = month_map.get(ym, {})
            if m.get("median_rent_10k"):
                row_vals.append(round(m["median_rent_10k"] / 10000.0, 2))
            else:
                row_vals.append(None)
        r = write_row(ws, r, row_vals)

    r += 1
    r = write_row(ws, r, ["[최근 전세가율/환산수익률]"] + [""] * (ncol - 1), fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["구", "전세가율(%)", "전세환산 수익률(%)"])
    for x in rows:
        r = write_row(
            ws,
            r,
            [
                x["district"],
                round(x["latest_jeonse_ratio"], 2) if x.get("latest_jeonse_ratio") else "-",
                round(x["latest_rent_yield"], 2) if x.get("latest_rent_yield") else "-",
            ],
        )
    set_widths(ws, [12] + [9] * len(months))


def add_sheet_value(wb, rows: list[dict[str, Any]]):
    ws = wb.create_sheet("밸류에이션")
    r = add_common_header(ws, "상대가치 분석", "서울 중앙값 대비 가격수준/전세가율/전세환산수익률", 10)
    headers = ["구", "최근매매(억)", "최근전세(억)", "서울대비가격배수", "전세가율(%)", "전세환산수익률(%)", "Value점수", "해석", "위험", "메모"]
    r = write_headers(ws, r, headers)

    for x in rows:
        pvs = x.get("price_vs_seoul")
        label = "중립"
        if pvs is not None and pvs < 0.9:
            label = "상대저평가"
        elif pvs is not None and pvs > 1.15:
            label = "상대고평가"

        risk = "보통"
        if (x.get("latest_jeonse_ratio") or 0) > 65:
            risk = "갭리스크"
        if (x.get("latest_jeonse_ratio") or 0) < 30:
            risk = "현금부담"

        r = write_row(
            ws,
            r,
            [
                x["district"],
                round(x["latest_price_eok"], 2) if x.get("latest_price_eok") else "-",
                round(x["latest_rent_eok"], 2) if x.get("latest_rent_eok") else "-",
                round(pvs, 3) if pvs else "-",
                round(x["latest_jeonse_ratio"], 2) if x.get("latest_jeonse_ratio") else "-",
                round(x["latest_rent_yield"], 2) if x.get("latest_rent_yield") else "-",
                x["value_score"],
                label,
                risk,
                "",
            ],
        )
        ws.cell(row=r - 1, column=7).fill = score_band_fill(float(x["value_score"]))

    set_widths(ws, [12, 10, 10, 12, 11, 14, 9, 12, 10, 12])


def add_sheet_liquidity(wb, rows: list[dict[str, Any]]):
    ws = wb.create_sheet("유동성")
    r = add_common_header(ws, "유동성 분석", "최근 거래량/3개월 평균 거래량 중심", 9)
    r = write_headers(ws, r, ["구", "최근매매표본", "3M평균표본", "최근전월세표본", "Liquidity점수", "정책충격흡수", "체결난이도", "권고", "비고"])

    for x in rows:
        liq_score = x["liquidity_score"]
        absorb = "높음" if liq_score >= 70 else "보통" if liq_score >= 55 else "낮음"
        exec_cost = "낮음" if liq_score >= 70 else "보통" if liq_score >= 50 else "높음"
        rec = "분할진입" if liq_score >= 70 else "선별진입" if liq_score >= 55 else "관망"

        r = write_row(
            ws,
            r,
            [
                x["district"],
                x["latest_trade_count"],
                round(x["avg_trade_3m"], 1),
                x["latest_rent_count"],
                liq_score,
                absorb,
                exec_cost,
                rec,
                "",
            ],
        )
        ws.cell(row=r - 1, column=5).fill = score_band_fill(float(liq_score))

    set_widths(ws, [12, 11, 11, 12, 10, 12, 10, 10, 12])


def add_sheet_risk(wb, rows: list[dict[str, Any]]):
    ws = wb.create_sheet("변동성리스크")
    r = add_common_header(ws, "변동성/낙폭 리스크", "변동성·낙폭·꼬리위험(ES)으로 하방위험 측정", 11)
    r = write_headers(ws, r, ["구", "12M성장", "3M성장", "변동성(연)", "최대낙폭", "월ES(20%)", "Stability점수", "Policy점수", "종합리스크", "LTPS강도", "코멘트"])

    for x in rows:
        risk_index = 100.0 - (0.65 * x["stability_score"] + 0.35 * x["policy_score"])
        if risk_index <= 35:
            bucket = "낮음"
        elif risk_index <= 55:
            bucket = "보통"
        else:
            bucket = "높음"
        intensity = "집중규제" if x["district"] in HIGH_INTENSITY_LTPS else "일반규제"

        r = write_row(
            ws,
            r,
            [
                x["district"],
                pct(x.get("growth_12m")),
                pct(x.get("growth_3m")),
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                f"{(x.get('max_drawdown') or 0)*100:.1f}%",
                f"{(x.get('tail_es_m20') or 0)*100:.1f}%",
                x["stability_score"],
                x["policy_score"],
                round(risk_index, 2),
                intensity,
                bucket,
            ],
        )
        ws.cell(row=r - 1, column=9).fill = score_band_fill(100 - risk_index)

    set_widths(ws, [12, 10, 10, 10, 10, 10, 11, 10, 10, 11, 10])


def add_sheet_benchmark(wb, rows: list[dict[str, Any]]):
    ws = wb.create_sheet("벤치마크")
    r = add_common_header(ws, "벤치마크 및 상대비교", "서울 중앙값/상하위 20% 비교", 10)

    # 상위/하위 20%
    n = len(rows)
    k = max(1, int(round(n * 0.2)))
    top = rows[:k]
    bottom = rows[-k:]

    def avg(items: list[dict[str, Any]], key: str) -> float:
        vals = [float(x[key]) for x in items if x.get(key) is not None]
        return statistics.mean(vals) if vals else 0.0

    r = write_row(ws, r, ["[요약]"] + [""] * 9, fill=FILL_GOLD, left=True)
    summary_rows = [
        ["그룹", "총점", "12M성장", "전세가율", "변동성", "거래량", "Trend", "Value", "Liquidity", "Stability"],
        ["상위20%", round(avg(top, "total_score"), 2), pct(avg(top, "growth_12m")), f"{avg(top, 'latest_jeonse_ratio'):.1f}%", f"{avg(top, 'vol_ann')*100:.1f}%", round(avg(top, "avg_trade_3m"), 1), round(avg(top, "trend_score"), 1), round(avg(top, "value_score"), 1), round(avg(top, "liquidity_score"), 1), round(avg(top, "stability_score"), 1)],
        ["하위20%", round(avg(bottom, "total_score"), 2), pct(avg(bottom, "growth_12m")), f"{avg(bottom, 'latest_jeonse_ratio'):.1f}%", f"{avg(bottom, 'vol_ann')*100:.1f}%", round(avg(bottom, "avg_trade_3m"), 1), round(avg(bottom, "trend_score"), 1), round(avg(bottom, "value_score"), 1), round(avg(bottom, "liquidity_score"), 1), round(avg(bottom, "stability_score"), 1)],
    ]
    r = write_headers(ws, r, summary_rows[0])
    r = write_row(ws, r, summary_rows[1])
    r = write_row(ws, r, summary_rows[2])

    r += 1
    r = write_row(ws, r, ["[강남3구+용산 vs 기타구]"] + [""] * 9, fill=FILL_GOLD, left=True)
    core_names = {"강남구", "서초구", "송파구", "용산구"}
    core = [x for x in rows if x["district"] in core_names]
    non_core = [x for x in rows if x["district"] not in core_names]
    r = write_headers(ws, r, ["그룹", "평균총점", "평균매매(억)", "12M성장", "전세가율", "변동성", "평균거래량", "정책점수", "리스크해석", "비고"])
    for nm, grp in [("강남3구+용산", core), ("기타21구", non_core)]:
        risk = "정책민감" if nm == "강남3구+용산" else "분산형"
        r = write_row(
            ws,
            r,
            [
                nm,
                round(avg(grp, "total_score"), 2),
                round(avg(grp, "latest_price_eok"), 2),
                pct(avg(grp, "growth_12m")),
                f"{avg(grp, 'latest_jeonse_ratio'):.1f}%",
                f"{avg(grp, 'vol_ann')*100:.1f}%",
                round(avg(grp, "avg_trade_3m"), 1),
                round(avg(grp, "policy_score"), 1),
                risk,
                "",
            ],
        )

    set_widths(ws, [14, 11, 12, 10, 10, 10, 10, 10, 12, 10])


def compute_walkforward_backtest(
    rows: list[dict[str, Any]],
    months: list[str],
    lookback: int = DEFAULT_BACKTEST_LOOKBACK,
    horizon: int = DEFAULT_BACKTEST_HORIZON,
    top_fraction: float = DEFAULT_BACKTEST_TOP_FRACTION,
) -> dict[str, Any]:
    if len(months) < lookback + horizon:
        return {"records": [], "config": None}

    slices: list[dict[str, Any]] = []
    for t in range(lookback - 1, len(months) - horizon):
        signal_rows: list[dict[str, Any]] = []
        all_forward: list[float] = []

        for r in rows:
            month_map = {m["year_month"]: m for m in r.get("months", [])}
            prices: list[float | None] = []
            counts: list[int] = []
            jeonse: list[float] = []
            yields: list[float] = []
            for ym in months:
                m = month_map.get(ym, {})
                p10k = m.get("median_trade_10k")
                prices.append((p10k / 10000.0) if p10k else None)
                counts.append(int(m.get("trade_count") or 0))
                jeonse.append(float(m.get("jeonse_ratio_pct") or 0.0))
                yields.append(float(m.get("rent_yield_proxy_pct") or 0.0))

            p_now = prices[t]
            p_ref = prices[t - lookback + 1]
            p_fwd = prices[t + horizon]
            if p_now is None or p_ref is None or p_fwd is None or p_ref <= 0 or p_now <= 0:
                continue

            momentum = p_now / p_ref - 1.0
            liq = statistics.mean(counts[max(0, t - lookback + 1): t + 1]) if counts else 0.0
            freshness = 1.0 if counts[t] > 0 else 0.0
            val = jeonse[t]
            yld = yields[t]
            rets: list[float] = []
            for i in range(max(1, t - lookback + 1), t + 1):
                p0 = prices[i - 1]
                p1 = prices[i]
                if p0 is None or p1 is None or p0 <= 0:
                    continue
                rets.append(math.log(p1 / p0))
            lowvol = stdev(rets) if len(rets) > 1 else 0.0
            fwd_ret = p_fwd / p_now - 1.0
            all_forward.append(fwd_ret)
            signal_rows.append(
                {
                    "district": r["district"],
                    "momentum": momentum,
                    "liq": liq,
                    "freshness": freshness,
                    "val": val,
                    "yld": yld,
                    "lowvol": lowvol,
                    "fwd_ret": fwd_ret,
                }
            )

        if len(signal_rows) < 8 or not all_forward:
            continue
        slices.append({"month": months[t], "rows": signal_rows, "all_forward": all_forward})

    if not slices:
        return {"records": [], "config": None}

    def eval_cfg(weights: dict[str, float], frac: float) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        recs: list[dict[str, Any]] = []
        cum_s = 1.0
        cum_b = 1.0
        for s in slices:
            rows_i = [dict(x) for x in s["rows"]]
            moms = [x["momentum"] for x in rows_i]
            liqs = [x["liq"] for x in rows_i]
            freshes = [x["freshness"] for x in rows_i]
            vals = [x["val"] for x in rows_i]
            ylds = [x["yld"] for x in rows_i]
            vols = [x["lowvol"] for x in rows_i]

            for x in rows_i:
                signal = 0.0
                if weights.get("mom"):
                    signal += weights["mom"] * percentile_rank(x["momentum"], moms)
                if weights.get("liq"):
                    signal += weights["liq"] * percentile_rank(x["liq"], liqs)
                if weights.get("fresh"):
                    signal += weights["fresh"] * percentile_rank(x["freshness"], freshes)
                if weights.get("val"):
                    signal += weights["val"] * percentile_rank(x["val"], vals)
                if weights.get("yld"):
                    signal += weights["yld"] * percentile_rank(x["yld"], ylds)
                if weights.get("lowvol"):
                    signal += weights["lowvol"] * percentile_rank(x["lowvol"], vols, reverse=True)
                x["signal"] = signal

            rows_i.sort(key=lambda x: x["signal"], reverse=True)
            k = max(3, int(round(len(rows_i) * frac)))
            k = min(k, len(rows_i))
            top = rows_i[:k]
            strategy = statistics.mean([x["fwd_ret"] for x in top]) if top else 0.0
            bench = statistics.mean(s["all_forward"]) if s["all_forward"] else 0.0
            excess = strategy - bench
            cum_s *= 1.0 + strategy
            cum_b *= 1.0 + bench
            recs.append(
                {
                    "rebalance_month": s["month"],
                    "strategy_ret": strategy,
                    "benchmark_ret": bench,
                    "excess_ret": excess,
                    "hit": 1 if excess > 0 else 0,
                    "sample_count": len(top),
                    "top_districts": ", ".join(x["district"] for x in top[:5]),
                }
            )
        avg_ex = statistics.mean([x["excess_ret"] for x in recs]) if recs else -1.0
        hit = safe_div(sum(x["hit"] for x in recs), float(len(recs))) if recs else 0.0
        return recs, {
            "avg_excess": avg_ex,
            "hit_ratio": hit,
            "cum_excess": cum_s - cum_b,
            "top_fraction": frac,
            "weights": weights,
            "samples": len(recs),
        }

    best_records: list[dict[str, Any]] = []
    best_cfg: dict[str, Any] | None = None
    candidates_tested = 0
    for weights in BACKTEST_SIGNAL_CANDIDATES:
        for frac in BACKTEST_TOP_FRACTION_CANDIDATES:
            candidates_tested += 1
            recs, cfg = eval_cfg(weights, frac)
            if not recs:
                continue
            if best_cfg is None:
                best_cfg = cfg
                best_records = recs
                continue
            lhs = (cfg["avg_excess"], cfg["hit_ratio"], cfg["cum_excess"])
            rhs = (best_cfg["avg_excess"], best_cfg["hit_ratio"], best_cfg["cum_excess"])
            if lhs > rhs:
                best_cfg = cfg
                best_records = recs

    if best_cfg is None:
        return {"records": [], "config": None}

    return {
        "records": best_records,
        "config": {
            **best_cfg,
            "lookback": lookback,
            "horizon": horizon,
            "candidates_tested": candidates_tested,
        },
    }


def add_sheet_backtest(
    wb,
    rows: list[dict[str, Any]],
    months: list[str],
    lookback: int = DEFAULT_BACKTEST_LOOKBACK,
    horizon: int = DEFAULT_BACKTEST_HORIZON,
    top_fraction: float = DEFAULT_BACKTEST_TOP_FRACTION,
):
    ws = wb.create_sheet("전략 백테스트")
    r = add_common_header(
        ws,
        "워크포워드 백테스트(구 단위)",
        f"lookback {lookback}M / forward {horizon}M / 다중신호 자동튜닝 전략",
        10,
    )
    r = write_headers(
        ws,
        r,
        ["리밸런스월", "전략수익", "벤치마크수익", "초과수익", "적중", "표본수", "누적전략", "누적벤치", "누적초과", "Top구(일부)"],
    )

    bt_obj = compute_walkforward_backtest(rows, months, lookback=lookback, horizon=horizon, top_fraction=top_fraction)
    bt = bt_obj.get("records", [])
    cfg = bt_obj.get("config")
    if not bt:
        write_row(ws, r, ["백테스트 표본이 부족합니다."] + [""] * 9, left=True)
        set_widths(ws, [10, 10, 12, 10, 6, 7, 10, 10, 10, 36])
        return

    cum_s = 1.0
    cum_b = 1.0
    for x in bt:
        cum_s *= 1.0 + x["strategy_ret"]
        cum_b *= 1.0 + x["benchmark_ret"]
        r = write_row(
            ws,
            r,
            [
                x["rebalance_month"],
                pct(x["strategy_ret"]),
                pct(x["benchmark_ret"]),
                pct(x["excess_ret"]),
                "Y" if x["hit"] else "N",
                x["sample_count"],
                pct(cum_s - 1.0),
                pct(cum_b - 1.0),
                pct((cum_s - cum_b)),
                x["top_districts"],
            ],
            left=True,
        )

    hit_ratio = safe_div(sum(x["hit"] for x in bt), float(len(bt)))
    avg_excess = statistics.mean([x["excess_ret"] for x in bt]) if bt else 0.0
    r += 1
    r = write_row(ws, r, ["[요약]"] + [""] * 9, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["평가구간", "횟수", "적중률", "평균초과수익", "누적전략", "누적벤치", "누적초과", "", "", ""])
    r = write_row(
        ws,
        r,
        [
            f"{bt[0]['rebalance_month']}~{bt[-1]['rebalance_month']}",
            len(bt),
            pct(hit_ratio),
            pct(avg_excess),
            pct(cum_s - 1.0),
            pct(cum_b - 1.0),
            pct(cum_s - cum_b),
            "",
            "",
            "",
        ],
    )
    if cfg:
        r += 1
        r = write_row(ws, r, ["[튜닝 설정]"] + [""] * 9, fill=FILL_GOLD, left=True)
        r = write_headers(ws, r, ["lookback", "horizon", "Top비중", "가중치", "탐색조합", "평균초과", "적중률", "누적초과", "", ""])
        write_row(
            ws,
            r,
            [
                cfg.get("lookback"),
                cfg.get("horizon"),
                f"{(cfg.get('top_fraction') or 0)*100:.0f}%",
                ", ".join(f"{k}:{v:.2f}" for k, v in (cfg.get("weights") or {}).items()),
                cfg.get("candidates_tested"),
                pct(cfg.get("avg_excess")),
                pct(cfg.get("hit_ratio")),
                pct(cfg.get("cum_excess")),
                "",
                "",
            ],
            left=True,
        )
    set_widths(ws, [10, 10, 12, 10, 6, 7, 10, 10, 10, 36])


def add_sheet_scenario(wb, rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("시나리오 분석")
    r = add_common_header(ws, f"확률기반 시나리오 분석 ({profile})", "P80/P50/P20 분위수 기반 누적수익률 + 레버리지 ROI", 13)
    r = write_headers(ws, r, ["순위", "구", "현재매매(억)", "기간(년)", "낙관누적", "중립누적", "비관누적", "낙관ROI", "중립ROI", "비관ROI", "기준연변동성", "ADV리스크", "비고"])

    for x in rows:
        sc = x["scenario"]
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                round(x["latest_price_eok"], 2) if x.get("latest_price_eok") else "-",
                sc.get("horizon_years"),
                pct(sc.get("bull_cum")),
                pct(sc.get("base_cum")),
                pct(sc.get("bear_cum")),
                pct(sc.get("bull_roi")),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
                f"{(sc.get('annual_vol') or 0)*100:.1f}%",
                x.get("adv_risk_composite"),
                f"보유수익률 {((sc.get('annual_utility_yield') or 0)*100):.2f}%",
            ],
        )

    set_widths(ws, [7, 12, 11, 8, 10, 10, 10, 10, 10, 10, 11, 9, 12])


def add_sheet_scenario_method(wb, profile: str):
    ws = wb.create_sheet("시나리오 방법론")
    r = add_common_header(ws, "시나리오 산정 방법론", "현실반영형 확률모형 + 정책/레버리지/세금 보정", 6)

    steps = [
        "1) 월별 매매중앙값 시계열로 12M/3M 성장률 및 연환산 변동성 산출",
        "2) 기초성장률 = 0.60*12M + 0.40*3M (상·하한 clipping으로 과최적화 방지)",
        "3) 월수익률은 winsorize(10%) + robust 평균/표준편차로 추정",
        "4) 월 하방꼬리위험(ES 20%)을 stability/시나리오 하방에 반영",
        "5) 정책/신선도 레짐(집중규제·거래공백월) 보정으로 상하방 비대칭 반영",
        "6) 단지 성장률은 구 성장률로 shrinkage(표본가중)해 노이즈 축소",
        "7) 단지 레벨 총점은 구 점수(거시) 20% + 단지 점수(미시) 80% 결합",
        "8) 전세가율 과열구간(갭리스크) 패널티를 value/policy/시나리오에 반영",
        "9) 프로필별 LTV/금리/세금/보유비용으로 레버리지 ROI 계산",
        "10) 금리충격은 고정 ±1%p가 아니라 단지별 가변충격폭으로 계산",
        "11) 워크포워드(lookback/forward 자동튜닝)로 전략 초과수익을 점검",
        "12) 비정량 합성점수(NQ)를 총점/시나리오 drift/민감도에 함께 반영",
        "13) 고급리스크(ADV)를 총점 패널티·하방 drift·금리민감도에 동시 반영",
    ]

    r = write_row(ws, r, ["[방법론 단계]"] + [""] * 5, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["단계", "설명", "수식/로직", "현실보정", "한계", "완화장치"])

    rows = [
        ["Step1", "기초통계", "median(price_t), median(rent_t)", "이상치 완화", "표본부족", "DataQ 경고"],
        ["Step2", "추세 산출", "g_base=0.6*g12+0.4*g3", "최근/중기 균형", "최근왜곡", "clipping"],
        ["Step3", "강건분포", "winsorize + robust(μ,σ) + empirical quantile(P20/P50/P80)", "꼬리값 완화", "구조변화", "주기 재추정"],
        ["Step4", "꼬리위험", "ES20 = avg(bottom20% monthly return)", "하방리스크 반영", "표본 민감", "최소표본 필터"],
        ["Step5", "레짐보정", "R'=R+regime(stale/policy)", "현실 충격 반영", "정책변화", "월별 업데이트"],
        ["Step6", "노이즈축소", "g_adj=w*g+(1-w)*g_district", "표본작은 단지 안정화", "과도한 평활", "w 상한/하한"],
        ["Step7", "재무반영", "ROI=(NetGain-이자-세금)/자기자본", "투자자별 현실성", "세법변경", "가정표 분리"],
        ["Step8", "민감도", "ROI(r±shock_i)", "단지별 금리충격", "단순 모수", "충격폭 가변화"],
        ["Step9", "비정량보정", "Score'=Score+NQoverlay", "정량신호의 현실보정", "prior 주관성", "prior+model 혼합"],
        ["Step10", "시나리오보정", "drift'=drift-f(nq_risk,nq_exec)", "하방 리스크 현실화", "과보수화", "상한/하한 캡"],
        ["Step11", "고급리스크보정", "Score'=Score-ADVpenalty, drift'=drift-f(adv_risk,supply,guarantee,mgmt)", "거래절벽/정책플립/실행실패 반영", "모형단순화", "캡/플래그/신뢰도 계층"],
    ]
    for x in rows:
        r = write_row(ws, r, x, left=True)

    r += 1
    r = write_row(ws, r, ["[프로필 설정]"] + [""] * 5, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["프로필", "핵심목표", "LTV", "대출금리", "취득세", "양도세"])
    for p, cfg in PROFILE_CONFIGS.items():
        r = write_row(
            ws,
            r,
            [
                p,
                cfg["display"],
                f"{cfg['ltv']*100:.0f}%",
                f"{cfg['loan_rate']*100:.2f}%",
                f"{cfg['buy_tax']*100:.2f}%",
                f"{cfg['capital_gain_tax']*100:.1f}%",
            ],
            left=True,
        )

    r += 1
    r = write_row(ws, r, ["[현재 생성 프로필]", profile, "", "", "", ""], fill=FILL_LIGHT, left=True)

    set_widths(ws, [11, 37, 20, 16, 14, 14])


def add_sheet_profile_weights(wb):
    ws = wb.create_sheet("프로필가중치")
    r = add_common_header(ws, "프로필별 멀티전문가 가중치", "실거주/투자자/법인 전략차이를 수치화", 8)
    r = write_headers(ws, r, ["프로필", "Trend", "Value", "Liquidity", "Stability", "Policy", "합계", "전략요약"])

    for p, cfg in PROFILE_CONFIGS.items():
        w = cfg["weights"]
        r = write_row(
            ws,
            r,
            [
                p,
                w["trend"],
                w["value"],
                w["liquidity"],
                w["stability"],
                w["policy"],
                sum(w.values()),
                cfg["display"],
            ],
            left=True,
        )

    set_widths(ws, [10, 8, 8, 10, 10, 8, 8, 32])


def add_sheet_tax_leverage(wb):
    ws = wb.create_sheet("세금레버리지가정")
    r = add_common_header(ws, "세금/레버리지 가정", "프로필별 보수적 가정치", 9)
    r = write_headers(ws, r, ["프로필", "LTV", "대출금리", "취득세", "연보유비용", "양도세", "투자기간(년)", "의미", "주의"])

    notes = {
        "실거주": "거주효용 포함, 세후수익보다 안정성 우선",
        "투자자": "현금흐름·회전율·세후ROI 균형",
        "법인": "세율/취득세 부담이 커 고수익·고확실성 지역 선호",
    }
    warns = {
        "실거주": "거주의무/이주비용 누락 금지",
        "투자자": "DSR/금리변동에 민감",
        "법인": "취득세 12%와 세후수익 급감 리스크",
    }

    for p, cfg in PROFILE_CONFIGS.items():
        r = write_row(
            ws,
            r,
            [
                p,
                f"{cfg['ltv']*100:.0f}%",
                f"{cfg['loan_rate']*100:.2f}%",
                f"{cfg['buy_tax']*100:.2f}%",
                f"{cfg['annual_hold_cost']*100:.2f}%",
                f"{cfg['capital_gain_tax']*100:.1f}%",
                cfg["horizon_years"],
                notes[p],
                warns[p],
            ],
            left=True,
        )

    set_widths(ws, [10, 8, 10, 9, 11, 8, 10, 28, 24])


def add_sheet_policy(wb, rows: list[dict[str, Any]]):
    ws = wb.create_sheet("정책규제체크")
    r = add_common_header(ws, "정책/규제 체크", "LTPS·금리·DSR 최신 공개자료 반영", 9)

    r = write_row(ws, r, ["[정책 이벤트]"] + [""] * 8, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["구분", "내용", "기준일", "시장영향", "점수반영", "리스크", "출처", "링크", "비고"])

    policy_rows = [
        ["LTPS", "서울시 전체 아파트 토지거래허가구역", "2025-10-20~2026-12-31", "유동성/레버리지 제약", "Policy score base=45", "중", "서울시", "https://land.seoul.go.kr/land/other/appointStatusSeoul.do", ""],
        ["LTPS", "강남·서초·송파·용산 집중규제", "2026-01-07 고시", "상방축소/하방확대", "집중규제구 -15", "중~상", "서울시", "https://land.seoul.go.kr/land/other/appointStatusSeoul.do", ""],
        ["금리", "한국은행 기준금리 2.50% 동결", "2026-01-15", "차입비용 기준선", "loan_rate anchor", "중", "한국은행", "https://www.bok.or.kr/portal/bbs/B0000169/view.do?menuNo=200059&nttId=10095737", ""],
        ["대출", "3단계 스트레스 DSR", "2025-07 시행", "대출가능액 축소", "LTV/금리 보수화", "중", "금융위원회", "https://www.fsc.go.kr/no010101/84617", ""],
    ]
    for p in policy_rows:
        r = write_row(ws, r, p, left=True)

    r += 1
    r = write_row(ws, r, ["[구별 규제강도 매핑]"] + [""] * 8, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["구", "정책점수", "강도", "집중규제", "유동성보정", "데이터품질", "권고", "메모", ""])
    for x in rows:
        intensity = "집중" if x["district"] in HIGH_INTENSITY_LTPS else "일반"
        rec = "보수진입" if intensity == "집중" else "분할검토"
        r = write_row(
            ws,
            r,
            [
                x["district"],
                x["policy_score"],
                intensity,
                "Y" if intensity == "집중" else "N",
                round(x["avg_trade_3m"], 1),
                x["data_quality"],
                rec,
                "",
                "",
            ],
        )

    set_widths(ws, [10, 9, 8, 9, 10, 10, 12, 15, 8])


def add_sheet_timing(wb, rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("매수타이밍신호")
    r = add_common_header(ws, f"매수 타이밍 신호 ({profile})", "추세·유동성·안정성 결합신호 + ADV 리스크 차감", 12)
    r = write_headers(ws, r, ["순위", "구", "타이밍점수", "Trend", "Liquidity", "Stability", "최근3M", "12M", "ADV리스크", "신호", "전략", "주의"])

    for x in rows:
        fresh_bonus = (x.get("market_freshness") or 0.0) * 8.0
        stale_penalty = (x.get("trade_stale_months") or 0.0) * 1.2
        adv_drag = clamp(safe_float(x.get("adv_risk_composite"), 50.0) - 45.0, 0.0, 55.0) * 0.12
        timing = 0.50 * x["trend_score"] + 0.30 * x["liquidity_score"] + 0.20 * x["stability_score"] + fresh_bonus - stale_penalty - adv_drag
        if timing >= 70:
            sig = "진입 우호"
            strat = "3회 분할매수"
        elif timing >= 55:
            sig = "중립"
            strat = "조건부 진입"
        else:
            sig = "관망"
            strat = "지표 개선 대기"

        caution = "정책민감" if x["district"] in HIGH_INTENSITY_LTPS else "일반"
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                round(timing, 2),
                x["trend_score"],
                x["liquidity_score"],
                x["stability_score"],
                pct(x.get("growth_3m")),
                pct(x.get("growth_12m")),
                round(safe_float(x.get("adv_risk_composite"), 50.0), 1),
                sig,
                strat,
                caution,
            ],
        )
        ws.cell(row=r - 1, column=3).fill = score_band_fill(float(timing))

    set_widths(ws, [7, 11, 10, 8, 9, 9, 9, 9, 9, 10, 12, 9])


def add_sheet_complex_overview(
    wb,
    complex_rows: list[dict[str, Any]],
    district_rows: list[dict[str, Any]],
    top_per_district: int,
):
    ws = wb.create_sheet("구내상위단지 요약")
    r = add_common_header(
        ws,
        "구 내부 상위단지 요약",
        f"구별 상위 {top_per_district}개 단지의 평균 점수/수익률 비교",
        11,
    )
    r = write_headers(
        ws,
        r,
        ["구", "구점수", "단지수(분석)", f"Top{top_per_district} 평균점수", f"Top{top_per_district} 평균중립ROI", f"Top{top_per_district} 평균비관ROI", "Top 단지명", "Top 단지점수", "Top단지 중립ROI", "전세가율", "비고"],
    )

    dmap = {d["district"]: d for d in district_rows}
    group: dict[str, list[dict[str, Any]]] = {}
    for c in complex_rows:
        group.setdefault(c["district"], []).append(c)

    for district, items in sorted(group.items(), key=lambda x: dmap.get(x[0], {}).get("rank", 999)):
        items_sorted = sorted(items, key=lambda x: x["total_score"], reverse=True)
        topn = items_sorted[:top_per_district]
        top1 = items_sorted[0]
        avg_score = statistics.mean([x["total_score"] for x in topn]) if topn else 0.0
        avg_base_roi = statistics.mean([x["scenario"]["base_roi"] for x in topn if x["scenario"].get("base_roi") is not None]) if topn else 0.0
        avg_bear_roi = statistics.mean([x["scenario"]["bear_roi"] for x in topn if x["scenario"].get("bear_roi") is not None]) if topn else 0.0
        top1_base = top1["scenario"].get("base_roi")
        note = "표본주의" if top1.get("data_quality", 0) < 55 else ""
        r = write_row(
            ws,
            r,
            [
                district,
                round((dmap.get(district, {}).get("total_score") or 0), 2),
                len(items_sorted),
                round(avg_score, 2),
                pct(avg_base_roi),
                pct(avg_bear_roi),
                top1["complex"],
                round(top1["total_score"], 2),
                pct(top1_base),
                f"{(top1.get('latest_jeonse_ratio') or 0):.1f}%" if top1.get("latest_jeonse_ratio") else "-",
                note,
            ],
            left=True,
        )

    set_widths(ws, [11, 9, 10, 11, 12, 12, 26, 10, 12, 9, 10])


def add_sheet_complex_rank(wb, complex_rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("서울단지 랭킹")
    r = add_common_header(
        ws,
        f"서울 단지 통합 랭킹 ({profile})",
        "구 단위 점수 + 단지 미시지표 결합",
        32,
    )
    r = write_headers(
        ws,
        r,
        [
            "전체순위",
            "구내순위",
            "구",
            "단지",
            "총점",
            "1순위테마",
            "1순위점수",
            "테마액션",
            "2순위테마",
            "퀀트최종",
            "퀀트레거시",
            "팩터순수",
            "NQ오버레이",
            "NQ합성",
            "NQ리스크",
            "ADV리스크",
            "ADV패널티",
            "Core점수",
            "최근매매(억)",
            "12M성장",
            "전세가율",
            "변동성",
            "최근거래",
            "거래공백월",
            "중립ROI",
            "비관ROI",
            "DataQ",
            "신뢰도",
            "매매표본합",
            "신선도패널티",
            "꼬리패널티",
            "NQ플래그패널티",
        ],
    )

    for x in complex_rows:
        sc = x["scenario"]
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district_rank"],
                x["district"],
                x["complex"],
                x["total_score"],
                x.get("theme_primary"),
                x.get("theme_primary_score"),
                x.get("theme_action"),
                x.get("theme_secondary"),
                x.get("quant_total_score"),
                x.get("quant_total_score_legacy"),
                x.get("quant_factor_score"),
                x.get("nq_overlay"),
                x.get("nq_composite"),
                x.get("nq_risk_index"),
                x.get("adv_risk_composite"),
                x.get("adv_risk_penalty"),
                x.get("core_score"),
                round(x["latest_price_eok"], 2) if x.get("latest_price_eok") else "-",
                pct(x.get("growth_12m")),
                f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                x.get("latest_trade_count") or 0,
                x.get("trade_stale_months") or 0,
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
                x.get("data_quality"),
                x.get("confidence_bucket"),
                x.get("total_trade_samples"),
                x.get("recency_penalty"),
                x.get("tail_penalty"),
                x.get("nq_flag_penalty"),
            ],
            left=True,
        )
        ws.cell(row=r - 1, column=5).fill = score_band_fill(float(x["total_score"]))
        ws.cell(row=r - 1, column=7).fill = score_band_fill(float(x.get("theme_primary_score") or 0))
        ws.cell(row=r - 1, column=14).fill = score_band_fill(float(x.get("nq_composite") or 0))
        ws.cell(row=r - 1, column=16).fill = score_band_fill(100.0 - float(x.get("adv_risk_composite") or 50.0))

    set_widths(
        ws,
        [8, 8, 11, 26, 9, 18, 10, 10, 18, 9, 9, 9, 9, 9, 8, 8, 8, 9, 9, 10, 10, 9, 9, 9, 10, 10, 10, 8, 8, 9, 11, 11],
    )


def add_sheet_complex_district_top(wb, complex_rows: list[dict[str, Any]], top_per_district: int):
    ws = wb.create_sheet("구별단지 TopN")
    r = add_common_header(
        ws,
        "구별 상위 단지 리스트",
        f"각 구당 상위 {top_per_district}개 단지 표시",
        12,
    )
    r = write_headers(
        ws,
        r,
        ["구", "구내순위", "단지", "총점", "최근매매(억)", "12M성장", "전세가율", "거래량(최근)", "중립ROI", "비관ROI", "DataQ", "비고"],
    )

    by_d: dict[str, list[dict[str, Any]]] = {}
    for c in complex_rows:
        by_d.setdefault(c["district"], []).append(c)

    for district in sorted(by_d):
        items = sorted(by_d[district], key=lambda x: x["total_score"], reverse=True)[:top_per_district]
        for x in items:
            sc = x["scenario"]
            note = "집중규제" if district in HIGH_INTENSITY_LTPS else ""
            r = write_row(
                ws,
                r,
                [
                    district,
                    x["district_rank"],
                    x["complex"],
                    x["total_score"],
                    round(x["latest_price_eok"], 2) if x.get("latest_price_eok") else "-",
                    pct(x.get("growth_12m")),
                    f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                    x.get("latest_trade_count") or 0,
                    pct(sc.get("base_roi")),
                    pct(sc.get("bear_roi")),
                    x.get("data_quality"),
                    note,
                ],
                left=True,
            )
        r = write_row(ws, r, [""] * 12)

    set_widths(ws, [10, 8, 26, 9, 10, 10, 9, 10, 10, 10, 8, 10])


def add_sheet_complex_scenario(wb, complex_rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("단지 시나리오")
    r = add_common_header(
        ws,
        f"단지별 시나리오 상세 ({profile})",
        "상위 단지 중심 낙관/중립/비관 수익률",
        13,
    )
    r = write_headers(
        ws,
        r,
        ["순위", "구", "단지", "현재가(억)", "기간(년)", "낙관누적", "중립누적", "비관누적", "낙관ROI", "중립ROI", "비관ROI", "ADV리스크", "가정(연보유수익률)"],
    )

    # 시나리오 시트는 상위 300개까지만 표기해 파일 크기/가독성 균형
    for x in complex_rows[:300]:
        sc = x["scenario"]
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["complex"],
                round(x["latest_price_eok"], 2) if x.get("latest_price_eok") else "-",
                sc.get("horizon_years"),
                pct(sc.get("bull_cum")),
                pct(sc.get("base_cum")),
                pct(sc.get("bear_cum")),
                pct(sc.get("bull_roi")),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
                x.get("adv_risk_composite"),
                f"{((sc.get('annual_utility_yield') or 0)*100):.2f}%",
            ],
            left=True,
        )

    set_widths(ws, [7, 10, 26, 10, 8, 10, 10, 10, 10, 10, 10, 9, 13])


def add_sheet_complex_confidence(wb, complex_rows: list[dict[str, Any]]):
    ws = wb.create_sheet("단지 신뢰도")
    r = add_common_header(
        ws,
        "단지 데이터 신뢰도",
        "표본수/커버리지/품질점수 기반 high·medium·low 등급",
        17,
    )
    r = write_headers(
        ws,
        r,
        [
            "순위",
            "구",
            "단지",
            "신뢰도",
            "NQ신뢰도",
            "ADV등급",
            "NQ합성",
            "DataQ",
            "매매표본합",
            "최근매매표본",
            "매매커버리지",
            "최종거래월",
            "거래공백월",
            "불확실성패널티",
            "샘플가중치",
            "중립ROI",
            "비고",
        ],
    )
    for x in complex_rows:
        sc = x["scenario"]
        bucket = x.get("confidence_bucket", "medium")
        note = "주의" if bucket == "low" else ""
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["complex"],
                bucket,
                x.get("nq_confidence_bucket"),
                x.get("adv_risk_bucket"),
                x.get("nq_composite"),
                x.get("data_quality"),
                x.get("total_trade_samples"),
                x.get("latest_trade_count"),
                f"{(x.get('coverage_trade') or 0)*100:.1f}%",
                x.get("last_trade_yyyymm") or "-",
                x.get("trade_stale_months") or 0,
                round(x.get("uncertainty_penalty") or 0.0, 3),
                x.get("sample_weight"),
                pct(sc.get("base_roi")),
                note,
            ],
            left=True,
        )
        score_col = 4
        bfill = (
            PatternFill("solid", fgColor="D5F5E3")
            if bucket == "high"
            else PatternFill("solid", fgColor="FCF3CF")
            if bucket == "medium"
            else PatternFill("solid", fgColor="FADBD8")
        )
        ws.cell(row=r - 1, column=score_col).fill = bfill
    set_widths(ws, [7, 10, 26, 9, 9, 9, 8, 8, 10, 10, 10, 10, 10, 12, 10, 10, 10])


def add_sheet_complex_sensitivity(wb, complex_rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("단지 민감도")
    r = add_common_header(
        ws,
        f"단지 금리 민감도 ({profile})",
        "중립 시나리오 기준 단지별 금리충격(가변) 영향",
        13,
    )
    r = write_headers(
        ws,
        r,
        ["순위", "구", "단지", "중립ROI", "금리상승 ROI", "금리하락 ROI", "민감도폭", "적용충격폭", "LTV", "기준금리", "ADV리스크", "신뢰도", "비고"],
    )
    for x in complex_rows[:400]:
        sc = x["scenario"]
        roi_base = sc.get("base_roi")
        roi_up = sc.get("base_roi_rate_up")
        roi_dn = sc.get("base_roi_rate_down")
        band = None
        if roi_up is not None and roi_dn is not None:
            band = roi_dn - roi_up
        asmp = sc.get("assumption", {})
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["complex"],
                pct(roi_base),
                pct(roi_up),
                pct(roi_dn),
                pct(band),
                f"{(asmp.get('effective_rate_shock') or 0)*100:.2f}%p",
                f"{(asmp.get('ltv') or 0)*100:.0f}%",
                f"{(asmp.get('loan_rate') or 0)*100:.2f}%",
                x.get("adv_risk_composite"),
                sc.get("confidence_bucket"),
                "",
            ],
            left=True,
        )
    set_widths(ws, [7, 10, 26, 10, 11, 11, 10, 10, 8, 9, 9, 8, 10])


def add_sheet_nonquant_overlay(
    wb,
    rows: list[dict[str, Any]],
    complex_rows: list[dict[str, Any]],
    profile: str,
):
    ws = wb.create_sheet("비정량 오버레이")
    r = add_common_header(
        ws,
        f"비정량 오버레이 점수 ({profile})",
        "거버넌스/생활환경/수요지속/관리품질/심리지표를 랭킹·시나리오에 결합",
        16,
    )

    r = write_row(ws, r, ["[구 단위 비정량 요약]"] + [""] * 15, fill=FILL_GOLD, left=True)
    r = write_headers(
        ws,
        r,
        [
            "순위",
            "구",
            "총점",
            "기존점수",
            "NQ오버레이",
            "NQ합성",
            "NQ리스크",
            "거버넌스",
            "생활환경",
            "수요지속",
            "관리품질",
            "심리지표",
            "NQ신뢰도",
            "NQ플래그",
            "중립ROI",
            "비관ROI",
        ],
    )
    for x in rows:
        sc = x.get("scenario", {})
        r = write_row(
            ws,
            r,
            [
                x.get("rank"),
                x.get("district"),
                x.get("total_score"),
                x.get("quant_total_score"),
                x.get("nq_overlay"),
                x.get("nq_composite"),
                x.get("nq_risk_index"),
                x.get("nq_governance"),
                x.get("nq_liveability"),
                x.get("nq_demand_persistence"),
                x.get("nq_management_quality"),
                x.get("nq_sentiment"),
                x.get("nq_confidence_bucket"),
                x.get("nq_flags"),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
            ],
            left=True,
        )
        ws.cell(row=r - 1, column=6).fill = score_band_fill(float(x.get("nq_composite") or 0))

    r += 1
    r = write_row(ws, r, ["[단지 비정량 요약 - 상위 300]"] + [""] * 15, fill=FILL_GOLD, left=True)
    r = write_headers(
        ws,
        r,
        [
            "전체순위",
            "구내순위",
            "구",
            "단지",
            "총점",
            "기존점수",
            "NQ오버레이",
            "NQ합성",
            "NQ리스크",
            "NQ신뢰도",
            "NQ플래그수",
            "NQ플래그",
            "중립ROI",
            "비관ROI",
            "거래공백월",
            "DataQ",
        ],
    )
    for x in complex_rows[:300]:
        sc = x.get("scenario", {})
        r = write_row(
            ws,
            r,
            [
                x.get("rank"),
                x.get("district_rank"),
                x.get("district"),
                x.get("complex"),
                x.get("total_score"),
                x.get("quant_total_score"),
                x.get("nq_overlay"),
                x.get("nq_composite"),
                x.get("nq_risk_index"),
                x.get("nq_confidence_bucket"),
                x.get("nq_flag_count"),
                x.get("nq_flags"),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
                x.get("trade_stale_months"),
                x.get("data_quality"),
            ],
            left=True,
        )
        ws.cell(row=r - 1, column=8).fill = score_band_fill(float(x.get("nq_composite") or 0))

    set_widths(ws, [8, 8, 10, 28, 9, 9, 9, 8, 8, 9, 9, 22, 10, 10, 10, 8])


def add_sheet_nonquant_method(wb, profile: str, nq_meta: dict[str, Any] | None = None):
    ws = wb.create_sheet("비정량 방법론")
    nq_meta = nq_meta or {}
    r = add_common_header(
        ws,
        f"비정량 방법론 ({profile})",
        "점수 산식과 현실 반영 기준(정량+정성 결합) 공개",
        5,
    )
    r = write_headers(ws, r, ["구분", "내용", "수식/규칙", "파라미터", "비고"])

    profile_weight = NQ_PROFILE_WEIGHTS.get(profile) or {}
    weight_line = ", ".join(
        f"{k.replace('nq_', '')}:{v:.2f}" for k, v in profile_weight.items()
    )
    prior_w = safe_float(nq_meta.get("prior_weight"), DEFAULT_NQ_PRIOR_WEIGHT)
    cap_d = safe_float(nq_meta.get("overlay_cap"), DEFAULT_NQ_OVERLAY_CAP)
    cap_c = safe_float(nq_meta.get("complex_overlay_cap"), DEFAULT_NQ_COMPLEX_OVERLAY_CAP)
    enabled_txt = "on" if nq_meta.get("enabled", True) else "off"

    rows = [
        ["입력계층", "시장모델 계층", "실거래 기반 프록시로 5개 비정량 차원 점수화", "model_dims", "거래신선도·거래량·변동성·전세구조 포함"],
        ["입력계층", "사전값 계층", "구 단위 prior JSON (외부 파일)과 블렌딩", f"prior_weight={prior_w:.2f}", "없으면 내장 priors 사용"],
        ["입력계층", "수동오버라이드 계층", "complex_overrides 존재 시 해당 단지 우선 반영", "override 65~80% 반영", "법적분쟁/하자/대형개발 등 이벤트용"],
        ["합성", "비정량 합성점수", "0.24*Gov + 0.24*Live + 0.22*Demand + 0.18*Mgmt + 0.12*Sent - FlagPenalty", "max penalty 14", "플래그 수·불확실성 패널티 동시 반영"],
        ["합성", "비정량 리스크지수", "100 - NQ합성", "0~100", "값이 높을수록 하방 리스크 큼"],
        ["랭킹반영", "프로필별 오버레이", "Overlay = cap * weighted_signal * confidence * (1-flag_drag)", weight_line, f"cap={cap_d:.1f}(구), {cap_c:.1f}(단지)"],
        ["시나리오반영", "하방보정", "base/bear drift에 nq_risk·nq_exec·nq_sent를 추가 차감", "flag_pen 포함", "낙관치도 소폭 보수화"],
        ["운영", "사전값 메타", f"as_of={nq_meta.get('as_of', '-')}, source={nq_meta.get('source', 'default')}", f"version={nq_meta.get('version', 'v7')}, enabled={enabled_txt}", str(nq_meta.get('method', '-'))],
        ["주의", "해석 원칙", "NQ는 단독 매수신호가 아니라 정량 신호의 현실 보정치", "최종 의사결정은 현장/법무/세무 검증 병행", "저신뢰(LOW) 단지는 추가 DD 필수"],
    ]
    for row in rows:
        r = write_row(ws, r, row, left=True)

    set_widths(ws, [10, 18, 56, 30, 24])


def add_sheet_advanced_risk_overlay(
    wb,
    rows: list[dict[str, Any]],
    complex_rows: list[dict[str, Any]],
    profile: str,
):
    ws = wb.create_sheet("고급리스크 오버레이")
    r = add_common_header(
        ws,
        f"고급리스크 오버레이 ({profile})",
        "유동성함정/군집쏠림/정책플립/실행리스크 + 공급/보증/관리DD를 총점·하방에 반영",
        17,
    )

    r = write_row(ws, r, ["[구 단위 고급리스크 요약]"] + [""] * 16, fill=FILL_GOLD, left=True)
    r = write_headers(
        ws,
        r,
        [
            "순위",
            "구",
            "총점",
            "ADV리스크",
            "유동성함정",
            "군집쏠림",
            "정책플립",
            "실행리스크",
            "공급압력",
            "보증리스크",
            "관리DD리스크",
            "ADV패널티",
            "ADV등급",
            "ADV플래그",
            "중립ROI",
            "비관ROI",
            "비고",
        ],
    )
    for x in rows:
        sc = x.get("scenario", {})
        r = write_row(
            ws,
            r,
            [
                x.get("rank"),
                x.get("district"),
                x.get("total_score"),
                x.get("adv_risk_composite"),
                x.get("adv_liquidity_trap"),
                x.get("adv_crowding"),
                x.get("adv_policy_flip"),
                x.get("adv_execution"),
                x.get("adv_supply_pressure"),
                x.get("adv_guarantee_risk"),
                x.get("adv_management_dd_risk"),
                x.get("adv_risk_penalty"),
                x.get("adv_risk_bucket"),
                x.get("adv_flags"),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
                f"{x.get('adv_source', 'model')} / w={safe_float(x.get('adv_signal_weight'), 0.0):.2f}",
            ],
            left=True,
        )
        ws.cell(row=r - 1, column=4).fill = score_band_fill(100.0 - float(x.get("adv_risk_composite") or 50.0))

    r += 1
    r = write_row(ws, r, ["[단지 고급리스크 요약 - 상위 300]"] + [""] * 16, fill=FILL_GOLD, left=True)
    r = write_headers(
        ws,
        r,
        [
            "전체순위",
            "구내순위",
            "구",
            "단지",
            "총점",
            "ADV리스크",
            "유동성함정",
            "군집쏠림",
            "정책플립",
            "실행리스크",
            "공급압력",
            "보증리스크",
            "관리DD리스크",
            "ADV패널티",
            "중립ROI",
            "비관ROI",
            "ADV플래그",
        ],
    )
    for x in complex_rows[:300]:
        sc = x.get("scenario", {})
        r = write_row(
            ws,
            r,
            [
                x.get("rank"),
                x.get("district_rank"),
                x.get("district"),
                x.get("complex"),
                x.get("total_score"),
                x.get("adv_risk_composite"),
                x.get("adv_liquidity_trap"),
                x.get("adv_crowding"),
                x.get("adv_policy_flip"),
                x.get("adv_execution"),
                x.get("adv_supply_pressure"),
                x.get("adv_guarantee_risk"),
                x.get("adv_management_dd_risk"),
                x.get("adv_risk_penalty"),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
                f"{x.get('adv_flags')}{' | ' + str(x.get('adv_source')) if x.get('adv_source') else ''}",
            ],
            left=True,
        )
        ws.cell(row=r - 1, column=6).fill = score_band_fill(100.0 - float(x.get("adv_risk_composite") or 50.0))

    set_widths(ws, [8, 8, 10, 28, 9, 8, 9, 9, 9, 9, 9, 10, 10, 9, 9, 10, 22])


def add_sheet_advanced_risk_method(wb, profile: str, adv_meta: dict[str, Any] | None = None):
    ws = wb.create_sheet("고급리스크 방법론")
    adv_meta = adv_meta or {}
    r = add_common_header(
        ws,
        f"고급리스크 방법론 ({profile})",
        "정량/NQ 외 실전 하방요인을 별도 계층으로 분리",
        5,
    )
    r = write_headers(ws, r, ["구분", "내용", "수식/규칙", "파라미터", "비고"])

    w = ADV_RISK_PROFILE_WEIGHTS.get(profile) or {}
    weight_line = ", ".join(f"{k.replace('adv_', '')}:{v:.2f}" for k, v in w.items())
    cap_d = safe_float(adv_meta.get("district_cap"), DEFAULT_ADV_RISK_DISTRICT_CAP)
    cap_c = safe_float(adv_meta.get("complex_cap"), DEFAULT_ADV_RISK_COMPLEX_CAP)
    enabled_txt = "on" if adv_meta.get("enabled", True) else "off"
    sig_source = str(adv_meta.get("source", "default"))
    sig_file = str(adv_meta.get("signal_file", "-") or "-")
    auto_enabled = "on" if adv_meta.get("autofill_enabled", False) else "off"
    auto_overwrite = "on" if adv_meta.get("autofill_overwrite", False) else "off"
    auto_run_id = str(adv_meta.get("autofill_run_id", "-") or "-")
    auto_hist = int(safe_float(adv_meta.get("autofill_history_size"), 0.0))
    auto_total_d = int(safe_float(adv_meta.get("autofill_total_rows_district"), 0.0))
    auto_total_c = int(safe_float(adv_meta.get("autofill_total_rows_complex"), 0.0))
    auto_signal_d = int(safe_float(adv_meta.get("autofill_signal_rows_district"), 0.0))
    auto_signal_c = int(safe_float(adv_meta.get("autofill_signal_rows_complex"), 0.0))
    auto_assigned_d = int(safe_float(adv_meta.get("autofill_assigned_rows_district"), 0.0))
    auto_assigned_c = int(safe_float(adv_meta.get("autofill_assigned_rows_complex"), 0.0))
    auto_rows_d = int(safe_float(adv_meta.get("autofill_updated_rows_district"), 0.0))
    auto_rows_c = int(safe_float(adv_meta.get("autofill_updated_rows_complex"), 0.0))
    auto_changed_d = int(safe_float(adv_meta.get("autofill_changed_rows_district"), 0.0))
    auto_changed_c = int(safe_float(adv_meta.get("autofill_changed_rows_complex"), 0.0))
    auto_range = str(adv_meta.get("autofill_month_range", "-"))

    rows = [
        ["리스크축", "유동성함정", "f(거래량부족, 거래공백, 데이터취약, 꼬리위험)", "avg_trade_3m, stale, DataQ, ES", "매도 실패·호가갭 리스크"],
        ["리스크축", "군집쏠림", "f(단기/중기 모멘텀 과열, 고가프리미엄, 갭과열)", "g12, g3, price_vs_seoul, jeonse", "추격매수 후 mean-reversion 리스크"],
        ["리스크축", "정책플립", "f(LTPS강도, 정책점수, 과열·고가중첩)", "district LTPS, policy_score", "규제 강화/완화 전환 민감도"],
        ["리스크축", "실행리스크", "f(불확실성·데이터품질·거버넌스·플래그)", "uncertainty_penalty, nq_gov", "법무/하자/체결실패 확률"],
        ["보조축", "공급압력", "f(중기둔화, 저유동성, 저임대수익, 고가프리미엄)", "trend12, liquidity, yield, premium", "입주·공급충격 대용지표"],
        ["보조축", "보증리스크", "f(전세가율, 꼬리위험, 거래공백, 유동성)", "jeonse, ES, stale", "전세보증 회수 스트레스 대용"],
        ["보조축", "관리DD리스크", "f(DataQ, 거버넌스, 불확실성, 관리품질)", "dq, nq_gov, nq_mgmt", "장기 보유/분쟁 리스크"],
        ["입력계층", "외생시그널", "district_signals/complex_signals JSON 블렌딩", "supply/guarantee/management + confidence", "없으면 model 단독"],
        ["합성", "ADV 리스크합성", "0.32*LiqTrap + 0.23*Crowd + 0.27*Policy + 0.18*Exec + flag_boost", "0~100", "높을수록 위험"],
        ["랭킹반영", "ADV 패널티", "Penalty = cap * weighted_risk * confidence * flag_mult", weight_line, f"cap={cap_d:.1f}(구), {cap_c:.1f}(단지)"],
        ["시나리오반영", "하방 drift 보정", "base/bear drift에서 adv_risk·liq·policy·exec·supply·guarantee 차감", "adv_flag_pen 포함", "낙관 시나리오도 소폭 보수화"],
        ["민감도반영", "금리충격 연동", "shock_scale += f(adv_risk, adv_liq, adv_guarantee)", "effective_rate_shock 확대", "취약자산의 금리탄력성 상향"],
        ["운영", "외생시그널 메타", f"as_of={adv_meta.get('as_of', '-')}, source={sig_source}", f"file={sig_file}", str(adv_meta.get("method", "-"))],
        ["운영", "외생시그널 자동채움", f"enabled={auto_enabled}, overwrite={auto_overwrite}, run={auto_run_id}", f"changed: d={auto_changed_d}/{auto_total_d}, c={auto_changed_c}/{auto_total_c} (assigned d={auto_assigned_d}, c={auto_assigned_c})", f"signal_rows: d={auto_signal_d}, c={auto_signal_c}, hist={auto_hist}, range={auto_range}"],
        ["운영", "활성화/캡", f"enabled={enabled_txt}", f"district_cap={cap_d:.1f}, complex_cap={cap_c:.1f}", "CLI에서 on/off 및 cap 조정"],
    ]
    for row in rows:
        r = write_row(ws, r, row, left=True)

    set_widths(ws, [10, 18, 58, 42, 24])


def add_sheet_advanced_signal_audit(
    wb,
    rows: list[dict[str, Any]],
    complex_rows: list[dict[str, Any]],
    profile: str,
    adv_meta: dict[str, Any] | None = None,
):
    ws = wb.create_sheet("ADV시그널감사로그")
    adv_meta = adv_meta or {}
    d_change_map = adv_meta.get("change_map_district") if isinstance(adv_meta.get("change_map_district"), dict) else {}
    c_change_map = adv_meta.get("change_map_complex") if isinstance(adv_meta.get("change_map_complex"), dict) else {}

    def _fmt_sig(v: Any) -> str:
        if v is None:
            return "-"
        try:
            fv = float(v)
            return f"{fv:.2f}"
        except Exception:  # noqa: BLE001
            return str(v)

    def _fmt_delta(entry: dict[str, Any] | None, key: str) -> str:
        if not entry:
            return "-"
        b = (entry.get("before") or {}).get(key)
        a = (entry.get("after") or {}).get(key)
        if b == a:
            return "-"
        return f"{_fmt_sig(b)} -> {_fmt_sig(a)}"

    def _entry_flag(entry: dict[str, Any] | None) -> str:
        if not entry:
            return "N"
        if bool(entry.get("touched")):
            return "Y"
        fields = entry.get("changed_fields")
        if isinstance(fields, list) and fields:
            return "Y"
        return "N"

    def _entry_fields(entry: dict[str, Any] | None) -> str:
        if not entry:
            return ""
        fields = entry.get("changed_fields")
        if isinstance(fields, list) and fields:
            return ", ".join(str(x) for x in fields if str(x).strip())
        if bool(entry.get("touched")):
            return "recomputed(no_change)"
        return ""

    r = add_common_header(
        ws,
        f"ADV 시그널 감사로그 ({profile})",
        "모델값/외생시그널/최종반영값 및 변경이력 추적",
        22,
    )
    r = write_row(
        ws,
        r,
        [
            "메타",
            f"source={adv_meta.get('source', '-')}",
            f"file={adv_meta.get('signal_file', '-')}",
            f"as_of={adv_meta.get('as_of', '-')}",
            f"autofill={adv_meta.get('autofill_enabled', False)}",
            f"overwrite={adv_meta.get('autofill_overwrite', False)}",
            f"run_id={adv_meta.get('autofill_run_id', '-')}",
            f"run_at={adv_meta.get('autofill_run_at', '-')}",
            f"history={adv_meta.get('autofill_history_size', 0)}",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        fill=FILL_GOLD,
        left=True,
    )

    r += 1
    r = write_row(ws, r, ["[구 단위 감사로그]"] + [""] * 21, fill=FILL_GOLD, left=True)
    r = write_headers(
        ws,
        r,
        [
            "순위",
            "구",
            "ADV출처",
            "시그널스코프",
            "시그널가중치",
            "모델공급",
            "시그널공급",
            "최종공급",
            "모델보증",
            "시그널보증",
            "최종보증",
            "모델관리DD",
            "시그널관리DD",
            "최종관리DD",
            "시그널신뢰도",
            "시그널플래그",
            "시그널출처",
            "변경여부",
            "변경필드",
            "공급변경(전→후)",
            "보증변경(전→후)",
            "관리DD변경(전→후)",
        ],
    )
    for x in rows:
        dkey = str(x.get("district") or "").strip()
        chg = d_change_map.get(dkey) if isinstance(d_change_map.get(dkey), dict) else None
        r = write_row(
            ws,
            r,
            [
                x.get("rank"),
                x.get("district"),
                x.get("adv_source"),
                x.get("adv_signal_scope"),
                safe_float(x.get("adv_signal_weight"), 0.0),
                x.get("adv_model_supply_pressure"),
                x.get("adv_signal_supply_pressure"),
                x.get("adv_supply_pressure"),
                x.get("adv_model_guarantee_risk"),
                x.get("adv_signal_guarantee_risk"),
                x.get("adv_guarantee_risk"),
                x.get("adv_model_management_dd_risk"),
                x.get("adv_signal_management_dd_risk"),
                x.get("adv_management_dd_risk"),
                x.get("adv_signal_confidence"),
                x.get("adv_signal_flags"),
                x.get("adv_signal_source"),
                _entry_flag(chg),
                _entry_fields(chg),
                _fmt_delta(chg, "supply_pressure"),
                _fmt_delta(chg, "guarantee_risk"),
                _fmt_delta(chg, "management_dd_risk"),
            ],
            left=True,
        )

    r += 1
    r = write_row(ws, r, ["[단지 감사로그 - 상위 300]"] + [""] * 21, fill=FILL_GOLD, left=True)
    r = write_headers(
        ws,
        r,
        [
            "전체순위",
            "구내순위",
            "구",
            "단지",
            "ADV출처",
            "시그널스코프",
            "시그널가중치",
            "모델공급",
            "시그널공급",
            "최종공급",
            "모델보증",
            "시그널보증",
            "최종보증",
            "모델관리DD",
            "시그널관리DD",
            "최종관리DD",
            "시그널출처",
            "변경여부",
            "변경필드",
            "공급변경(전→후)",
            "보증변경(전→후)",
            "관리DD변경(전→후)",
        ],
    )
    for x in complex_rows[:300]:
        ckey = str(x.get("complex_key") or "").strip()
        if not ckey:
            ckey = f"{str(x.get('district') or '').strip()}::{str(x.get('complex') or '').strip()}"
        chg = c_change_map.get(ckey) if isinstance(c_change_map.get(ckey), dict) else None
        r = write_row(
            ws,
            r,
            [
                x.get("rank"),
                x.get("district_rank"),
                x.get("district"),
                x.get("complex"),
                x.get("adv_source"),
                x.get("adv_signal_scope"),
                safe_float(x.get("adv_signal_weight"), 0.0),
                x.get("adv_model_supply_pressure"),
                x.get("adv_signal_supply_pressure"),
                x.get("adv_supply_pressure"),
                x.get("adv_model_guarantee_risk"),
                x.get("adv_signal_guarantee_risk"),
                x.get("adv_guarantee_risk"),
                x.get("adv_model_management_dd_risk"),
                x.get("adv_signal_management_dd_risk"),
                x.get("adv_management_dd_risk"),
                x.get("adv_signal_source"),
                _entry_flag(chg),
                _entry_fields(chg),
                _fmt_delta(chg, "supply_pressure"),
                _fmt_delta(chg, "guarantee_risk"),
                _fmt_delta(chg, "management_dd_risk"),
            ],
            left=True,
        )

    set_widths(ws, [8, 8, 10, 26, 11, 11, 10, 10, 11, 10, 10, 11, 10, 11, 10, 24, 28, 8, 22, 16, 16, 16])


def add_sheet_advanced_signal_history(
    wb,
    profile: str,
    adv_meta: dict[str, Any] | None = None,
):
    ws = wb.create_sheet("ADV시그널히스토리")
    adv_meta = adv_meta or {}
    history = adv_meta.get("autofill_run_history") if isinstance(adv_meta.get("autofill_run_history"), list) else []

    r = add_common_header(
        ws,
        f"ADV 시그널 자동채움 히스토리 ({profile})",
        "실행 회차별 변경/할당 요약",
        16,
    )
    r = write_row(
        ws,
        r,
        [
            "메타",
            f"history_count={len(history)}",
            f"source={adv_meta.get('source', '-')}",
            f"file={adv_meta.get('signal_file', '-')}",
            f"as_of={adv_meta.get('as_of', '-')}",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        fill=FILL_GOLD,
        left=True,
    )

    r += 1
    r = write_headers(
        ws,
        r,
        [
            "No",
            "run_id",
            "run_at",
            "month_range",
            "overwrite",
            "d_total",
            "d_signal",
            "d_assigned",
            "d_changed_rows",
            "d_changed_dims",
            "c_total",
            "c_signal",
            "c_assigned",
            "c_changed_rows",
            "c_changed_dims",
            "비고",
        ],
    )

    rows_h: list[dict[str, Any]] = []
    for x in history:
        if isinstance(x, dict):
            rows_h.append(x)
    rows_h = list(reversed(rows_h))
    if not rows_h:
        r = write_row(ws, r, ["-", "-", "-", "-", "-", "-", "-", "-", "-", "-", "-", "-", "-", "-", "-", "히스토리 없음"], left=True)
    else:
        for i, h in enumerate(rows_h, 1):
            r = write_row(
                ws,
                r,
                [
                    i,
                    h.get("run_id"),
                    h.get("run_at"),
                    h.get("month_range"),
                    "Y" if h.get("overwrite") else "N",
                    h.get("district_total"),
                    h.get("district_signal_rows"),
                    h.get("district_assigned_rows"),
                    h.get("district_changed_rows"),
                    h.get("district_changed_dims"),
                    h.get("complex_total"),
                    h.get("complex_signal_rows"),
                    h.get("complex_assigned_rows"),
                    h.get("complex_changed_rows"),
                    h.get("complex_changed_dims"),
                    "latest" if i == 1 else "",
                ],
                left=True,
            )

    set_widths(ws, [6, 18, 21, 14, 9, 8, 8, 10, 12, 12, 8, 8, 10, 12, 12, 10])


def add_sheet_theme_method(wb, profile: str):
    ws = wb.create_sheet("테마방법론")
    r = add_common_header(
        ws,
        f"투자 테마 분류 방법론 ({profile})",
        "정량·비정량·고급리스크를 테마별 기회/리스크 점수로 재조합",
        9,
    )
    r = write_headers(
        ws,
        r,
        [
            "테마코드",
            "테마명",
            "핵심아이디어",
            "기회신호(가중치)",
            "리스크신호(가중치)",
            "프로필바이어스",
            "프로필별 실행가이드",
            "핵심리스크",
            "모니터링포인트",
        ],
    )

    bias_map = THEME_PROFILE_BIAS.get(profile) or {}

    def fmt_w(w: dict[str, float]) -> str:
        items = sorted(w.items(), key=lambda kv: safe_float(kv[1], 0.0), reverse=True)
        return ", ".join([f"{THEME_FEATURE_LABELS.get(k, k)}:{safe_float(v, 0.0):.2f}" for k, v in items[:5]])

    for t in THEME_CATALOG:
        key = str(t.get("key") or "")
        tactic, risk, monitor = _theme_strategy_for_profile(key, profile, "기본전략")
        r = write_row(
            ws,
            r,
            [
                key,
                t.get("name"),
                t.get("desc"),
                fmt_w(t.get("opp_weights") or {}),
                fmt_w(t.get("risk_weights") or {}),
                f"{safe_float(bias_map.get(key), 1.0):.2f}",
                tactic,
                risk,
                monitor,
            ],
            left=True,
        )

    set_widths(ws, [14, 22, 30, 40, 40, 10, 36, 24, 32])


def add_sheet_theme_classification_district(wb, rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("구테마분류")
    r = add_common_header(ws, f"구 단위 테마 분류 ({profile})", "1순위/2순위 테마와 집중도(스프레드) 확인", 15)
    r = write_headers(
        ws,
        r,
        [
            "순위",
            "구",
            "총점",
            "1순위테마",
            "1순위점수",
            "1순위리스크",
            "2순위테마",
            "2순위점수",
            "2순위리스크",
            "테마스프레드",
            "테마주의",
            "테마액션",
            "테마TOP3",
            "중립ROI",
            "핵심근거",
        ],
    )
    for x in rows:
        sc = x.get("scenario", {})
        r = write_row(
            ws,
            r,
            [
                x.get("rank"),
                x.get("district"),
                x.get("total_score"),
                x.get("theme_primary"),
                x.get("theme_primary_score"),
                x.get("theme_primary_risk"),
                x.get("theme_secondary"),
                x.get("theme_secondary_score"),
                x.get("theme_secondary_risk"),
                x.get("theme_spread"),
                x.get("theme_caution"),
                x.get("theme_action"),
                x.get("theme_candidates"),
                pct(sc.get("base_roi")),
                x.get("theme_primary_reason"),
            ],
            left=True,
        )
        ws.cell(row=r - 1, column=5).fill = score_band_fill(safe_float(x.get("theme_primary_score"), 0.0))

    set_widths(ws, [7, 10, 9, 20, 10, 10, 20, 10, 10, 10, 8, 10, 30, 10, 40])


def add_sheet_theme_classification_complex(wb, complex_rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("단지테마분류")
    r = add_common_header(ws, f"단지 테마 분류 ({profile})", "상위 단지의 테마별 분류 및 집중도", 17)
    r = write_headers(
        ws,
        r,
        [
            "전체순위",
            "구내순위",
            "구",
            "단지",
            "총점",
            "1순위테마",
            "1순위점수",
            "1순위리스크",
            "2순위테마",
            "2순위점수",
            "2순위리스크",
            "테마스프레드",
            "테마주의",
            "테마액션",
            "중립ROI",
            "테마TOP3",
            "핵심근거",
        ],
    )
    for x in complex_rows[:600]:
        sc = x.get("scenario", {})
        r = write_row(
            ws,
            r,
            [
                x.get("rank"),
                x.get("district_rank"),
                x.get("district"),
                x.get("complex"),
                x.get("total_score"),
                x.get("theme_primary"),
                x.get("theme_primary_score"),
                x.get("theme_primary_risk"),
                x.get("theme_secondary"),
                x.get("theme_secondary_score"),
                x.get("theme_secondary_risk"),
                x.get("theme_spread"),
                x.get("theme_caution"),
                x.get("theme_action"),
                pct(sc.get("base_roi")),
                x.get("theme_candidates"),
                x.get("theme_primary_reason"),
            ],
            left=True,
        )
        ws.cell(row=r - 1, column=7).fill = score_band_fill(safe_float(x.get("theme_primary_score"), 0.0))

    set_widths(ws, [8, 8, 10, 28, 9, 20, 10, 10, 20, 10, 10, 10, 8, 10, 10, 30, 36])


def add_sheet_theme_ranking(wb, rows: list[dict[str, Any]], complex_rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("테마별랭킹")
    r = add_common_header(ws, f"테마별 랭킹 ({profile})", "테마 점수 기준 Top 후보를 테마별로 정리", 10)

    r = write_row(ws, r, ["[구 테마 Top5]"] + [""] * 9, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["테마", "순위", "구", "테마점수", "테마리스크", "테마액션", "총점", "중립ROI", "비관ROI", "핵심근거"])
    for t in THEME_CATALOG:
        key = str(t.get("key") or "")
        name = str(t.get("name") or key)
        sorted_rows = sorted(
            rows,
            key=lambda x: safe_float(((x.get("theme_score_map") or {}).get(key, {}) or {}).get("conviction"), 0.0),
            reverse=True,
        )[:5]
        for i, x in enumerate(sorted_rows, 1):
            ent = ((x.get("theme_score_map") or {}).get(key, {}) or {})
            sc = x.get("scenario", {})
            r = write_row(
                ws,
                r,
                [
                    name,
                    i,
                    x.get("district"),
                    ent.get("conviction"),
                    ent.get("risk"),
                    x.get("theme_action"),
                    x.get("total_score"),
                    pct(sc.get("base_roi")),
                    pct(sc.get("bear_roi")),
                    ent.get("driver"),
                ],
                left=True,
            )
        r = write_row(ws, r, [""] * 10)

    r += 1
    r = write_row(ws, r, ["[단지 테마 Top8]"] + [""] * 9, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["테마", "순위", "구", "단지", "테마점수", "테마리스크", "테마액션", "총점", "중립ROI", "핵심근거"])
    for t in THEME_CATALOG:
        key = str(t.get("key") or "")
        name = str(t.get("name") or key)
        sorted_rows = sorted(
            complex_rows,
            key=lambda x: safe_float(((x.get("theme_score_map") or {}).get(key, {}) or {}).get("conviction"), 0.0),
            reverse=True,
        )[:8]
        for i, x in enumerate(sorted_rows, 1):
            ent = ((x.get("theme_score_map") or {}).get(key, {}) or {})
            sc = x.get("scenario", {})
            r = write_row(
                ws,
                r,
                [
                    name,
                    i,
                    x.get("district"),
                    x.get("complex"),
                    ent.get("conviction"),
                    ent.get("risk"),
                    x.get("theme_action"),
                    x.get("total_score"),
                    pct(sc.get("base_roi")),
                    ent.get("driver"),
                ],
                left=True,
            )
        r = write_row(ws, r, [""] * 10)

    set_widths(ws, [20, 7, 10, 26, 10, 10, 10, 9, 10, 40])


def add_sheet_theme_matrix_district(wb, rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("구테마매트릭스")
    theme_keys = [str(t.get("key") or "") for t in THEME_CATALOG]
    theme_names = [str(t.get("name") or k) for t, k in zip(THEME_CATALOG, theme_keys)]
    ncol = 6 + len(theme_names)
    r = add_common_header(
        ws,
        f"구 테마 매트릭스 ({profile})",
        "구별 대표테마와 모든 테마 Conviction 점수(필터/소팅용)",
        ncol,
    )
    headers = ["순위", "구", "총점", "대표테마", "대표점수", "테마액션"] + theme_names
    r = write_headers(ws, r, headers)
    for x in rows:
        smap = x.get("theme_score_map") or {}
        row_vals = [
            x.get("rank"),
            x.get("district"),
            x.get("total_score"),
            x.get("theme_primary"),
            x.get("theme_primary_score"),
            x.get("theme_action"),
        ]
        for key in theme_keys:
            ent = (smap.get(key, {}) or {})
            row_vals.append(ent.get("conviction"))
        r = write_row(ws, r, row_vals)
        ws.cell(row=r - 1, column=5).fill = score_band_fill(safe_float(x.get("theme_primary_score"), 0.0))

    widths = [7, 10, 9, 20, 10, 10] + [11] * len(theme_names)
    set_widths(ws, widths)


def add_sheet_theme_matrix_complex(wb, complex_rows: list[dict[str, Any]], profile: str, top_n: int = 800):
    ws = wb.create_sheet("단지테마매트릭스")
    theme_keys = [str(t.get("key") or "") for t in THEME_CATALOG]
    theme_names = [str(t.get("name") or k) for t, k in zip(THEME_CATALOG, theme_keys)]
    ncol = 8 + len(theme_names)
    r = add_common_header(
        ws,
        f"단지 테마 매트릭스 ({profile})",
        f"상위 {top_n}개 단지의 대표테마와 전체 테마 점수(실전 필터용)",
        ncol,
    )
    headers = ["전체순위", "구내순위", "구", "단지", "총점", "대표테마", "대표점수", "테마액션"] + theme_names
    r = write_headers(ws, r, headers)
    for x in complex_rows[:top_n]:
        smap = x.get("theme_score_map") or {}
        row_vals = [
            x.get("rank"),
            x.get("district_rank"),
            x.get("district"),
            x.get("complex"),
            x.get("total_score"),
            x.get("theme_primary"),
            x.get("theme_primary_score"),
            x.get("theme_action"),
        ]
        for key in theme_keys:
            ent = (smap.get(key, {}) or {})
            row_vals.append(ent.get("conviction"))
        r = write_row(ws, r, row_vals, left=True)
        ws.cell(row=r - 1, column=7).fill = score_band_fill(safe_float(x.get("theme_primary_score"), 0.0))

    widths = [8, 8, 10, 26, 9, 20, 10, 10] + [11] * len(theme_names)
    set_widths(ws, widths)


def add_sheet_quant_factor_method(wb, profile: str):
    ws = wb.create_sheet("퀀트팩터방법론")
    r = add_common_header(
        ws,
        f"퀀트 멀티팩터 방법론 ({profile})",
        "논문 기반 팩터 + 상관축소 + 데이터품질 계수 + 레거시모델 블렌딩",
        8,
    )
    r = write_headers(
        ws,
        r,
        [
            "팩터코드",
            "팩터명",
            "계산식(요약)",
            "방향",
            "프로필기본가중치",
            "가중조정",
            "근거논문",
            "비고",
        ],
    )
    base_w = QUANT_PROFILE_WEIGHTS.get(profile) or {}
    for f in QUANT_FACTOR_CATALOG:
        key = str(f.get("key") or "")
        r = write_row(
            ws,
            r,
            [
                key,
                f.get("name"),
                f.get("formula"),
                f.get("direction"),
                round(safe_float(base_w.get(key), 0.0), 4),
                "상관축소(corr)+품질계수(q_mult)",
                f.get("evidence"),
                "",
            ],
            left=True,
        )

    r += 1
    r = write_row(ws, r, ["[스코어 계산식]"] + [""] * 7, fill=FILL_GOLD, left=True)
    alpha = clamp(safe_float(QUANT_FACTOR_BLEND.get(profile), 0.60), 0.35, 0.80)
    r = write_row(ws, r, [f"1) Robust Z: winsor(10%) 후 median/MAD z-score (clip ±3)", "", "", "", "", "", "", ""], left=True)
    r = write_row(ws, r, [f"2) FactorScore = 50 + 12.5*z", "", "", "", "", "", "", ""], left=True)
    r = write_row(ws, r, [f"3) QuantFactor = Sum(weight_i * factor_i) * q_mult", "", "", "", "", "", "", ""], left=True)
    r = write_row(ws, r, [f"4) QuantFinal = {(1.0-alpha):.2f}*Legacy + {alpha:.2f}*QuantFactor", "", "", "", "", "", "", ""], left=True)

    set_widths(ws, [12, 16, 36, 8, 12, 24, 38, 20])


def add_sheet_quant_factor_district(wb, rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("구팩터스코어")
    keys = [str(x.get("key") or "") for x in QUANT_FACTOR_CATALOG]
    names = [str(x.get("name") or k) for x, k in zip(QUANT_FACTOR_CATALOG, keys)]
    ncol = 8 + len(keys)
    r = add_common_header(
        ws,
        f"구 팩터 스코어 ({profile})",
        "팩터별 점수와 최종 퀀트점수(블렌드) 확인",
        ncol,
    )
    headers = ["순위", "구", "총점", "퀀트최종", "퀀트레거시", "팩터순수", "품질계수", "팩터드라이버"] + names
    r = write_headers(ws, r, headers)
    for x in rows:
        row_vals = [
            x.get("rank"),
            x.get("district"),
            x.get("total_score"),
            x.get("quant_total_score"),
            x.get("quant_total_score_legacy"),
            x.get("quant_factor_score"),
            x.get("qf_quality_mult"),
            x.get("qf_driver"),
        ]
        for k in keys:
            row_vals.append(x.get(f"qf_{k}"))
        r = write_row(ws, r, row_vals, left=True)
        ws.cell(row=r - 1, column=4).fill = score_band_fill(safe_float(x.get("quant_total_score"), 0.0))
        ws.cell(row=r - 1, column=6).fill = score_band_fill(safe_float(x.get("quant_factor_score"), 0.0))

    set_widths(ws, [7, 10, 9, 10, 10, 10, 9, 32] + [10] * len(keys))


def add_sheet_quant_factor_complex(wb, complex_rows: list[dict[str, Any]], profile: str, top_n: int = 700):
    ws = wb.create_sheet("단지팩터스코어")
    keys = [str(x.get("key") or "") for x in QUANT_FACTOR_CATALOG]
    names = [str(x.get("name") or k) for x, k in zip(QUANT_FACTOR_CATALOG, keys)]
    ncol = 10 + len(keys)
    r = add_common_header(
        ws,
        f"단지 팩터 스코어 ({profile})",
        f"상위 {top_n}개 단지 팩터 분해(실전 필터링)",
        ncol,
    )
    headers = ["전체순위", "구내순위", "구", "단지", "총점", "퀀트최종", "퀀트레거시", "팩터순수", "품질계수", "팩터드라이버"] + names
    r = write_headers(ws, r, headers)
    for x in complex_rows[:top_n]:
        row_vals = [
            x.get("rank"),
            x.get("district_rank"),
            x.get("district"),
            x.get("complex"),
            x.get("total_score"),
            x.get("quant_total_score"),
            x.get("quant_total_score_legacy"),
            x.get("quant_factor_score"),
            x.get("qf_quality_mult"),
            x.get("qf_driver"),
        ]
        for k in keys:
            row_vals.append(x.get(f"qf_{k}"))
        r = write_row(ws, r, row_vals, left=True)
        ws.cell(row=r - 1, column=6).fill = score_band_fill(safe_float(x.get("quant_total_score"), 0.0))
        ws.cell(row=r - 1, column=8).fill = score_band_fill(safe_float(x.get("quant_factor_score"), 0.0))

    set_widths(ws, [8, 8, 10, 24, 9, 10, 10, 10, 9, 32] + [10] * len(keys))


def add_sheet_universe_meta(
    wb,
    complex_rows: list[dict[str, Any]],
    universe_label: str,
    original_complex_count: int | None,
):
    ws = wb.create_sheet("유니버스 설정")
    r = add_common_header(
        ws,
        "실전매수 유니버스 설정",
        "신뢰도 필터 적용 내역",
        8,
    )
    r = write_headers(ws, r, ["항목", "값", "비고", "", "", "", "", ""])
    total_now = len(complex_rows)
    total_org = original_complex_count if original_complex_count is not None else total_now
    excluded = max(0, total_org - total_now)
    cnt_high = sum(1 for x in complex_rows if x.get("confidence_bucket") == "high")
    cnt_med = sum(1 for x in complex_rows if x.get("confidence_bucket") == "medium")
    cnt_low = sum(1 for x in complex_rows if x.get("confidence_bucket") == "low")
    rows = [
        ["유니버스 타입", universe_label, "예: High+Medium", "", "", "", "", ""],
        ["원본 단지 수", total_org, "필터 전", "", "", "", "", ""],
        ["유니버스 단지 수", total_now, "필터 후", "", "", "", "", ""],
        ["제외 단지 수", excluded, "low 신뢰도 제외", "", "", "", "", ""],
        ["신뢰도 분포", f"high={cnt_high}, medium={cnt_med}, low={cnt_low}", "유니버스 기준", "", "", "", "", ""],
    ]
    for row in rows:
        r = write_row(ws, r, row, left=True)
    set_widths(ws, [18, 40, 22, 8, 8, 8, 8, 8])


def filter_complex_universe(
    complex_rows: list[dict[str, Any]],
    allowed_buckets: tuple[str, ...] = ("high", "medium"),
) -> list[dict[str, Any]]:
    allowed = set(allowed_buckets)
    filtered = [dict(x) for x in complex_rows if x.get("confidence_bucket") in allowed]
    filtered.sort(key=lambda r: r["total_score"], reverse=True)
    for i, r in enumerate(filtered, 1):
        r["rank"] = i
    group: dict[str, int] = {}
    for r in filtered:
        d = r["district"]
        group[d] = group.get(d, 0) + 1
        r["district_rank"] = group[d]
    return filtered


def add_sheet_profile_recommendation(wb, rows: list[dict[str, Any]], profile: str):
    ws_name = {
        "실거주": "실거주 추천",
        "투자자": "투자자 추천",
        "법인": "법인 추천",
    }[profile]
    ws = wb.create_sheet(ws_name)
    r = add_common_header(ws, f"{profile} 관점 Top 10", PROFILE_CONFIGS[profile]["display"], 16)
    r = write_headers(
        ws,
        r,
        [
            "순위",
            "구",
            "총점",
            "대표테마",
            "테마점수",
            "테마액션",
            "NQ합성",
            "ADV리스크",
            "중립ROI",
            "비관ROI",
            "전세가율",
            "변동성",
            "정책",
            "추천전략",
            "핵심리스크",
            "한줄코멘트",
        ],
    )

    def rec_score(x: dict[str, Any]) -> float:
        sc = x["scenario"]
        base = float(sc.get("base_roi") or -1.0)
        bear = float(sc.get("bear_roi") or -1.0)
        fresh = clamp(1.0 - safe_div(float(x.get("trade_stale_months") or 0), 6.0), 0.0, 1.0)
        utility = float(sc.get("annual_utility_yield") or 0.0) * 100.0
        vol_pen = max(0.0, (x.get("vol_ann") or 0.0) - 0.22) * 100.0
        nq = float(x.get("nq_composite") or 50.0)
        nq_pen = float(x.get("nq_flag_count") or 0.0) * 0.8
        adv_risk = clamp(float(x.get("adv_risk_composite") or 50.0) / 100.0, 0.0, 1.0)
        adv_pen = float(x.get("adv_risk_penalty") or 0.0)
        adv_crowd = clamp(float(x.get("adv_crowding") or 50.0) / 100.0, 0.0, 1.0)
        adv_policy = clamp(float(x.get("adv_policy_flip") or 50.0) / 100.0, 0.0, 1.0)
        if profile == "실거주":
            return (
                0.22 * x["total_score"]
                + 0.34 * x["stability_score"]
                + 0.18 * x["value_score"]
                + 8.0 * fresh
                + 0.20 * utility
                + 0.14 * nq
                - 0.90 * vol_pen
                - nq_pen
                - 12.0 * adv_risk
                - 1.2 * adv_pen
            )
        if profile == "투자자":
            return (
                0.25 * x["total_score"]
                + 0.30 * x["trend_score"]
                + 0.20 * x["liquidity_score"]
                + 10.0 * base
                + 6.0 * fresh
                + 0.15 * nq
                - nq_pen
                - 10.0 * adv_risk
                - 1.5 * adv_pen
                - 4.0 * max(0.0, adv_crowd - 0.70)
            )
        return (
            0.20 * x["total_score"]
            + 0.30 * x["policy_score"]
            + 0.18 * x["liquidity_score"]
            + 9.0 * base
            + 5.0 * fresh
            + 0.18 * nq
            - 4.0 * max(0.0, -0.75 - bear)
            - nq_pen
            - 11.0 * adv_risk
            - 1.4 * adv_pen
            - 5.0 * max(0.0, adv_policy - 0.65)
        )

    def hard_filter(x: dict[str, Any]) -> bool:
        stale = float(x.get("trade_stale_months") or 99)
        dq = float(x.get("data_quality") or 0)
        vol = float(x.get("vol_ann") or 0)
        latest = float(x.get("latest_trade_count") or 0)
        nq = float(x.get("nq_composite") or 0)
        adv = float(x.get("adv_risk_composite") or 50)
        adv_policy = float(x.get("adv_policy_flip") or 50)
        if profile == "실거주":
            return dq >= 60 and stale <= 2 and vol <= 0.40 and nq >= 50 and adv <= 68
        if profile == "투자자":
            return dq >= 58 and stale <= 2 and latest > 0 and nq >= 48 and adv <= 72
        return dq >= 58 and stale <= 3 and x.get("policy_score", 0) >= 38 and nq >= 50 and adv <= 70 and adv_policy <= 75

    ranked = sorted(rows, key=rec_score, reverse=True)
    filtered = [x for x in ranked if hard_filter(x)]
    if len(filtered) < 10:
        seen = {x["district"] for x in filtered}
        for x in ranked:
            if x["district"] in seen:
                continue
            filtered.append(x)
            seen.add(x["district"])
            if len(filtered) >= 10:
                break
    top = filtered[:10]

    for i, x in enumerate(top, 1):
        sc = x["scenario"]
        theme_key = x.get("theme_primary_key")
        theme_action = x.get("theme_action") or _theme_action_tag(
            safe_float(x.get("theme_primary_score"), 0.0),
            safe_float(x.get("theme_primary_risk"), 50.0),
            safe_float(x.get("theme_spread"), 0.0),
        )
        strategy, risk, comment = _theme_strategy_for_profile(theme_key, profile, str(theme_action))

        r = write_row(
            ws,
            r,
            [
                i,
                x["district"],
                x["total_score"],
                x.get("theme_primary"),
                x.get("theme_primary_score"),
                theme_action,
                x.get("nq_composite"),
                x.get("adv_risk_composite"),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
                f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                x["policy_score"],
                strategy,
                risk,
                comment,
            ],
            left=True,
        )
        ws.cell(row=r - 1, column=5).fill = score_band_fill(safe_float(x.get("theme_primary_score"), 0.0))
        ws.cell(row=r - 1, column=8).fill = score_band_fill(100.0 - safe_float(x.get("adv_risk_composite"), 50.0))

    set_widths(ws, [7, 10, 9, 18, 10, 10, 8, 9, 10, 10, 10, 9, 8, 34, 22, 32])


def add_sheet_references(wb):
    ws = wb.create_sheet("참고문헌")
    r = add_common_header(ws, "근거자료(공식/논문/커뮤니티)", "본 분석의 현실반영 근거", 6)
    r = write_headers(ws, r, ["분류", "자료명", "기준시점", "활용목적", "URL", "비고"])
    for s in SOURCES:
        r = write_row(
            ws,
            r,
            [
                s["category"],
                s["title"],
                s["as_of"],
                s["usage"],
                s["url"],
                "",
            ],
            left=True,
        )
    set_widths(ws, [10, 40, 12, 42, 65, 10])


def add_sheet_checklist(wb, profile: str):
    ws = wb.create_sheet("실행체크리스트")
    r = add_common_header(ws, f"실행 체크리스트 ({profile})", "실제 투자 실행 전 필수 검증", 6)
    r = write_headers(ws, r, ["No", "체크항목", "왜 중요한가", "데이터/문서", "판단기준", "완료"])

    checklist = [
        [1, "자금계획", "LTV/DSR 충족 여부", "대출사전심사", "DSR 허용범위", ""],
        [2, "세금시뮬레이션", "세후수익 급변 방지", "세무사 검토", "보수세율 적용", ""],
        [3, "정책확인", "허가구역/규제 변경", "서울시/금융위 공지", "최근 공고 기준", ""],
        [4, "표본신뢰도", "왜곡 방지", "DataQ/표본수", "DataQ 60 이상 권장", ""],
        [5, "비정량점검", "정량 외 리스크 누락 방지", "NQ합성/NQ플래그", "NQ 50 이상·중대플래그 없음", ""],
        [6, "시나리오점검", "하방손실 관리", "Bear ROI", "손실허용범위 내", ""],
        [7, "유동성", "매도/갈아타기 가능성", "최근 거래량", "상위 50% 권장", ""],
        [8, "현장검증", "입지/소음/학군 체감", "현장방문", "핵심 하자 무", ""],
        [9, "출구전략", "매도 타이밍", "목표수익률/기간", "사전 규칙화", ""],
        [10, "공급/입주물량", "국지 과공급 구간 회피", "입주예정 물량표", "2년 순증공급 급증 구간 주의", ""],
        [11, "전세보증리스크", "임차인 보증사고/회수 리스크", "HUG/HF 통계", "보증사고 급증 지역 보수화", ""],
        [12, "관리품질/DD", "하자·관리비·민원으로 실수익 훼손", "K-apt/관리규약/회의록", "중대 하자·분쟁 이력 없음", ""],
    ]
    for row in checklist:
        r = write_row(ws, r, row, left=True)

    set_widths(ws, [6, 14, 24, 22, 20, 8])


def add_sheet_ranking_breakdown(
    wb,
    profile_rows: list[dict[str, Any]],
    complex_rows: list[dict[str, Any]],
    profile: str,
):
    ws = wb.create_sheet("랭킹산정상세")
    w = PROFILE_CONFIGS[profile]["weights"]
    r = add_common_header(ws, f"랭킹 산정 상세 ({profile})", "총점 기여도(레거시정량+팩터+NQ+ADV) 분해표", 18)

    r = write_row(ws, r, ["[구 랭킹 기여도]"] + [""] * 17, fill=FILL_GOLD, left=True)
    r = write_headers(
        ws,
        r,
        [
            "순위",
            "구",
            "총점",
            "Trend기여",
            "Value기여",
            "Liquidity기여",
            "Stability기여",
            "Policy기여",
            "퀀트최종",
            "퀀트레거시",
            "팩터순수",
            "품질계수",
            "NQ오버레이",
            "ADV패널티",
            "NQ합성",
            "ADV리스크",
            "팩터드라이버",
            "비고",
        ],
    )
    for x in profile_rows:
        t = safe_float(x.get("trend_score")) * w["trend"] / 100.0
        v = safe_float(x.get("value_score")) * w["value"] / 100.0
        l = safe_float(x.get("liquidity_score")) * w["liquidity"] / 100.0
        s = safe_float(x.get("stability_score")) * w["stability"] / 100.0
        p = safe_float(x.get("policy_score")) * w["policy"] / 100.0
        r = write_row(
            ws,
            r,
            [
                x.get("rank"),
                x.get("district"),
                x.get("total_score"),
                round(t, 2),
                round(v, 2),
                round(l, 2),
                round(s, 2),
                round(p, 2),
                x.get("quant_total_score"),
                x.get("quant_total_score_legacy"),
                x.get("quant_factor_score"),
                x.get("qf_quality_mult"),
                x.get("nq_overlay"),
                x.get("adv_risk_penalty"),
                x.get("nq_composite"),
                x.get("adv_risk_composite"),
                x.get("qf_driver"),
                "",
            ],
            left=True,
        )
        ws.cell(row=r - 1, column=3).fill = score_band_fill(float(x.get("total_score") or 0.0))

    r += 1
    r = write_row(ws, r, ["[단지 랭킹 기여도 - 상위 400]"] + [""] * 17, fill=FILL_GOLD, left=True)
    r = write_headers(
        ws,
        r,
        [
            "전체순위",
            "구내순위",
            "구",
            "단지",
            "총점",
            "Core",
            "퀀트최종",
            "퀀트레거시",
            "팩터순수",
            "품질계수",
            "NQ오버레이",
            "ADV패널티",
            "신선도패널티",
            "꼬리패널티",
            "NQ플래그패널티",
            "ADV리스크",
            "팩터드라이버",
            "비고",
        ],
    )
    for x in complex_rows[:400]:
        r = write_row(
            ws,
            r,
            [
                x.get("rank"),
                x.get("district_rank"),
                x.get("district"),
                x.get("complex"),
                x.get("total_score"),
                x.get("core_score"),
                x.get("quant_total_score"),
                x.get("quant_total_score_legacy"),
                x.get("quant_factor_score"),
                x.get("qf_quality_mult"),
                x.get("nq_overlay"),
                x.get("adv_risk_penalty"),
                x.get("recency_penalty"),
                x.get("tail_penalty"),
                x.get("nq_flag_penalty"),
                x.get("adv_risk_composite"),
                x.get("qf_driver"),
                "",
            ],
            left=True,
        )
        ws.cell(row=r - 1, column=5).fill = score_band_fill(float(x.get("total_score") or 0.0))

    set_widths(ws, [8, 8, 11, 26, 9, 9, 9, 9, 9, 9, 9, 9, 10, 10, 11, 9, 30, 12])


def add_sheet_risk_watchlist(
    wb,
    profile_rows: list[dict[str, Any]],
    complex_rows: list[dict[str, Any]],
    profile: str,
):
    ws = wb.create_sheet("리스크워치리스트")
    r = add_common_header(ws, f"리스크 워치리스트 ({profile})", "총점은 높지만 하방 리스크가 큰 후보", 12)

    def district_reason(x: dict[str, Any]) -> str:
        reasons: list[str] = []
        if safe_float(x.get("adv_risk_composite"), 0.0) >= 65:
            reasons.append("ADV고위험")
        if safe_float(x.get("data_quality"), 100.0) < 60:
            reasons.append("DataQ낮음")
        if safe_float(x.get("nq_flag_count"), 0.0) >= 3:
            reasons.append("NQ플래그다수")
        sc = x.get("scenario", {})
        if safe_float(sc.get("bear_roi"), 0.0) < -0.55:
            reasons.append("비관ROI취약")
        return ", ".join(reasons)

    risky_d = [x for x in profile_rows if safe_float(x.get("total_score"), 0.0) >= 60 and district_reason(x)]
    risky_d.sort(key=lambda x: (safe_float(x.get("adv_risk_composite"), 0.0), -safe_float(x.get("total_score"), 0.0)), reverse=True)

    r = write_row(ws, r, ["[구 워치리스트]"] + [""] * 11, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["순위", "구", "총점", "ADV리스크", "ADV패널티", "NQ합성", "DataQ", "중립ROI", "비관ROI", "전세가율", "거래공백월", "리스크사유"])
    for x in risky_d[:25]:
        sc = x.get("scenario", {})
        r = write_row(
            ws,
            r,
            [
                x.get("rank"),
                x.get("district"),
                x.get("total_score"),
                x.get("adv_risk_composite"),
                x.get("adv_risk_penalty"),
                x.get("nq_composite"),
                x.get("data_quality"),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
                f"{safe_float(x.get('latest_jeonse_ratio'), 0.0):.1f}%",
                x.get("trade_stale_months"),
                district_reason(x),
            ],
            left=True,
        )

    def complex_reason(x: dict[str, Any]) -> str:
        reasons: list[str] = []
        if safe_float(x.get("adv_risk_composite"), 0.0) >= 68:
            reasons.append("ADV고위험")
        if safe_float(x.get("trade_stale_months"), 0.0) >= 3:
            reasons.append("거래공백")
        if safe_float(x.get("tail_es_abs"), 0.0) >= 0.06:
            reasons.append("꼬리위험")
        if safe_float(x.get("data_quality"), 100.0) < 58:
            reasons.append("DataQ낮음")
        return ", ".join(reasons)

    risky_c = [x for x in complex_rows if safe_float(x.get("total_score"), 0.0) >= 68 and complex_reason(x)]
    risky_c.sort(key=lambda x: (safe_float(x.get("adv_risk_composite"), 0.0), -safe_float(x.get("total_score"), 0.0)), reverse=True)

    r += 1
    r = write_row(ws, r, ["[단지 워치리스트]"] + [""] * 11, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["전체순위", "구내순위", "구", "단지", "총점", "ADV리스크", "ADV패널티", "DataQ", "중립ROI", "비관ROI", "거래공백월", "리스크사유"])
    for x in risky_c[:40]:
        sc = x.get("scenario", {})
        r = write_row(
            ws,
            r,
            [
                x.get("rank"),
                x.get("district_rank"),
                x.get("district"),
                x.get("complex"),
                x.get("total_score"),
                x.get("adv_risk_composite"),
                x.get("adv_risk_penalty"),
                x.get("data_quality"),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
                x.get("trade_stale_months"),
                complex_reason(x),
            ],
            left=True,
        )

    set_widths(ws, [8, 8, 11, 26, 9, 9, 9, 8, 10, 10, 10, 22])


def add_sheet_cross_profile_consensus(
    wb,
    rows_by_profile: dict[str, list[dict[str, Any]]],
    complex_rows_by_profile: dict[str, list[dict[str, Any]]],
):
    ws = wb.create_sheet("프로필교집합")
    r = add_common_header(ws, "프로필 교집합 분석", "실거주/투자자/법인 공통 상위 후보", 10)

    profiles = ["실거주", "투자자", "법인"]
    district_rank_map: dict[str, dict[str, int]] = {}
    top_sets: dict[str, set[str]] = {}
    for p in profiles:
        rows = rows_by_profile.get(p) or []
        district_rank_map[p] = {x["district"]: int(x.get("rank") or 999) for x in rows}
        top_sets[p] = {x["district"] for x in rows[:10]}

    common = sorted(top_sets["실거주"] & top_sets["투자자"] & top_sets["법인"])
    union = sorted(set().union(*top_sets.values()))

    r = write_row(ws, r, ["[구 공통상위]"] + [""] * 9, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["구", "실거주순위", "투자자순위", "법인순위", "평균순위", "최고순위", "최저순위", "편차", "공통Top10", "비고"])
    for d in union:
        rk = [district_rank_map[p].get(d, 999) for p in profiles]
        avg_rk = statistics.mean(rk)
        st = statistics.pstdev(rk) if len(rk) > 1 else 0.0
        r = write_row(
            ws,
            r,
            [
                d,
                rk[0],
                rk[1],
                rk[2],
                round(avg_rk, 2),
                min(rk),
                max(rk),
                round(st, 2),
                "Y" if d in common else "N",
                "",
            ],
            left=True,
        )

    # 단지 교집합은 표본을 넓혀 Top80 기준으로 계산
    complex_top: dict[str, set[str]] = {}
    complex_rank_map: dict[str, dict[str, int]] = {}
    for p in profiles:
        rows = complex_rows_by_profile.get(p) or []
        complex_top[p] = {f"{x['district']}::{x['complex']}" for x in rows[:80]}
        complex_rank_map[p] = {f"{x['district']}::{x['complex']}": int(x.get("rank") or 9999) for x in rows}
    common_c = sorted(complex_top["실거주"] & complex_top["투자자"] & complex_top["법인"])

    r += 1
    r = write_row(ws, r, ["[단지 공통상위 Top80 교집합]"] + [""] * 9, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["구", "단지", "실거주순위", "투자자순위", "법인순위", "평균순위", "편차", "중립(투자자)ROI", "ADV리스크(투자자)", "비고"])
    inv_map = {f"{x['district']}::{x['complex']}": x for x in (complex_rows_by_profile.get("투자자") or [])}
    for key in common_c[:120]:
        d, c = key.split("::", 1)
        rk = [complex_rank_map[p].get(key, 9999) for p in profiles]
        avg_rk = statistics.mean(rk)
        st = statistics.pstdev(rk) if len(rk) > 1 else 0.0
        ix = inv_map.get(key, {})
        sc = ix.get("scenario", {})
        r = write_row(
            ws,
            r,
            [
                d,
                c,
                rk[0],
                rk[1],
                rk[2],
                round(avg_rk, 2),
                round(st, 2),
                pct(sc.get("base_roi")),
                ix.get("adv_risk_composite"),
                "",
            ],
            left=True,
        )

    set_widths(ws, [11, 26, 9, 9, 9, 9, 8, 12, 10, 12])


def build_workbook(
    profile_rows: list[dict[str, Any]],
    complex_rows: list[dict[str, Any]],
    profile: str,
    months: list[str],
    rows_by_profile: dict[str, list[dict[str, Any]]] | None = None,
    complex_rows_by_profile: dict[str, list[dict[str, Any]]] | None = None,
    top_per_district: int = DEFAULT_TOP_COMPLEX_PER_DISTRICT,
    universe_label: str = "전체(ALL)",
    original_complex_count: int | None = None,
    backtest_lookback: int = DEFAULT_BACKTEST_LOOKBACK,
    backtest_horizon: int = DEFAULT_BACKTEST_HORIZON,
    backtest_top_fraction: float = DEFAULT_BACKTEST_TOP_FRACTION,
    nq_meta: dict[str, Any] | None = None,
    adv_meta: dict[str, Any] | None = None,
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    add_sheet_dashboard(wb, profile_rows, profile, months)
    add_sheet_scorecard(wb, profile_rows, profile, months)
    add_sheet_quant_factor_method(wb, profile)
    add_sheet_quant_factor_district(wb, profile_rows, profile)
    add_sheet_quant_factor_complex(wb, complex_rows, profile, top_n=700)
    add_sheet_coverage(wb, profile_rows, months)
    add_sheet_trade_trend(wb, profile_rows, months)
    add_sheet_rent_trend(wb, profile_rows, months)
    add_sheet_value(wb, profile_rows)
    add_sheet_liquidity(wb, profile_rows)
    add_sheet_risk(wb, profile_rows)
    add_sheet_benchmark(wb, profile_rows)
    add_sheet_backtest(
        wb,
        profile_rows,
        months,
        lookback=backtest_lookback,
        horizon=backtest_horizon,
        top_fraction=backtest_top_fraction,
    )
    add_sheet_scenario(wb, profile_rows, profile)
    add_sheet_scenario_method(wb, profile)
    add_sheet_profile_weights(wb)
    add_sheet_tax_leverage(wb)
    add_sheet_policy(wb, profile_rows)
    add_sheet_timing(wb, profile_rows, profile)
    add_sheet_complex_overview(wb, complex_rows, profile_rows, top_per_district)
    add_sheet_complex_rank(wb, complex_rows, profile)
    add_sheet_complex_district_top(wb, complex_rows, top_per_district)
    add_sheet_complex_scenario(wb, complex_rows, profile)
    add_sheet_complex_confidence(wb, complex_rows)
    add_sheet_complex_sensitivity(wb, complex_rows, profile)
    add_sheet_nonquant_overlay(wb, profile_rows, complex_rows, profile)
    add_sheet_nonquant_method(wb, profile, nq_meta)
    add_sheet_advanced_risk_overlay(wb, profile_rows, complex_rows, profile)
    add_sheet_advanced_risk_method(wb, profile, adv_meta)
    add_sheet_advanced_signal_audit(wb, profile_rows, complex_rows, profile, adv_meta)
    add_sheet_advanced_signal_history(wb, profile, adv_meta)
    add_sheet_theme_method(wb, profile)
    add_sheet_theme_classification_district(wb, profile_rows, profile)
    add_sheet_theme_classification_complex(wb, complex_rows, profile)
    add_sheet_theme_ranking(wb, profile_rows, complex_rows, profile)
    add_sheet_theme_matrix_district(wb, profile_rows, profile)
    add_sheet_theme_matrix_complex(wb, complex_rows, profile, top_n=800)
    add_sheet_universe_meta(wb, complex_rows, universe_label, original_complex_count)
    add_sheet_ranking_breakdown(wb, profile_rows, complex_rows, profile)
    add_sheet_risk_watchlist(wb, profile_rows, complex_rows, profile)

    # 프로필별 추천 시트 3개를 모두 포함(비교 가능)
    rows_by_profile = rows_by_profile or {profile: profile_rows}
    complex_rows_by_profile = complex_rows_by_profile or {profile: complex_rows}
    add_sheet_cross_profile_consensus(wb, rows_by_profile, complex_rows_by_profile)
    for p in ["실거주", "투자자", "법인"]:
        src = rows_by_profile.get(p) or profile_rows
        add_sheet_profile_recommendation(wb, src, p)

    add_sheet_references(wb)
    add_sheet_checklist(wb, profile)

    return wb


# ---------------------------------------------------------------------------
# Cache and main flow
# ---------------------------------------------------------------------------


def save_cache(path: Path, payload: dict[str, Any]):
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def save_workbook_safe(wb: openpyxl.Workbook, path: Path) -> Path:
    try:
        wb.save(path)
        return path
    except PermissionError:
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        alt = path.with_name(f"{path.stem}_locked_{ts}{path.suffix}")
        wb.save(alt)
        return alt


def load_cache(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def serialize_panel(panel: dict[str, dict[str, dict[str, Any]]]) -> dict[str, Any]:
    return panel


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="서울 전체 아파트 심층분석 엑셀 생성기")
    p.add_argument("--months", type=int, default=12, help="수집 월 수 (기본 12)")
    p.add_argument("--end-yyyymm", default="", help="종료월 YYYYMM (미지정 시 직전 완료월)")
    p.add_argument("--include-current-month", action="store_true", help="당월(부분월 가능)까지 포함")
    p.add_argument("--max-workers", type=int, default=16, help="병렬 작업 수")
    p.add_argument("--num-rows", type=int, default=1000, help="API 페이지 행수")
    p.add_argument("--area-min", type=float, default=DEFAULT_AREA_MIN)
    p.add_argument("--area-max", type=float, default=DEFAULT_AREA_MAX)
    p.add_argument("--deposit-yield-pct", type=float, default=DEFAULT_DEPOSIT_YIELD_PCT)
    p.add_argument("--top-complex-per-district", type=int, default=DEFAULT_TOP_COMPLEX_PER_DISTRICT, help="구별 상위 단지 표시 개수")
    p.add_argument("--complex-min-total-trade", type=int, default=DEFAULT_COMPLEX_MIN_TOTAL_TRADE, help="단지 분석 최소 누적 매매표본")
    p.add_argument("--complex-min-month-coverage", type=int, default=DEFAULT_COMPLEX_MIN_MONTH_COVERAGE, help="단지 분석 최소 거래월 수")
    p.add_argument("--complex-shrink-trade-ref", type=int, default=DEFAULT_COMPLEX_SHRINK_TRADE_REF, help="단지 성장률 shrinkage 기준 표본수")
    p.add_argument("--no-universe", action="store_true", help="High+Medium 실전유니버스 파일 생성 비활성화")
    p.add_argument("--api-key", default="", help="data.go.kr API key (없으면 env/내장키 사용)")
    p.add_argument("--refresh-cache", action="store_true", help="캐시 무시하고 재수집")
    p.add_argument("--profile", choices=list(PROFILE_CONFIGS.keys()), default="투자자")
    p.add_argument("--all-profiles", action="store_true", help="3개 프로필 모두 생성")
    p.add_argument("--backtest-lookback", type=int, default=DEFAULT_BACKTEST_LOOKBACK, help="백테스트 시그널 lookback 월수")
    p.add_argument("--backtest-horizon", type=int, default=DEFAULT_BACKTEST_HORIZON, help="백테스트 forward 월수")
    p.add_argument("--backtest-top-frac", type=float, default=DEFAULT_BACKTEST_TOP_FRACTION, help="백테스트 Top 선택 비중(0~1)")
    p.add_argument("--nonquant-off", action="store_true", help="비정량 오버레이 비활성화")
    p.add_argument("--nonquant-prior-file", default=DEFAULT_NONQUANT_PRIOR_FILE, help="비정량 prior/override JSON 경로")
    p.add_argument("--nonquant-prior-weight", type=float, default=DEFAULT_NQ_PRIOR_WEIGHT, help="prior 가중치(0~1)")
    p.add_argument("--nonquant-overlay-cap", type=float, default=DEFAULT_NQ_OVERLAY_CAP, help="구 랭킹 비정량 오버레이 최대 절대값")
    p.add_argument("--nonquant-complex-overlay-cap", type=float, default=DEFAULT_NQ_COMPLEX_OVERLAY_CAP, help="단지 랭킹 비정량 오버레이 최대 절대값")
    p.add_argument("--advanced-risk-off", action="store_true", help="고급리스크 오버레이 비활성화")
    p.add_argument("--advanced-risk-cap", type=float, default=DEFAULT_ADV_RISK_DISTRICT_CAP, help="구 랭킹 고급리스크 패널티 최대값")
    p.add_argument("--advanced-risk-complex-cap", type=float, default=DEFAULT_ADV_RISK_COMPLEX_CAP, help="단지 랭킹 고급리스크 패널티 최대값")
    p.add_argument("--advanced-signal-file", default=DEFAULT_ADV_SIGNAL_FILE, help="고급리스크 외생 시그널 JSON 경로")
    p.add_argument("--advanced-signal-autofill-off", action="store_true", help="외생 시그널 자동채움 비활성화")
    p.add_argument("--advanced-signal-overwrite", action="store_true", help="외생 시그널 자동채움 시 기존 값을 덮어쓰기")
    p.add_argument("--output-dir", default="C:/Devs/KRE")
    p.add_argument(
        "--region-code-file",
        default="C:/Devs/KRE/real-estate-mcp/src/real_estate/resources/region_codes.txt",
        help="region_codes.txt 경로",
    )
    return p.parse_args()


def main() -> int:
    args = parse_args()

    if args.months < 6:
        raise SystemExit("months는 최소 6 이상을 권장합니다.")
    if args.area_min >= args.area_max:
        raise SystemExit("area-min은 area-max보다 작아야 합니다.")
    if args.top_complex_per_district < 1:
        raise SystemExit("top-complex-per-district는 1 이상이어야 합니다.")
    if args.complex_min_total_trade < 1:
        raise SystemExit("complex-min-total-trade는 1 이상이어야 합니다.")
    if args.complex_min_month_coverage < 1:
        raise SystemExit("complex-min-month-coverage는 1 이상이어야 합니다.")
    if args.complex_shrink_trade_ref < 10:
        raise SystemExit("complex-shrink-trade-ref는 10 이상을 권장합니다.")
    if args.backtest_lookback < 3:
        raise SystemExit("backtest-lookback는 3 이상이어야 합니다.")
    if args.backtest_horizon < 1:
        raise SystemExit("backtest-horizon은 1 이상이어야 합니다.")
    if not (0.05 <= args.backtest_top_frac <= 0.50):
        raise SystemExit("backtest-top-frac는 0.05~0.50 범위여야 합니다.")
    if not (0.0 <= args.nonquant_prior_weight <= 1.0):
        raise SystemExit("nonquant-prior-weight는 0~1 범위여야 합니다.")
    if not (0.0 <= args.nonquant_overlay_cap <= 30.0):
        raise SystemExit("nonquant-overlay-cap는 0~30 범위여야 합니다.")
    if not (0.0 <= args.nonquant_complex_overlay_cap <= 30.0):
        raise SystemExit("nonquant-complex-overlay-cap는 0~30 범위여야 합니다.")
    if not (0.0 <= args.advanced_risk_cap <= 20.0):
        raise SystemExit("advanced-risk-cap는 0~20 범위여야 합니다.")
    if not (0.0 <= args.advanced_risk_complex_cap <= 20.0):
        raise SystemExit("advanced-risk-complex-cap는 0~20 범위여야 합니다.")

    api_key = get_api_key(args.api_key)
    if not api_key:
        raise SystemExit("API key를 찾지 못했습니다.")

    end_yyyymm = (args.end_yyyymm or "").strip()
    if not end_yyyymm:
        end_yyyymm = now_yyyymm() if args.include_current_month else latest_complete_yyyymm()
    if len(end_yyyymm) != 6 or not end_yyyymm.isdigit():
        raise SystemExit("end-yyyymm 형식은 YYYYMM 이어야 합니다.")
    if (not args.include_current_month) and end_yyyymm == now_yyyymm():
        print("[warn] 당월은 부분월일 수 있습니다. 직전 완료월 사용을 권장합니다.")

    months = iter_months(end_yyyymm, args.months)
    if args.months < args.backtest_lookback + args.backtest_horizon:
        print("[warn] 백테스트 표본이 부족할 수 있습니다. months 증가를 권장합니다.")
    region_code_file = Path(args.region_code_file)
    districts = load_seoul_gu_list(region_code_file)

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    nq_enabled = not args.nonquant_off
    nq_prior_path = Path(args.nonquant_prior_file) if args.nonquant_prior_file else None
    if nq_enabled and nq_prior_path:
        ensure_nonquant_prior_file(nq_prior_path)
    nq_payload = load_nonquant_prior_payload(nq_prior_path if nq_enabled else None)
    nq_meta = dict(nq_payload.get("meta") or {})
    nq_meta["enabled"] = nq_enabled
    nq_meta["prior_weight"] = args.nonquant_prior_weight
    nq_meta["overlay_cap"] = args.nonquant_overlay_cap
    nq_meta["complex_overlay_cap"] = args.nonquant_complex_overlay_cap
    if not nq_enabled:
        nq_meta["source"] = "disabled"

    adv_enabled = not args.advanced_risk_off
    adv_autofill_enabled = adv_enabled and (not args.advanced_signal_autofill_off)
    adv_autofill_run_id = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    adv_signal_path = Path(args.advanced_signal_file) if args.advanced_signal_file else None
    if adv_enabled and adv_signal_path:
        ensure_advanced_signal_file(adv_signal_path)
    adv_signal_payload = load_advanced_signal_payload(adv_signal_path if adv_enabled else None)
    adv_signal_meta = dict(adv_signal_payload.get("meta") or {})
    adv_last_changes = adv_signal_payload.get("autofill_last_changes") if isinstance(adv_signal_payload.get("autofill_last_changes"), dict) else {}
    adv_history = adv_signal_payload.get("autofill_run_history") if isinstance(adv_signal_payload.get("autofill_run_history"), list) else []
    adv_meta = {
        "enabled": adv_enabled,
        "district_cap": args.advanced_risk_cap,
        "complex_cap": args.advanced_risk_complex_cap,
        "as_of": adv_signal_meta.get("as_of", ""),
        "version": adv_signal_meta.get("version", "v7"),
        "method": adv_signal_meta.get("method", ""),
        "source": adv_signal_meta.get("source", "default"),
        "signal_file": str(adv_signal_path) if adv_signal_path else "",
        "autofill_enabled": adv_autofill_enabled,
        "autofill_overwrite": bool(args.advanced_signal_overwrite),
        "autofill_month_range": f"{months[0]}~{months[-1]}",
        "autofill_run_id": str(adv_signal_meta.get("autofill_run_id", "")),
        "autofill_run_at": str(adv_signal_meta.get("autofill_run_at", "")),
        "autofill_total_rows_district": int(safe_float(adv_signal_meta.get("autofill_total_rows_district"), 0.0)),
        "autofill_signal_rows_district": int(safe_float(adv_signal_meta.get("autofill_signal_rows_district"), 0.0)),
        "autofill_assigned_rows_district": int(safe_float(adv_signal_meta.get("autofill_assigned_rows_district"), 0.0)),
        "autofill_updated_rows_district": int(safe_float(adv_signal_meta.get("autofill_updated_rows_district"), 0.0)),
        "autofill_changed_rows_district": int(safe_float(adv_signal_meta.get("autofill_changed_rows_district"), 0.0)),
        "autofill_changed_dims_district": int(safe_float(adv_signal_meta.get("autofill_changed_dims_district"), 0.0)),
        "autofill_total_rows_complex": int(safe_float(adv_signal_meta.get("autofill_total_rows_complex"), 0.0)),
        "autofill_signal_rows_complex": int(safe_float(adv_signal_meta.get("autofill_signal_rows_complex"), 0.0)),
        "autofill_assigned_rows_complex": int(safe_float(adv_signal_meta.get("autofill_assigned_rows_complex"), 0.0)),
        "autofill_updated_rows_complex": int(safe_float(adv_signal_meta.get("autofill_updated_rows_complex"), 0.0)),
        "autofill_changed_rows_complex": int(safe_float(adv_signal_meta.get("autofill_changed_rows_complex"), 0.0)),
        "autofill_changed_dims_complex": int(safe_float(adv_signal_meta.get("autofill_changed_dims_complex"), 0.0)),
        "change_map_district": adv_last_changes.get("district") if isinstance(adv_last_changes.get("district"), dict) else {},
        "change_map_complex": adv_last_changes.get("complex") if isinstance(adv_last_changes.get("complex"), dict) else {},
        "autofill_history_size": len(adv_history),
        "autofill_run_history": adv_history,
    }
    if not adv_enabled:
        adv_meta["source"] = "disabled"

    print(
        f"[period] {months[0]}~{months[-1]} / include_current_month={args.include_current_month} "
        f"/ nonquant={'on' if nq_enabled else 'off'} / adv={'on' if adv_enabled else 'off'} "
        f"/ adv_signal={adv_meta.get('source', 'default')} / adv_autofill={'on' if adv_autofill_enabled else 'off'}"
    )
    cache_path = out_dir / f"seoul_market_cache_v7_{months[0]}_{months[-1]}_{int(args.area_min)}_{int(args.area_max)}.json"

    if cache_path.exists() and not args.refresh_cache:
        print(f"[cache] 로드: {cache_path}")
        cache = load_cache(cache_path)
        panel = cache["panel"]
        logs = cache.get("logs", [])
    else:
        print(f"[collect] 서울 25개구 x {len(months)}개월 수집 시작")
        panel, logs = collect_market_panel(
            districts=districts,
            months=months,
            api_key=api_key,
            num_rows=args.num_rows,
            max_workers=args.max_workers,
            area_min=args.area_min,
            area_max=args.area_max,
            deposit_yield_pct=args.deposit_yield_pct,
        )
        cache = {
            "generated_at": dt.datetime.now().isoformat(),
            "months": months,
            "area_filter": [args.area_min, args.area_max],
            "panel": serialize_panel(panel),
            "logs": logs,
        }
        save_cache(cache_path, cache)
        print(f"[cache] 저장: {cache_path}")

    print("[analyze] 지표 계산")
    metrics = build_district_metrics(districts, months, panel)
    metrics = run_multi_expert_scoring(metrics)
    metrics = apply_nonquant_to_district_metrics(
        metrics,
        nq_payload,
        prior_weight=args.nonquant_prior_weight,
        enabled=nq_enabled,
    )
    if adv_autofill_enabled:
        adv_signal_payload = autofill_advanced_signal_districts(
            metrics,
            adv_signal_payload,
            months=months,
            overwrite=bool(args.advanced_signal_overwrite),
            run_id=adv_autofill_run_id,
        )

    metrics = apply_advanced_risk_to_metrics(
        metrics,
        enabled=adv_enabled,
        signal_payload=adv_signal_payload,
        level="district",
    )
    complex_metrics = build_complex_metrics(
        months=months,
        panel=panel,
        district_metrics=metrics,
        min_total_trade=args.complex_min_total_trade,
        min_month_coverage=min(args.complex_min_month_coverage, len(months)),
        shrink_trade_ref=args.complex_shrink_trade_ref,
    )
    complex_metrics = run_complex_expert_scoring(complex_metrics)
    complex_metrics = apply_nonquant_to_complex_metrics(
        complex_metrics,
        metrics,
        nq_payload,
        prior_weight=args.nonquant_prior_weight,
        enabled=nq_enabled,
    )
    if adv_autofill_enabled:
        adv_signal_payload = autofill_advanced_signal_complexes(
            complex_metrics,
            adv_signal_payload,
            months=months,
            overwrite=bool(args.advanced_signal_overwrite),
            run_id=adv_autofill_run_id,
        )
        if adv_signal_path:
            save_advanced_signal_payload(adv_signal_path, adv_signal_payload)
        adv_signal_meta = dict(adv_signal_payload.get("meta") or {})
        adv_meta["as_of"] = adv_signal_meta.get("as_of", adv_meta.get("as_of", ""))
        adv_meta["version"] = adv_signal_meta.get("version", adv_meta.get("version", "v7"))
        adv_meta["method"] = adv_signal_meta.get("method", adv_meta.get("method", ""))
        adv_meta["autofill_run_id"] = str(adv_signal_meta.get("autofill_run_id", ""))
        adv_meta["autofill_run_at"] = str(adv_signal_meta.get("autofill_run_at", ""))
        adv_meta["autofill_total_rows_district"] = int(safe_float(adv_signal_meta.get("autofill_total_rows_district"), 0.0))
        adv_meta["autofill_signal_rows_district"] = int(safe_float(adv_signal_meta.get("autofill_signal_rows_district"), 0.0))
        adv_meta["autofill_assigned_rows_district"] = int(safe_float(adv_signal_meta.get("autofill_assigned_rows_district"), 0.0))
        adv_meta["autofill_updated_rows_district"] = int(safe_float(adv_signal_meta.get("autofill_updated_rows_district"), 0.0))
        adv_meta["autofill_changed_rows_district"] = int(safe_float(adv_signal_meta.get("autofill_changed_rows_district"), 0.0))
        adv_meta["autofill_changed_dims_district"] = int(safe_float(adv_signal_meta.get("autofill_changed_dims_district"), 0.0))
        adv_meta["autofill_total_rows_complex"] = int(safe_float(adv_signal_meta.get("autofill_total_rows_complex"), 0.0))
        adv_meta["autofill_signal_rows_complex"] = int(safe_float(adv_signal_meta.get("autofill_signal_rows_complex"), 0.0))
        adv_meta["autofill_assigned_rows_complex"] = int(safe_float(adv_signal_meta.get("autofill_assigned_rows_complex"), 0.0))
        adv_meta["autofill_updated_rows_complex"] = int(safe_float(adv_signal_meta.get("autofill_updated_rows_complex"), 0.0))
        adv_meta["autofill_changed_rows_complex"] = int(safe_float(adv_signal_meta.get("autofill_changed_rows_complex"), 0.0))
        adv_meta["autofill_changed_dims_complex"] = int(safe_float(adv_signal_meta.get("autofill_changed_dims_complex"), 0.0))
        adv_meta["autofill_month_range"] = str(adv_signal_meta.get("autofill_month_range", f"{months[0]}~{months[-1]}"))
        last_changes = adv_signal_payload.get("autofill_last_changes") if isinstance(adv_signal_payload.get("autofill_last_changes"), dict) else {}
        adv_meta["change_map_district"] = last_changes.get("district") if isinstance(last_changes.get("district"), dict) else {}
        adv_meta["change_map_complex"] = last_changes.get("complex") if isinstance(last_changes.get("complex"), dict) else {}
        run_history = adv_signal_payload.get("autofill_run_history") if isinstance(adv_signal_payload.get("autofill_run_history"), list) else []
        adv_meta["autofill_history_size"] = len(run_history)
        adv_meta["autofill_run_history"] = run_history

    complex_metrics = apply_advanced_risk_to_metrics(
        complex_metrics,
        enabled=adv_enabled,
        signal_payload=adv_signal_payload,
        level="complex",
    )
    if not complex_metrics:
        raise SystemExit("단지 메트릭이 비어 있습니다. area/min-trade 조건을 완화해 주세요.")

    profiles = ["실거주", "투자자", "법인"] if args.all_profiles else [args.profile]
    generated: list[Path] = []
    district_overlay_cap = args.nonquant_overlay_cap if nq_enabled else 0.0
    complex_overlay_cap = args.nonquant_complex_overlay_cap if nq_enabled else 0.0
    district_adv_cap = args.advanced_risk_cap if adv_enabled else 0.0
    complex_adv_cap = args.advanced_risk_complex_cap if adv_enabled else 0.0

    rows_by_profile: dict[str, list[dict[str, Any]]] = {}
    complex_rows_by_profile: dict[str, list[dict[str, Any]]] = {}
    for p in ["실거주", "투자자", "법인"]:
        rows_p = score_by_profile(metrics, p, nq_overlay_cap=district_overlay_cap, adv_risk_cap=district_adv_cap)
        rows_p = attach_scenarios(rows_p, p)
        rows_by_profile[p] = apply_theme_classification(rows_p, profile=p, level="district")
        crows_p = score_complex_by_profile(
            complex_metrics,
            p,
            nq_overlay_cap=complex_overlay_cap,
            adv_risk_cap=complex_adv_cap,
        )
        crows_p = attach_scenarios(crows_p, p)
        complex_rows_by_profile[p] = apply_theme_classification(crows_p, profile=p, level="complex")

    for profile in profiles:
        rows = rows_by_profile[profile]
        crows = complex_rows_by_profile[profile]
        wb = build_workbook(
            rows,
            crows,
            profile,
            months,
            rows_by_profile=rows_by_profile,
            complex_rows_by_profile=complex_rows_by_profile,
            top_per_district=args.top_complex_per_district,
            universe_label="전체(ALL)",
            original_complex_count=len(crows),
            backtest_lookback=args.backtest_lookback,
            backtest_horizon=args.backtest_horizon,
            backtest_top_fraction=args.backtest_top_frac,
            nq_meta=nq_meta,
            adv_meta=adv_meta,
        )
        out_path = out_dir / f"서울_전체_아파트_심층분석_v7_{profile}.xlsx"
        actual = save_workbook_safe(wb, out_path)
        generated.append(actual)
        print(f"[done] {profile}: {actual}")

    # 통합본(투자자 기준)도 추가 생성
    rows_i = rows_by_profile["투자자"]
    crows_i = complex_rows_by_profile["투자자"]
    wb_i = build_workbook(
        rows_i,
        crows_i,
        "투자자",
        months,
        rows_by_profile=rows_by_profile,
        complex_rows_by_profile=complex_rows_by_profile,
        top_per_district=args.top_complex_per_district,
        universe_label="전체(ALL)",
        original_complex_count=len(crows_i),
        backtest_lookback=args.backtest_lookback,
        backtest_horizon=args.backtest_horizon,
        backtest_top_fraction=args.backtest_top_frac,
        nq_meta=nq_meta,
        adv_meta=adv_meta,
    )
    master_path = out_dir / "서울_전체_아파트_심층분석_v7_통합.xlsx"
    actual_master = save_workbook_safe(wb_i, master_path)
    generated.append(actual_master)
    print(f"[done] 통합본: {actual_master}")

    if not args.no_universe:
        print("[universe] High+Medium 실전유니버스 생성")
        for profile in profiles:
            rows = rows_by_profile[profile]
            crows_all = complex_rows_by_profile[profile]
            crows_uni = filter_complex_universe(crows_all, ("high", "medium"))
            if not crows_uni:
                print(f"[skip] {profile} 유니버스가 비어 생성 생략")
                continue
            wb_u = build_workbook(
                rows,
                crows_uni,
                profile,
                months,
                rows_by_profile=rows_by_profile,
                complex_rows_by_profile=complex_rows_by_profile,
                top_per_district=args.top_complex_per_district,
                universe_label="High+Medium",
                original_complex_count=len(crows_all),
                backtest_lookback=args.backtest_lookback,
                backtest_horizon=args.backtest_horizon,
                backtest_top_fraction=args.backtest_top_frac,
                nq_meta=nq_meta,
                adv_meta=adv_meta,
            )
            out_u = out_dir / f"서울_전체_아파트_실전유니버스_v7_{profile}.xlsx"
            actual_u = save_workbook_safe(wb_u, out_u)
            generated.append(actual_u)
            print(f"[done] 유니버스 {profile}: {actual_u}")

        crows_all_i = complex_rows_by_profile["투자자"]
        crows_uni_i = filter_complex_universe(crows_all_i, ("high", "medium"))
        if crows_uni_i:
            wb_ui = build_workbook(
                rows_i,
                crows_uni_i,
                "투자자",
                months,
                rows_by_profile=rows_by_profile,
                complex_rows_by_profile=complex_rows_by_profile,
                top_per_district=args.top_complex_per_district,
                universe_label="High+Medium",
                original_complex_count=len(crows_all_i),
                backtest_lookback=args.backtest_lookback,
                backtest_horizon=args.backtest_horizon,
                backtest_top_fraction=args.backtest_top_frac,
                nq_meta=nq_meta,
                adv_meta=adv_meta,
            )
            out_ui = out_dir / "서울_전체_아파트_실전유니버스_v7_통합.xlsx"
            actual_ui = save_workbook_safe(wb_ui, out_ui)
            generated.append(actual_ui)
            print(f"[done] 유니버스 통합본: {actual_ui}")

    print("\n생성 완료 파일:")
    for p in generated:
        print(f"- {p}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
