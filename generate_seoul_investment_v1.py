#!/usr/bin/env python3
"""서울 25개 자치구 아파트 심층분석 생성기 (v1)

핵심 기능
- 국토부 실거래 API(아파트 매매/전월세) 월별 수집
- 서울 25개 구 병렬 수집 + 병렬 다중전문가 스코어링
- 확률(분위수) 기반 낙관/중립/비관 시나리오
- 실거주/투자자/법인 3개 프로필 별도 엑셀 생성

실행 예시
python generate_seoul_investment_v1.py --months 12 --all-profiles
"""

from __future__ import annotations

import argparse
import concurrent.futures as cf
import datetime as dt
import json
import math
import os
import random
import statistics
import time
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DEFAULT_DATA_GO_KR_API_KEY = (
    "1deaf1b26dc45dd6f26d5b481d28d2fe784327eb82cc8ccdb727a51baf462419"
)

APT_TRADE_URL = "https://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade"
APT_RENT_URL = "https://apis.data.go.kr/1613000/RTMSDataSvcAptRent/getRTMSDataSvcAptRent"

DEFAULT_AREA_MIN = 59.0
DEFAULT_AREA_MAX = 99.0
DEFAULT_DEPOSIT_YIELD_PCT = 3.0

HIGH_INTENSITY_LTPS = {"강남구", "서초구", "송파구", "용산구"}

PROFILE_CONFIGS: dict[str, dict[str, Any]] = {
    "실거주": {
        "weights": {
            "trend": 17,
            "value": 23,
            "liquidity": 18,
            "stability": 27,
            "policy": 15,
        },
        "ltv": 0.40,
        "loan_rate": 0.037,
        "buy_tax": 0.015,
        "annual_hold_cost": 0.004,
        "capital_gain_tax": 0.15,
        "horizon_years": 5,
        "display": "실거주 안정성 + 생활비/현금흐름 방어",
    },
    "투자자": {
        "weights": {
            "trend": 29,
            "value": 24,
            "liquidity": 22,
            "stability": 10,
            "policy": 15,
        },
        "ltv": 0.50,
        "loan_rate": 0.041,
        "buy_tax": 0.025,
        "annual_hold_cost": 0.006,
        "capital_gain_tax": 0.30,
        "horizon_years": 3,
        "display": "모멘텀 + 밸류 + 유동성 중심",
    },
    "법인": {
        "weights": {
            "trend": 24,
            "value": 17,
            "liquidity": 20,
            "stability": 9,
            "policy": 30,
        },
        "ltv": 0.55,
        "loan_rate": 0.044,
        "buy_tax": 0.120,
        "annual_hold_cost": 0.009,
        "capital_gain_tax": 0.45,
        "horizon_years": 3,
        "display": "세후수익/규제리스크 중심",
    },
}

# 분위수 기반 z-score (현실적 범위: P20/P50/P80)
SCENARIO_Z = {
    "bull": 0.8416,
    "base": 0.0,
    "bear": -0.8416,
}

SOURCES: list[dict[str, str]] = [
    {
        "category": "official",
        "title": "서울시 토지거래허가구역 지정현황(2026-01-07 기준)",
        "url": "https://land.seoul.go.kr/land/other/appointStatusSeoul.do",
        "as_of": "2026-01-07",
        "usage": "정책/규제 리스크(서울 전체 아파트 LTPS, 강남·서초·송파·용산 집중규제) 반영",
    },
    {
        "category": "official",
        "title": "한국은행 통화정책방향 관련 기자간담회(2026-01)",
        "url": "https://www.bok.or.kr/portal/bbs/B0000169/view.do?menuNo=200059&nttId=10095737",
        "as_of": "2026-01-15",
        "usage": "기준금리 2.50% 동결을 금리 가정의 기준점으로 사용",
    },
    {
        "category": "official",
        "title": "금융위 3단계 스트레스 DSR 시행방안(2025-05-20)",
        "url": "https://www.fsc.go.kr/no010101/84617",
        "as_of": "2025-05-20",
        "usage": "대출한도 제약 시나리오(레버리지 민감도) 반영",
    },
    {
        "category": "official",
        "title": "금융위 2026 상반기 스트레스 DSR 운영방향",
        "url": "https://www.fsc.go.kr/po010101/85824",
        "as_of": "2025-12-00",
        "usage": "수도권/지방 차등, 스트레스금리 하한 구조를 대출가정 보정에 반영",
    },
    {
        "category": "official",
        "title": "IMF Article IV Consultation with Korea (2025-02-06)",
        "url": "https://www.imf.org/en/news/articles/2025/02/06/pr-2530-korea-imf-concludes-2024-art-iv-consultation-with-republic-of-korea",
        "as_of": "2025-02-06",
        "usage": "부동산 PF/비은행권 리스크를 하방 시나리오 리스크 프리미엄으로 반영",
    },
    {
        "category": "paper",
        "title": "Micro-Analysis of Price Spillover Effect (Land, 2021)",
        "url": "https://www.mdpi.com/2073-445X/10/8/879",
        "as_of": "2021-08-21",
        "usage": "서울권역 내 가격 파급효과(상위가격지 주도 전이) 해석 근거",
    },
    {
        "category": "paper",
        "title": "Household lending, interest rates and housing price bubbles in Korea (2011)",
        "url": "https://www.sciencedirect.com/science/article/abs/pii/S0264999311000198",
        "as_of": "2011-05-01",
        "usage": "금리/가계대출 변화가 버블 국면 변동성에 미치는 영향의 고전적 근거",
    },
    {
        "category": "community",
        "title": "Blind: 잠실 주공5단지 가능?",
        "url": "https://www.teamblind.com/kr/post/%EC%9E%A0%EC%8B%A4-%EC%A3%BC%EA%B3%B55%EB%8B%A8%EC%A7%80-%EA%B0%80%EB%8A%A5-rGt1xuvk",
        "as_of": "2024-08-18",
        "usage": "실수요/투자 의사결정에서 레버리지·사업지연 체감리스크 정성검증",
    },
    {
        "category": "community",
        "title": "Blind: 잠5 잠실 주공 5단지",
        "url": "https://www.teamblind.com/kr/post/%EC%9E%A05-%EC%9E%A0%EC%8B%A4-%EC%A3%BC%EA%B3%B5-5%EB%8B%A8%EC%A7%80-da2xdoeB",
        "as_of": "2023-04-26",
        "usage": "재건축 지연/행정 리스크를 보수 시나리오 기간 변수에 반영",
    },
]


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

DARK = "1B2A4A"
BLUE = "2C3E6B"
GOLD = "C9A84C"
WHITE = "FFFFFF"
LIGHT = "F3F5F9"
LINE = "D9D9D9"
GREEN = "27AE60"
RED = "E74C3C"
ORANGE = "F39C12"

FONT_TITLE = Font(name="맑은 고딕", size=14, bold=True, color=WHITE)
FONT_HEADER = Font(name="맑은 고딕", size=10, bold=True, color=WHITE)
FONT_CELL = Font(name="맑은 고딕", size=10)
FONT_SMALL = Font(name="맑은 고딕", size=9, color="666666")

FILL_TITLE = PatternFill("solid", fgColor=DARK)
FILL_HEADER = PatternFill("solid", fgColor=BLUE)
FILL_LIGHT = PatternFill("solid", fgColor=LIGHT)
FILL_GOLD = PatternFill("solid", fgColor=GOLD)
FILL_WHITE = PatternFill("solid", fgColor=WHITE)

BORDER = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class DistrictInfo:
    code: str
    name: str


@dataclass
class MonthSummary:
    year_month: str
    trade_count: int
    rent_count: int
    median_trade_10k: int | None
    median_rent_10k: int | None
    median_monthly_rent_10k: int | None
    jeonse_ratio_pct: float | None
    rent_yield_proxy_pct: float | None


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------


def clamp(v: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, v))


def pct(v: float | None) -> str:
    if v is None:
        return "-"
    return f"{v * 100:.1f}%"


def eok_from_10k(v: int | None) -> float | None:
    if v is None:
        return None
    return v / 10000.0


def now_yyyymm() -> str:
    t = dt.date.today()
    return f"{t.year:04d}{t.month:02d}"


def shift_yyyymm(yyyymm: str, delta_months: int) -> str:
    y = int(yyyymm[:4])
    m = int(yyyymm[4:])
    total = y * 12 + (m - 1) + delta_months
    ny = total // 12
    nm = total % 12 + 1
    return f"{ny:04d}{nm:02d}"


def iter_months(end_yyyymm: str, months: int) -> list[str]:
    start_delta = -(months - 1)
    return [shift_yyyymm(end_yyyymm, i) for i in range(start_delta, 1)]


def median_int(values: list[int]) -> int | None:
    if not values:
        return None
    return int(statistics.median(values))


def stdev(values: list[float]) -> float:
    if len(values) <= 1:
        return 0.0
    return statistics.stdev(values)


def safe_div(a: float, b: float, default: float = 0.0) -> float:
    if not b:
        return default
    return a / b


def percentile_rank(value: float | None, arr: list[float], reverse: bool = False) -> float:
    valid = [x for x in arr if x is not None and not math.isnan(x)]
    if value is None or not valid:
        return 50.0
    sorted_vals = sorted(valid)
    n = len(sorted_vals)
    less = sum(1 for x in sorted_vals if x < value)
    equal = sum(1 for x in sorted_vals if x == value)
    rank = (less + 0.5 * equal) / n * 100.0
    if reverse:
        rank = 100.0 - rank
    return clamp(rank, 0.0, 100.0)


def style_cell(cell, *, font=None, fill=None, align=None):
    cell.font = font or FONT_CELL
    cell.fill = fill or FILL_WHITE
    cell.alignment = align or ALIGN_CENTER
    cell.border = BORDER


def write_title(ws, row: int, title: str, col_end: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=1, value=title)
    c.font = FONT_TITLE
    c.fill = FILL_TITLE
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = 30
    return row + 1


def write_headers(ws, row: int, headers: list[str]) -> int:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        style_cell(c, font=FONT_HEADER, fill=FILL_HEADER)
    ws.row_dimensions[row].height = 22
    return row + 1


def write_row(ws, row: int, values: list[Any], fill: PatternFill | None = None, left: bool = False) -> int:
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        style_cell(c, fill=fill or FILL_WHITE, align=ALIGN_LEFT if left else ALIGN_CENTER)
    return row + 1


def set_widths(ws, widths: list[float]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def score_band_fill(score: float) -> PatternFill:
    if score >= 75:
        return PatternFill("solid", fgColor="D5F5E3")
    if score >= 55:
        return PatternFill("solid", fgColor="FCF3CF")
    return PatternFill("solid", fgColor="FADBD8")


# ---------------------------------------------------------------------------
# Region / API helpers
# ---------------------------------------------------------------------------


def get_api_key(explicit: str | None = None) -> str:
    if explicit:
        return explicit.strip()
    env = os.getenv("DATA_GO_KR_API_KEY", "").strip()
    if env:
        return env
    return DEFAULT_DATA_GO_KR_API_KEY


def load_seoul_gu_list(region_code_file: Path) -> list[DistrictInfo]:
    text = region_code_file.read_text(encoding="utf-8")
    rows: list[DistrictInfo] = []
    prefix = "\uc11c\uc6b8\ud2b9\ubcc4\uc2dc "
    for line in text.splitlines():
        parts = line.split("\t")
        if len(parts) < 2:
            continue
        code10 = parts[0].strip()
        name = parts[1].strip()
        if not (code10.isdigit() and len(code10) == 10):
            continue
        if code10 == "1100000000":
            continue
        if not code10.startswith("11"):
            continue
        if not code10.endswith("00000"):
            continue
        if not name.startswith(prefix):
            continue
        gu_name = name.replace(prefix, "")
        rows.append(DistrictInfo(code=code10[:5], name=gu_name))

    rows.sort(key=lambda x: x.code)
    if len(rows) != 25:
        raise RuntimeError(f"서울 자치구 코드 25개를 찾지 못했습니다. 현재: {len(rows)}")
    return rows


def build_url(base_url: str, api_key: str, region_code: str, yyyymm: str, num_rows: int, page_no: int) -> str:
    params = {
        "serviceKey": api_key,
        "LAWD_CD": region_code,
        "DEAL_YMD": yyyymm,
        "numOfRows": str(num_rows),
        "pageNo": str(page_no),
    }
    return f"{base_url}?{urllib.parse.urlencode(params)}"


def fetch_xml(url: str, timeout_sec: int = 25, retries: int = 4) -> str:
    backoff = 0.6
    last_err: Exception | None = None
    for _ in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "KRE-Seoul-Analyzer/1.0"})
            with urllib.request.urlopen(req, timeout=timeout_sec) as resp:
                return resp.read().decode("utf-8", errors="ignore")
        except Exception as e:  # noqa: BLE001
            last_err = e
            time.sleep(backoff + random.random() * 0.3)
            backoff *= 1.9
    raise RuntimeError(f"API 요청 실패: {last_err}")


def parse_xml_items_trade(xml_text: str) -> tuple[list[dict[str, Any]], int, str | None]:
    root = ET.fromstring(xml_text)
    code = (root.findtext(".//resultCode") or "").strip()
    if code != "000":
        return [], 0, code

    total_count = int((root.findtext(".//totalCount") or "0").strip() or "0")
    out: list[dict[str, Any]] = []
    for item in root.findall(".//item"):
        if ((item.findtext("cdealType") or "").strip()) == "O":
            continue
        raw_price = (item.findtext("dealAmount") or "").replace(",", "").strip()
        if not raw_price.isdigit():
            continue
        area_raw = (item.findtext("excluUseAr") or "0").strip()
        try:
            area = float(area_raw)
        except ValueError:
            area = 0.0
        out.append(
            {
                "apt_name": (item.findtext("aptNm") or "").strip(),
                "dong": (item.findtext("umdNm") or "").strip(),
                "area_sqm": area,
                "price_10k": int(raw_price),
                "year": int((item.findtext("dealYear") or "0").strip() or 0),
                "month": int((item.findtext("dealMonth") or "0").strip() or 0),
                "day": int((item.findtext("dealDay") or "0").strip() or 0),
            }
        )
    return out, total_count, None


def parse_xml_items_rent(xml_text: str) -> tuple[list[dict[str, Any]], int, str | None]:
    root = ET.fromstring(xml_text)
    code = (root.findtext(".//resultCode") or "").strip()
    if code != "000":
        return [], 0, code

    total_count = int((root.findtext(".//totalCount") or "0").strip() or "0")
    out: list[dict[str, Any]] = []
    for item in root.findall(".//item"):
        if ((item.findtext("cdealType") or "").strip()) == "O":
            continue
        raw_dep = (item.findtext("deposit") or "").replace(",", "").strip()
        raw_mon = (item.findtext("monthlyRent") or "0").replace(",", "").strip() or "0"
        if not raw_dep.isdigit():
            continue
        area_raw = (item.findtext("excluUseAr") or "0").strip()
        try:
            area = float(area_raw)
        except ValueError:
            area = 0.0
        out.append(
            {
                "apt_name": (item.findtext("aptNm") or "").strip(),
                "dong": (item.findtext("umdNm") or "").strip(),
                "area_sqm": area,
                "deposit_10k": int(raw_dep),
                "monthly_rent_10k": int(raw_mon) if raw_mon.isdigit() else 0,
                "contract_type": (item.findtext("contractType") or "").strip(),
                "year": int((item.findtext("dealYear") or "0").strip() or 0),
                "month": int((item.findtext("dealMonth") or "0").strip() or 0),
                "day": int((item.findtext("dealDay") or "0").strip() or 0),
            }
        )
    return out, total_count, None


def fetch_all_pages(
    *,
    base_url: str,
    api_key: str,
    region_code: str,
    yyyymm: str,
    num_rows: int,
    parse_kind: str,
    max_pages: int = 80,
) -> tuple[list[dict[str, Any]], int, str | None]:
    parser = parse_xml_items_trade if parse_kind == "trade" else parse_xml_items_rent
    all_items: list[dict[str, Any]] = []
    total_count = 0

    for page in range(1, max_pages + 1):
        url = build_url(base_url, api_key, region_code, yyyymm, num_rows, page)
        xml_text = fetch_xml(url)
        items, tcount, err = parser(xml_text)
        if err is not None:
            return [], 0, err

        total_count = max(total_count, tcount)
        all_items.extend(items)

        if not items:
            break
        if len(items) < num_rows:
            break
        if total_count and len(all_items) >= total_count:
            break

    return all_items, total_count, None


def filter_area(items: list[dict[str, Any]], area_min: float, area_max: float) -> list[dict[str, Any]]:
    return [x for x in items if area_min <= float(x.get("area_sqm", 0.0)) <= area_max]


def summarize_month(
    *,
    yyyymm: str,
    trade_items: list[dict[str, Any]],
    rent_items: list[dict[str, Any]],
    deposit_yield_pct: float,
) -> MonthSummary:
    prices = [int(x["price_10k"]) for x in trade_items if int(x.get("price_10k", 0)) > 0]
    deposits = [int(x["deposit_10k"]) for x in rent_items if int(x.get("deposit_10k", 0)) > 0]
    monthly_rents = [
        int(x["monthly_rent_10k"])
        for x in rent_items
        if int(x.get("monthly_rent_10k", 0)) > 0
    ]

    med_price = median_int(prices)
    med_deposit = median_int(deposits)
    med_monthly = median_int(monthly_rents)

    jeonse_ratio = None
    if med_price and med_deposit:
        jeonse_ratio = med_deposit / med_price * 100.0

    rent_yield_proxy = None
    if med_price and (med_deposit or med_monthly):
        annual_income_10k = (med_monthly or 0) * 12 + int((med_deposit or 0) * (deposit_yield_pct / 100.0))
        rent_yield_proxy = annual_income_10k / med_price * 100.0

    return MonthSummary(
        year_month=yyyymm,
        trade_count=len(prices),
        rent_count=len(deposits),
        median_trade_10k=med_price,
        median_rent_10k=med_deposit,
        median_monthly_rent_10k=med_monthly,
        jeonse_ratio_pct=jeonse_ratio,
        rent_yield_proxy_pct=rent_yield_proxy,
    )

# ---------------------------------------------------------------------------
# Collection and analytics
# ---------------------------------------------------------------------------


def collect_one_district_month(
    district: DistrictInfo,
    yyyymm: str,
    api_key: str,
    num_rows: int,
    area_min: float,
    area_max: float,
    deposit_yield_pct: float,
) -> tuple[str, str, dict[str, Any]]:
    t0 = time.time()

    trade_items, _, trade_err = fetch_all_pages(
        base_url=APT_TRADE_URL,
        api_key=api_key,
        region_code=district.code,
        yyyymm=yyyymm,
        num_rows=num_rows,
        parse_kind="trade",
    )
    if trade_err:
        return district.name, yyyymm, {
            "error": f"trade_api_error:{trade_err}",
            "elapsed": round(time.time() - t0, 2),
        }

    rent_items, _, rent_err = fetch_all_pages(
        base_url=APT_RENT_URL,
        api_key=api_key,
        region_code=district.code,
        yyyymm=yyyymm,
        num_rows=num_rows,
        parse_kind="rent",
    )
    if rent_err:
        return district.name, yyyymm, {
            "error": f"rent_api_error:{rent_err}",
            "elapsed": round(time.time() - t0, 2),
        }

    trade_f = filter_area(trade_items, area_min, area_max)
    rent_f = filter_area(rent_items, area_min, area_max)
    summary = summarize_month(
        yyyymm=yyyymm,
        trade_items=trade_f,
        rent_items=rent_f,
        deposit_yield_pct=deposit_yield_pct,
    )

    return district.name, yyyymm, {
        "error": None,
        "elapsed": round(time.time() - t0, 2),
        "summary": summary.__dict__,
        "raw_counts": {
            "trade_total": len(trade_items),
            "trade_filtered": len(trade_f),
            "rent_total": len(rent_items),
            "rent_filtered": len(rent_f),
        },
    }


def collect_market_panel(
    *,
    districts: list[DistrictInfo],
    months: list[str],
    api_key: str,
    num_rows: int,
    max_workers: int,
    area_min: float,
    area_max: float,
    deposit_yield_pct: float,
) -> tuple[dict[str, dict[str, dict[str, Any]]], list[dict[str, Any]]]:
    panel: dict[str, dict[str, dict[str, Any]]] = {d.name: {} for d in districts}
    logs: list[dict[str, Any]] = []

    with cf.ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {}
        for d in districts:
            for ym in months:
                fut = ex.submit(
                    collect_one_district_month,
                    d,
                    ym,
                    api_key,
                    num_rows,
                    area_min,
                    area_max,
                    deposit_yield_pct,
                )
                futures[fut] = (d.name, ym)

        total = len(futures)
        done = 0
        for fut in cf.as_completed(futures):
            name, ym = futures[fut]
            done += 1
            try:
                n, y, payload = fut.result()
            except Exception as e:  # noqa: BLE001
                payload = {"error": f"exception:{e}", "elapsed": None}
                n, y = name, ym

            panel[n][y] = payload
            logs.append({"district": n, "year_month": y, **payload})
            if done % 40 == 0 or done == total:
                print(f"[collect] {done}/{total} 완료")

    return panel, logs


def ffill(values: list[float | None]) -> list[float | None]:
    out: list[float | None] = []
    last = None
    for v in values:
        if v is None:
            out.append(last)
        else:
            out.append(v)
            last = v
    return out


def compute_max_drawdown(series: list[float]) -> float:
    if not series:
        return 0.0
    peak = series[0]
    mdd = 0.0
    for x in series:
        peak = max(peak, x)
        if peak > 0:
            dd = (x / peak) - 1.0
            mdd = min(mdd, dd)
    return abs(mdd)


def annualized_vol_from_prices(prices: list[float]) -> float:
    rets: list[float] = []
    for i in range(1, len(prices)):
        p0 = prices[i - 1]
        p1 = prices[i]
        if p0 > 0 and p1 > 0:
            rets.append(math.log(p1 / p0))
    if len(rets) <= 1:
        return 0.0
    return stdev(rets) * math.sqrt(12.0)


def growth(prices: list[float | None], lag: int) -> float | None:
    if len(prices) <= lag or lag <= 0:
        return None
    p1 = prices[-1]
    p0 = prices[-1 - lag]
    if p1 is None or p0 is None:
        return None
    if p0 <= 0:
        return None
    return p1 / p0 - 1.0


def build_district_metrics(
    districts: list[DistrictInfo],
    months: list[str],
    panel: dict[str, dict[str, dict[str, Any]]],
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []

    for d in districts:
        month_rows = []
        for ym in months:
            payload = panel.get(d.name, {}).get(ym, {})
            if payload.get("error"):
                month_rows.append({"year_month": ym, "error": payload.get("error")})
                continue
            summary = payload.get("summary") or {}
            raw_counts = payload.get("raw_counts") or {}
            month_rows.append(
                {
                    "year_month": ym,
                    **summary,
                    **raw_counts,
                }
            )

        trade_series: list[float | None] = [
            (r.get("median_trade_10k") / 10000.0) if r.get("median_trade_10k") else None
            for r in month_rows
        ]
        rent_series: list[float | None] = [
            (r.get("median_rent_10k") / 10000.0) if r.get("median_rent_10k") else None
            for r in month_rows
        ]
        jeonse_series: list[float | None] = [r.get("jeonse_ratio_pct") for r in month_rows]
        yield_series: list[float | None] = [r.get("rent_yield_proxy_pct") for r in month_rows]

        trade_counts = [int(r.get("trade_count") or 0) for r in month_rows]
        rent_counts = [int(r.get("rent_count") or 0) for r in month_rows]

        trade_ffill = ffill(trade_series)
        rent_ffill = ffill(rent_series)

        valid_price = [x for x in trade_ffill if x is not None and x > 0]
        latest_price = valid_price[-1] if valid_price else None

        valid_rent = [x for x in rent_ffill if x is not None and x > 0]
        latest_rent = valid_rent[-1] if valid_rent else None

        lag_3m = min(3, max(1, len(trade_ffill) - 1))
        lag_6m = min(6, max(1, len(trade_ffill) - 1))
        # 12개월 데이터가 있으면 끝-처음(11간격)으로 YoY에 준하는 성장률을 계산한다.
        lag_12m = min(11, max(1, len(trade_ffill) - 1))
        g3 = growth(trade_ffill, lag_3m)
        g6 = growth(trade_ffill, lag_6m)
        g12 = growth(trade_ffill, lag_12m)

        vols = annualized_vol_from_prices([x for x in trade_ffill if x is not None])
        mdd = compute_max_drawdown([x for x in trade_ffill if x is not None])

        latest_jeonse = next((x for x in reversed(jeonse_series) if x is not None), None)
        latest_yield = next((x for x in reversed(yield_series) if x is not None), None)

        latest_trade_count = trade_counts[-1] if trade_counts else 0
        latest_rent_count = rent_counts[-1] if rent_counts else 0
        avg_trade_3m = statistics.mean(trade_counts[-3:]) if len(trade_counts) >= 3 else statistics.mean(trade_counts) if trade_counts else 0.0

        months_with_trade = sum(1 for x in trade_counts if x > 0)
        months_with_rent = sum(1 for x in rent_counts if x > 0)
        coverage_trade = months_with_trade / len(months) if months else 0.0
        coverage_rent = months_with_rent / len(months) if months else 0.0

        total_trade_samples = sum(trade_counts)
        total_rent_samples = sum(rent_counts)

        quality = (
            40.0 * clamp(coverage_trade, 0.0, 1.0)
            + 25.0 * clamp(coverage_rent, 0.0, 1.0)
            + 20.0 * clamp(safe_div(latest_trade_count, 60.0), 0.0, 1.0)
            + 15.0 * clamp(safe_div(latest_rent_count, 120.0), 0.0, 1.0)
        )

        out.append(
            {
                "district": d.name,
                "code": d.code,
                "months": month_rows,
                "latest_price_eok": latest_price,
                "latest_rent_eok": latest_rent,
                "growth_3m": g3,
                "growth_6m": g6,
                "growth_12m": g12,
                "vol_ann": vols,
                "max_drawdown": mdd,
                "latest_jeonse_ratio": latest_jeonse,
                "latest_rent_yield": latest_yield,
                "latest_trade_count": latest_trade_count,
                "latest_rent_count": latest_rent_count,
                "avg_trade_3m": float(avg_trade_3m),
                "coverage_trade": coverage_trade,
                "coverage_rent": coverage_rent,
                "total_trade_samples": total_trade_samples,
                "total_rent_samples": total_rent_samples,
                "data_quality": round(quality, 1),
            }
        )

    # 서울 전체 상대가치
    latest_prices = [m["latest_price_eok"] for m in out if m["latest_price_eok"]]
    seoul_median_price = statistics.median(latest_prices) if latest_prices else 0.0

    for m in out:
        p = m["latest_price_eok"]
        if p and seoul_median_price > 0:
            m["price_vs_seoul"] = p / seoul_median_price
        else:
            m["price_vs_seoul"] = None

    return out


def run_multi_expert_scoring(metrics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    # 병렬 전문가 에이전트 입력 벡터
    growth12_all = [m.get("growth_12m") for m in metrics if m.get("growth_12m") is not None]
    growth3_all = [m.get("growth_3m") for m in metrics if m.get("growth_3m") is not None]
    jeonse_all = [m.get("latest_jeonse_ratio") for m in metrics if m.get("latest_jeonse_ratio") is not None]
    yield_all = [m.get("latest_rent_yield") for m in metrics if m.get("latest_rent_yield") is not None]
    pvs_all = [m.get("price_vs_seoul") for m in metrics if m.get("price_vs_seoul") is not None]
    liq_all = [m.get("avg_trade_3m") for m in metrics if m.get("avg_trade_3m") is not None]
    vol_all = [m.get("vol_ann") for m in metrics if m.get("vol_ann") is not None]
    dd_all = [m.get("max_drawdown") for m in metrics if m.get("max_drawdown") is not None]

    def trend_expert(m: dict[str, Any]) -> float:
        s12 = percentile_rank(m.get("growth_12m"), growth12_all)
        s3 = percentile_rank(m.get("growth_3m"), growth3_all)
        return 0.65 * s12 + 0.35 * s3

    def value_expert(m: dict[str, Any]) -> float:
        s_jeonse = percentile_rank(m.get("latest_jeonse_ratio"), jeonse_all)
        s_yield = percentile_rank(m.get("latest_rent_yield"), yield_all)
        s_pvs = percentile_rank(m.get("price_vs_seoul"), pvs_all, reverse=True)
        return 0.40 * s_jeonse + 0.35 * s_yield + 0.25 * s_pvs

    def liquidity_expert(m: dict[str, Any]) -> float:
        s_liq = percentile_rank(m.get("avg_trade_3m"), liq_all)
        s_latest = percentile_rank(float(m.get("latest_trade_count") or 0), liq_all)
        return 0.65 * s_liq + 0.35 * s_latest

    def stability_expert(m: dict[str, Any]) -> float:
        s_vol = percentile_rank(m.get("vol_ann"), vol_all, reverse=True)
        s_dd = percentile_rank(m.get("max_drawdown"), dd_all, reverse=True)
        dq = m.get("data_quality") or 0.0
        return 0.40 * s_vol + 0.35 * s_dd + 0.25 * dq

    def policy_expert(m: dict[str, Any]) -> float:
        base = 45.0  # 서울 전체 LTPS 가정
        if m["district"] in HIGH_INTENSITY_LTPS:
            base -= 15.0
        # 거래가 충분하면 정책 충격 흡수력이 상대적으로 높다고 가정
        base += clamp((m.get("avg_trade_3m") or 0) / 12.0, 0.0, 10.0)
        base += clamp((m.get("data_quality") or 0) / 20.0, 0.0, 5.0)
        return clamp(base, 0.0, 100.0)

    with cf.ThreadPoolExecutor(max_workers=5) as ex:
        f_trend = {ex.submit(trend_expert, m): m for m in metrics}
        f_value = {ex.submit(value_expert, m): m for m in metrics}
        f_liq = {ex.submit(liquidity_expert, m): m for m in metrics}
        f_stb = {ex.submit(stability_expert, m): m for m in metrics}
        f_pol = {ex.submit(policy_expert, m): m for m in metrics}

        for fut_map, key in [
            (f_trend, "trend_score"),
            (f_value, "value_score"),
            (f_liq, "liquidity_score"),
            (f_stb, "stability_score"),
            (f_pol, "policy_score"),
        ]:
            for fut in cf.as_completed(fut_map):
                m = fut_map[fut]
                try:
                    m[key] = round(float(fut.result()), 2)
                except Exception:  # noqa: BLE001
                    m[key] = 50.0

    return metrics


def score_by_profile(metrics: list[dict[str, Any]], profile: str) -> list[dict[str, Any]]:
    conf = PROFILE_CONFIGS[profile]
    w = conf["weights"]

    rows = []
    for m in metrics:
        total = (
            m["trend_score"] * w["trend"]
            + m["value_score"] * w["value"]
            + m["liquidity_score"] * w["liquidity"]
            + m["stability_score"] * w["stability"]
            + m["policy_score"] * w["policy"]
        ) / 100.0

        x = dict(m)
        x["profile"] = profile
        x["total_score"] = round(total, 2)
        rows.append(x)

    rows.sort(key=lambda r: r["total_score"], reverse=True)
    for i, r in enumerate(rows, 1):
        r["rank"] = i
    return rows


def scenario_for_district(m: dict[str, Any], profile: str) -> dict[str, Any]:
    cfg = PROFILE_CONFIGS[profile]

    price = m.get("latest_price_eok") or 0.0
    if price <= 0:
        return {
            "bull_cum": None,
            "base_cum": None,
            "bear_cum": None,
            "bull_roi": None,
            "base_roi": None,
            "bear_roi": None,
            "horizon_years": cfg["horizon_years"],
        }

    g12 = m.get("growth_12m")
    g3 = m.get("growth_3m")
    # 결측 대비: 장기평균 성장률 2.0% 가정
    g12 = 0.02 if g12 is None else g12
    g3 = 0.005 if g3 is None else g3
    base_annual = clamp(0.60 * g12 + 0.40 * g3, -0.06, 0.10)

    vol_ann = m.get("vol_ann") or 0.08
    vol_ann = clamp(vol_ann, 0.04, 0.25)
    vol_m = vol_ann / math.sqrt(12.0)

    h = int(cfg["horizon_years"])
    t_months = h * 12

    # 정책 레짐 보정: LTPS 강화지역은 상방축소/하방확대
    regime_up = -0.006 if m["district"] in HIGH_INTENSITY_LTPS else -0.003
    regime_down = -0.012 if m["district"] in HIGH_INTENSITY_LTPS else -0.008

    def annual_from_z(z: float) -> float:
        mu_m = math.log1p(base_annual) / 12.0
        drift = (mu_m - 0.5 * vol_m * vol_m) * 12.0
        shock = z * vol_m * math.sqrt(12.0)
        return math.exp(drift + shock) - 1.0

    bull_a = annual_from_z(SCENARIO_Z["bull"]) + regime_up
    base_a = annual_from_z(SCENARIO_Z["base"]) - 0.002
    bear_a = annual_from_z(SCENARIO_Z["bear"]) + regime_down

    bull_a = clamp(bull_a, -0.15, 0.20)
    base_a = clamp(base_a, -0.12, 0.14)
    bear_a = clamp(bear_a, -0.25, 0.10)

    bull_c = (1.0 + bull_a) ** h - 1.0
    base_c = (1.0 + base_a) ** h - 1.0
    bear_c = (1.0 + bear_a) ** h - 1.0

    buy_tax = cfg["buy_tax"]
    hold_cost = cfg["annual_hold_cost"]
    cg_tax = cfg["capital_gain_tax"]
    ltv = cfg["ltv"]
    loan_rate = cfg["loan_rate"]

    debt = price * ltv
    equity = price - debt
    interest_total = debt * loan_rate * h
    carry_total = price * hold_cost * h
    # 보유기간 임대/거주 효용 수익률(현실 반영): 전세환산수익률을 보수적으로 반영
    rent_yield = clamp((m.get("latest_rent_yield") or 1.8) / 100.0, 0.008, 0.05)
    if profile == "실거주":
        annual_utility_yield = max(0.012, rent_yield * 0.60)
    elif profile == "투자자":
        annual_utility_yield = rent_yield * 0.85
    else:
        annual_utility_yield = rent_yield * 0.80
    utility_income_total = price * annual_utility_yield * h

    def leveraged_roi(cum_ret: float) -> float:
        exit_price = price * (1.0 + cum_ret)
        gross_gain = exit_price - price
        tx = max(gross_gain, 0.0) * cg_tax
        net_gain = (
            gross_gain
            + utility_income_total
            - tx
            - (price * buy_tax)
            - carry_total
            - interest_total
        )
        if equity <= 0:
            return 0.0
        return net_gain / equity

    return {
        "bull_cum": bull_c,
        "base_cum": base_c,
        "bear_cum": bear_c,
        "bull_roi": leveraged_roi(bull_c),
        "base_roi": leveraged_roi(base_c),
        "bear_roi": leveraged_roi(bear_c),
        "horizon_years": h,
        "annual_base": base_a,
        "annual_vol": vol_ann,
        "annual_utility_yield": annual_utility_yield,
        "assumption": {
            "buy_tax": buy_tax,
            "hold_cost": hold_cost,
            "capital_gain_tax": cg_tax,
            "ltv": ltv,
            "loan_rate": loan_rate,
            "utility_income_total": utility_income_total,
            "t_months": t_months,
        },
    }


def attach_scenarios(rows: list[dict[str, Any]], profile: str) -> list[dict[str, Any]]:
    for r in rows:
        r["scenario"] = scenario_for_district(r, profile)
    return rows

# ---------------------------------------------------------------------------
# Workbook builders
# ---------------------------------------------------------------------------


def add_common_header(ws, title: str, subtitle: str, cols: int) -> int:
    r = write_title(ws, 1, title, cols)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    c = ws.cell(row=r, column=1, value=subtitle)
    c.font = FONT_SMALL
    c.alignment = ALIGN_LEFT
    c.fill = FILL_LIGHT
    c.border = BORDER
    ws.row_dimensions[r].height = 20
    return r + 1


def add_sheet_dashboard(wb, rows: list[dict[str, Any]], profile: str, months: list[str]):
    ws = wb.create_sheet("종합 대시보드")
    r = add_common_header(
        ws,
        f"서울 25개구 아파트 심층분석 대시보드 ({profile})",
        f"기준: {months[0]}~{months[-1]} / 면적필터 59~99m² / 생성시각 {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}",
        10,
    )

    top5 = rows[:5]
    bot5 = rows[-5:]

    r = write_row(ws, r, ["[핵심 가정]"] + [""] * 9, fill=FILL_GOLD, left=True)
    assumptions = [
        "1) 월별 중앙값 기반(거래 outlier 완화), 구 단위 비교",
        "2) 낙관/중립/비관은 P80/P50/P20 분위수 기반 확률 시나리오",
        "3) 레버리지 ROI는 프로필별 LTV/금리/세율/보유비용 반영",
        "4) 서울 LTPS 규제와 4개 집중규제구(강남·서초·송파·용산) 정책 패널티 반영",
        "5) 데이터 품질점수(표본수+커버리지) 낮으면 점수 자동 보수화",
    ]
    for line in assumptions:
        r = write_row(ws, r, [line] + [""] * 9, left=True)

    r += 1
    r = write_row(ws, r, ["[TOP 5 추천]"] + [""] * 9, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["순위", "구", "총점", "12M성장", "전세가율", "변동성", "거래량(최근)", "낙관ROI", "중립ROI", "비관ROI"])
    for x in top5:
        sc = x["scenario"]
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["total_score"],
                pct(x.get("growth_12m")),
                f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                x.get("latest_trade_count") or 0,
                pct(sc.get("bull_roi")),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
            ],
        )

    r += 1
    r = write_row(ws, r, ["[BOTTOM 5 (관망/주의)]"] + [""] * 9, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["순위", "구", "총점", "12M성장", "전세가율", "변동성", "거래량(최근)", "낙관ROI", "중립ROI", "비관ROI"])
    for x in bot5:
        sc = x["scenario"]
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["total_score"],
                pct(x.get("growth_12m")),
                f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                x.get("latest_trade_count") or 0,
                pct(sc.get("bull_roi")),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
            ],
        )

    set_widths(ws, [8, 13, 9, 11, 11, 10, 12, 11, 11, 11])


def add_sheet_scorecard(wb, rows: list[dict[str, Any]], profile: str, months: list[str]):
    ws = wb.create_sheet("서울25구 스코어카드")
    r = add_common_header(ws, f"서울 25개구 스코어카드 ({profile})", f"분석기간: {months[0]}~{months[-1]}", 15)
    headers = [
        "순위", "구", "총점", "Trend", "Value", "Liquidity", "Stability", "Policy", "DataQ",
        "최근매매(억)", "최근전세(억)", "12M성장", "전세가율", "변동성", "최근거래수",
    ]
    r = write_headers(ws, r, headers)
    for x in rows:
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["total_score"],
                x["trend_score"],
                x["value_score"],
                x["liquidity_score"],
                x["stability_score"],
                x["policy_score"],
                x["data_quality"],
                round(x["latest_price_eok"], 2) if x.get("latest_price_eok") else "-",
                round(x["latest_rent_eok"], 2) if x.get("latest_rent_eok") else "-",
                pct(x.get("growth_12m")),
                f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                x.get("latest_trade_count") or 0,
            ],
        )
        ws.cell(row=r - 1, column=3).fill = score_band_fill(float(x["total_score"]))

    set_widths(ws, [6, 12, 8, 8, 8, 9, 9, 8, 8, 10, 10, 10, 9, 9, 9])


def add_sheet_coverage(wb, rows: list[dict[str, Any]], months: list[str]):
    ws = wb.create_sheet("데이터 커버리지")
    r = add_common_header(ws, "데이터 품질/커버리지", "표본수와 월별 커버리지를 점수화하여 신뢰도 반영", 11)
    headers = [
        "구", "분석월수", "매매커버리지", "전월세커버리지", "총매매표본", "총전월세표본", "최근매매표본", "최근전월세표본", "평균거래(3M)", "DataQ", "비고"
    ]
    r = write_headers(ws, r, headers)
    for x in rows:
        note = "양호"
        if x["data_quality"] < 55:
            note = "주의(표본 부족)"
        elif x["data_quality"] < 70:
            note = "보통"
        r = write_row(
            ws,
            r,
            [
                x["district"],
                len(months),
                f"{x['coverage_trade']*100:.1f}%",
                f"{x['coverage_rent']*100:.1f}%",
                x["total_trade_samples"],
                x["total_rent_samples"],
                x["latest_trade_count"],
                x["latest_rent_count"],
                round(x["avg_trade_3m"], 1),
                x["data_quality"],
                note,
            ],
        )
        ws.cell(row=r - 1, column=10).fill = score_band_fill(float(x["data_quality"]))
    set_widths(ws, [12, 9, 11, 12, 11, 12, 11, 12, 11, 8, 16])


def add_sheet_trade_trend(wb, rows: list[dict[str, Any]], months: list[str]):
    ws = wb.create_sheet("매매 추세")
    ncol = 1 + len(months)
    r = add_common_header(ws, "구별 매매 중앙값 추세(억)", "면적필터 59~99m², 월별 중앙값", ncol)
    r = write_headers(ws, r, ["구"] + months)
    for x in rows:
        row_vals = [x["district"]]
        month_map = {m["year_month"]: m for m in x["months"]}
        for ym in months:
            m = month_map.get(ym, {})
            if m.get("median_trade_10k"):
                row_vals.append(round(m["median_trade_10k"] / 10000.0, 2))
            else:
                row_vals.append(None)
        r = write_row(ws, r, row_vals)
    set_widths(ws, [12] + [9] * len(months))


def add_sheet_rent_trend(wb, rows: list[dict[str, Any]], months: list[str]):
    ws = wb.create_sheet("전세 추세")
    ncol = 1 + len(months)
    r = add_common_header(ws, "구별 전세 중앙값 추세(억)", "면적필터 59~99m², 월별 중앙값", ncol)
    r = write_headers(ws, r, ["구"] + months)
    for x in rows:
        row_vals = [x["district"]]
        month_map = {m["year_month"]: m for m in x["months"]}
        for ym in months:
            m = month_map.get(ym, {})
            if m.get("median_rent_10k"):
                row_vals.append(round(m["median_rent_10k"] / 10000.0, 2))
            else:
                row_vals.append(None)
        r = write_row(ws, r, row_vals)

    r += 1
    r = write_row(ws, r, ["[최근 전세가율/환산수익률]"] + [""] * (ncol - 1), fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["구", "전세가율(%)", "전세환산 수익률(%)"])
    for x in rows:
        r = write_row(
            ws,
            r,
            [
                x["district"],
                round(x["latest_jeonse_ratio"], 2) if x.get("latest_jeonse_ratio") else "-",
                round(x["latest_rent_yield"], 2) if x.get("latest_rent_yield") else "-",
            ],
        )
    set_widths(ws, [12] + [9] * len(months))


def add_sheet_value(wb, rows: list[dict[str, Any]]):
    ws = wb.create_sheet("밸류에이션")
    r = add_common_header(ws, "상대가치 분석", "서울 중앙값 대비 가격수준/전세가율/전세환산수익률", 10)
    headers = ["구", "최근매매(억)", "최근전세(억)", "서울대비가격배수", "전세가율(%)", "전세환산수익률(%)", "Value점수", "해석", "위험", "메모"]
    r = write_headers(ws, r, headers)

    for x in rows:
        pvs = x.get("price_vs_seoul")
        label = "중립"
        if pvs is not None and pvs < 0.9:
            label = "상대저평가"
        elif pvs is not None and pvs > 1.15:
            label = "상대고평가"

        risk = "보통"
        if (x.get("latest_jeonse_ratio") or 0) > 65:
            risk = "갭리스크"
        if (x.get("latest_jeonse_ratio") or 0) < 30:
            risk = "현금부담"

        r = write_row(
            ws,
            r,
            [
                x["district"],
                round(x["latest_price_eok"], 2) if x.get("latest_price_eok") else "-",
                round(x["latest_rent_eok"], 2) if x.get("latest_rent_eok") else "-",
                round(pvs, 3) if pvs else "-",
                round(x["latest_jeonse_ratio"], 2) if x.get("latest_jeonse_ratio") else "-",
                round(x["latest_rent_yield"], 2) if x.get("latest_rent_yield") else "-",
                x["value_score"],
                label,
                risk,
                "",
            ],
        )
        ws.cell(row=r - 1, column=7).fill = score_band_fill(float(x["value_score"]))

    set_widths(ws, [12, 10, 10, 12, 11, 14, 9, 12, 10, 12])


def add_sheet_liquidity(wb, rows: list[dict[str, Any]]):
    ws = wb.create_sheet("유동성")
    r = add_common_header(ws, "유동성 분석", "최근 거래량/3개월 평균 거래량 중심", 9)
    r = write_headers(ws, r, ["구", "최근매매표본", "3M평균표본", "최근전월세표본", "Liquidity점수", "정책충격흡수", "체결난이도", "권고", "비고"])

    for x in rows:
        liq_score = x["liquidity_score"]
        absorb = "높음" if liq_score >= 70 else "보통" if liq_score >= 55 else "낮음"
        exec_cost = "낮음" if liq_score >= 70 else "보통" if liq_score >= 50 else "높음"
        rec = "분할진입" if liq_score >= 70 else "선별진입" if liq_score >= 55 else "관망"

        r = write_row(
            ws,
            r,
            [
                x["district"],
                x["latest_trade_count"],
                round(x["avg_trade_3m"], 1),
                x["latest_rent_count"],
                liq_score,
                absorb,
                exec_cost,
                rec,
                "",
            ],
        )
        ws.cell(row=r - 1, column=5).fill = score_band_fill(float(liq_score))

    set_widths(ws, [12, 11, 11, 12, 10, 12, 10, 10, 12])


def add_sheet_risk(wb, rows: list[dict[str, Any]]):
    ws = wb.create_sheet("변동성리스크")
    r = add_common_header(ws, "변동성/낙폭 리스크", "연환산 변동성과 최대낙폭으로 하방위험 측정", 10)
    r = write_headers(ws, r, ["구", "12M성장", "3M성장", "변동성(연)", "최대낙폭", "Stability점수", "Policy점수", "종합리스크", "LTPS강도", "코멘트"])

    for x in rows:
        risk_index = 100.0 - (0.65 * x["stability_score"] + 0.35 * x["policy_score"])
        if risk_index <= 35:
            bucket = "낮음"
        elif risk_index <= 55:
            bucket = "보통"
        else:
            bucket = "높음"
        intensity = "집중규제" if x["district"] in HIGH_INTENSITY_LTPS else "일반규제"

        r = write_row(
            ws,
            r,
            [
                x["district"],
                pct(x.get("growth_12m")),
                pct(x.get("growth_3m")),
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                f"{(x.get('max_drawdown') or 0)*100:.1f}%",
                x["stability_score"],
                x["policy_score"],
                round(risk_index, 2),
                intensity,
                bucket,
            ],
        )
        ws.cell(row=r - 1, column=8).fill = score_band_fill(100 - risk_index)

    set_widths(ws, [12, 10, 10, 10, 10, 11, 10, 10, 11, 10])


def add_sheet_benchmark(wb, rows: list[dict[str, Any]]):
    ws = wb.create_sheet("벤치마크")
    r = add_common_header(ws, "벤치마크 및 상대비교", "서울 중앙값/상하위 20% 비교", 10)

    # 상위/하위 20%
    n = len(rows)
    k = max(1, int(round(n * 0.2)))
    top = rows[:k]
    bottom = rows[-k:]

    def avg(items: list[dict[str, Any]], key: str) -> float:
        vals = [float(x[key]) for x in items if x.get(key) is not None]
        return statistics.mean(vals) if vals else 0.0

    r = write_row(ws, r, ["[요약]"] + [""] * 9, fill=FILL_GOLD, left=True)
    summary_rows = [
        ["그룹", "총점", "12M성장", "전세가율", "변동성", "거래량", "Trend", "Value", "Liquidity", "Stability"],
        ["상위20%", round(avg(top, "total_score"), 2), pct(avg(top, "growth_12m")), f"{avg(top, 'latest_jeonse_ratio'):.1f}%", f"{avg(top, 'vol_ann')*100:.1f}%", round(avg(top, "avg_trade_3m"), 1), round(avg(top, "trend_score"), 1), round(avg(top, "value_score"), 1), round(avg(top, "liquidity_score"), 1), round(avg(top, "stability_score"), 1)],
        ["하위20%", round(avg(bottom, "total_score"), 2), pct(avg(bottom, "growth_12m")), f"{avg(bottom, 'latest_jeonse_ratio'):.1f}%", f"{avg(bottom, 'vol_ann')*100:.1f}%", round(avg(bottom, "avg_trade_3m"), 1), round(avg(bottom, "trend_score"), 1), round(avg(bottom, "value_score"), 1), round(avg(bottom, "liquidity_score"), 1), round(avg(bottom, "stability_score"), 1)],
    ]
    r = write_headers(ws, r, summary_rows[0])
    r = write_row(ws, r, summary_rows[1])
    r = write_row(ws, r, summary_rows[2])

    r += 1
    r = write_row(ws, r, ["[강남3구+용산 vs 기타구]"] + [""] * 9, fill=FILL_GOLD, left=True)
    core_names = {"강남구", "서초구", "송파구", "용산구"}
    core = [x for x in rows if x["district"] in core_names]
    non_core = [x for x in rows if x["district"] not in core_names]
    r = write_headers(ws, r, ["그룹", "평균총점", "평균매매(억)", "12M성장", "전세가율", "변동성", "평균거래량", "정책점수", "리스크해석", "비고"])
    for nm, grp in [("강남3구+용산", core), ("기타21구", non_core)]:
        risk = "정책민감" if nm == "강남3구+용산" else "분산형"
        r = write_row(
            ws,
            r,
            [
                nm,
                round(avg(grp, "total_score"), 2),
                round(avg(grp, "latest_price_eok"), 2),
                pct(avg(grp, "growth_12m")),
                f"{avg(grp, 'latest_jeonse_ratio'):.1f}%",
                f"{avg(grp, 'vol_ann')*100:.1f}%",
                round(avg(grp, "avg_trade_3m"), 1),
                round(avg(grp, "policy_score"), 1),
                risk,
                "",
            ],
        )

    set_widths(ws, [14, 11, 12, 10, 10, 10, 10, 10, 12, 10])


def add_sheet_scenario(wb, rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("시나리오 분석")
    r = add_common_header(ws, f"확률기반 시나리오 분석 ({profile})", "P80/P50/P20 분위수 기반 누적수익률 + 레버리지 ROI", 12)
    r = write_headers(ws, r, ["순위", "구", "현재매매(억)", "기간(년)", "낙관누적", "중립누적", "비관누적", "낙관ROI", "중립ROI", "비관ROI", "기준연변동성", "비고"])

    for x in rows:
        sc = x["scenario"]
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                round(x["latest_price_eok"], 2) if x.get("latest_price_eok") else "-",
                sc.get("horizon_years"),
                pct(sc.get("bull_cum")),
                pct(sc.get("base_cum")),
                pct(sc.get("bear_cum")),
                pct(sc.get("bull_roi")),
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
                f"{(sc.get('annual_vol') or 0)*100:.1f}%",
                f"보유수익률 {((sc.get('annual_utility_yield') or 0)*100):.2f}%",
            ],
        )

    set_widths(ws, [7, 12, 11, 8, 10, 10, 10, 10, 10, 10, 11, 12])


def add_sheet_scenario_method(wb, profile: str):
    ws = wb.create_sheet("시나리오 방법론")
    r = add_common_header(ws, "시나리오 산정 방법론", "현실반영형 확률모형 + 정책/레버리지/세금 보정", 6)

    steps = [
        "1) 월별 매매중앙값 시계열로 12M/3M 성장률 및 연환산 변동성 산출",
        "2) 기초성장률 = 0.60*12M + 0.40*3M (과최적화 방지를 위해 상하한 clipping)",
        "3) 연수익률 분포는 로그정규 가정: exp((mu-0.5*sigma^2)+z*sigma)-1",
        "4) z값은 P80/P50/P20 (= +0.8416 / 0 / -0.8416) 사용",
        "5) LTPS 규제강도에 따라 상방/하방 레짐 보정(상방축소, 하방확대)",
        "6) 전세환산 수익률(또는 실거주 효용)을 보유수익으로 반영해 과도한 비관 편향을 완화",
        "7) 프로필별 LTV/대출금리/취득세/보유비용/양도세를 적용해 레버리지 ROI 계산",
        "8) 데이터 품질점수(DataQ)로 표본이 낮은 구의 해석신뢰도 경고",
    ]

    r = write_row(ws, r, ["[방법론 단계]"] + [""] * 5, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["단계", "설명", "수식/로직", "현실보정", "한계", "완화장치"])

    rows = [
        ["Step1", "기초통계", "median(price_t), median(rent_t)", "이상치 완화", "표본부족", "DataQ 경고"],
        ["Step2", "추세 산출", "g_base=0.6*g12+0.4*g3", "최근/중기 균형", "최근왜곡", "clipping"],
        ["Step3", "확률분포", "R=exp((mu-0.5σ²)+zσ)-1", "변동성 반영", "정규가정", "z 완화(P80/P20)"],
        ["Step4", "레짐보정", "R'=R+regime_adj", "규제강도 반영", "정책변화", "정기 갱신"],
        ["Step5", "재무반영", "ROI=(NetGain-이자-세금)/자기자본", "투자자별 현실성", "세법변경", "가정표 분리"],
    ]
    for x in rows:
        r = write_row(ws, r, x, left=True)

    r += 1
    r = write_row(ws, r, ["[프로필 설정]"] + [""] * 5, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["프로필", "핵심목표", "LTV", "대출금리", "취득세", "양도세"])
    for p, cfg in PROFILE_CONFIGS.items():
        r = write_row(
            ws,
            r,
            [
                p,
                cfg["display"],
                f"{cfg['ltv']*100:.0f}%",
                f"{cfg['loan_rate']*100:.2f}%",
                f"{cfg['buy_tax']*100:.2f}%",
                f"{cfg['capital_gain_tax']*100:.1f}%",
            ],
            left=True,
        )

    r += 1
    r = write_row(ws, r, ["[현재 생성 프로필]", profile, "", "", "", ""], fill=FILL_LIGHT, left=True)

    set_widths(ws, [11, 37, 20, 16, 14, 14])


def add_sheet_profile_weights(wb):
    ws = wb.create_sheet("프로필가중치")
    r = add_common_header(ws, "프로필별 멀티전문가 가중치", "실거주/투자자/법인 전략차이를 수치화", 8)
    r = write_headers(ws, r, ["프로필", "Trend", "Value", "Liquidity", "Stability", "Policy", "합계", "전략요약"])

    for p, cfg in PROFILE_CONFIGS.items():
        w = cfg["weights"]
        r = write_row(
            ws,
            r,
            [
                p,
                w["trend"],
                w["value"],
                w["liquidity"],
                w["stability"],
                w["policy"],
                sum(w.values()),
                cfg["display"],
            ],
            left=True,
        )

    set_widths(ws, [10, 8, 8, 10, 10, 8, 8, 32])


def add_sheet_tax_leverage(wb):
    ws = wb.create_sheet("세금레버리지가정")
    r = add_common_header(ws, "세금/레버리지 가정", "프로필별 보수적 가정치", 9)
    r = write_headers(ws, r, ["프로필", "LTV", "대출금리", "취득세", "연보유비용", "양도세", "투자기간(년)", "의미", "주의"])

    notes = {
        "실거주": "거주효용 포함, 세후수익보다 안정성 우선",
        "투자자": "현금흐름·회전율·세후ROI 균형",
        "법인": "세율/취득세 부담이 커 고수익·고확실성 지역 선호",
    }
    warns = {
        "실거주": "거주의무/이주비용 누락 금지",
        "투자자": "DSR/금리변동에 민감",
        "법인": "취득세 12%와 세후수익 급감 리스크",
    }

    for p, cfg in PROFILE_CONFIGS.items():
        r = write_row(
            ws,
            r,
            [
                p,
                f"{cfg['ltv']*100:.0f}%",
                f"{cfg['loan_rate']*100:.2f}%",
                f"{cfg['buy_tax']*100:.2f}%",
                f"{cfg['annual_hold_cost']*100:.2f}%",
                f"{cfg['capital_gain_tax']*100:.1f}%",
                cfg["horizon_years"],
                notes[p],
                warns[p],
            ],
            left=True,
        )

    set_widths(ws, [10, 8, 10, 9, 11, 8, 10, 28, 24])


def add_sheet_policy(wb, rows: list[dict[str, Any]]):
    ws = wb.create_sheet("정책규제체크")
    r = add_common_header(ws, "정책/규제 체크", "LTPS·금리·DSR 최신 공개자료 반영", 9)

    r = write_row(ws, r, ["[정책 이벤트]"] + [""] * 8, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["구분", "내용", "기준일", "시장영향", "점수반영", "리스크", "출처", "링크", "비고"])

    policy_rows = [
        ["LTPS", "서울시 전체 아파트 토지거래허가구역", "2025-10-20~2026-12-31", "유동성/레버리지 제약", "Policy score base=45", "중", "서울시", "https://land.seoul.go.kr/land/other/appointStatusSeoul.do", ""],
        ["LTPS", "강남·서초·송파·용산 집중규제", "2026-01-07 고시", "상방축소/하방확대", "집중규제구 -15", "중~상", "서울시", "https://land.seoul.go.kr/land/other/appointStatusSeoul.do", ""],
        ["금리", "한국은행 기준금리 2.50% 동결", "2026-01-15", "차입비용 기준선", "loan_rate anchor", "중", "한국은행", "https://www.bok.or.kr/portal/bbs/B0000169/view.do?menuNo=200059&nttId=10095737", ""],
        ["대출", "3단계 스트레스 DSR", "2025-07 시행", "대출가능액 축소", "LTV/금리 보수화", "중", "금융위원회", "https://www.fsc.go.kr/no010101/84617", ""],
    ]
    for p in policy_rows:
        r = write_row(ws, r, p, left=True)

    r += 1
    r = write_row(ws, r, ["[구별 규제강도 매핑]"] + [""] * 8, fill=FILL_GOLD, left=True)
    r = write_headers(ws, r, ["구", "정책점수", "강도", "집중규제", "유동성보정", "데이터품질", "권고", "메모", ""])
    for x in rows:
        intensity = "집중" if x["district"] in HIGH_INTENSITY_LTPS else "일반"
        rec = "보수진입" if intensity == "집중" else "분할검토"
        r = write_row(
            ws,
            r,
            [
                x["district"],
                x["policy_score"],
                intensity,
                "Y" if intensity == "집중" else "N",
                round(x["avg_trade_3m"], 1),
                x["data_quality"],
                rec,
                "",
                "",
            ],
        )

    set_widths(ws, [10, 9, 8, 9, 10, 10, 12, 15, 8])


def add_sheet_timing(wb, rows: list[dict[str, Any]], profile: str):
    ws = wb.create_sheet("매수타이밍신호")
    r = add_common_header(ws, f"매수 타이밍 신호 ({profile})", "추세·유동성·안정성 결합신호", 11)
    r = write_headers(ws, r, ["순위", "구", "타이밍점수", "Trend", "Liquidity", "Stability", "최근3M", "12M", "신호", "전략", "주의"])

    for x in rows:
        timing = 0.50 * x["trend_score"] + 0.30 * x["liquidity_score"] + 0.20 * x["stability_score"]
        if timing >= 70:
            sig = "진입 우호"
            strat = "3회 분할매수"
        elif timing >= 55:
            sig = "중립"
            strat = "조건부 진입"
        else:
            sig = "관망"
            strat = "지표 개선 대기"

        caution = "정책민감" if x["district"] in HIGH_INTENSITY_LTPS else "일반"
        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                round(timing, 2),
                x["trend_score"],
                x["liquidity_score"],
                x["stability_score"],
                pct(x.get("growth_3m")),
                pct(x.get("growth_12m")),
                sig,
                strat,
                caution,
            ],
        )
        ws.cell(row=r - 1, column=3).fill = score_band_fill(float(timing))

    set_widths(ws, [7, 11, 10, 8, 9, 9, 9, 9, 10, 12, 9])


def add_sheet_profile_recommendation(wb, rows: list[dict[str, Any]], profile: str):
    ws_name = {
        "실거주": "실거주 추천",
        "투자자": "투자자 추천",
        "법인": "법인 추천",
    }[profile]
    ws = wb.create_sheet(ws_name)
    r = add_common_header(ws, f"{profile} 관점 Top 10", PROFILE_CONFIGS[profile]["display"], 11)
    r = write_headers(ws, r, ["순위", "구", "총점", "중립ROI", "비관ROI", "전세가율", "변동성", "정책", "추천전략", "핵심리스크", "한줄코멘트"])

    top = rows[:10]
    for x in top:
        sc = x["scenario"]
        if profile == "실거주":
            strategy = "실거주+자산방어"
            risk = "생활권/이주비"
        elif profile == "투자자":
            strategy = "분할매수+회전관리"
            risk = "금리/레버리지"
        else:
            strategy = "세후ROI 선별진입"
            risk = "취득세/법인세"

        r = write_row(
            ws,
            r,
            [
                x["rank"],
                x["district"],
                x["total_score"],
                pct(sc.get("base_roi")),
                pct(sc.get("bear_roi")),
                f"{(x.get('latest_jeonse_ratio') or 0):.1f}%" if x.get("latest_jeonse_ratio") else "-",
                f"{(x.get('vol_ann') or 0)*100:.1f}%",
                x["policy_score"],
                strategy,
                risk,
                "",
            ],
        )

    set_widths(ws, [7, 11, 9, 10, 10, 10, 9, 8, 14, 12, 12])


def add_sheet_references(wb):
    ws = wb.create_sheet("참고문헌")
    r = add_common_header(ws, "근거자료(공식/논문/커뮤니티)", "본 분석의 현실반영 근거", 6)
    r = write_headers(ws, r, ["분류", "자료명", "기준시점", "활용목적", "URL", "비고"])
    for s in SOURCES:
        r = write_row(
            ws,
            r,
            [
                s["category"],
                s["title"],
                s["as_of"],
                s["usage"],
                s["url"],
                "",
            ],
            left=True,
        )
    set_widths(ws, [10, 40, 12, 42, 65, 10])


def add_sheet_checklist(wb, profile: str):
    ws = wb.create_sheet("실행체크리스트")
    r = add_common_header(ws, f"실행 체크리스트 ({profile})", "실제 투자 실행 전 필수 검증", 6)
    r = write_headers(ws, r, ["No", "체크항목", "왜 중요한가", "데이터/문서", "판단기준", "완료"])

    checklist = [
        [1, "자금계획", "LTV/DSR 충족 여부", "대출사전심사", "DSR 허용범위", ""],
        [2, "세금시뮬레이션", "세후수익 급변 방지", "세무사 검토", "보수세율 적용", ""],
        [3, "정책확인", "허가구역/규제 변경", "서울시/금융위 공지", "최근 공고 기준", ""],
        [4, "표본신뢰도", "왜곡 방지", "DataQ/표본수", "DataQ 60 이상 권장", ""],
        [5, "시나리오점검", "하방손실 관리", "Bear ROI", "손실허용범위 내", ""],
        [6, "유동성", "매도/갈아타기 가능성", "최근 거래량", "상위 50% 권장", ""],
        [7, "현장검증", "입지/소음/학군 체감", "현장방문", "핵심 하자 무", ""],
        [8, "출구전략", "매도 타이밍", "목표수익률/기간", "사전 규칙화", ""],
    ]
    for row in checklist:
        r = write_row(ws, r, row, left=True)

    set_widths(ws, [6, 14, 24, 22, 20, 8])


def build_workbook(
    profile_rows: list[dict[str, Any]],
    profile: str,
    months: list[str],
    rows_by_profile: dict[str, list[dict[str, Any]]] | None = None,
) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    add_sheet_dashboard(wb, profile_rows, profile, months)
    add_sheet_scorecard(wb, profile_rows, profile, months)
    add_sheet_coverage(wb, profile_rows, months)
    add_sheet_trade_trend(wb, profile_rows, months)
    add_sheet_rent_trend(wb, profile_rows, months)
    add_sheet_value(wb, profile_rows)
    add_sheet_liquidity(wb, profile_rows)
    add_sheet_risk(wb, profile_rows)
    add_sheet_benchmark(wb, profile_rows)
    add_sheet_scenario(wb, profile_rows, profile)
    add_sheet_scenario_method(wb, profile)
    add_sheet_profile_weights(wb)
    add_sheet_tax_leverage(wb)
    add_sheet_policy(wb, profile_rows)
    add_sheet_timing(wb, profile_rows, profile)

    # 프로필별 추천 시트 3개를 모두 포함(비교 가능)
    rows_by_profile = rows_by_profile or {profile: profile_rows}
    for p in ["실거주", "투자자", "법인"]:
        src = rows_by_profile.get(p) or profile_rows
        add_sheet_profile_recommendation(wb, src, p)

    add_sheet_references(wb)
    add_sheet_checklist(wb, profile)

    return wb


# ---------------------------------------------------------------------------
# Cache and main flow
# ---------------------------------------------------------------------------


def save_cache(path: Path, payload: dict[str, Any]):
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_cache(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def serialize_panel(panel: dict[str, dict[str, dict[str, Any]]]) -> dict[str, Any]:
    return panel


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="서울 전체 아파트 심층분석 엑셀 생성기")
    p.add_argument("--months", type=int, default=12, help="수집 월 수 (기본 12)")
    p.add_argument("--end-yyyymm", default=now_yyyymm(), help="종료월 YYYYMM (기본: 현재월)")
    p.add_argument("--max-workers", type=int, default=16, help="병렬 작업 수")
    p.add_argument("--num-rows", type=int, default=1000, help="API 페이지 행수")
    p.add_argument("--area-min", type=float, default=DEFAULT_AREA_MIN)
    p.add_argument("--area-max", type=float, default=DEFAULT_AREA_MAX)
    p.add_argument("--deposit-yield-pct", type=float, default=DEFAULT_DEPOSIT_YIELD_PCT)
    p.add_argument("--api-key", default="", help="data.go.kr API key (없으면 env/내장키 사용)")
    p.add_argument("--refresh-cache", action="store_true", help="캐시 무시하고 재수집")
    p.add_argument("--profile", choices=list(PROFILE_CONFIGS.keys()), default="투자자")
    p.add_argument("--all-profiles", action="store_true", help="3개 프로필 모두 생성")
    p.add_argument("--output-dir", default="C:/Devs/KRE")
    p.add_argument(
        "--region-code-file",
        default="C:/Devs/KRE/real-estate-mcp/src/real_estate/resources/region_codes.txt",
        help="region_codes.txt 경로",
    )
    return p.parse_args()


def main() -> int:
    args = parse_args()

    if args.months < 6:
        raise SystemExit("months는 최소 6 이상을 권장합니다.")
    if args.area_min >= args.area_max:
        raise SystemExit("area-min은 area-max보다 작아야 합니다.")

    api_key = get_api_key(args.api_key)
    if not api_key:
        raise SystemExit("API key를 찾지 못했습니다.")

    months = iter_months(args.end_yyyymm, args.months)
    region_code_file = Path(args.region_code_file)
    districts = load_seoul_gu_list(region_code_file)

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    cache_path = out_dir / f"seoul_market_cache_{months[0]}_{months[-1]}_{int(args.area_min)}_{int(args.area_max)}.json"

    if cache_path.exists() and not args.refresh_cache:
        print(f"[cache] 로드: {cache_path}")
        cache = load_cache(cache_path)
        panel = cache["panel"]
        logs = cache.get("logs", [])
    else:
        print(f"[collect] 서울 25개구 x {len(months)}개월 수집 시작")
        panel, logs = collect_market_panel(
            districts=districts,
            months=months,
            api_key=api_key,
            num_rows=args.num_rows,
            max_workers=args.max_workers,
            area_min=args.area_min,
            area_max=args.area_max,
            deposit_yield_pct=args.deposit_yield_pct,
        )
        cache = {
            "generated_at": dt.datetime.now().isoformat(),
            "months": months,
            "area_filter": [args.area_min, args.area_max],
            "panel": serialize_panel(panel),
            "logs": logs,
        }
        save_cache(cache_path, cache)
        print(f"[cache] 저장: {cache_path}")

    print("[analyze] 지표 계산")
    metrics = build_district_metrics(districts, months, panel)
    metrics = run_multi_expert_scoring(metrics)

    profiles = ["실거주", "투자자", "법인"] if args.all_profiles else [args.profile]
    generated: list[Path] = []

    rows_by_profile: dict[str, list[dict[str, Any]]] = {}
    for p in ["실거주", "투자자", "법인"]:
        rows_p = score_by_profile(metrics, p)
        rows_by_profile[p] = attach_scenarios(rows_p, p)

    for profile in profiles:
        rows = rows_by_profile[profile]
        wb = build_workbook(rows, profile, months, rows_by_profile=rows_by_profile)
        out_path = out_dir / f"서울_전체_아파트_심층분석_v1_{profile}.xlsx"
        wb.save(out_path)
        generated.append(out_path)
        print(f"[done] {profile}: {out_path}")

    # 통합본(투자자 기준)도 추가 생성
    rows_i = rows_by_profile["투자자"]
    wb_i = build_workbook(rows_i, "투자자", months, rows_by_profile=rows_by_profile)
    master_path = out_dir / "서울_전체_아파트_심층분석_v1_통합.xlsx"
    wb_i.save(master_path)
    generated.append(master_path)
    print(f"[done] 통합본: {master_path}")

    print("\n생성 완료 파일:")
    for p in generated:
        print(f"- {p}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
