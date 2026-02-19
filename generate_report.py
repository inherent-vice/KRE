import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from statistics import median

wb = Workbook()

# === 데이터 ===
trades_202602 = [{"apt_name":"현대1차","dong":"거여동","area_sqm":84.64,"floor":11,"price_10k":145000,"trade_date":"2026-02-04","build_year":1992,"deal_type":"직거래"},{"apt_name":"대림","dong":"방이동","area_sqm":125.075,"floor":5,"price_10k":265000,"trade_date":"2026-02-09","build_year":1985,"deal_type":"중개거래"},{"apt_name":"거여1단지","dong":"거여동","area_sqm":39.6,"floor":2,"price_10k":80000,"trade_date":"2026-02-06","build_year":1997,"deal_type":"중개거래"},{"apt_name":"문정가락현대1차","dong":"문정동","area_sqm":84.98,"floor":4,"price_10k":199000,"trade_date":"2026-02-02","build_year":1984,"deal_type":"중개거래"},{"apt_name":"장미2","dong":"신천동","area_sqm":120,"floor":9,"price_10k":380000,"trade_date":"2026-02-11","build_year":1979,"deal_type":"중개거래"},{"apt_name":"테라펠리체송파","dong":"가락동","area_sqm":44.0351,"floor":8,"price_10k":82000,"trade_date":"2026-02-11","build_year":2021,"deal_type":"중개거래"},{"apt_name":"대림","dong":"오금동","area_sqm":125.79,"floor":6,"price_10k":250000,"trade_date":"2026-02-07","build_year":1988,"deal_type":"중개거래"},{"apt_name":"파크리오","dong":"신천동","area_sqm":121.63,"floor":17,"price_10k":372000,"trade_date":"2026-02-05","build_year":2008,"deal_type":"중개거래"},{"apt_name":"거여4단지","dong":"거여동","area_sqm":59.73,"floor":10,"price_10k":126000,"trade_date":"2026-02-09","build_year":1997,"deal_type":"중개거래"},{"apt_name":"주공아파트 5단지","dong":"잠실동","area_sqm":76.5,"floor":4,"price_10k":424700,"trade_date":"2026-02-07","build_year":1978,"deal_type":"중개거래"},{"apt_name":"주성파크빌","dong":"가락동","area_sqm":67.03,"floor":5,"price_10k":64000,"trade_date":"2026-02-06","build_year":2002,"deal_type":"중개거래"},{"apt_name":"현대2차","dong":"거여동","area_sqm":82.96,"floor":12,"price_10k":131000,"trade_date":"2026-02-04","build_year":1992,"deal_type":"중개거래"},{"apt_name":"송파파밀리에더퍼스트","dong":"송파동","area_sqm":59.935,"floor":5,"price_10k":150000,"trade_date":"2026-02-03","build_year":2023,"deal_type":"중개거래"},{"apt_name":"대림","dong":"방이동","area_sqm":84.915,"floor":15,"price_10k":249300,"trade_date":"2026-02-07","build_year":1985,"deal_type":"중개거래"},{"apt_name":"헬리오시티","dong":"가락동","area_sqm":150.09,"floor":29,"price_10k":450000,"trade_date":"2026-02-09","build_year":2018,"deal_type":"중개거래"},{"apt_name":"미성","dong":"풍납동","area_sqm":116.97,"floor":7,"price_10k":173000,"trade_date":"2026-02-06","build_year":1985,"deal_type":"중개거래"},{"apt_name":"트윈레이크팰리스빌","dong":"송파동","area_sqm":28.11,"floor":5,"price_10k":36500,"trade_date":"2026-02-07","build_year":2018,"deal_type":"중개거래"},{"apt_name":"동아한가람1","dong":"풍납동","area_sqm":84.38,"floor":9,"price_10k":149900,"trade_date":"2026-02-09","build_year":1995,"deal_type":"중개거래"},{"apt_name":"우성","dong":"가락동","area_sqm":109.25,"floor":15,"price_10k":211000,"trade_date":"2026-02-09","build_year":1986,"deal_type":"중개거래"},{"apt_name":"현대리버빌2지구","dong":"풍납동","area_sqm":84.97,"floor":8,"price_10k":149500,"trade_date":"2026-02-08","build_year":1999,"deal_type":"중개거래"},{"apt_name":"문정시영","dong":"문정동","area_sqm":39.69,"floor":10,"price_10k":103000,"trade_date":"2026-02-07","build_year":1989,"deal_type":"중개거래"},{"apt_name":"잠실더샵스타파크","dong":"신천동","area_sqm":173.17,"floor":28,"price_10k":283000,"trade_date":"2026-02-03","build_year":2008,"deal_type":"중개거래"},{"apt_name":"미켈란호수가","dong":"석촌동","area_sqm":129.45,"floor":15,"price_10k":215000,"trade_date":"2026-02-07","build_year":2007,"deal_type":"중개거래"},{"apt_name":"문정시영","dong":"문정동","area_sqm":35.46,"floor":11,"price_10k":80000,"trade_date":"2026-02-04","build_year":1989,"deal_type":"중개거래"},{"apt_name":"현대백조","dong":"오금동","area_sqm":84.92,"floor":7,"price_10k":149500,"trade_date":"2026-02-01","build_year":1997,"deal_type":"중개거래"},{"apt_name":"e편한세상송파파크센트럴","dong":"거여동","area_sqm":59.96,"floor":23,"price_10k":180000,"trade_date":"2026-02-04","build_year":2021,"deal_type":"중개거래"},{"apt_name":"송파시그니처롯데캐슬","dong":"거여동","area_sqm":84.98,"floor":22,"price_10k":212500,"trade_date":"2026-02-07","build_year":2021,"deal_type":"중개거래"},{"apt_name":"선경쉐르빌(106-23)","dong":"문정동","area_sqm":76.47,"floor":7,"price_10k":73000,"trade_date":"2026-02-07","build_year":2002,"deal_type":"중개거래"},{"apt_name":"리센츠","dong":"잠실동","area_sqm":27.68,"floor":15,"price_10k":158000,"trade_date":"2026-02-07","build_year":2008,"deal_type":"중개거래"},{"apt_name":"송파위례리슈빌","dong":"거여동","area_sqm":105.4624,"floor":9,"price_10k":215000,"trade_date":"2026-02-04","build_year":2021,"deal_type":"중개거래"},{"apt_name":"한강극동","dong":"풍납동","area_sqm":114.76,"floor":5,"price_10k":167000,"trade_date":"2026-02-06","build_year":1995,"deal_type":"중개거래"},{"apt_name":"쌍용스윗닷홈(105동)","dong":"가락동","area_sqm":122.43,"floor":3,"price_10k":162000,"trade_date":"2026-02-04","build_year":2005,"deal_type":"중개거래"},{"apt_name":"가락현대투웰브(101동)","dong":"가락동","area_sqm":105.02,"floor":7,"price_10k":120000,"trade_date":"2026-02-05","build_year":2005,"deal_type":"중개거래"},{"apt_name":"현대리버빌2지구","dong":"풍납동","area_sqm":40.95,"floor":17,"price_10k":67500,"trade_date":"2026-02-03","build_year":1999,"deal_type":"중개거래"},{"apt_name":"송파호반베르디움더퍼스트","dong":"오금동","area_sqm":101.4632,"floor":19,"price_10k":167000,"trade_date":"2026-02-03","build_year":2018,"deal_type":"중개거래"},{"apt_name":"마천금호어울림1차아파트","dong":"마천동","area_sqm":114.81,"floor":2,"price_10k":110000,"trade_date":"2026-02-02","build_year":2002,"deal_type":"중개거래"},{"apt_name":"이트리움송파(281-4)","dong":"풍납동","area_sqm":84.7,"floor":5,"price_10k":84645,"trade_date":"2026-02-05","build_year":2024,"deal_type":"직거래"},{"apt_name":"현대(2-4차)","dong":"오금동","area_sqm":84.98,"floor":5,"price_10k":260000,"trade_date":"2026-02-04","build_year":1984,"deal_type":"중개거래"},{"apt_name":"우성","dong":"가락동","area_sqm":43.79,"floor":9,"price_10k":159000,"trade_date":"2026-02-03","build_year":1986,"deal_type":"중개거래"},{"apt_name":"잠실엘스","dong":"잠실동","area_sqm":84.8,"floor":6,"price_10k":345000,"trade_date":"2026-02-04","build_year":2008,"deal_type":"중개거래"},{"apt_name":"송파레이크파크호반써밋Ⅱ","dong":"거여동","area_sqm":108.8119,"floor":11,"price_10k":226000,"trade_date":"2026-02-04","build_year":2022,"deal_type":"중개거래"},{"apt_name":"잠실아이스페이스","dong":"신천동","area_sqm":76.55,"floor":22,"price_10k":120000,"trade_date":"2026-02-03","build_year":2001,"deal_type":"직거래"},{"apt_name":"극동","dong":"풍납동","area_sqm":79.945,"floor":2,"price_10k":165000,"trade_date":"2026-02-01","build_year":1987,"deal_type":"중개거래"}]

trades_202601 = [{"apt_name":"문정시영","dong":"문정동","area_sqm":39.69,"floor":14,"price_10k":100000,"trade_date":"2026-01-31","build_year":1989,"deal_type":"중개거래"},{"apt_name":"송파레이크파크호반써밋Ⅰ","dong":"거여동","area_sqm":108.8119,"floor":6,"price_10k":195700,"trade_date":"2026-01-28","build_year":2022,"deal_type":"중개거래"},{"apt_name":"우성아파트","dong":"잠실동","area_sqm":80.35,"floor":5,"price_10k":301500,"trade_date":"2026-01-24","build_year":1981,"deal_type":"중개거래"},{"apt_name":"한강극동","dong":"풍납동","area_sqm":114.76,"floor":2,"price_10k":173000,"trade_date":"2026-01-31","build_year":1995,"deal_type":"중개거래"},{"apt_name":"장미2","dong":"신천동","area_sqm":71.2,"floor":11,"price_10k":301000,"trade_date":"2026-01-23","build_year":1979,"deal_type":"중개거래"},{"apt_name":"위례24단지(꿈에그린)","dong":"장지동","area_sqm":51.77,"floor":4,"price_10k":150000,"trade_date":"2026-01-28","build_year":2013,"deal_type":"중개거래"},{"apt_name":"거여4단지","dong":"거여동","area_sqm":59.73,"floor":3,"price_10k":130000,"trade_date":"2026-01-31","build_year":1997,"deal_type":"중개거래"},{"apt_name":"건영1","dong":"문정동","area_sqm":84.39,"floor":8,"price_10k":182000,"trade_date":"2026-01-28","build_year":1996,"deal_type":"중개거래"},{"apt_name":"동아한가람1","dong":"풍납동","area_sqm":84.38,"floor":15,"price_10k":154000,"trade_date":"2026-01-12","build_year":1995,"deal_type":"중개거래"},{"apt_name":"신아","dong":"풍납동","area_sqm":130.05,"floor":8,"price_10k":85500,"trade_date":"2026-01-23","build_year":1999,"deal_type":"중개거래"},{"apt_name":"방이금호어울림","dong":"방이동","area_sqm":84.3,"floor":3,"price_10k":119000,"trade_date":"2026-01-29","build_year":2001,"deal_type":"중개거래"},{"apt_name":"신동아","dong":"마천동","area_sqm":59.88,"floor":8,"price_10k":64000,"trade_date":"2026-01-16","build_year":2001,"deal_type":"중개거래"},{"apt_name":"파크리오","dong":"신천동","area_sqm":144.77,"floor":32,"price_10k":363000,"trade_date":"2026-01-14","build_year":2008,"deal_type":"중개거래"},{"apt_name":"호수임광","dong":"송파동","area_sqm":57.72,"floor":10,"price_10k":107000,"trade_date":"2026-01-17","build_year":1995,"deal_type":"중개거래"},{"apt_name":"주공아파트 5단지","dong":"잠실동","area_sqm":76.5,"floor":1,"price_10k":402700,"trade_date":"2026-01-21","build_year":1978,"deal_type":"중개거래"},{"apt_name":"파크하비오","dong":"문정동","area_sqm":84.9725,"floor":14,"price_10k":188000,"trade_date":"2026-01-25","build_year":2016,"deal_type":"중개거래"},{"apt_name":"현대리버빌1지구","dong":"풍납동","area_sqm":43.2,"floor":3,"price_10k":84000,"trade_date":"2026-01-22","build_year":1999,"deal_type":"중개거래"},{"apt_name":"갤러리아팰리스","dong":"잠실동","area_sqm":163.935,"floor":41,"price_10k":376000,"trade_date":"2026-01-20","build_year":2005,"deal_type":"중개거래"},{"apt_name":"리센츠","dong":"잠실동","area_sqm":124.22,"floor":18,"price_10k":470000,"trade_date":"2026-01-26","build_year":2008,"deal_type":"중개거래"},{"apt_name":"거여1단지","dong":"거여동","area_sqm":49.8,"floor":4,"price_10k":110000,"trade_date":"2026-01-31","build_year":1997,"deal_type":"중개거래"},{"apt_name":"송파두산위브","dong":"오금동","area_sqm":59.9823,"floor":11,"price_10k":130000,"trade_date":"2026-01-30","build_year":2019,"deal_type":"중개거래"},{"apt_name":"강변현대","dong":"풍납동","area_sqm":83.19,"floor":5,"price_10k":116500,"trade_date":"2026-01-13","build_year":1991,"deal_type":"중개거래"},{"apt_name":"주공아파트 5단지","dong":"잠실동","area_sqm":76.5,"floor":6,"price_10k":419800,"trade_date":"2026-01-28","build_year":1978,"deal_type":"중개거래"},{"apt_name":"올림픽선수기자촌3단지","dong":"방이동","area_sqm":83.06,"floor":6,"price_10k":300000,"trade_date":"2026-01-22","build_year":1988,"deal_type":"중개거래"},{"apt_name":"파크하비오","dong":"문정동","area_sqm":84.9342,"floor":4,"price_10k":185000,"trade_date":"2026-01-06","build_year":2016,"deal_type":"중개거래"},{"apt_name":"쌍용스윗닷홈2차","dong":"거여동","area_sqm":104.14,"floor":8,"price_10k":129500,"trade_date":"2026-01-22","build_year":2006,"deal_type":"중개거래"},{"apt_name":"거여4단지","dong":"거여동","area_sqm":39.6,"floor":2,"price_10k":72000,"trade_date":"2026-01-16","build_year":1997,"deal_type":"중개거래"},{"apt_name":"송파파인타운8단지","dong":"장지동","area_sqm":59.97,"floor":9,"price_10k":137250,"trade_date":"2026-01-31","build_year":2008,"deal_type":"중개거래"},{"apt_name":"잠실엘스","dong":"잠실동","area_sqm":119.93,"floor":13,"price_10k":460000,"trade_date":"2026-01-16","build_year":2008,"deal_type":"중개거래"},{"apt_name":"위례24단지(꿈에그린)","dong":"장지동","area_sqm":51.78,"floor":3,"price_10k":155000,"trade_date":"2026-01-29","build_year":2013,"deal_type":"중개거래"},{"apt_name":"송파파인타운3단지","dong":"장지동","area_sqm":84.94,"floor":10,"price_10k":185000,"trade_date":"2026-01-15","build_year":2008,"deal_type":"중개거래"},{"apt_name":"가락삼익맨숀","dong":"송파동","area_sqm":108.338,"floor":12,"price_10k":270000,"trade_date":"2026-01-17","build_year":1984,"deal_type":"중개거래"},{"apt_name":"우성아파트","dong":"잠실동","area_sqm":96.65,"floor":9,"price_10k":330000,"trade_date":"2026-01-24","build_year":1981,"deal_type":"중개거래"},{"apt_name":"거여1단지","dong":"거여동","area_sqm":39.6,"floor":2,"price_10k":89950,"trade_date":"2026-01-31","build_year":1997,"deal_type":"중개거래"},{"apt_name":"트리지움","dong":"잠실동","area_sqm":84.95,"floor":4,"price_10k":319000,"trade_date":"2026-01-30","build_year":2007,"deal_type":"중개거래"},{"apt_name":"예명","dong":"석촌동","area_sqm":68.93,"floor":6,"price_10k":106000,"trade_date":"2026-01-27","build_year":2001,"deal_type":"중개거래"},{"apt_name":"테라펠리체송파","dong":"가락동","area_sqm":44.0351,"floor":7,"price_10k":80500,"trade_date":"2026-01-31","build_year":2021,"deal_type":"중개거래"},{"apt_name":"현대1","dong":"풍납동","area_sqm":84,"floor":8,"price_10k":181000,"trade_date":"2026-01-09","build_year":1995,"deal_type":"중개거래"},{"apt_name":"신성노바빌아파트","dong":"풍납동","area_sqm":84.96,"floor":6,"price_10k":134500,"trade_date":"2026-01-31","build_year":2000,"deal_type":"중개거래"},{"apt_name":"문정시영","dong":"문정동","area_sqm":35.46,"floor":5,"price_10k":76800,"trade_date":"2026-01-28","build_year":1989,"deal_type":"중개거래"},{"apt_name":"파크리오","dong":"신천동","area_sqm":84.97,"floor":9,"price_10k":318000,"trade_date":"2026-01-31","build_year":2008,"deal_type":"중개거래"},{"apt_name":"이스턴잠실","dong":"방이동","area_sqm":23.93,"floor":6,"price_10k":37000,"trade_date":"2026-01-31","build_year":2020,"deal_type":"중개거래"},{"apt_name":"위례24단지(꿈에그린)","dong":"장지동","area_sqm":84.96,"floor":1,"price_10k":195000,"trade_date":"2026-01-31","build_year":2013,"deal_type":"중개거래"},{"apt_name":"송파더센트레","dong":"장지동","area_sqm":51.77,"floor":1,"price_10k":143000,"trade_date":"2026-01-31","build_year":2013,"deal_type":"중개거래"},{"apt_name":"송파레이크파크호반써밋Ⅰ","dong":"거여동","area_sqm":108.8119,"floor":11,"price_10k":204000,"trade_date":"2026-01-30","build_year":2022,"deal_type":"중개거래"},{"apt_name":"파크리오","dong":"신천동","area_sqm":84.9,"floor":17,"price_10k":298000,"trade_date":"2026-01-30","build_year":2008,"deal_type":"중개거래"},{"apt_name":"동아한가람1","dong":"풍납동","area_sqm":59.98,"floor":3,"price_10k":118000,"trade_date":"2026-01-27","build_year":1995,"deal_type":"중개거래"},{"apt_name":"잠실엘스","dong":"잠실동","area_sqm":59.96,"floor":9,"price_10k":297500,"trade_date":"2026-01-31","build_year":2008,"deal_type":"중개거래"},{"apt_name":"대림","dong":"오금동","area_sqm":125.79,"floor":10,"price_10k":252000,"trade_date":"2026-01-30","build_year":1988,"deal_type":"중개거래"},{"apt_name":"대림","dong":"가락동","area_sqm":76.49,"floor":9,"price_10k":153000,"trade_date":"2026-01-24","build_year":1988,"deal_type":"중개거래"},{"apt_name":"대림","dong":"가락동","area_sqm":59.24,"floor":10,"price_10k":131000,"trade_date":"2026-01-29","build_year":1988,"deal_type":"중개거래"},{"apt_name":"송파파인타운6단지","dong":"장지동","area_sqm":84.98,"floor":5,"price_10k":179000,"trade_date":"2026-01-20","build_year":2008,"deal_type":"중개거래"},{"apt_name":"파크하비오","dong":"문정동","area_sqm":84.9725,"floor":9,"price_10k":210000,"trade_date":"2026-01-30","build_year":2016,"deal_type":"중개거래"},{"apt_name":"동신","dong":"삼전동","area_sqm":61.04,"floor":4,"price_10k":75000,"trade_date":"2026-01-17","build_year":2003,"deal_type":"중개거래"},{"apt_name":"호수임광","dong":"송파동","area_sqm":56.6,"floor":2,"price_10k":104500,"trade_date":"2026-01-30","build_year":1995,"deal_type":"중개거래"},{"apt_name":"송파파인타운4단지","dong":"장지동","area_sqm":84.95,"floor":12,"price_10k":199500,"trade_date":"2026-01-08","build_year":2008,"deal_type":"중개거래"},{"apt_name":"우성아파트","dong":"잠실동","area_sqm":96.65,"floor":7,"price_10k":250000,"trade_date":"2026-01-15","build_year":1981,"deal_type":"직거래"},{"apt_name":"삼익1","dong":"마천동","area_sqm":84.9,"floor":9,"price_10k":120000,"trade_date":"2026-01-29","build_year":1996,"deal_type":"중개거래"},{"apt_name":"레이크팰리스","dong":"잠실동","area_sqm":84.82,"floor":9,"price_10k":330000,"trade_date":"2026-01-21","build_year":2006,"deal_type":"중개거래"},{"apt_name":"건영1","dong":"문정동","area_sqm":77.68,"floor":1,"price_10k":162000,"trade_date":"2026-01-28","build_year":1996,"deal_type":"중개거래"},{"apt_name":"대림","dong":"오금동","area_sqm":125.79,"floor":9,"price_10k":259000,"trade_date":"2026-01-05","build_year":1988,"deal_type":"중개거래"},{"apt_name":"우성","dong":"가락동","area_sqm":43.79,"floor":10,"price_10k":140000,"trade_date":"2026-01-28","build_year":1986,"deal_type":"중개거래"},{"apt_name":"장미1","dong":"신천동","area_sqm":120,"floor":6,"price_10k":389000,"trade_date":"2026-01-11","build_year":1979,"deal_type":"중개거래"},{"apt_name":"리센츠","dong":"잠실동","area_sqm":84.99,"floor":4,"price_10k":330000,"trade_date":"2026-01-24","build_year":2008,"deal_type":"중개거래"},{"apt_name":"문정시영","dong":"문정동","area_sqm":39.69,"floor":4,"price_10k":98000,"trade_date":"2026-01-28","build_year":1989,"deal_type":"중개거래"},{"apt_name":"석촌호수효성해링턴타워","dong":"석촌동","area_sqm":15.8307,"floor":11,"price_10k":25700,"trade_date":"2026-01-28","build_year":2015,"deal_type":"중개거래"},{"apt_name":"파크하비오","dong":"문정동","area_sqm":84.9342,"floor":5,"price_10k":188500,"trade_date":"2026-01-09","build_year":2016,"deal_type":"중개거래"},{"apt_name":"쌍용","dong":"풍납동","area_sqm":84.86,"floor":12,"price_10k":144000,"trade_date":"2026-01-07","build_year":1994,"deal_type":"중개거래"},{"apt_name":"헬리오시티","dong":"가락동","area_sqm":84.99,"floor":11,"price_10k":314000,"trade_date":"2026-01-02","build_year":2018,"deal_type":"중개거래"},{"apt_name":"현대3차","dong":"거여동","area_sqm":84.94,"floor":12,"price_10k":124800,"trade_date":"2026-01-24","build_year":1993,"deal_type":"중개거래"},{"apt_name":"가락3차쌍용스윗닷홈101동,102동","dong":"가락동","area_sqm":59.94,"floor":7,"price_10k":148000,"trade_date":"2026-01-28","build_year":2005,"deal_type":"중개거래"},{"apt_name":"헬리오시티","dong":"가락동","area_sqm":84.99,"floor":12,"price_10k":300000,"trade_date":"2026-01-13","build_year":2018,"deal_type":"중개거래"},{"apt_name":"리센츠","dong":"잠실동","area_sqm":124.22,"floor":7,"price_10k":410000,"trade_date":"2026-01-24","build_year":2008,"deal_type":"중개거래"},{"apt_name":"잠실엘스","dong":"잠실동","area_sqm":84.8,"floor":13,"price_10k":337000,"trade_date":"2026-01-26","build_year":2008,"deal_type":"중개거래"},{"apt_name":"거여4단지","dong":"거여동","area_sqm":49.8,"floor":2,"price_10k":97000,"trade_date":"2026-01-14","build_year":1997,"deal_type":"중개거래"},{"apt_name":"위례24단지(꿈에그린)","dong":"장지동","area_sqm":51.77,"floor":7,"price_10k":115000,"trade_date":"2026-01-10","build_year":2013,"deal_type":"직거래"},{"apt_name":"트리지움","dong":"잠실동","area_sqm":84.83,"floor":9,"price_10k":335000,"trade_date":"2026-01-24","build_year":2007,"deal_type":"중개거래"},{"apt_name":"리센츠","dong":"잠실동","area_sqm":27.68,"floor":11,"price_10k":155000,"trade_date":"2026-01-27","build_year":2008,"deal_type":"중개거래"},{"apt_name":"가락금호아파트","dong":"가락동","area_sqm":59.91,"floor":3,"price_10k":169000,"trade_date":"2026-01-28","build_year":1997,"deal_type":"중개거래"},{"apt_name":"송파시그니처롯데캐슬","dong":"거여동","area_sqm":59.98,"floor":23,"price_10k":183500,"trade_date":"2026-01-27","build_year":2021,"deal_type":"중개거래"},{"apt_name":"파크리오","dong":"신천동","area_sqm":84.97,"floor":20,"price_10k":322000,"trade_date":"2026-01-26","build_year":2008,"deal_type":"중개거래"},{"apt_name":"헬리오시티","dong":"가락동","area_sqm":84.98,"floor":24,"price_10k":290000,"trade_date":"2026-01-09","build_year":2018,"deal_type":"중개거래"},{"apt_name":"헬리오시티","dong":"가락동","area_sqm":84.97,"floor":18,"price_10k":304000,"trade_date":"2026-01-27","build_year":2018,"deal_type":"중개거래"},{"apt_name":"테라펠리체송파","dong":"가락동","area_sqm":44.0351,"floor":15,"price_10k":82000,"trade_date":"2026-01-09","build_year":2021,"deal_type":"중개거래"},{"apt_name":"상아2차아파트","dong":"오금동","area_sqm":60,"floor":7,"price_10k":150000,"trade_date":"2026-01-13","build_year":1988,"deal_type":"중개거래"},{"apt_name":"상아2차아파트","dong":"오금동","area_sqm":84.98,"floor":1,"price_10k":165000,"trade_date":"2026-01-16","build_year":1988,"deal_type":"중개거래"},{"apt_name":"올림픽카운티","dong":"방이동","area_sqm":56.22,"floor":7,"price_10k":60000,"trade_date":"2026-01-22","build_year":2003,"deal_type":"중개거래"},{"apt_name":"주공아파트 5단지","dong":"잠실동","area_sqm":82.51,"floor":14,"price_10k":457500,"trade_date":"2026-01-21","build_year":1978,"deal_type":"중개거래"},{"apt_name":"현대2차","dong":"거여동","area_sqm":82.96,"floor":7,"price_10k":103000,"trade_date":"2026-01-12","build_year":1992,"deal_type":"중개거래"},{"apt_name":"트리지움","dong":"잠실동","area_sqm":84.83,"floor":19,"price_10k":320000,"trade_date":"2026-01-26","build_year":2007,"deal_type":"중개거래"},{"apt_name":"건영1","dong":"문정동","area_sqm":77.68,"floor":2,"price_10k":157500,"trade_date":"2026-01-24","build_year":1996,"deal_type":"중개거래"},{"apt_name":"현대1","dong":"풍납동","area_sqm":83.02,"floor":6,"price_10k":173000,"trade_date":"2026-01-10","build_year":1995,"deal_type":"중개거래"},{"apt_name":"올림픽레아아파트","dong":"풍납동","area_sqm":52.3564,"floor":6,"price_10k":80000,"trade_date":"2026-01-26","build_year":2024,"deal_type":"중개거래"},{"apt_name":"극동","dong":"풍납동","area_sqm":79.945,"floor":8,"price_10k":145000,"trade_date":"2026-01-26","build_year":1987,"deal_type":"중개거래"},{"apt_name":"건영1","dong":"문정동","area_sqm":84.39,"floor":3,"price_10k":164000,"trade_date":"2026-01-25","build_year":1996,"deal_type":"중개거래"},{"apt_name":"헬리오시티","dong":"가락동","area_sqm":49.21,"floor":29,"price_10k":239000,"trade_date":"2026-01-27","build_year":2018,"deal_type":"중개거래"},{"apt_name":"현대(2-4차)","dong":"오금동","area_sqm":130.93,"floor":10,"price_10k":291500,"trade_date":"2026-01-13","build_year":1984,"deal_type":"중개거래"},{"apt_name":"문정시영","dong":"문정동","area_sqm":35.46,"floor":9,"price_10k":78000,"trade_date":"2026-01-24","build_year":1989,"deal_type":"중개거래"},{"apt_name":"송파시그니처롯데캐슬","dong":"거여동","area_sqm":84.98,"floor":5,"price_10k":204000,"trade_date":"2026-01-23","build_year":2021,"deal_type":"중개거래"}]

trades_202512 = [{"apt_name":"파크리오","dong":"신천동","area_sqm":84.79,"floor":36,"price_10k":295000,"trade_date":"2025-12-02","build_year":2008,"deal_type":"중개거래"},{"apt_name":"파크리오","dong":"신천동","area_sqm":121.63,"floor":28,"price_10k":354000,"trade_date":"2025-12-29","build_year":2008,"deal_type":"중개거래"},{"apt_name":"한양아파트","dong":"송파동","area_sqm":119.68,"floor":10,"price_10k":244000,"trade_date":"2025-12-24","build_year":1983,"deal_type":"중개거래"},{"apt_name":"헬리오시티","dong":"가락동","area_sqm":84.96,"floor":7,"price_10k":291000,"trade_date":"2025-12-27","build_year":2018,"deal_type":"중개거래"},{"apt_name":"롯데캐슬골드","dong":"신천동","area_sqm":166.7,"floor":17,"price_10k":250000,"trade_date":"2025-12-30","build_year":2005,"deal_type":"중개거래"},{"apt_name":"송파파크데일2단지","dong":"마천동","area_sqm":114.91,"floor":15,"price_10k":137000,"trade_date":"2025-12-31","build_year":2011,"deal_type":"중개거래"},{"apt_name":"미성맨션","dong":"송파동","area_sqm":84.95,"floor":2,"price_10k":220000,"trade_date":"2025-12-27","build_year":1985,"deal_type":"중개거래"},{"apt_name":"상아2차아파트","dong":"오금동","area_sqm":45.77,"floor":2,"price_10k":119200,"trade_date":"2025-12-31","build_year":1988,"deal_type":"중개거래"},{"apt_name":"동승","dong":"삼전동","area_sqm":53.04,"floor":1,"price_10k":54500,"trade_date":"2025-12-12","build_year":1993,"deal_type":"중개거래"},{"apt_name":"거여1단지","dong":"거여동","area_sqm":39.6,"floor":3,"price_10k":75000,"trade_date":"2025-12-30","build_year":1997,"deal_type":"중개거래"},{"apt_name":"한화오벨리스크","dong":"가락동","area_sqm":25.32,"floor":20,"price_10k":19300,"trade_date":"2025-12-23","build_year":2004,"deal_type":"중개거래"},{"apt_name":"천마그랑밸리","dong":"마천동","area_sqm":84.95,"floor":4,"price_10k":99000,"trade_date":"2025-12-05","build_year":2004,"deal_type":"중개거래"},{"apt_name":"우성아파트","dong":"잠실동","area_sqm":131.08,"floor":4,"price_10k":380000,"trade_date":"2025-12-19","build_year":1981,"deal_type":"중개거래"},{"apt_name":"송파레미니스2단지","dong":"오금동","area_sqm":84.94,"floor":19,"price_10k":171300,"trade_date":"2025-12-31","build_year":2018,"deal_type":"중개거래"},{"apt_name":"송파파인타운11단지","dong":"장지동","area_sqm":84.94,"floor":2,"price_10k":167000,"trade_date":"2025-12-30","build_year":2007,"deal_type":"중개거래"},{"apt_name":"건영1","dong":"문정동","area_sqm":84.39,"floor":2,"price_10k":158000,"trade_date":"2025-12-23","build_year":1996,"deal_type":"중개거래"},{"apt_name":"문정시영","dong":"문정동","area_sqm":39.69,"floor":13,"price_10k":95000,"trade_date":"2025-12-18","build_year":1989,"deal_type":"중개거래"},{"apt_name":"리센츠","dong":"잠실동","area_sqm":27.68,"floor":4,"price_10k":130000,"trade_date":"2025-12-15","build_year":2008,"deal_type":"중개거래"},{"apt_name":"한신잠실코아","dong":"신천동","area_sqm":52.51,"floor":2,"price_10k":145000,"trade_date":"2025-12-31","build_year":1988,"deal_type":"중개거래"},{"apt_name":"잠실한솔","dong":"석촌동","area_sqm":59.86,"floor":16,"price_10k":158500,"trade_date":"2025-12-17","build_year":2000,"deal_type":"중개거래"},{"apt_name":"파크리오","dong":"신천동","area_sqm":84.9,"floor":30,"price_10k":305000,"trade_date":"2025-12-23","build_year":2008,"deal_type":"중개거래"},{"apt_name":"헬리오시티","dong":"가락동","area_sqm":130.06,"floor":14,"price_10k":369000,"trade_date":"2025-12-22","build_year":2018,"deal_type":"중개거래"},{"apt_name":"인텍스빌","dong":"방이동","area_sqm":117.61,"floor":2,"price_10k":105000,"trade_date":"2025-12-29","build_year":2000,"deal_type":"중개거래"},{"apt_name":"올림픽훼밀리타운","dong":"문정동","area_sqm":117.585,"floor":9,"price_10k":277000,"trade_date":"2025-12-27","build_year":1988,"deal_type":"중개거래"},{"apt_name":"한신잠실코아","dong":"신천동","area_sqm":52.51,"floor":8,"price_10k":141500,"trade_date":"2025-12-24","build_year":1988,"deal_type":"중개거래"},{"apt_name":"현대리버빌2지구","dong":"풍납동","area_sqm":84.97,"floor":4,"price_10k":125250,"trade_date":"2025-12-17","build_year":1999,"deal_type":"중개거래"},{"apt_name":"삼익1","dong":"마천동","area_sqm":84.9,"floor":10,"price_10k":85000,"trade_date":"2025-12-30","build_year":1996,"deal_type":"직거래"},{"apt_name":"한성나이스빌","dong":"방이동","area_sqm":74.83,"floor":2,"price_10k":63000,"trade_date":"2025-12-15","build_year":2002,"deal_type":"중개거래"},{"apt_name":"송파파인타운3단지","dong":"장지동","area_sqm":84.94,"floor":9,"price_10k":170000,"trade_date":"2025-12-31","build_year":2008,"deal_type":"중개거래"},{"apt_name":"가락(1차)쌍용아파트","dong":"가락동","area_sqm":59.92,"floor":10,"price_10k":196000,"trade_date":"2025-12-26","build_year":1997,"deal_type":"중개거래"},{"apt_name":"강변현대","dong":"풍납동","area_sqm":83.19,"floor":3,"price_10k":100000,"trade_date":"2025-12-31","build_year":1991,"deal_type":"중개거래"},{"apt_name":"거여1단지","dong":"거여동","area_sqm":59.73,"floor":12,"price_10k":109700,"trade_date":"2025-12-17","build_year":1997,"deal_type":"중개거래"},{"apt_name":"마천금호어울림1차아파트","dong":"마천동","area_sqm":101.91,"floor":19,"price_10k":111000,"trade_date":"2025-12-19","build_year":2002,"deal_type":"중개거래"},{"apt_name":"가락3차쌍용스윗닷홈(103동)","dong":"가락동","area_sqm":84.94,"floor":2,"price_10k":154000,"trade_date":"2025-12-27","build_year":2005,"deal_type":"중개거래"},{"apt_name":"송파","dong":"송파동","area_sqm":83.22,"floor":6,"price_10k":119000,"trade_date":"2025-12-30","build_year":2003,"deal_type":"중개거래"},{"apt_name":"한양아파트","dong":"송파동","area_sqm":64.26,"floor":6,"price_10k":203000,"trade_date":"2025-12-19","build_year":1983,"deal_type":"중개거래"},{"apt_name":"롯데캐슬골드","dong":"신천동","area_sqm":166.7,"floor":28,"price_10k":335000,"trade_date":"2025-12-09","build_year":2005,"deal_type":"중개거래"},{"apt_name":"주공아파트 5단지","dong":"잠실동","area_sqm":82.51,"floor":11,"price_10k":462500,"trade_date":"2025-12-27","build_year":1978,"deal_type":"중개거래"},{"apt_name":"상아2차아파트","dong":"오금동","area_sqm":84.98,"floor":8,"price_10k":160000,"trade_date":"2025-12-13","build_year":1988,"deal_type":"중개거래"},{"apt_name":"상아2차아파트","dong":"오금동","area_sqm":45.77,"floor":8,"price_10k":119000,"trade_date":"2025-12-27","build_year":1988,"deal_type":"중개거래"},{"apt_name":"올림픽선수기자촌3단지","dong":"방이동","area_sqm":100.31,"floor":10,"price_10k":320000,"trade_date":"2025-12-18","build_year":1988,"deal_type":"중개거래"},{"apt_name":"헬리오시티","dong":"가락동","area_sqm":39.1,"floor":7,"price_10k":172000,"trade_date":"2025-12-16","build_year":2018,"deal_type":"중개거래"},{"apt_name":"리센츠","dong":"잠실동","area_sqm":84.99,"floor":21,"price_10k":355000,"trade_date":"2025-12-27","build_year":2008,"deal_type":"중개거래"},{"apt_name":"더샵스타리버","dong":"신천동","area_sqm":142.17,"floor":15,"price_10k":220000,"trade_date":"2025-12-29","build_year":2006,"deal_type":"중개거래"},{"apt_name":"옥산크리스탈","dong":"방이동","area_sqm":62.5,"floor":5,"price_10k":76000,"trade_date":"2025-12-30","build_year":2001,"deal_type":"중개거래"},{"apt_name":"이연파레스","dong":"가락동","area_sqm":96,"floor":5,"price_10k":120000,"trade_date":"2025-12-26","build_year":1999,"deal_type":"직거래"},{"apt_name":"헬리오시티","dong":"가락동","area_sqm":49.21,"floor":27,"price_10k":234500,"trade_date":"2025-12-31","build_year":2018,"deal_type":"중개거래"},{"apt_name":"파크리오","dong":"신천동","area_sqm":84.79,"floor":12,"price_10k":296000,"trade_date":"2025-12-22","build_year":2008,"deal_type":"중개거래"},{"apt_name":"동아한가람1","dong":"풍납동","area_sqm":84.38,"floor":1,"price_10k":114000,"trade_date":"2025-12-12","build_year":1995,"deal_type":"중개거래"},{"apt_name":"거여1단지","dong":"거여동","area_sqm":49.8,"floor":9,"price_10k":96000,"trade_date":"2025-12-20","build_year":1997,"deal_type":"중개거래"},{"apt_name":"주공아파트 5단지","dong":"잠실동","area_sqm":82.51,"floor":9,"price_10k":457500,"trade_date":"2025-12-29","build_year":1978,"deal_type":"중개거래"},{"apt_name":"올림픽훼밀리타운","dong":"문정동","area_sqm":84.751,"floor":1,"price_10k":155000,"trade_date":"2025-12-29","build_year":1988,"deal_type":"직거래"},{"apt_name":"장미2","dong":"신천동","area_sqm":99,"floor":9,"price_10k":358000,"trade_date":"2025-12-23","build_year":1979,"deal_type":"중개거래"},{"apt_name":"파크리오","dong":"신천동","area_sqm":59.95,"floor":22,"price_10k":274000,"trade_date":"2025-12-16","build_year":2008,"deal_type":"중개거래"},{"apt_name":"잠실올림픽공원아이파크","dong":"풍납동","area_sqm":59.97,"floor":13,"price_10k":203000,"trade_date":"2025-12-31","build_year":2019,"deal_type":"중개거래"},{"apt_name":"아키죤아파트1동","dong":"가락동","area_sqm":75.29,"floor":2,"price_10k":84500,"trade_date":"2025-12-23","build_year":2002,"deal_type":"중개거래"},{"apt_name":"잠실월드메르디앙","dong":"잠실동","area_sqm":72.33,"floor":15,"price_10k":147000,"trade_date":"2025-12-22","build_year":2005,"deal_type":"중개거래"},{"apt_name":"장미1","dong":"신천동","area_sqm":82.45,"floor":7,"price_10k":324000,"trade_date":"2025-12-10","build_year":1979,"deal_type":"중개거래"},{"apt_name":"대림","dong":"오금동","area_sqm":125.79,"floor":8,"price_10k":250000,"trade_date":"2025-12-31","build_year":1988,"deal_type":"중개거래"},{"apt_name":"잠실올림픽공원아이파크","dong":"풍납동","area_sqm":84.99,"floor":7,"price_10k":225000,"trade_date":"2025-12-31","build_year":2019,"deal_type":"중개거래"},{"apt_name":"가락금호아파트","dong":"가락동","area_sqm":59.91,"floor":20,"price_10k":175000,"trade_date":"2025-12-31","build_year":1997,"deal_type":"중개거래"},{"apt_name":"가락(1차)쌍용아파트","dong":"가락동","area_sqm":84.69,"floor":2,"price_10k":200000,"trade_date":"2025-12-20","build_year":1997,"deal_type":"중개거래"},{"apt_name":"송파시그니처롯데캐슬","dong":"거여동","area_sqm":59.98,"floor":23,"price_10k":174500,"trade_date":"2025-12-31","build_year":2021,"deal_type":"중개거래"},{"apt_name":"SK","dong":"송파동","area_sqm":83.8,"floor":6,"price_10k":150000,"trade_date":"2025-12-31","build_year":2001,"deal_type":"중개거래"},{"apt_name":"신성노바빌아파트","dong":"풍납동","area_sqm":84.96,"floor":5,"price_10k":125000,"trade_date":"2025-12-30","build_year":2000,"deal_type":"중개거래"},{"apt_name":"리센츠","dong":"잠실동","area_sqm":27.68,"floor":11,"price_10k":170000,"trade_date":"2025-12-27","build_year":2008,"deal_type":"중개거래"},{"apt_name":"송파시그니처롯데캐슬","dong":"거여동","area_sqm":84.98,"floor":13,"price_10k":199000,"trade_date":"2025-12-31","build_year":2021,"deal_type":"중개거래"},{"apt_name":"두산위브","dong":"가락동","area_sqm":30.5,"floor":13,"price_10k":25500,"trade_date":"2025-12-22","build_year":2004,"deal_type":"중개거래"},{"apt_name":"거여4단지","dong":"거여동","area_sqm":49.8,"floor":10,"price_10k":105000,"trade_date":"2025-12-30","build_year":1997,"deal_type":"중개거래"},{"apt_name":"거여1단지","dong":"거여동","area_sqm":49.8,"floor":12,"price_10k":95000,"trade_date":"2025-12-30","build_year":1997,"deal_type":"중개거래"},{"apt_name":"파크리오","dong":"신천동","area_sqm":84.97,"floor":1,"price_10k":286000,"trade_date":"2025-12-26","build_year":2008,"deal_type":"중개거래"},{"apt_name":"송파파인타운9단지","dong":"장지동","area_sqm":59.95,"floor":13,"price_10k":163000,"trade_date":"2025-12-30","build_year":2007,"deal_type":"중개거래"},{"apt_name":"코오롱아파트","dong":"방이동","area_sqm":84.95,"floor":8,"price_10k":205000,"trade_date":"2025-12-27","build_year":1991,"deal_type":"중개거래"},{"apt_name":"현대1","dong":"풍납동","area_sqm":83.02,"floor":18,"price_10k":174000,"trade_date":"2025-12-29","build_year":1995,"deal_type":"중개거래"},{"apt_name":"송파건원여미지아파트","dong":"마천동","area_sqm":51.764,"floor":10,"price_10k":72000,"trade_date":"2025-12-27","build_year":2018,"deal_type":"중개거래"},{"apt_name":"현대(2-4차)","dong":"오금동","area_sqm":84.98,"floor":6,"price_10k":182000,"trade_date":"2025-12-26","build_year":1984,"deal_type":"직거래"},{"apt_name":"현대(6차)","dong":"가락동","area_sqm":84.8,"floor":1,"price_10k":140000,"trade_date":"2025-12-03","build_year":1991,"deal_type":"중개거래"},{"apt_name":"잠실아이스페이스","dong":"신천동","area_sqm":67.11,"floor":16,"price_10k":87500,"trade_date":"2025-12-17","build_year":2001,"deal_type":"중개거래"},{"apt_name":"주공아파트 5단지","dong":"잠실동","area_sqm":76.5,"floor":15,"price_10k":387700,"trade_date":"2025-12-08","build_year":1978,"deal_type":"중개거래"},{"apt_name":"헬리오시티","dong":"가락동","area_sqm":39.12,"floor":16,"price_10k":170000,"trade_date":"2025-12-20","build_year":2018,"deal_type":"중개거래"},{"apt_name":"가락삼익맨숀","dong":"송파동","area_sqm":108.338,"floor":2,"price_10k":269000,"trade_date":"2025-12-27","build_year":1984,"deal_type":"중개거래"},{"apt_name":"가락미륭아파트","dong":"가락동","area_sqm":50,"floor":1,"price_10k":144000,"trade_date":"2025-12-09","build_year":1986,"deal_type":"중개거래"},{"apt_name":"아시아선수촌아파트","dong":"잠실동","area_sqm":151.008,"floor":3,"price_10k":545000,"trade_date":"2025-12-15","build_year":1986,"deal_type":"중개거래"},{"apt_name":"올림픽훼밀리타운","dong":"문정동","area_sqm":84.705,"floor":13,"price_10k":268000,"trade_date":"2025-12-30","build_year":1988,"deal_type":"중개거래"},{"apt_name":"잠실엘스","dong":"잠실동","area_sqm":59.96,"floor":5,"price_10k":294000,"trade_date":"2025-12-26","build_year":2008,"deal_type":"중개거래"},{"apt_name":"송파더센트레","dong":"장지동","area_sqm":51.77,"floor":13,"price_10k":144000,"trade_date":"2025-12-29","build_year":2013,"deal_type":"중개거래"},{"apt_name":"위례신도시송파푸르지오","dong":"장지동","area_sqm":112.9476,"floor":2,"price_10k":195000,"trade_date":"2025-12-27","build_year":2015,"deal_type":"중개거래"},{"apt_name":"송파파크데일2단지","dong":"마천동","area_sqm":114.81,"floor":2,"price_10k":132000,"trade_date":"2025-12-29","build_year":2011,"deal_type":"중개거래"},{"apt_name":"현대","dong":"문정동","area_sqm":84.03,"floor":3,"price_10k":97500,"trade_date":"2025-12-29","build_year":1991,"deal_type":"중개거래"},{"apt_name":"우성아파트","dong":"잠실동","area_sqm":131.08,"floor":1,"price_10k":374000,"trade_date":"2025-12-13","build_year":1981,"deal_type":"중개거래"},{"apt_name":"헬리오시티","dong":"가락동","area_sqm":99.6,"floor":22,"price_10k":325000,"trade_date":"2025-12-26","build_year":2018,"deal_type":"중개거래"},{"apt_name":"올림픽선수기자촌2단지","dong":"방이동","area_sqm":100.82,"floor":3,"price_10k":325000,"trade_date":"2025-12-27","build_year":1988,"deal_type":"중개거래"},{"apt_name":"문정래미안","dong":"문정동","area_sqm":170.7,"floor":9,"price_10k":254500,"trade_date":"2025-12-23","build_year":2004,"deal_type":"중개거래"},{"apt_name":"송파파인타운1단지","dong":"장지동","area_sqm":84.95,"floor":2,"price_10k":155000,"trade_date":"2025-12-23","build_year":2010,"deal_type":"중개거래"},{"apt_name":"현대리버빌2지구","dong":"풍납동","area_sqm":40.95,"floor":7,"price_10k":60000,"trade_date":"2025-12-25","build_year":1999,"deal_type":"중개거래"},{"apt_name":"주공아파트 5단지","dong":"잠실동","area_sqm":76.5,"floor":9,"price_10k":407700,"trade_date":"2025-12-18","build_year":1978,"deal_type":"중개거래"},{"apt_name":"한신잠실코아","dong":"신천동","area_sqm":50.76,"floor":7,"price_10k":138000,"trade_date":"2025-12-23","build_year":1988,"deal_type":"중개거래"}]

all_trades = {
    "2025-12": trades_202512,
    "2026-01": trades_202601,
    "2026-02": trades_202602,
}

# === 스타일 ===
header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
sub_header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
sub_header_font = Font(name="맑은 고딕", bold=True, size=10)
data_font = Font(name="맑은 고딕", size=10)
price_high_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
price_low_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")
num_fmt_price = '#,##0'

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_data_cell(ws, row, col, align=center_align):
    cell = ws.cell(row=row, column=col)
    cell.font = data_font
    cell.alignment = align
    cell.border = thin_border
    return cell

# ===== Sheet 1: 전체 거래 내역 =====
ws1 = wb.active
ws1.title = "전체 거래내역"
headers1 = ["거래월", "아파트명", "동", "면적(m²)", "층", "거래가(만원)", "거래가(억원)", "거래일", "건축년도", "연식", "거래유형", "평당가(만원)"]
for c, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=c, value=h)
style_header_row(ws1, 1, len(headers1))

row = 2
for month, items in sorted(all_trades.items()):
    for t in sorted(items, key=lambda x: x["trade_date"]):
        pyeong = t["area_sqm"] / 3.305785
        price_eok = t["price_10k"] / 10000
        age = 2026 - t["build_year"]
        price_per_pyeong = round(t["price_10k"] / pyeong) if pyeong > 0 else 0

        ws1.cell(row=row, column=1, value=month)
        ws1.cell(row=row, column=2, value=t["apt_name"])
        ws1.cell(row=row, column=3, value=t["dong"])
        ws1.cell(row=row, column=4, value=round(t["area_sqm"], 2))
        ws1.cell(row=row, column=5, value=t["floor"])
        c6 = ws1.cell(row=row, column=6, value=t["price_10k"])
        c6.number_format = num_fmt_price
        c7 = ws1.cell(row=row, column=7, value=round(price_eok, 2))
        ws1.cell(row=row, column=8, value=t["trade_date"])
        ws1.cell(row=row, column=9, value=t["build_year"])
        ws1.cell(row=row, column=10, value=age)
        ws1.cell(row=row, column=11, value=t["deal_type"])
        c12 = ws1.cell(row=row, column=12, value=price_per_pyeong)
        c12.number_format = num_fmt_price

        for col in range(1, len(headers1) + 1):
            style_data_cell(ws1, row, col)
        row += 1

for col in range(1, len(headers1) + 1):
    ws1.column_dimensions[get_column_letter(col)].width = max(12, len(str(headers1[col-1])) * 2)
ws1.column_dimensions["B"].width = 28
ws1.column_dimensions["C"].width = 10
ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers1))}{row-1}"

# ===== Sheet 2: 월별 요약 통계 =====
ws2 = wb.create_sheet("월별 요약통계")
headers2 = ["구분", "2025-12", "2026-01", "2026-02", "추이"]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=c, value=h)
style_header_row(ws2, 1, len(headers2))

stats = {}
for month, items in sorted(all_trades.items()):
    prices = [t["price_10k"] for t in items]
    prices_84 = [t["price_10k"] for t in items if 80 <= t["area_sqm"] <= 90]
    stats[month] = {
        "count": len(items),
        "median": median(prices),
        "min": min(prices),
        "max": max(prices),
        "avg": round(sum(prices) / len(prices)),
        "median_84": median(prices_84) if prices_84 else 0,
        "count_84": len(prices_84),
    }

stat_rows = [
    ("총 거래건수", "count"),
    ("중위가격(만원)", "median"),
    ("평균가격(만원)", "avg"),
    ("최저가격(만원)", "min"),
    ("최고가격(만원)", "max"),
    ("84m² 중위가(만원)", "median_84"),
    ("84m² 거래건수", "count_84"),
]

for i, (label, key) in enumerate(stat_rows, 2):
    ws2.cell(row=i, column=1, value=label)
    months_sorted = ["2025-12", "2026-01", "2026-02"]
    for j, m in enumerate(months_sorted, 2):
        val = stats.get(m, {}).get(key, 0)
        c = ws2.cell(row=i, column=j, value=val)
        if "가격" in label:
            c.number_format = num_fmt_price
    # 추이 (2025-12 vs 2026-02)
    v_dec = stats.get("2025-12", {}).get(key, 0)
    v_feb = stats.get("2026-02", {}).get(key, 0)
    if v_dec and v_dec != 0:
        change = round((v_feb - v_dec) / v_dec * 100, 1)
        ws2.cell(row=i, column=5, value=f"{'+' if change > 0 else ''}{change}%")
    for col in range(1, 6):
        style_data_cell(ws2, i, col)

for col in range(1, 6):
    ws2.column_dimensions[get_column_letter(col)].width = 18

# ===== Sheet 3: 동별 분석 =====
ws3 = wb.create_sheet("동별 분석")
headers3 = ["동", "거래건수", "중위가(만원)", "평균가(만원)", "최저가(만원)", "최고가(만원)", "84m²중위가(만원)", "평균연식"]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=c, value=h)
style_header_row(ws3, 1, len(headers3))

all_items = []
for items in all_trades.values():
    all_items.extend(items)

dong_data = {}
for t in all_items:
    dong = t["dong"]
    if dong not in dong_data:
        dong_data[dong] = []
    dong_data[dong].append(t)

row = 2
for dong in sorted(dong_data.keys()):
    items = dong_data[dong]
    prices = [t["price_10k"] for t in items]
    prices_84 = [t["price_10k"] for t in items if 80 <= t["area_sqm"] <= 90]
    ages = [2026 - t["build_year"] for t in items]

    ws3.cell(row=row, column=1, value=dong)
    ws3.cell(row=row, column=2, value=len(items))
    c3 = ws3.cell(row=row, column=3, value=median(prices)); c3.number_format = num_fmt_price
    c4 = ws3.cell(row=row, column=4, value=round(sum(prices)/len(prices))); c4.number_format = num_fmt_price
    c5 = ws3.cell(row=row, column=5, value=min(prices)); c5.number_format = num_fmt_price
    c6 = ws3.cell(row=row, column=6, value=max(prices)); c6.number_format = num_fmt_price
    c7 = ws3.cell(row=row, column=7, value=median(prices_84) if prices_84 else "N/A")
    if prices_84:
        c7.number_format = num_fmt_price
    ws3.cell(row=row, column=8, value=round(sum(ages)/len(ages), 1))

    for col in range(1, len(headers3) + 1):
        style_data_cell(ws3, row, col)
    row += 1

for col in range(1, len(headers3) + 1):
    ws3.column_dimensions[get_column_letter(col)].width = 16

# ===== Sheet 4: 주요 단지별 분석 =====
ws4 = wb.create_sheet("주요단지 분석")
headers4 = ["아파트명", "동", "거래건수", "중위가(만원)", "최저가(만원)", "최고가(만원)", "대표면적(m²)", "건축년도", "평균평당가(만원)"]
for c, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=c, value=h)
style_header_row(ws4, 1, len(headers4))

apt_data = {}
for t in all_items:
    key = t["apt_name"]
    if key not in apt_data:
        apt_data[key] = []
    apt_data[key].append(t)

# 거래 2건 이상인 단지만
major_apts = {k: v for k, v in apt_data.items() if len(v) >= 2}
row = 2
for apt_name in sorted(major_apts.keys(), key=lambda k: -median([t["price_10k"] for t in major_apts[k]])):
    items = major_apts[apt_name]
    prices = [t["price_10k"] for t in items]
    areas = [t["area_sqm"] for t in items]
    ppps = []
    for t in items:
        pyeong = t["area_sqm"] / 3.305785
        if pyeong > 0:
            ppps.append(t["price_10k"] / pyeong)

    ws4.cell(row=row, column=1, value=apt_name)
    ws4.cell(row=row, column=2, value=items[0]["dong"])
    ws4.cell(row=row, column=3, value=len(items))
    c4 = ws4.cell(row=row, column=4, value=median(prices)); c4.number_format = num_fmt_price
    c5 = ws4.cell(row=row, column=5, value=min(prices)); c5.number_format = num_fmt_price
    c6 = ws4.cell(row=row, column=6, value=max(prices)); c6.number_format = num_fmt_price
    ws4.cell(row=row, column=7, value=round(median(areas), 2))
    ws4.cell(row=row, column=8, value=items[0]["build_year"])
    c9 = ws4.cell(row=row, column=9, value=round(median(ppps)) if ppps else 0); c9.number_format = num_fmt_price

    for col in range(1, len(headers4) + 1):
        style_data_cell(ws4, row, col)
    row += 1

ws4.column_dimensions["A"].width = 30
for col in range(2, len(headers4) + 1):
    ws4.column_dimensions[get_column_letter(col)].width = 16
ws4.auto_filter.ref = f"A1:{get_column_letter(len(headers4))}{row-1}"

# ===== Sheet 5: 분석 요약 =====
ws5 = wb.create_sheet("분석 요약")
ws5.sheet_properties.tabColor = "FF5722"

title_font = Font(name="맑은 고딕", bold=True, size=14, color="2F5496")
section_font = Font(name="맑은 고딕", bold=True, size=11, color="2F5496")
body_font = Font(name="맑은 고딕", size=10)

ws5.merge_cells("A1:F1")
c = ws5.cell(row=1, column=1, value="송파구(잠실역 주변) 아파트 매매 실거래가 심층 분석 리포트")
c.font = title_font
c.alignment = Alignment(horizontal="center", vertical="center")

ws5.merge_cells("A2:F2")
ws5.cell(row=2, column=1, value="분석 기간: 2025년 12월 ~ 2026년 2월 | 데이터 출처: 국토교통부 실거래가 공개시스템").font = body_font

analysis_lines = [
    ("", ""),
    ("1. 시장 개요", "section"),
    (f"  - 3개월 총 거래건수: {len(all_items)}건 (12월 {len(trades_202512)}건, 1월 {len(trades_202601)}건, 2월 {len(trades_202602)}건)", ""),
    (f"  - 2월은 아직 월 중반으로 거래 집계 중 (전월 대비 거래량 비교 시 주의)", ""),
    (f"  - 송파구 전체 중위가: {median([t['price_10k'] for t in all_items]):,.0f}만원 ({median([t['price_10k'] for t in all_items])/10000:.1f}억원)", ""),
    ("", ""),
    ("2. 잠실동 핵심 단지 현황 (84m² 기준)", "section"),
]

jamsil_apts = [t for t in all_items if t["dong"] == "잠실동"]
jamsil_84 = [t for t in jamsil_apts if 80 <= t["area_sqm"] <= 90]
if jamsil_84:
    jamsil_84_prices = [t["price_10k"] for t in jamsil_84]
    analysis_lines.append((f"  - 잠실동 84m² 중위가: {median(jamsil_84_prices):,.0f}만원 ({median(jamsil_84_prices)/10000:.1f}억원)", ""))
    analysis_lines.append((f"  - 가격 범위: {min(jamsil_84_prices):,.0f} ~ {max(jamsil_84_prices):,.0f}만원", ""))

jamsil_majors = ["잠실엘스", "리센츠", "트리지움", "파크리오", "레이크팰리스", "주공아파트 5단지", "헬리오시티"]
for name in jamsil_majors:
    if name in apt_data:
        items = apt_data[name]
        med = median([t["price_10k"] for t in items])
        analysis_lines.append((f"  - {name}: 중위가 {med:,.0f}만원 ({med/10000:.1f}억원), {len(items)}건 거래", ""))

analysis_lines += [
    ("", ""),
    ("3. 동별 가격 서열", "section"),
]

dong_medians = []
for dong, items in dong_data.items():
    prices_84 = [t["price_10k"] for t in items if 80 <= t["area_sqm"] <= 90]
    if prices_84:
        dong_medians.append((dong, median(prices_84), len(prices_84)))
dong_medians.sort(key=lambda x: -x[1])
for dong, med, cnt in dong_medians:
    analysis_lines.append((f"  - {dong}: 84m² 중위가 {med:,.0f}만원 ({med/10000:.1f}억원), {cnt}건", ""))

analysis_lines += [
    ("", ""),
    ("4. 전월세 데이터", "section"),
    ("  - 전월세 API 접근 불가 (HTTP 403). 공공데이터포털에서 '국토교통부_아파트 전월세 자료' 활용 신청 필요", ""),
    ("  - 신청 완료 후 전세가율 분석, 갭투자 리스크 분석 가능", ""),
    ("", ""),
    ("5. 투자 참고사항", "section"),
    ("  - 잠실동 주공5단지: 재건축 프리미엄으로 76m²가 42~46억원대, 송파구 내 최고 평당가", ""),
    ("  - 잠실엘스/리센츠/트리지움: 84m² 기준 31~35억원대 안정적 시세 형성", ""),
    ("  - 헬리오시티: 84m² 기준 29~31억원대, 대단지 프리미엄 보유", ""),
    ("  - 거여동/마천동: 8~21억원대로 송파구 내 상대적 저가 지역", ""),
]

row = 4
for text, style in analysis_lines:
    ws5.merge_cells(f"A{row}:F{row}")
    c = ws5.cell(row=row, column=1, value=text)
    if style == "section":
        c.font = section_font
    else:
        c.font = body_font
    row += 1

ws5.column_dimensions["A"].width = 80

# === 저장 ===
output_path = "C:/Devs/KRE/송파구_잠실_아파트_실거래가_분석.xlsx"
wb.save(output_path)
print(f"엑셀 파일 생성 완료: {output_path}")
