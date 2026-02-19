"""잠실/송파 재건축 투자 다차원 심층분석 엑셀 생성기 v2
- 8개 핵심 단지 + 송파구 전체 정비사업 현황
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, RadarChart, LineChart

# ── 스타일 ──
DARK_BLUE = "1B2A4A"
GOLD = "C9A84C"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
MID_GRAY = "D9D9D9"
RED = "E74C3C"
GREEN = "27AE60"
ORANGE = "F39C12"
BLUE = "3498DB"

title_font = Font(name="맑은 고딕", size=16, bold=True, color=WHITE)
header_font = Font(name="맑은 고딕", size=11, bold=True, color=WHITE)
sub_header_font = Font(name="맑은 고딕", size=10, bold=True, color=DARK_BLUE)
data_font = Font(name="맑은 고딕", size=10)
small_font = Font(name="맑은 고딕", size=9, color="666666")
score_font = Font(name="맑은 고딕", size=12, bold=True)

title_fill = PatternFill("solid", fgColor=DARK_BLUE)
header_fill = PatternFill("solid", fgColor="2C3E6B")
sub_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
gold_fill = PatternFill("solid", fgColor=GOLD)
green_fill = PatternFill("solid", fgColor="D5F5E3")
red_fill = PatternFill("solid", fgColor="FADBD8")
orange_fill = PatternFill("solid", fgColor="FDEBD0")
white_fill = PatternFill("solid", fgColor=WHITE)

thin_border = Border(
    left=Side(style='thin', color=MID_GRAY), right=Side(style='thin', color=MID_GRAY),
    top=Side(style='thin', color=MID_GRAY), bottom=Side(style='thin', color=MID_GRAY))
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_range(ws, row, c1, c2, font=None, fill=None, alignment=None):
    for c in range(c1, c2+1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment

def write_title(ws, row, title, col_end=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=1, value=title)
    c.font = title_font; c.fill = title_fill; c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 36
    return row + 1

def write_headers(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start+i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border
    ws.row_dimensions[row].height = 28
    return row + 1

def write_row(ws, row, data, col_start=1, font=None, fill=None):
    for i, val in enumerate(data):
        c = ws.cell(row=row, column=col_start+i, value=val)
        c.font = font or data_font; c.fill = fill or white_fill
        c.alignment = center_align; c.border = thin_border
    return row + 1

def score_fill(s):
    if s >= 8: return green_fill
    elif s >= 6: return PatternFill("solid", fgColor="EAFAF1")
    elif s >= 4: return orange_fill
    else: return red_fill

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ========== 데이터 ==========
COMPLEXES = [
    "잠실주공5단지", "장미1·2차", "장미3차",
    "우성1·2·3차", "우성4차",
    "아시아선수촌", "올림픽선수촌", "올림픽훼밀리"
]
SHORT = ["주공5", "장미1·2", "장미3", "우성1·2·3", "우성4", "아시아", "선수촌", "훼밀리"]

BASIC = {
    "잠실주공5단지": {"준공": 1978, "연식": 48, "세대수": 3930, "동수": 30, "층": 15, "면적": "76~84",
        "주소": "잠실동 40", "재건축세대": 6491, "재건축층": 70, "시공사": "삼성물산+GS+현대산업",
        "브랜드": "래미안+자이+아이파크", "용적률기존": 169, "용적률계획": 390},
    "장미1·2차": {"준공": 1979, "연식": 47, "세대수": 2262, "동수": 26, "층": 15, "면적": "71~82",
        "주소": "잠실동 16-5", "재건축세대": 5165, "재건축층": 50, "시공사": "미선정",
        "브랜드": "미정", "용적률기존": 179, "용적률계획": 360},
    "장미3차": {"준공": 1983, "연식": 43, "세대수": 900, "동수": 10, "층": 15, "면적": "69~84",
        "주소": "잠실동 221", "재건축세대": 5165, "재건축층": 50, "시공사": "미선정(장미통합)",
        "브랜드": "미정", "용적률기존": 185, "용적률계획": 360},
    "우성1·2·3차": {"준공": 1981, "연식": 45, "세대수": 1842, "동수": 20, "층": 15, "면적": "59~121",
        "주소": "잠실동 22", "재건축세대": 2680, "재건축층": 49, "시공사": "GS건설",
        "브랜드": "자이", "용적률기존": 174, "용적률계획": 350},
    "우성4차": {"준공": 1982, "연식": 44, "세대수": 555, "동수": 7, "층": 15, "면적": "59~84",
        "주소": "잠실동 19", "재건축세대": 825, "재건축층": 49, "시공사": "DL이앤씨",
        "브랜드": "아크로", "용적률기존": 168, "용적률계획": 299},
    "아시아선수촌": {"준공": 1986, "연식": 40, "세대수": 1356, "동수": 18, "층": 18, "면적": "84~184",
        "주소": "잠실동", "재건축세대": 3483, "재건축층": 75, "시공사": "미선정",
        "브랜드": "미정", "용적률기존": 160, "용적률계획": 340},
    "올림픽선수촌": {"준공": 1988, "연식": 38, "세대수": 5540, "동수": 122, "층": 24, "면적": "82~211",
        "주소": "방이동 89", "재건축세대": 9200, "재건축층": 49, "시공사": "미선정",
        "브랜드": "미정", "용적률기존": 137, "용적률계획": 270},
    "올림픽훼밀리": {"준공": 1988, "연식": 38, "세대수": 4494, "동수": 56, "층": 15, "면적": "84~192",
        "주소": "문정동 150", "재건축세대": 6620, "재건축층": 26, "시공사": "미선정",
        "브랜드": "미정", "용적률기존": 194, "용적률계획": 300},
}

PROGRESS = {
    "잠실주공5단지": {"단계": "사업시행계획인가 신청", "진행률": 65,
        "안전진단": "완료(D)", "정비구역": "완료", "조합설립": "완료(2020.12)",
        "사업시행인가": "신청완료(2025)", "관리처분": "미착수(2026말예상)",
        "착공": "2027년", "준공": "2030~2031", "이주": "2026말~2027초"},
    "장미1·2차": {"단계": "사업시행인가 진행중", "진행률": 45,
        "안전진단": "완료(D)", "정비구역": "완료", "조합설립": "완료",
        "사업시행인가": "진행중", "관리처분": "미착수",
        "착공": "2028~2029", "준공": "2031~2032", "이주": "2028"},
    "장미3차": {"단계": "장미통합 사업시행인가 진행중", "진행률": 40,
        "안전진단": "완료", "정비구역": "완료", "조합설립": "완료",
        "사업시행인가": "진행중(통합)", "관리처분": "미착수",
        "착공": "2028~2029", "준공": "2031~2032", "이주": "2028"},
    "우성1·2·3차": {"단계": "사업시행인가 진행중", "진행률": 55,
        "안전진단": "완료(D)", "정비구역": "완료", "조합설립": "완료",
        "사업시행인가": "진행중(GS건설확정)", "관리처분": "미착수",
        "착공": "2028", "준공": "2031~2032", "이주": "2027~2028"},
    "우성4차": {"단계": "관리처분인가 준비중", "진행률": 70,
        "안전진단": "완료", "정비구역": "완료", "조합설립": "완료",
        "사업시행인가": "완료", "관리처분": "준비중(2026상반기)",
        "착공": "2026.09", "준공": "2029~2030", "이주": "2026하반기"},
    "아시아선수촌": {"단계": "추진위 승인 완료", "진행률": 25,
        "안전진단": "완료", "정비구역": "추진중", "조합설립": "2026상반기목표",
        "사업시행인가": "미착수", "관리처분": "미착수",
        "착공": "2030~2031", "준공": "2033~2035", "이주": "2030"},
    "올림픽선수촌": {"단계": "추진위 승인 완료", "진행률": 20,
        "안전진단": "완료(D)", "정비구역": "추진중", "조합설립": "2026목표",
        "사업시행인가": "미착수", "관리처분": "미착수",
        "착공": "2029~2030", "준공": "2033~2035", "이주": "2029~2030"},
    "올림픽훼밀리": {"단계": "추진위 승인 완료", "진행률": 20,
        "안전진단": "완료", "정비구역": "추진중", "조합설립": "2026하반기",
        "사업시행인가": "미착수", "관리처분": "미착수",
        "착공": "2029~2030", "준공": "2033~2035", "이주": "2029~2030"},
}

# 시세 (전용 84m² 기준, 억원)
PRICE = {
    "잠실주공5단지": {"매매가": 45.0, "전세가": 12.5, "전세가율": 27.8, "1년전": 38.0, "3년전": 25.0, "5년전": 20.0,
        "신고가": 45.5, "최저1년": 36.0, "호가": 47.0, "프리미엄": 15.0},
    "장미1·2차": {"매매가": 26.5, "전세가": 9.0, "전세가율": 34.0, "1년전": 22.0, "3년전": 16.0, "5년전": 13.0,
        "신고가": 26.5, "최저1년": 20.0, "호가": 28.0, "프리미엄": 6.5},
    "장미3차": {"매매가": 25.0, "전세가": 8.5, "전세가율": 34.0, "1년전": 21.0, "3년전": 15.5, "5년전": 12.0,
        "신고가": 25.0, "최저1년": 19.0, "호가": 26.0, "프리미엄": 5.0},
    "우성1·2·3차": {"매매가": 24.0, "전세가": 8.0, "전세가율": 33.3, "1년전": 20.0, "3년전": 14.5, "5년전": 11.0,
        "신고가": 25.0, "최저1년": 18.0, "호가": 26.0, "프리미엄": 4.0},
    "우성4차": {"매매가": 22.0, "전세가": 7.5, "전세가율": 34.1, "1년전": 18.5, "3년전": 13.0, "5년전": 10.0,
        "신고가": 22.5, "최저1년": 16.5, "호가": 24.0, "프리미엄": 2.0},
    "아시아선수촌": {"매매가": 28.0, "전세가": 7.0, "전세가율": 25.0, "1년전": 22.0, "3년전": 16.0, "5년전": 13.0,
        "신고가": 30.0, "최저1년": 20.0, "호가": 32.0, "프리미엄": 8.0},
    "올림픽선수촌": {"매매가": 29.0, "전세가": 6.0, "전세가율": 20.7, "1년전": 22.0, "3년전": 17.0, "5년전": 15.0,
        "신고가": 29.3, "최저1년": 21.0, "호가": 32.0, "프리미엄": 7.0},
    "올림픽훼밀리": {"매매가": 21.5, "전세가": 5.5, "전세가율": 25.6, "1년전": 18.0, "3년전": 15.0, "5년전": 13.0,
        "신고가": 22.0, "최저1년": 17.0, "호가": 23.0, "프리미엄": 3.5},
}

# 다차원 팩터 점수 (1-10)
FACTORS = {
    "잠실주공5단지": {
        "입지 프리미엄": 10, "역세권 접근성": 10, "한강 조망": 9,
        "학군": 8, "생활 인프라": 9, "대형 개발호재": 10,
        "재건축 진행도": 7, "시공사 신뢰도": 10, "브랜드 가치": 10,
        "단지 규모": 10, "가격 매력도": 3, "전세가율(갭)": 2,
        "기대 수익률": 3, "유동성": 4, "리스크 안정성": 7,
        "조합 안정성": 8, "정책 수혜": 8,
        "랜드마크 잠재력": 10, "임대수익": 3, "보유비용 효율": 3,
    },
    "장미1·2차": {
        "입지 프리미엄": 9, "역세권 접근성": 8, "한강 조망": 10,
        "학군": 8, "생활 인프라": 8, "대형 개발호재": 8,
        "재건축 진행도": 5, "시공사 신뢰도": 5, "브랜드 가치": 5,
        "단지 규모": 8, "가격 매력도": 6, "전세가율(갭)": 5,
        "기대 수익률": 7, "유동성": 6, "리스크 안정성": 5,
        "조합 안정성": 6, "정책 수혜": 7,
        "랜드마크 잠재력": 8, "임대수익": 5, "보유비용 효율": 6,
    },
    "장미3차": {
        "입지 프리미엄": 8, "역세권 접근성": 7, "한강 조망": 9,
        "학군": 8, "생활 인프라": 8, "대형 개발호재": 8,
        "재건축 진행도": 4, "시공사 신뢰도": 5, "브랜드 가치": 5,
        "단지 규모": 7, "가격 매력도": 7, "전세가율(갭)": 5,
        "기대 수익률": 7, "유동성": 5, "리스크 안정성": 5,
        "조합 안정성": 5, "정책 수혜": 7,
        "랜드마크 잠재력": 7, "임대수익": 5, "보유비용 효율": 6,
    },
    "우성1·2·3차": {
        "입지 프리미엄": 8, "역세권 접근성": 7, "한강 조망": 6,
        "학군": 8, "생활 인프라": 8, "대형 개발호재": 7,
        "재건축 진행도": 6, "시공사 신뢰도": 8, "브랜드 가치": 7,
        "단지 규모": 7, "가격 매력도": 7, "전세가율(갭)": 5,
        "기대 수익률": 6, "유동성": 6, "리스크 안정성": 6,
        "조합 안정성": 6, "정책 수혜": 7,
        "랜드마크 잠재력": 6, "임대수익": 5, "보유비용 효율": 7,
    },
    "우성4차": {
        "입지 프리미엄": 7, "역세권 접근성": 7, "한강 조망": 5,
        "학군": 8, "생활 인프라": 8, "대형 개발호재": 7,
        "재건축 진행도": 8, "시공사 신뢰도": 7, "브랜드 가치": 7,
        "단지 규모": 4, "가격 매력도": 8, "전세가율(갭)": 6,
        "기대 수익률": 8, "유동성": 7, "리스크 안정성": 7,
        "조합 안정성": 7, "정책 수혜": 7,
        "랜드마크 잠재력": 5, "임대수익": 6, "보유비용 효율": 8,
    },
    "아시아선수촌": {
        "입지 프리미엄": 9, "역세권 접근성": 8, "한강 조망": 5,
        "학군": 8, "생활 인프라": 8, "대형 개발호재": 9,
        "재건축 진행도": 3, "시공사 신뢰도": 4, "브랜드 가치": 4,
        "단지 규모": 8, "가격 매력도": 5, "전세가율(갭)": 3,
        "기대 수익률": 6, "유동성": 4, "리스크 안정성": 4,
        "조합 안정성": 5, "정책 수혜": 7,
        "랜드마크 잠재력": 9, "임대수익": 3, "보유비용 효율": 4,
    },
    "올림픽선수촌": {
        "입지 프리미엄": 8, "역세권 접근성": 7, "한강 조망": 4,
        "학군": 7, "생활 인프라": 8, "대형 개발호재": 8,
        "재건축 진행도": 2, "시공사 신뢰도": 4, "브랜드 가치": 4,
        "단지 규모": 10, "가격 매력도": 5, "전세가율(갭)": 2,
        "기대 수익률": 7, "유동성": 4, "리스크 안정성": 4,
        "조합 안정성": 4, "정책 수혜": 7,
        "랜드마크 잠재력": 8, "임대수익": 3, "보유비용 효율": 4,
    },
    "올림픽훼밀리": {
        "입지 프리미엄": 7, "역세권 접근성": 8, "한강 조망": 3,
        "학군": 7, "생활 인프라": 7, "대형 개발호재": 6,
        "재건축 진행도": 2, "시공사 신뢰도": 4, "브랜드 가치": 4,
        "단지 규모": 9, "가격 매력도": 7, "전세가율(갭)": 3,
        "기대 수익률": 6, "유동성": 5, "리스크 안정성": 4,
        "조합 안정성": 4, "정책 수혜": 6,
        "랜드마크 잠재력": 5, "임대수익": 4, "보유비용 효율": 6,
    },
}

WEIGHTS = {
    "입지 프리미엄": 12, "역세권 접근성": 8, "한강 조망": 6,
    "학군": 5, "생활 인프라": 5, "대형 개발호재": 8,
    "재건축 진행도": 10, "시공사 신뢰도": 7, "브랜드 가치": 5,
    "단지 규모": 4, "가격 매력도": 8, "전세가율(갭)": 3,
    "기대 수익률": 8, "유동성": 3, "리스크 안정성": 7,
    "조합 안정성": 4, "정책 수혜": 3,
    "랜드마크 잠재력": 4, "임대수익": 2, "보유비용 효율": 3,
}

ROI = {
    "잠실주공5단지": {"매입가": 45.0, "취득세": 1.35, "중개": 0.225, "분담금": 3.5,
        "이주기간": 4.5, "전세비용연": 0.35, "보유세연": 0.32, "이자연": 0.80,
        "신축가_하": 48, "신축가_중": 52, "신축가_상": 58, "준공": 2031},
    "장미1·2차": {"매입가": 26.5, "취득세": 0.80, "중개": 0.13, "분담금": 2.5,
        "이주기간": 4.0, "전세비용연": 0.32, "보유세연": 0.18, "이자연": 0.80,
        "신축가_하": 38, "신축가_중": 42, "신축가_상": 48, "준공": 2032},
    "장미3차": {"매입가": 25.0, "취득세": 0.75, "중개": 0.13, "분담금": 2.5,
        "이주기간": 4.0, "전세비용연": 0.30, "보유세연": 0.16, "이자연": 0.80,
        "신축가_하": 36, "신축가_중": 40, "신축가_상": 46, "준공": 2032},
    "우성1·2·3차": {"매입가": 24.0, "취득세": 0.72, "중개": 0.12, "분담금": 2.5,
        "이주기간": 4.0, "전세비용연": 0.28, "보유세연": 0.15, "이자연": 0.80,
        "신축가_하": 35, "신축가_중": 40, "신축가_상": 45, "준공": 2032},
    "우성4차": {"매입가": 22.0, "취득세": 0.66, "중개": 0.11, "분담금": 2.0,
        "이주기간": 3.0, "전세비용연": 0.26, "보유세연": 0.13, "이자연": 0.80,
        "신축가_하": 30, "신축가_중": 35, "신축가_상": 40, "준공": 2029},
    "아시아선수촌": {"매입가": 28.0, "취득세": 0.84, "중개": 0.14, "분담금": 3.0,
        "이주기간": 5.5, "전세비용연": 0.25, "보유세연": 0.20, "이자연": 0.80,
        "신축가_하": 38, "신축가_중": 45, "신축가_상": 52, "준공": 2034},
    "올림픽선수촌": {"매입가": 29.0, "취득세": 0.87, "중개": 0.15, "분담금": 1.5,
        "이주기간": 6.0, "전세비용연": 0.21, "보유세연": 0.22, "이자연": 0.80,
        "신축가_하": 38, "신축가_중": 43, "신축가_상": 50, "준공": 2034},
    "올림픽훼밀리": {"매입가": 21.5, "취득세": 0.65, "중개": 0.11, "분담금": 3.6,
        "이주기간": 6.0, "전세비용연": 0.19, "보유세연": 0.14, "이자연": 0.80,
        "신축가_하": 30, "신축가_중": 36, "신축가_상": 42, "준공": 2034},
}

def calc_total_cost(d, nv):
    return d["매입가"]+d["취득세"]+d["중개"]+d["분담금"]+(d["전세비용연"]+d["보유세연"]+d["이자연"])*d["이주기간"]+nv*0.0296

def calc_roi(d, nv):
    tc = calc_total_cost(d, nv)
    base = d["매입가"]+d["취득세"]+d["중개"]
    return round((nv - tc) / base * 100, 1) if base else 0

RISKS = [
    ("추가공사비", [(8,7),(7,6),(7,6),(7,6),(6,5),(7,6),(8,7),(7,6)]),
    ("사업지연(인허가)", [(6,4),(8,7),(8,7),(7,6),(5,3),(8,7),(9,8),(8,7)]),
    ("조합 내부 갈등", [(7,4),(6,5),(6,6),(6,5),(4,3),(6,5),(8,7),(7,6)]),
    ("부동산시장 하락", [(9,5),(7,5),(7,5),(7,5),(6,5),(8,5),(8,5),(7,5)]),
    ("금리 상승", [(8,4),(6,4),(6,4),(6,4),(5,4),(7,4),(7,4),(6,4)]),
    ("정책/규제 강화", [(8,6),(7,6),(7,6),(7,6),(6,5),(7,6),(7,6),(7,6)]),
    ("재초환 부담금", [(9,8),(7,6),(6,6),(6,5),(5,4),(7,6),(7,6),(5,5)]),
    ("토지거래허가 연장", [(7,8),(5,6),(5,6),(5,6),(4,5),(5,6),(5,6),(4,5)]),
    ("분담금 급등", [(8,5),(7,6),(7,6),(7,6),(5,4),(7,5),(5,3),(8,7)]),
    ("시공사 리스크", [(6,2),(7,7),(7,7),(5,3),(5,3),(8,8),(8,8),(8,8)]),
    ("양도세 중과", [(8,5),(6,5),(6,5),(6,5),(5,5),(7,5),(7,5),(6,5)]),
    ("보유세 증가", [(9,7),(6,5),(6,5),(6,5),(5,4),(7,6),(8,6),(6,5)]),
]

BENCHMARKS = [
    ("둔촌주공→올림픽파크포레온", "강동구", 5930, 12032, 16, 13.0, 24.0, 11.0, "200~300%", "약1조원"),
    ("반포주공1→래미안원베일리", "서초구", 2444, 2990, 15, 20.0, 42.0, 22.0, "150~250%", "일부"),
    ("개포주공1→디에이치포레센트", "강남구", 1980, 6702, 14, 15.0, 38.0, 23.0, "200~300%", "해결"),
    ("개포주공4→래미안포레스트", "강남구", 1100, 3375, 12, 14.0, 35.0, 21.0, "180~280%", "없음"),
    ("잠실주공1~4→잠실르엘", "송파구", 5678, 6300, 13, 13.5, 30.0, 16.5, "150~250%", "일부"),
]

POLICIES = [
    ("토지거래허가제", "잠실동 전체(2026.06.22까지)", "실거주 의무, 전세불가", "매우높음"),
    ("재건축초과이익환수제", "2026년 본격 적용", "초과이익 최대50% 환수", "매우높음"),
    ("분양가상한제", "과열지역 적용", "일반분양 수익 제한", "높음"),
    ("안전진단 완화", "구조안전성 비중 30%로 하향", "재건축 착수 용이", "긍정적"),
    ("종부세 강화", "2026 공정시장가액비율100%", "보유세 대폭 증가", "높음"),
    ("다주택 양도세 중과", "조정지역+20~30%p", "단기매도시 세부담 극대화", "높음"),
    ("DSR 규제", "총부채원리금상환비율40%", "대출한도 제한", "중간"),
    ("GTX-A 개통", "2025말 개통", "교통 프리미엄 강화", "긍정적"),
    ("MICE 복합단지", "잠실종합운동장 일대", "잠실 미래가치 상승", "긍정적"),
    ("3호선 송파하남연장", "2032년 개통 예정", "올림픽선수촌 수혜", "긍정적"),
]

# 송파구 전체 정비사업 목록
ALL_PROJECTS = [
    ("잠실주공5단지", "잠실동", "재건축", 1978, 3930, 6491, "사업시행인가 신청", "상"),
    ("장미1·2·3차(통합)", "잠실동", "재건축", 1979, 3162, 5165, "사업시행인가 진행중", "상"),
    ("우성1·2·3차", "잠실동", "재건축", 1981, 1842, 2680, "사업시행인가 진행중", "상"),
    ("우성4차", "잠실동", "재건축", 1982, 555, 825, "관리처분인가 준비중", "상"),
    ("아시아선수촌", "잠실동", "재건축", 1986, 1356, 3483, "추진위 승인", "상"),
    ("올림픽선수기자촌", "방이동", "재건축", 1988, 5540, 9200, "추진위 승인", "상"),
    ("올림픽훼밀리타운", "문정동", "재건축", 1988, 4494, 6620, "추진위 승인", "상"),
    ("잠실르엘(미성크로바)", "잠실동", "재건축", 1978, 1400, 1865, "시공중(2026분양)", "완료단계"),
    ("잠실래미안아이파크(진주)", "잠실동", "재건축", 1980, 1507, 2678, "준공(2025.12)", "완료단계"),
    ("가락극동(르엘)", "가락동", "재건축", 1984, 588, 999, "시공사선정(롯데)", "중상"),
    ("가락프라자", "가락동", "재건축", 1985, 670, 1068, "관리처분완료/이주시작", "중상"),
    ("삼환가락", "가락동", "재건축", 1984, 660, 1101, "사업시행인가 완료", "중"),
    ("가락미륭", "가락동", "재건축", 1985, 450, 614, "사업시행인가 완료", "중"),
    ("가락상아1차", "가락동", "재건축", 1984, 226, 405, "이주완료/착공준비", "중"),
    ("가락쌍용1차", "가락동", "리모델링", 1987, 2064, 2348, "사전자문 통과", "중"),
    ("가락쌍용2차(래미안아펠릭스)", "가락동", "리모델링", 1988, 492, 565, "조합설립/시공사선정", "중"),
    ("한양3차+대림가락(래미안비아채)", "방이동", "재건축", 1984, 1060, 1374, "시공사선정(삼성물산)", "중상"),
    ("방이코오롱", "방이동", "재건축", 1991, 758, 1011, "안전진단통과/추진위구성", "중"),
    ("오금현대2·3·4차", "오금동", "재건축", 1984, 1316, 2436, "정비구역확정/조합설립중", "중상"),
    ("오금대림", "오금동", "재건축", 1988, 749, None, "안전진단E등급/재건축확정", "중"),
    ("송파한양1차", "송파동", "재건축", 1983, 748, 954, "신통기획 자문완료", "중"),
    ("송파한양2차", "송파동", "재건축", 1983, 930, 1346, "정비구역확정/시공사선정중", "중상"),
    ("풍납미성", "풍납동", "재건축", 1987, 311, 413, "신통기획 자문완료", "중상"),
    ("풍납극동", "풍납동", "재건축", 1986, 400, 642, "신통기획 자문완료", "중상"),
    ("마천1구역", "마천동", "재개발", None, None, 2900, "추진중", "중"),
    ("마천2구역", "마천동", "재개발", None, None, 1729, "촉진구역확정", "중"),
    ("마천4구역", "마천동", "재개발", None, None, 1254, "사업시행진행(현대건설)", "중"),
    ("거여새마을", "거여동", "재개발", None, None, 1678, "시공사선정(삼성+GS)", "중"),
]


# ========== 시트 생성 ==========
def sheet_scorecard(wb):
    ws = wb.active
    ws.title = "종합 스코어카드"
    ncol = len(COMPLEXES) + 3
    set_widths(ws, [22] + [4] + [13]*len(COMPLEXES) + [12])
    r = write_title(ws, 1, "잠실/송파 재건축 다차원 투자 스코어카드 (8개 단지)", ncol)
    r += 1

    factors = list(WEIGHTS.keys())
    headers = ["평가 팩터", "가중치"] + SHORT + ["비고"]
    r = write_headers(ws, r, headers)

    for f in factors:
        w = WEIGHTS[f]
        rd = [f, w] + [FACTORS[cpx][f] for cpx in COMPLEXES] + [""]
        fill_r = sub_fill if factors.index(f) % 2 == 0 else white_fill
        r = write_row(ws, r, rd, fill=fill_r)
        for ci, cpx in enumerate(COMPLEXES):
            cell = ws.cell(row=r-1, column=3+ci)
            cell.fill = score_fill(FACTORS[cpx][f])
            cell.font = Font(name="맑은 고딕", size=10, bold=True)

    # 가중평균 종합
    r += 1
    ws.cell(row=r, column=1, value="가중평균 종합점수").font = Font(name="맑은 고딕", size=12, bold=True)
    ws.cell(row=r, column=1).fill = gold_fill
    ws.cell(row=r, column=2, value="100%").font = score_font
    ws.cell(row=r, column=2).fill = gold_fill

    total_w = sum(WEIGHTS.values())
    scores = {}
    for ci, cpx in enumerate(COMPLEXES):
        sc = sum(FACTORS[cpx][f] * WEIGHTS[f] for f in factors) / total_w
        scores[cpx] = sc
        cell = ws.cell(row=r, column=3+ci, value=round(sc, 2))
        cell.font = Font(name="맑은 고딕", size=13, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = center_align
    ws.cell(row=r, column=ncol).fill = gold_fill
    ws.row_dimensions[r].height = 32

    # 순위
    ranked = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    r += 1
    ws.cell(row=r, column=1, value="투자 추천 순위").font = Font(name="맑은 고딕", size=11, bold=True)
    ws.cell(row=r, column=1).fill = gold_fill
    ws.cell(row=r, column=2).fill = gold_fill
    medals = {1: "FFD700", 2: "C0C0C0", 3: "CD7F32"}
    for ci, cpx in enumerate(COMPLEXES):
        rank = [i+1 for i, (c, _) in enumerate(ranked) if c == cpx][0]
        cell = ws.cell(row=r, column=3+ci, value=f"{rank}위")
        cell.font = Font(name="맑은 고딕", size=12, bold=True)
        cell.alignment = center_align
        cell.fill = PatternFill("solid", fgColor=medals.get(rank, LIGHT_GRAY))
    ws.cell(row=r, column=ncol).fill = gold_fill

    # 카테고리별 소계
    categories = {
        "입지/환경": ["입지 프리미엄", "역세권 접근성", "한강 조망", "학군", "생활 인프라"],
        "재건축 사업성": ["재건축 진행도", "시공사 신뢰도", "브랜드 가치", "단지 규모", "대형 개발호재"],
        "투자 수익성": ["가격 매력도", "전세가율(갭)", "기대 수익률", "유동성", "임대수익"],
        "안정성/리스크": ["리스크 안정성", "조합 안정성", "정책 수혜", "보유비용 효율"],
        "미래 가치": ["랜드마크 잠재력"],
    }
    r += 2
    headers2 = ["카테고리", "가중치"] + SHORT + [""]
    r = write_headers(ws, r, headers2)
    for cat, cfs in categories.items():
        cw = sum(WEIGHTS[f] for f in cfs)
        rd = [cat, cw]
        for cpx in COMPLEXES:
            cs = sum(FACTORS[cpx][f]*WEIGHTS[f] for f in cfs) / cw
            rd.append(round(cs, 1))
        rd.append("")
        fill_r = sub_fill if list(categories.keys()).index(cat) % 2 == 0 else white_fill
        r = write_row(ws, r, rd, fill=fill_r)

    # 레이더 차트 데이터
    r += 2
    chart_r = r
    ws.cell(row=r, column=1, value="카테고리").font = sub_header_font
    for ci, cpx in enumerate(COMPLEXES):
        ws.cell(row=r, column=2+ci, value=SHORT[ci]).font = sub_header_font
    r += 1
    for cat, cfs in categories.items():
        cw = sum(WEIGHTS[f] for f in cfs)
        ws.cell(row=r, column=1, value=cat)
        for ci, cpx in enumerate(COMPLEXES):
            cs = sum(FACTORS[cpx][f]*WEIGHTS[f] for f in cfs) / cw
            ws.cell(row=r, column=2+ci, value=round(cs, 1))
        r += 1

    radar = RadarChart()
    radar.type = "filled"
    radar.title = "카테고리별 투자 역량 비교 (8개 단지)"
    radar.style = 26; radar.width = 22; radar.height = 15
    labels = Reference(ws, min_col=1, min_row=chart_r+1, max_row=chart_r+len(categories))
    for ci in range(len(COMPLEXES)):
        vals = Reference(ws, min_col=2+ci, min_row=chart_r, max_row=chart_r+len(categories))
        radar.add_data(vals, titles_from_data=True)
    radar.set_categories(labels)
    ws.add_chart(radar, f"A{r+1}")
    return ws


def sheet_basic(wb):
    ws = wb.create_sheet("단지별 기본정보")
    ncol = len(COMPLEXES) + 1
    set_widths(ws, [20] + [14]*len(COMPLEXES))
    r = write_title(ws, 1, "8개 재건축 단지 기본정보 비교", ncol)
    r += 1

    fields = [
        ("준공년도", lambda d: d["준공"]), ("연식(년)", lambda d: d["연식"]),
        ("현재 세대수", lambda d: d["세대수"]), ("동 수", lambda d: d["동수"]),
        ("최고층", lambda d: d["층"]), ("전용면적(m²)", lambda d: d["면적"]),
        ("기존 용적률(%)", lambda d: d["용적률기존"]),
        ("── 재건축 후 ──", lambda d: "─────"),
        ("재건축 세대수", lambda d: d["재건축세대"]), ("재건축 최고층", lambda d: d["재건축층"]),
        ("계획 용적률(%)", lambda d: d["용적률계획"]),
        ("용적률 증가(%p)", lambda d: d["용적률계획"]-d["용적률기존"]),
        ("세대수 증가율", lambda d: f"{(d['재건축세대']/d['세대수']-1)*100:.0f}%"),
        ("시공사", lambda d: d["시공사"]), ("브랜드", lambda d: d["브랜드"]),
    ]
    r = write_headers(ws, r, ["항목"] + SHORT)
    for fname, fn in fields:
        rd = [fname] + [fn(BASIC[cpx]) for cpx in COMPLEXES]
        fill_r = gold_fill if fname.startswith("──") else (sub_fill if fields.index((fname,fn))%2==0 else white_fill)
        r = write_row(ws, r, rd, fill=fill_r)
    return ws


def sheet_progress(wb):
    ws = wb.create_sheet("재건축 진행현황")
    ncol = len(COMPLEXES) + 1
    set_widths(ws, [22] + [14]*len(COMPLEXES))
    r = write_title(ws, 1, "재건축 단계별 진행현황", ncol)
    r += 1

    pf = ["단계","진행률","안전진단","정비구역","조합설립","사업시행인가","관리처분","착공","준공","이주"]
    r = write_headers(ws, r, ["항목"] + SHORT)
    for field in pf:
        rd = [field] + [str(PROGRESS[cpx].get(field,"")) if field != "진행률" else f"{PROGRESS[cpx]['진행률']}%" for cpx in COMPLEXES]
        fill_r = sub_fill if pf.index(field) % 2 == 0 else white_fill
        r = write_row(ws, r, rd, fill=fill_r)
        if field == "진행률":
            for ci, cpx in enumerate(COMPLEXES):
                pv = PROGRESS[cpx]["진행률"]
                cell = ws.cell(row=r-1, column=2+ci)
                if pv >= 70: cell.fill = green_fill
                elif pv >= 50: cell.fill = PatternFill("solid", fgColor="EAFAF1")
                elif pv >= 30: cell.fill = orange_fill
                else: cell.fill = red_fill
                cell.font = Font(name="맑은 고딕", size=10, bold=True)

    # 진행률 차트
    r += 2
    cr = r
    ws.cell(row=r, column=1, value="단지명").font = sub_header_font
    ws.cell(row=r, column=2, value="진행률(%)").font = sub_header_font
    r += 1
    for cpx in COMPLEXES:
        ws.cell(row=r, column=1, value=SHORT[COMPLEXES.index(cpx)])
        ws.cell(row=r, column=2, value=PROGRESS[cpx]["진행률"])
        r += 1
    chart = BarChart()
    chart.type = "col"; chart.title = "단지별 재건축 진행률"; chart.y_axis.title = "%"
    chart.style = 10; chart.width = 20; chart.height = 12
    cats = Reference(ws, min_col=1, min_row=cr+1, max_row=cr+len(COMPLEXES))
    vals = Reference(ws, min_col=2, min_row=cr, max_row=cr+len(COMPLEXES))
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"D{cr}")
    return ws


def sheet_price(wb):
    ws = wb.create_sheet("시세 및 가격분석")
    ncol = len(COMPLEXES) + 1
    set_widths(ws, [24] + [13]*len(COMPLEXES))
    r = write_title(ws, 1, "시세 및 가격 다차원 분석 (전용 84m² 기준, 억원)", ncol)
    r += 1

    fields = [
        ("현재 매매가", lambda d: d["매매가"]),
        ("현재 전세가", lambda d: d["전세가"]),
        ("전세가율(%)", lambda d: d["전세가율"]),
        ("호가", lambda d: d["호가"]),
        ("신고가", lambda d: d["신고가"]),
        ("최저가(1년)", lambda d: d["최저1년"]),
        ("1년전", lambda d: d["1년전"]),
        ("3년전", lambda d: d["3년전"]),
        ("5년전", lambda d: d["5년전"]),
        ("── 분석지표 ──", lambda d: "─────"),
        ("1년 상승률(%)", lambda d: round((d["매매가"]/d["1년전"]-1)*100, 1)),
        ("3년 상승률(%)", lambda d: round((d["매매가"]/d["3년전"]-1)*100, 1)),
        ("5년 상승률(%)", lambda d: round((d["매매가"]/d["5년전"]-1)*100, 1)),
        ("연평균 상승률(5Y)", lambda d: round(((d["매매가"]/d["5년전"])**(1/5)-1)*100, 1)),
        ("재건축 프리미엄(억)", lambda d: d["프리미엄"]),
        ("갭(매매-전세)(억)", lambda d: round(d["매매가"]-d["전세가"], 1)),
        ("호가-실거래 괴리(%)", lambda d: round((d["호가"]/d["매매가"]-1)*100, 1)),
    ]
    r = write_headers(ws, r, ["항목"] + SHORT)
    for fname, fn in fields:
        rd = [fname] + [fn(PRICE[cpx]) if not fname.startswith("──") else "─────" for cpx in COMPLEXES]
        fill_r = gold_fill if fname.startswith("──") else (sub_fill if fields.index((fname,fn))%2==0 else white_fill)
        r = write_row(ws, r, rd, fill=fill_r)

    # 가격추이 차트
    r += 2
    cr = r
    years = ["5년전","3년전","1년전","현재"]
    ws.cell(row=r, column=1, value="시점").font = sub_header_font
    for ci, cpx in enumerate(COMPLEXES):
        ws.cell(row=r, column=2+ci, value=SHORT[ci]).font = sub_header_font
    r += 1
    for yr in years:
        ws.cell(row=r, column=1, value=yr)
        for ci, cpx in enumerate(COMPLEXES):
            val = {"5년전": PRICE[cpx]["5년전"], "3년전": PRICE[cpx]["3년전"],
                   "1년전": PRICE[cpx]["1년전"], "현재": PRICE[cpx]["매매가"]}[yr]
            ws.cell(row=r, column=2+ci, value=val)
        r += 1
    line = LineChart()
    line.title = "단지별 매매가 추이 (전용 84m², 억원)"
    line.style = 10; line.width = 22; line.height = 14; line.y_axis.title = "억원"
    cats = Reference(ws, min_col=1, min_row=cr+1, max_row=cr+len(years))
    for ci in range(len(COMPLEXES)):
        vals = Reference(ws, min_col=2+ci, min_row=cr, max_row=cr+len(years))
        line.add_data(vals, titles_from_data=True)
    line.set_categories(cats)
    ws.add_chart(line, f"A{r+1}")
    return ws


def sheet_roi(wb):
    ws = wb.create_sheet("수익성 시뮬레이션")
    ncol = len(COMPLEXES) + 1
    set_widths(ws, [26] + [13]*len(COMPLEXES))
    r = write_title(ws, 1, "단지별 투자 수익성 시뮬레이션 (전용 84m², 억원)", ncol)
    r += 1

    fields = [
        ("── 매입비용 ──", None),
        ("매입가", lambda d: d["매입가"]),
        ("취득세(3%)", lambda d: d["취득세"]),
        ("중개수수료(0.5%)", lambda d: d["중개"]),
        ("매입 총비용", lambda d: round(d["매입가"]+d["취득세"]+d["중개"], 2)),
        ("── 보유비용 ──", None),
        ("추정 분담금", lambda d: d["분담금"]),
        ("이주기간(년)", lambda d: d["이주기간"]),
        ("전세기회비용/년", lambda d: d["전세비용연"]),
        ("보유세/년", lambda d: d["보유세연"]),
        ("대출이자/년(20억)", lambda d: d["이자연"]),
        ("보유기간 총비용", lambda d: round(d["분담금"]+(d["전세비용연"]+d["보유세연"]+d["이자연"])*d["이주기간"], 1)),
        ("── 예상수익 ──", None),
        ("예상 신축가(비관)", lambda d: d["신축가_하"]),
        ("예상 신축가(중립)", lambda d: d["신축가_중"]),
        ("예상 신축가(낙관)", lambda d: d["신축가_상"]),
        ("준공 예정연도", lambda d: d["준공"]),
        ("── 수익률 ──", None),
        ("총투자비(중립)", lambda d: round(calc_total_cost(d, d["신축가_중"]), 1)),
        ("예상차익(비관)", lambda d: round(d["신축가_하"]-calc_total_cost(d, d["신축가_하"]), 1)),
        ("예상차익(중립)", lambda d: round(d["신축가_중"]-calc_total_cost(d, d["신축가_중"]), 1)),
        ("예상차익(낙관)", lambda d: round(d["신축가_상"]-calc_total_cost(d, d["신축가_상"]), 1)),
        ("ROI(중립)(%)", lambda d: calc_roi(d, d["신축가_중"])),
        ("연환산ROI(%)", lambda d: round(((d["신축가_중"]/calc_total_cost(d, d["신축가_중"]))**(1/max(d["이주기간"]+1,1))-1)*100, 1)),
    ]
    r = write_headers(ws, r, ["항목"] + SHORT)
    for fname, fn in fields:
        rd = [fname] + (["─────"]*len(COMPLEXES) if fn is None else [fn(ROI[cpx]) for cpx in COMPLEXES])
        fill_r = gold_fill if fname.startswith("──") else (sub_fill if fields.index((fname,fn))%2==0 else white_fill)
        r = write_row(ws, r, rd, fill=fill_r)

    # ROI 차트
    r += 2; cr = r
    ws.cell(row=r,column=1,value="단지명").font = sub_header_font
    for si, s in enumerate(["비관ROI","중립ROI","낙관ROI"]):
        ws.cell(row=r, column=2+si, value=s).font = sub_header_font
    r += 1
    for ci, cpx in enumerate(COMPLEXES):
        ws.cell(row=r, column=1, value=SHORT[ci])
        for si, key in enumerate(["신축가_하","신축가_중","신축가_상"]):
            ws.cell(row=r, column=2+si, value=calc_roi(ROI[cpx], ROI[cpx][key]))
        r += 1
    chart = BarChart()
    chart.type = "col"; chart.title = "시나리오별 ROI 비교(%)"; chart.y_axis.title = "ROI(%)"
    chart.style = 10; chart.width = 22; chart.height = 14
    cats = Reference(ws, min_col=1, min_row=cr+1, max_row=cr+len(COMPLEXES))
    for si in range(3):
        vals = Reference(ws, min_col=2+si, min_row=cr, max_row=cr+len(COMPLEXES))
        chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{r+1}")
    return ws


def sheet_cost(wb):
    ws = wb.create_sheet("비용구조 분석")
    set_widths(ws, [28, 16, 16, 16])
    r = write_title(ws, 1, "재건축 투자 비용 구조 상세", 4)
    r += 1

    sections = [
        ("[1] 취득세 세율", ["구분","세율","조건","비고"], [
            ("1주택(6억이하)","1.0%","비규제","지방교육세 별도"),
            ("1주택(6~9억)","1~3%","비규제","구간별 차등"),
            ("1주택(9억초과)","3.0%","비규제",""),
            ("2주택(조정지역)","8.0%","규제중과","+지방교육세+농특세"),
            ("3주택이상","12.0%","규제중과",""),
            ("멸실후 토지매수","4.6%","토지취득세",""),
            ("원시취득(85m²이하)","2.8%","신축취득","+0.16%=2.96%"),
            ("원시취득(85m²초과)","2.8%","신축취득","+0.2%=3.16%"),
        ]),
        ("[2] 중개수수료", ["거래금액","상한요율","예시(40억)","실제협의선"], [
            ("5억미만","0.6%","-","-"),("5~9억","0.5%","-","-"),
            ("9~12억","0.5%","-","-"),("12~15억","0.6%","-","-"),
            ("15억이상","0.7%","최대2800만","0.4~0.5%"),
        ]),
        ("[3] 보유세(1주택,공시가30억)", ["항목","금액(만/년)","비고","변동요인"], [
            ("재산세","800~1,200","공시가기준","공시가변동"),
            ("종부세","1,500~2,500","공제12억","가액비율"),
            ("합계","2,300~3,700","","2026증가"),
            ("다주택종부세","3,000~6,000+","중과세율","상한폐지"),
        ]),
        ("[4] 대출금리(2026.02)", ["구분","금리","비고","월이자(20억)"], [
            ("기준금리","2.50%","2026.01인하","-"),
            ("변동금리","3.8~4.3%","주담대","633~717만"),
            ("고정금리","4.0~4.5%","주담대","667~750만"),
            ("보금자리론","3.9~4.2%","일반","650~700만"),
            ("보금자리론(우대)","2.9~3.2%","우대","483~533만"),
        ]),
    ]
    for title, hdrs, rows in sections:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.cell(row=r, column=1, value=title).font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
        r += 1
        r = write_headers(ws, r, hdrs)
        for rd in rows:
            r = write_row(ws, r, rd, fill=sub_fill if rows.index(rd)%2==0 else white_fill)
        r += 1
    return ws


def sheet_risk(wb):
    ws = wb.create_sheet("리스크 매트릭스")
    ncol = 1 + len(COMPLEXES)*2 + 1
    set_widths(ws, [20] + [7,7]*len(COMPLEXES) + [10])
    r = write_title(ws, 1, "단지별 리스크 매트릭스 (영향도×확률)", ncol)
    r += 1

    hdrs = ["리스크 요인"]
    for s in SHORT:
        hdrs += [f"{s[:3]}영향", f"{s[:3]}확률"]
    hdrs.append("등급")
    r = write_headers(ws, r, hdrs)

    for rname, rdata in RISKS:
        rd = [rname]
        max_s = 0
        for ci in range(len(COMPLEXES)):
            imp, prob = rdata[ci]
            rd += [imp, prob]
            max_s = max(max_s, imp*prob)
        grade = "매우높음" if max_s >= 48 else ("높음" if max_s >= 30 else ("중간" if max_s >= 16 else "낮음"))
        rd.append(grade)
        r = write_row(ws, r, rd)
        cell = ws.cell(row=r-1, column=ncol)
        if grade == "매우높음": cell.fill = red_fill
        elif grade == "높음": cell.fill = orange_fill
        elif grade == "중간": cell.fill = PatternFill("solid", fgColor="FEF9E7")
        else: cell.fill = green_fill

    # 종합 리스크 점수
    r += 1
    ws.cell(row=r, column=1, value="종합 리스크 점수").font = Font(name="맑은 고딕", size=11, bold=True)
    ws.cell(row=r, column=1).fill = gold_fill
    for ci in range(len(COMPLEXES)):
        total = sum(RISKS[ri][1][ci][0]*RISKS[ri][1][ci][1] for ri in range(len(RISKS)))
        avg = round(total/len(RISKS), 1)
        col = 2+ci*2
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)
        cell = ws.cell(row=r, column=col, value=avg)
        cell.font = Font(name="맑은 고딕", size=12, bold=True)
        cell.alignment = center_align
        if avg >= 40: cell.fill = red_fill
        elif avg >= 30: cell.fill = orange_fill
        else: cell.fill = green_fill
    return ws


def sheet_benchmark(wb):
    ws = wb.create_sheet("벤치마크(과거사례)")
    set_widths(ws, [28, 10, 10, 10, 10, 12, 12, 10, 14, 14])
    r = write_title(ws, 1, "서울 주요 재건축 완료 사례 벤치마크", 10)
    r += 1
    hdrs = ["사업명","위치","기존세대","신축세대","소요기간","분양가(84m²)","현시세(84m²)","차익","조합원수익률","추가공사비"]
    r = write_headers(ws, r, hdrs)
    for bm in BENCHMARKS:
        r = write_row(ws, r, list(bm), fill=sub_fill if BENCHMARKS.index(bm)%2==0 else white_fill)

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.cell(row=r, column=1, value="[벤치마크 시사점]").font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
    r += 1
    for ins in [
        "1. 서울 재건축 평균 소요기간: 조합설립~입주 약 10~16년",
        "2. 분양가 대비 현시세 평균 프리미엄: 100~200% 상승",
        "3. 추가공사비 이슈는 5천세대+ 대단지에서 빈번 → 주공5/선수촌 주의",
        "4. 초기 매입가 낮을수록 ROI 극대화 → 우성4차/훼밀리 유리",
        "5. 한강변+강남권 입지 = 완공후 시세 프리미엄 극대화",
        "6. 시공사 브랜드가 완공후 5~10% 추가 프리미엄",
    ]:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        ws.cell(row=r, column=1, value=ins).font = data_font
        r += 1
    return ws


def sheet_scenario(wb):
    ws = wb.create_sheet("시나리오 분석")
    ncol = len(COMPLEXES) + 1
    set_widths(ws, [24] + [13]*len(COMPLEXES))
    r = write_title(ws, 1, "3개 시나리오별 투자 수익 시뮬레이션", ncol)
    r += 1

    scenarios = [
        ("낙관 (Bull)", GREEN, "금리인하, 규제완화, 부동산 상승기",
         {"잠실주공5단지": (58,2.0,4.0), "장미1·2차": (48,1.5,3.5), "장미3차": (46,1.5,3.5),
          "우성1·2·3차": (45,1.5,3.5), "우성4차": (40,1.0,2.5), "아시아선수촌": (52,2.0,5.0),
          "올림픽선수촌": (50,1.0,5.5), "올림픽훼밀리": (42,2.5,5.5)}),
        ("중립 (Base)", BLUE, "현 추세 유지, 완만한 상승",
         {"잠실주공5단지": (52,3.5,5.0), "장미1·2차": (42,2.5,4.5), "장미3차": (40,2.5,4.5),
          "우성1·2·3차": (40,2.5,4.5), "우성4차": (35,2.0,3.5), "아시아선수촌": (45,3.0,6.0),
          "올림픽선수촌": (43,1.5,6.5), "올림픽훼밀리": (36,3.6,6.5)}),
        ("비관 (Bear)", RED, "금리상승, 규제강화, 경기침체",
         {"잠실주공5단지": (48,5.0,6.5), "장미1·2차": (38,4.0,6.0), "장미3차": (36,4.0,6.0),
          "우성1·2·3차": (35,4.0,6.0), "우성4차": (30,3.0,4.5), "아시아선수촌": (38,4.0,7.5),
          "올림픽선수촌": (38,2.0,8.0), "올림픽훼밀리": (30,5.0,8.0)}),
    ]

    for sname, color, premise, sdata in scenarios:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        cell = ws.cell(row=r, column=1, value=f"시나리오: {sname}")
        cell.font = Font(name="맑은 고딕", size=13, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=color)
        ws.row_dimensions[r].height = 30
        r += 1
        ws.cell(row=r, column=1, value=f"전제: {premise}").font = small_font
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        r += 1
        r = write_headers(ws, r, ["항목"] + SHORT)

        items = [
            ("예상신축가(억)", lambda cpx: sdata[cpx][0]),
            ("분담금(억)", lambda cpx: sdata[cpx][1]),
            ("이주기간(년)", lambda cpx: sdata[cpx][2]),
            ("총투자비(억)", lambda cpx: round(
                PRICE[cpx]["매매가"]*1.035 + sdata[cpx][1] + 1.25*sdata[cpx][2] + sdata[cpx][0]*0.0296, 1)),
            ("예상차익(억)", lambda cpx: round(
                sdata[cpx][0] - (PRICE[cpx]["매매가"]*1.035 + sdata[cpx][1] + 1.25*sdata[cpx][2] + sdata[cpx][0]*0.0296), 1)),
            ("ROI(%)", lambda cpx: round(
                (sdata[cpx][0] - (PRICE[cpx]["매매가"]*1.035 + sdata[cpx][1] + 1.25*sdata[cpx][2] + sdata[cpx][0]*0.0296))
                / (PRICE[cpx]["매매가"]*1.035) * 100, 1)),
        ]
        for iname, fn in items:
            rd = [iname] + [fn(cpx) for cpx in COMPLEXES]
            r = write_row(ws, r, rd, fill=sub_fill if items.index((iname,fn))%2==0 else white_fill)
            if "ROI" in iname or "차익" in iname:
                for ci in range(len(COMPLEXES)):
                    cell = ws.cell(row=r-1, column=2+ci)
                    try:
                        v = float(cell.value)
                        if v > 0: cell.fill = green_fill; cell.font = Font(name="맑은 고딕", size=10, bold=True, color=GREEN)
                        elif v < 0: cell.fill = red_fill; cell.font = Font(name="맑은 고딕", size=10, bold=True, color=RED)
                    except: pass
        r += 1
    return ws


def sheet_cashflow(wb):
    ws = wb.create_sheet("현금흐름 분석")
    set_widths(ws, [16, 14, 14, 14, 14, 14, 14])
    r = write_title(ws, 1, "연도별 현금흐름 비교 (주공5 vs 우성4 vs 올림픽선수촌)", 7)
    r += 1

    cf_hdrs = ["연도", "구분", "지출(억)", "수입(억)", "누적투자(억)", "자산가치(억)", "순자산(억)"]

    cases = [
        ("[A] 잠실주공5단지 (매입가 45억)", [
            (2026,"매입+취득세+중개",-46.6,0,-46.6,45.0,-1.6),
            (2027,"이주+보유비용",-1.5,0,-48.1,46.0,-2.1),
            (2028,"분담금1차+보유비용",-3.0,0,-51.1,47.0,-4.1),
            (2029,"보유비용+분담금잔금",-1.9,0,-53.0,48.0,-5.0),
            (2030,"보유비용",-1.4,0,-54.4,50.0,-4.4),
            (2031,"준공+신축취득세",-1.5,0,-55.9,52.0,-3.9),
        ]),
        ("[B] 우성4차 (매입가 22억)", [
            (2026,"매입+취득세+중개+이주",-23.7,0,-23.7,22.0,-1.7),
            (2027,"분담금1차+보유비용",-2.4,0,-26.1,24.0,-2.1),
            (2028,"보유비용+분담금잔금",-1.2,0,-27.3,28.0,0.7),
            (2029,"준공+신축취득세",-1.0,0,-28.3,35.0,6.7),
        ]),
        ("[C] 올림픽선수촌 (매입가 29억)", [
            (2026,"매입+취득세+중개",-30.0,0,-30.0,29.0,-1.0),
            (2027,"보유비용",-1.2,0,-31.2,30.0,-1.2),
            (2028,"보유비용",-1.2,0,-32.4,31.0,-1.4),
            (2029,"이주+보유비용",-1.4,0,-33.8,32.0,-1.8),
            (2030,"분담금+보유비용",-2.7,0,-36.5,35.0,-1.5),
            (2031,"보유비용",-1.2,0,-37.7,37.0,-0.7),
            (2032,"보유비용",-1.2,0,-38.9,39.0,0.1),
            (2033,"보유비용",-1.2,0,-40.1,41.0,0.9),
            (2034,"준공+신축취득세",-1.3,0,-41.4,43.0,1.6),
        ]),
    ]
    for title, rows in cases:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.cell(row=r, column=1, value=title).font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
        r += 1
        r = write_headers(ws, r, cf_hdrs)
        for rd in rows:
            r = write_row(ws, r, list(rd))
            cell = ws.cell(row=r-1, column=7)
            if rd[6] < 0: cell.fill = red_fill; cell.font = Font(name="맑은 고딕", size=10, bold=True, color=RED)
            else: cell.fill = green_fill; cell.font = Font(name="맑은 고딕", size=10, bold=True, color=GREEN)
        r += 2
    return ws


def sheet_policy(wb):
    ws = wb.create_sheet("정책규제 영향")
    set_widths(ws, [22, 26, 26, 12])
    r = write_title(ws, 1, "정책/규제 영향 분석", 4)
    r += 1
    r = write_headers(ws, r, ["정책/규제", "현황", "투자 영향", "영향도"])
    for pol in POLICIES:
        r = write_row(ws, r, pol, fill=sub_fill if POLICIES.index(pol)%2==0 else white_fill)
        cell = ws.cell(row=r-1, column=4)
        if "매우" in str(pol[3]): cell.fill = red_fill
        elif "높음" in str(pol[3]): cell.fill = orange_fill
        elif "긍정" in str(pol[3]): cell.fill = green_fill
    return ws


def sheet_all_projects(wb):
    """송파구 전체 정비사업 현황"""
    ws = wb.create_sheet("송파구 전체 정비사업")
    set_widths(ws, [28, 10, 10, 10, 10, 10, 22, 8])
    r = write_title(ws, 1, "송파구 전체 재건축/재개발/리모델링 현황 (28개 사업)", 8)
    r += 1

    hdrs = ["단지명", "위치", "유형", "준공", "기존세대", "재건축세대", "진행단계", "유망도"]
    r = write_headers(ws, r, hdrs)
    for proj in ALL_PROJECTS:
        r = write_row(ws, r, list(proj), fill=sub_fill if ALL_PROJECTS.index(proj)%2==0 else white_fill)
        cell = ws.cell(row=r-1, column=8)
        if proj[7] == "상": cell.fill = green_fill
        elif proj[7] == "중상": cell.fill = PatternFill("solid", fgColor="EAFAF1")
        elif proj[7] == "완료단계": cell.fill = PatternFill("solid", fgColor="D6EAF8")

    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.cell(row=r, column=1, value="※ 총 41개 정비사업 중 주요 28개 사업 표시. 소규모 정비/가로주택 등 제외.").font = small_font
    return ws


def sheet_final(wb):
    ws = wb.create_sheet("최종 투자 판단")
    ncol = len(COMPLEXES) + 1
    set_widths(ws, [24] + [13]*len(COMPLEXES))
    r = write_title(ws, 1, "최종 투자 판단 종합 매트릭스 (8개 단지)", ncol)
    r += 1

    # 투자자 유형별 추천
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value="[투자자 유형별 추천]").font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
    r += 1
    r = write_headers(ws, r, ["투자자 유형", "1순위", "2순위", "이유", "주의사항", ""])
    types = [
        ("공격적(고수익)", "우성4차", "올림픽선수촌", "빠른사업/최대사업성", "소규모/초기단계"),
        ("안정적(자산보전)", "잠실주공5단지", "장미1·2차", "최고입지/브랜드", "높은매입가"),
        ("중장기(5년+)", "올림픽선수촌", "아시아선수촌", "용적률137%/잠재력", "8~10년소요"),
        ("가성비(진입가대비)", "올림픽훼밀리", "우성4차", "낮은매입가", "분담금/고도제한"),
        ("실거주+투자", "잠실주공5단지", "우성1·2·3차", "잠실핵심입지", "이주기간비용"),
    ]
    for t in types:
        r = write_row(ws, r, list(t) + [""])

    # 종합평가
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    ws.cell(row=r, column=1, value="[종합 평가 매트릭스]").font = Font(name="맑은 고딕", size=12, bold=True, color=DARK_BLUE)
    r += 1
    r = write_headers(ws, r, ["평가 항목"] + SHORT)

    eval_items = [
        ("입지/환경(30%)", [9.4, 8.6, 8.0, 7.4, 7.0, 7.6, 6.8, 6.4]),
        ("재건축사업성(25%)", [8.8, 6.2, 5.8, 6.8, 6.6, 5.6, 5.2, 4.6]),
        ("투자수익성(25%)", [3.6, 5.8, 5.8, 5.6, 7.0, 4.2, 4.2, 5.0]),
        ("안정성(10%)", [6.5, 6.0, 5.5, 6.3, 7.3, 4.8, 4.2, 4.5]),
        ("미래가치(10%)", [10.0, 8.0, 7.0, 6.0, 5.0, 9.0, 8.0, 5.0]),
    ]
    for iname, scores in eval_items:
        r = write_row(ws, r, [iname]+scores, fill=sub_fill if eval_items.index((iname,scores))%2==0 else white_fill)
        for ci in range(len(scores)):
            cell = ws.cell(row=r-1, column=2+ci)
            cell.fill = score_fill(scores[ci])
            cell.font = Font(name="맑은 고딕", size=10, bold=True)

    # 가중종합
    weights = [0.30, 0.25, 0.25, 0.10, 0.10]
    ws.cell(row=r, column=1, value="가중종합점수(100점)").font = Font(name="맑은 고딕", size=12, bold=True)
    ws.cell(row=r, column=1).fill = gold_fill
    final_scores = []
    for ci in range(len(COMPLEXES)):
        ws_val = sum(eval_items[ei][1][ci]*weights[ei] for ei in range(len(eval_items)))
        final_scores.append((COMPLEXES[ci], ws_val))
        cell = ws.cell(row=r, column=2+ci, value=round(ws_val*10, 1))
        cell.font = Font(name="맑은 고딕", size=13, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = center_align
    r += 1

    ranked = sorted(final_scores, key=lambda x: x[1], reverse=True)
    ws.cell(row=r, column=1, value="최종 순위").font = Font(name="맑은 고딕", size=12, bold=True)
    ws.cell(row=r, column=1).fill = gold_fill
    medals = {1: "FFD700", 2: "C0C0C0", 3: "CD7F32"}
    for ci, cpx in enumerate(COMPLEXES):
        rank = [i+1 for i, (c,_) in enumerate(ranked) if c == cpx][0]
        cell = ws.cell(row=r, column=2+ci, value=f"{rank}위")
        cell.font = Font(name="맑은 고딕", size=13, bold=True)
        cell.alignment = center_align
        cell.fill = PatternFill("solid", fgColor=medals.get(rank, LIGHT_GRAY))
    r += 2

    # 결론
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    ws.cell(row=r, column=1, value="[핵심 결론 및 전략적 제언]").font = Font(name="맑은 고딕", size=13, bold=True, color=WHITE)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=DARK_BLUE)
    r += 1
    conclusions = [
        "1. [단기 최고ROI] 우성4차: 2029 준공, 22억→35억 (연환산 6.8%)",
        "   → 사업진행 가장 빠름, 49층 초고층 확정, 착공 임박",
        "",
        "2. [초대형 잠재력] 올림픽선수촌: 용적률137%, 분담금0~2억, 9200세대",
        "   → 서울 최대급 재건축, 장기투자 시 최대 수혜 가능성",
        "   → 단, 사업 초기(추진위 단계), 준공까지 8~10년 소요",
        "",
        "3. [밸런스] 장미1·2차: 한강조망+적정가+5165세대 대단지",
        "   → 26.5억 매입, 42억+ 기대, 리스크 대비 수익 균형",
        "",
        "4. [확정시공사] 우성1·2·3차: GS건설 확정, 자이 브랜드",
        "   → 24억 매입, 시공사 리스크 해소, 안정적 중기 투자",
        "",
        "5. [트로피자산] 잠실주공5단지: 최고입지이나 프리미엄 과반영",
        "   → 45억 매입, 순수익 제한적, 자산가치 보전 목적에 적합",
        "",
        "6. [가성비] 올림픽훼밀리: 21.5억 진입, 가락시장역 초역세권",
        "   → 분담금3.6억+고도제한26층이 약점, 장기 보유 전략 필요",
        "",
        "7. [사업초기] 아시아선수촌: 잠실동 핵심+3483세대, 추진위 단계",
        "   → 초기진입 기회있으나 준공까지 8~10년, 자금여력 필수",
        "",
        "8. [분산전략] 우성4차(22억)+장미1·2차(26.5억) 동시매입 고려",
        "   → 총48.5억 투자, 리스크 분산+수익 극대화 포트폴리오",
        "",
        "※ 토지거래허가제(2026.06)+재초환+보유세강화 = 공통 리스크",
        "※ 매입 전 실거주 계획 및 세금 시뮬레이션 필수",
    ]
    for conc in conclusions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        ws.cell(row=r, column=1, value=conc).font = Font(name="맑은 고딕", size=10)
        ws.cell(row=r, column=1).alignment = left_align
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    ws.cell(row=r, column=1, value="※ 2026.02 기준 공개데이터 및 시장정보 기반. 투자판단 참고자료이며, 전문가 상담 권장.").font = small_font
    return ws


# ========== 메인 ==========
def main():
    wb = openpyxl.Workbook()
    sheets = [
        ("1/14 종합 스코어카드", sheet_scorecard),
        ("2/14 단지별 기본정보", sheet_basic),
        ("3/14 재건축 진행현황", sheet_progress),
        ("4/14 시세/가격분석", sheet_price),
        ("5/14 수익성 시뮬레이션", sheet_roi),
        ("6/14 비용구조 분석", sheet_cost),
        ("7/14 리스크 매트릭스", sheet_risk),
        ("8/14 벤치마크(과거사례)", sheet_benchmark),
        ("9/14 시나리오 분석", sheet_scenario),
        ("10/14 현금흐름 분석", sheet_cashflow),
        ("11/14 정책규제 영향", sheet_policy),
        ("12/14 송파구 전체 정비사업", sheet_all_projects),
        ("13/14 최종 투자 판단", sheet_final),
    ]
    for name, fn in sheets:
        print(f"{name} 생성...")
        fn(wb)

    output = "C:/Devs/KRE/잠실_송파_재건축_투자분석_v2.xlsx"
    wb.save(output)
    print(f"\n완료! {output}")
    print(f"총 {len(wb.sheetnames)}개 시트: {', '.join(wb.sheetnames)}")

if __name__ == "__main__":
    main()
